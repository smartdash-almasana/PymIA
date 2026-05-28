from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Literal

from openpyxl import load_workbook

from pymia.telegram_runtime import SENTINEL

DEFAULT_DOCUMENTS_DIR = Path(".runtime") / "telegram_documents"
MAX_PREVIEW_ROWS = 3
MAX_TEXT_LENGTH = 4000


@dataclass(frozen=True)
class ColumnSummary:
    index: int
    name: str


@dataclass(frozen=True)
class SheetSummary:
    name: str
    rows: int
    cols: int
    columns: tuple[ColumnSummary, ...]
    empty: bool


@dataclass(frozen=True)
class ExcelSummary:
    file_path: str
    sheets: tuple[SheetSummary, ...]


@dataclass(frozen=True)
class ExcelSummaryResult:
    text: str
    source: Literal["pymia"] = "pymia"
    mode: Literal["summary", "blocked", "not_found", "error"] = "summary"
    file_path: str | None = None


def resolve_latest_excel(documents_dir: str | Path = DEFAULT_DOCUMENTS_DIR) -> Path | None:
    base = Path(documents_dir)
    if not base.exists() or not base.is_dir():
        return None
    candidates = [p for p in base.iterdir() if p.is_file() and p.suffix.lower() == ".xlsx"]
    if not candidates:
        return None
    return max(candidates, key=lambda p: p.stat().st_mtime)


def summarize_excel(path: str | Path) -> ExcelSummary:
    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        sheets: list[SheetSummary] = []
        for ws in workbook.worksheets:
            rows = ws.max_row or 0
            cols = ws.max_column or 0
            first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
            columns: list[ColumnSummary] = []
            for i in range(1, cols + 1):
                raw_name = first_row[i - 1] if i - 1 < len(first_row) else None
                name = str(raw_name).strip() if raw_name is not None and str(raw_name).strip() else f"column_{i}"
                columns.append(ColumnSummary(index=i, name=name))
            sheets.append(
                SheetSummary(
                    name=ws.title,
                    rows=rows,
                    cols=cols,
                    columns=tuple(columns),
                    empty=(rows == 0 or cols == 0),
                )
            )
        return ExcelSummary(file_path=str(Path(path).resolve()), sheets=tuple(sheets))
    finally:
        workbook.close()


def render_summary_text(summary: ExcelSummary) -> str:
    lines = [
        f"{SENTINEL} Resumen estructural del Excel:",
        f"Archivo: {Path(summary.file_path).name}",
        f"Hojas: {len(summary.sheets)}",
    ]
    for sheet in summary.sheets:
        lines.append(
            f"- Hoja '{sheet.name}': rows={sheet.rows}, cols={sheet.cols}, empty={'yes' if sheet.empty else 'no'}"
        )
        preview = ", ".join(col.name for col in sheet.columns[:MAX_PREVIEW_ROWS]) if sheet.columns else "(sin columnas)"
        lines.append(f"  Columns preview: {preview}")
    text = "\n".join(lines)
    return text[:MAX_TEXT_LENGTH]


def analyze_latest_excel(documents_dir: str | Path = DEFAULT_DOCUMENTS_DIR) -> ExcelSummaryResult:
    latest = resolve_latest_excel(documents_dir)
    if latest is None:
        return ExcelSummaryResult(
            text=f"{SENTINEL} No encontre archivos .xlsx en .runtime/telegram_documents.",
            mode="not_found",
            file_path=None,
        )

    try:
        summary = summarize_excel(latest)
        text = render_summary_text(summary)
        if SENTINEL not in text:
            text = f"{SENTINEL} {text}"
        return ExcelSummaryResult(text=text, mode="summary", file_path=summary.file_path)
    except Exception:
        return ExcelSummaryResult(
            text=f"{SENTINEL} No pude leer el Excel en este momento. Reintenta con otro archivo .xlsx.",
            mode="error",
            file_path=str(latest.resolve()),
        )

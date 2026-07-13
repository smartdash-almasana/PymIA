"""
Service 1 XLSX Structure Reader V1

Reads structural metadata from XLSX files without performing:
- Business metric calculations
- Accounting semantic interpretation
- Diagnostics
- Runtime authorization

This is a pure structural introspection module.
"""

from __future__ import annotations

import os
from typing import Any


def read_service_1_xlsx_structure_v1(path: str) -> dict[str, Any]:
    """
    Read structural metadata from an XLSX file.

    Args:
        path: Filesystem path to the XLSX file.

    Returns:
        Dictionary containing structural metadata with schema_version "1.0".

    Raises:
        FileNotFoundError: If the file does not exist.
        ValueError: If the file is not an XLSX file.
        OSError: If the file cannot be opened or read.
    """
    if not os.path.exists(path):
        raise FileNotFoundError(f"XLSX file not found: {path}")

    if not path.lower().endswith(".xlsx"):
        raise ValueError(f"File is not an XLSX file: {path}")

    try:
        import openpyxl
    except ImportError as e:
        raise ImportError("openpyxl is required to read XLSX files") from e

    try:
        wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    except Exception as e:
        raise OSError(f"Failed to open XLSX file: {e}") from e

    warnings = []
    sheets = []

    try:
        sheet_names = wb.sheetnames
        sheet_count = len(sheet_names)

        if sheet_count == 0:
            warnings.append("Workbook has no sheets")

        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            max_row = ws.max_row or 0
            max_column = ws.max_column or 0

            # Extract headers from first row
            headers = []
            empty_header_count = 0

            if max_row > 0 and max_column > 0:
                for col_idx in range(1, max_column + 1):
                    cell_value = ws.cell(row=1, column=col_idx).value
                    if cell_value is not None:
                        headers.append(str(cell_value))
                    else:
                        headers.append("")
                        empty_header_count += 1

            # Count sample rows (excluding header)
            sample_rows_count = max(0, max_row - 1)

            sheets.append({
                "name": sheet_name,
                "max_row": max_row,
                "max_column": max_column,
                "headers": headers,
                "empty_header_count": empty_header_count,
                "sample_rows_count": sample_rows_count,
            })

    finally:
        wb.close()

    return {
        "schema_version": "1.0",
        "service_name": "SERVICE_1",
        "source_path_basename": os.path.basename(path),
        "workbook": {
            "sheet_count": sheet_count,
            "sheets": sheets,
        },
        "warnings": warnings,
        "runtime_authorized": False,
    }

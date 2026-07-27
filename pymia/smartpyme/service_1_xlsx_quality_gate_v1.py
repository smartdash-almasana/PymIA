from __future__ import annotations

from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Final, Iterable
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook

SCHEMA_VERSION: Final[str] = "SERVICE_1_XLSX_QUALITY_GATE_V1"
STATUS_PASS: Final[str] = "PASS"
STATUS_FAIL: Final[str] = "FAIL"

EXCEL_ERROR_TOKENS: Final[frozenset[str]] = frozenset(
    {
        "#REF!",
        "#DIV/0!",
        "#VALUE!",
        "#NAME?",
        "#NUM!",
        "#NULL!",
        "#N/A",
    }
)


@dataclass(frozen=True)
class Service1XlsxQualityResultV1:
    schema_version: str
    verdict: str
    file_path: str
    zip_integrity: bool
    workbook_readable: bool
    sheet_names: tuple[str, ...]
    expected_sheet_names: tuple[str, ...]
    missing_expected_sheets: tuple[str, ...]
    formula_cell_count: int
    excel_error_cells: tuple[str, ...]
    external_link_count: int
    macro_archive_present: bool
    failures: tuple[str, ...]
    runtime_authorized: bool = False
    delivery_authorized: bool = False
    product_ready: bool = False

    def to_dict(self) -> dict[str, object]:
        return asdict(self)


def evaluate_service_1_xlsx_quality_v1(
    *,
    xlsx_path: str | Path,
    expected_sheet_names: Iterable[str] = (),
) -> Service1XlsxQualityResultV1:
    path = Path(xlsx_path)
    expected = tuple(str(name) for name in expected_sheet_names)
    failures: list[str] = []

    if not path.is_file():
        return _failed_result(path, expected, "FILE_NOT_FOUND")

    zip_integrity = False
    try:
        with ZipFile(path, "r") as archive:
            zip_integrity = archive.testzip() is None
            if not zip_integrity:
                failures.append("ZIP_INTEGRITY_FAILED")
    except (BadZipFile, OSError):
        failures.append("INVALID_XLSX_ZIP_PACKAGE")

    sheet_names: tuple[str, ...] = ()
    missing_expected: tuple[str, ...] = expected
    formula_cell_count = 0
    error_cells: list[str] = []
    external_link_count = 0
    macro_archive_present = False
    workbook_readable = False

    if zip_integrity:
        try:
            workbook = load_workbook(path, data_only=False, keep_links=True)
            workbook_readable = True
            sheet_names = tuple(workbook.sheetnames)
            missing_expected = tuple(name for name in expected if name not in sheet_names)
            if missing_expected:
                failures.append("EXPECTED_SHEET_MISSING")

            formula_cell_count = _count_formula_cells(workbook)
            error_cells.extend(_find_error_cells(workbook))
            external_link_count = len(getattr(workbook, "_external_links", ()) or ())
            macro_archive_present = workbook.vba_archive is not None
            workbook.close()

            # A second load with data_only=True inspects cached values, when present.
            cached = load_workbook(path, data_only=True, keep_links=True)
            error_cells.extend(_find_error_cells(cached))
            cached.close()
        except Exception as exc:  # Fail closed at artifact boundary.
            failures.append(f"WORKBOOK_READ_FAILED:{type(exc).__name__}")

    unique_errors = tuple(sorted(set(error_cells)))
    if unique_errors:
        failures.append("EXCEL_ERROR_CELLS_PRESENT")

    verdict = STATUS_PASS if not failures else STATUS_FAIL
    return Service1XlsxQualityResultV1(
        schema_version=SCHEMA_VERSION,
        verdict=verdict,
        file_path=str(path.resolve()),
        zip_integrity=zip_integrity,
        workbook_readable=workbook_readable,
        sheet_names=sheet_names,
        expected_sheet_names=expected,
        missing_expected_sheets=missing_expected,
        formula_cell_count=formula_cell_count,
        excel_error_cells=unique_errors,
        external_link_count=external_link_count,
        macro_archive_present=macro_archive_present,
        failures=tuple(failures),
        runtime_authorized=False,
        delivery_authorized=False,
        product_ready=False,
    )


def _count_formula_cells(workbook) -> int:
    count = 0
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.data_type == "f" or (isinstance(cell.value, str) and cell.value.startswith("=")):
                    count += 1
    return count


def _find_error_cells(workbook) -> list[str]:
    errors: list[str] = []
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                value = cell.value
                if cell.data_type == "e" or (isinstance(value, str) and value.strip().upper() in EXCEL_ERROR_TOKENS):
                    errors.append(f"{worksheet.title}!{cell.coordinate}:{value}")
    return errors


def _failed_result(
    path: Path,
    expected: tuple[str, ...],
    failure: str,
) -> Service1XlsxQualityResultV1:
    return Service1XlsxQualityResultV1(
        schema_version=SCHEMA_VERSION,
        verdict=STATUS_FAIL,
        file_path=str(path.resolve()),
        zip_integrity=False,
        workbook_readable=False,
        sheet_names=(),
        expected_sheet_names=expected,
        missing_expected_sheets=expected,
        formula_cell_count=0,
        excel_error_cells=(),
        external_link_count=0,
        macro_archive_present=False,
        failures=(failure,),
        runtime_authorized=False,
        delivery_authorized=False,
        product_ready=False,
    )


__all__ = [
    "SCHEMA_VERSION",
    "STATUS_PASS",
    "STATUS_FAIL",
    "EXCEL_ERROR_TOKENS",
    "Service1XlsxQualityResultV1",
    "evaluate_service_1_xlsx_quality_v1",
]

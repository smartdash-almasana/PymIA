"""Downloadable reconciliation workpaper for Servicio 1.

Builds an XLSX from an already-governed reconciliation result plus explicit
human review decisions. It does not mutate source movements, certify evidence,
create accounting entries, or close a reconciliation.
"""
from __future__ import annotations

import hashlib
import json
import re
from io import BytesIO
from typing import Any, Final, Mapping, Sequence

from openpyxl import Workbook

from pymia.smartpyme.service_1_reconciliation_request_gate_v1 import (
    BANK_RECONCILIATION,
    MERCADO_PAGO_BANK_RECONCILIATION,
)

SCHEMA_VERSION: Final[str] = "SERVICE_1_RECONCILIATION_WORKPAPER_XLSX_V1"
PACKET_TYPE: Final[str] = "RECONCILIATION_WORKPAPER_XLSX"
SHEET_NAMES: Final[tuple[str, ...]] = (
    "Resumen",
    "Casos",
    "Decisiones",
    "Pendientes",
    "Trazabilidad",
    "Limites",
)

_LIMITATIONS: Final[tuple[str, ...]] = (
    "Documento de trabajo para revisión; no es certificación, auditoría ni cierre contable.",
    "No modifica los movimientos ni los archivos fuente.",
    "No genera asientos contables ni autoriza ejecución automática.",
    "Las coincidencias algorítmicas no equivalen a aceptación humana.",
    "Los casos sin decisión o con decisión PENDING permanecen pendientes.",
)


def build_service_1_reconciliation_workpaper_xlsx_v1(
    *,
    reconciliation_packet: Mapping[str, Any],
    human_decisions: Sequence[Mapping[str, Any]],
) -> dict[str, Any]:
    if not isinstance(reconciliation_packet, Mapping):
        raise ValueError("reconciliation_packet is required")
    if isinstance(human_decisions, (str, bytes)) or not isinstance(human_decisions, Sequence):
        raise ValueError("human_decisions must be a sequence")

    run_raw = reconciliation_packet.get("reconciliation_run")
    run = run_raw if isinstance(run_raw, Mapping) else {}
    case_id = _text(run.get("case_id"))
    reconciliation_type = _text(run.get("reconciliation_type"))
    assisted_raw = run.get("assisted_review")
    assisted = assisted_raw if isinstance(assisted_raw, Mapping) else {}
    review_result_raw = assisted.get("review_result")
    review_result = review_result_raw if isinstance(review_result_raw, Mapping) else {}

    if not case_id:
        raise ValueError("reconciliation case_id is required")
    if reconciliation_type not in {BANK_RECONCILIATION, MERCADO_PAGO_BANK_RECONCILIATION}:
        raise ValueError("unsupported reconciliation_type")
    if not review_result:
        raise ValueError("assisted reconciliation review_result is required")

    cases = _flatten_review_cases(reconciliation_type, review_result)
    known_refs = {item["review_item_ref"] for item in cases}
    decisions = _validated_decisions(
        human_decisions=human_decisions,
        case_id=case_id,
        reconciliation_type=reconciliation_type,
        known_refs=known_refs,
    )
    latest = _latest_decisions(decisions)

    for item in cases:
        current = latest.get(item["review_item_ref"])
        item["human_decision"] = _text(current.get("decision")) if current else "PENDING_REVIEW"
        item["reviewed_by"] = _text(current.get("reviewed_by")) if current else ""
        item["decided_at"] = _text(current.get("decided_at")) if current else ""
        item["observation"] = _text(current.get("observation")) if current else ""

    pending = [
        item for item in cases if item["human_decision"] in {"PENDING", "PENDING_REVIEW"}
    ]
    confirmed = sum(item["human_decision"] == "CONFIRM" for item in cases)
    rejected = sum(item["human_decision"] == "REJECT" for item in cases)

    workbook = Workbook()
    workbook.properties.creator = "SERVICE_1"
    workbook.properties.lastModifiedBy = "SERVICE_1"
    _write_summary(
        workbook.active,
        case_id=case_id,
        reconciliation_type=reconciliation_type,
        case_count=len(cases),
        decision_event_count=len(decisions),
        confirmed=confirmed,
        rejected=rejected,
        pending=len(pending),
    )
    _write_cases(workbook.create_sheet(SHEET_NAMES[1]), cases)
    _write_decisions(workbook.create_sheet(SHEET_NAMES[2]), decisions)
    _write_pending(workbook.create_sheet(SHEET_NAMES[3]), pending)
    _write_traceability(
        workbook.create_sheet(SHEET_NAMES[4]),
        reconciliation_packet=reconciliation_packet,
        run=run,
        assisted=assisted,
    )
    _write_limits(workbook.create_sheet(SHEET_NAMES[5]))

    buffer = BytesIO()
    workbook.save(buffer)
    content = buffer.getvalue()
    workpaper_id = _workpaper_id(
        case_id=case_id,
        reconciliation_type=reconciliation_type,
        cases=cases,
        decisions=decisions,
    )
    return {
        "schema_version": SCHEMA_VERSION,
        "packet_type": PACKET_TYPE,
        "workpaper_id": workpaper_id,
        "case_id": case_id,
        "reconciliation_type": reconciliation_type,
        "filename": f"conciliacion_{_safe_filename(case_id)}.xlsx",
        "content": content,
        "sheet_names": list(SHEET_NAMES),
        "case_count": len(cases),
        "decision_event_count": len(decisions),
        "confirmed_count": confirmed,
        "rejected_count": rejected,
        "pending_count": len(pending),
        "source_data_modified": False,
        "accounting_closure_authorized": False,
        "runtime_authorized": False,
        "llm_used": False,
    }


def _flatten_review_cases(
    reconciliation_type: str,
    review_result: Mapping[str, Any],
) -> list[dict[str, Any]]:
    sections: list[tuple[str, list[Any]]]
    if reconciliation_type == BANK_RECONCILIATION:
        sections = [
            ("exact", _summary_items(review_result, "exact_matches_summary")),
            ("probable", _summary_items(review_result, "probable_matches_summary")),
            ("ambiguous", _summary_items(review_result, "ambiguous_matches_summary")),
            ("amount_difference", _summary_items(review_result, "amount_differences_summary")),
            ("date_difference", _summary_items(review_result, "date_differences_summary")),
            ("bank_pending", _summary_items(review_result, "bank_pending_summary")),
            ("internal_pending", _summary_items(review_result, "internal_pending_summary")),
            ("missing_evidence", _summary_items(review_result, "missing_evidence_summary")),
        ]
    else:
        sections = [
            ("exact", _list_value(review_result, "conciliaciones")),
            ("ambiguous", _list_value(review_result, "ambiguos")),
            ("amount_difference", _list_value(review_result, "diferencias_importe")),
            ("bank_pending", _list_value(review_result, "movimientos_banco_sin_operacion_mp")),
            ("internal_pending", _list_value(review_result, "operaciones_mp_sin_acreditacion")),
            ("calculation_inconsistency", _list_value(review_result, "inconsistencias_calculo")),
            ("missing_evidence", _list_value(review_result, "faltantes_evidencia")),
        ]

    cases: list[dict[str, Any]] = []
    for category, items in sections:
        for position, raw_item in enumerate(items, start=1):
            item = raw_item if isinstance(raw_item, Mapping) else {"value": raw_item}
            cases.append(
                {
                    "review_item_ref": f"{category}:{position}",
                    "review_category": category,
                    "review_item": dict(item),
                }
            )
    return cases


def _validated_decisions(
    *,
    human_decisions: Sequence[Mapping[str, Any]],
    case_id: str,
    reconciliation_type: str,
    known_refs: set[str],
) -> list[dict[str, Any]]:
    decisions: list[dict[str, Any]] = []
    for raw in human_decisions:
        if not isinstance(raw, Mapping):
            raise ValueError("human decision must be a mapping")
        record = dict(raw)
        if _text(record.get("case_id")) != case_id:
            raise ValueError("human decision belongs to another case")
        if _text(record.get("reconciliation_type")) != reconciliation_type:
            raise ValueError("human decision belongs to another reconciliation type")
        item_ref = _text(record.get("review_item_ref"))
        if item_ref not in known_refs:
            raise ValueError("human decision references an unknown review item")
        if _text(record.get("decision")) not in {"CONFIRM", "REJECT", "PENDING"}:
            raise ValueError("human decision has invalid decision")
        if not _text(record.get("reviewed_by")):
            raise ValueError("human decision reviewed_by is required")
        decisions.append(record)
    return decisions


def _latest_decisions(decisions: Sequence[Mapping[str, Any]]) -> dict[str, Mapping[str, Any]]:
    latest: dict[str, Mapping[str, Any]] = {}
    for record in decisions:
        latest[_text(record.get("review_item_ref"))] = record
    return latest


def _write_summary(
    worksheet,
    *,
    case_id: str,
    reconciliation_type: str,
    case_count: int,
    decision_event_count: int,
    confirmed: int,
    rejected: int,
    pending: int,
) -> None:
    worksheet.title = SHEET_NAMES[0]
    rows = [
        ("schema_version", SCHEMA_VERSION),
        ("case_id", case_id),
        ("reconciliation_type", reconciliation_type),
        ("casos_totales", case_count),
        ("eventos_decision", decision_event_count),
        ("confirmados", confirmed),
        ("rechazados", rejected),
        ("pendientes", pending),
        ("estado_documento", "PAPEL_DE_TRABAJO_NO_FINAL"),
    ]
    worksheet.append(["campo", "valor"])
    for key, value in rows:
        worksheet.append([key, value])


def _write_cases(worksheet, cases: Sequence[Mapping[str, Any]]) -> None:
    worksheet.append(
        [
            "review_item_ref",
            "categoria",
            "decision_actual",
            "reviewed_by",
            "decided_at",
            "observacion",
            "caso",
        ]
    )
    for item in cases:
        worksheet.append(
            [
                _safe_cell(item.get("review_item_ref")),
                _safe_cell(item.get("review_category")),
                _safe_cell(item.get("human_decision")),
                _safe_cell(item.get("reviewed_by")),
                _safe_cell(item.get("decided_at")),
                _safe_cell(item.get("observation")),
                _safe_cell(item.get("review_item")),
            ]
        )


def _write_decisions(worksheet, decisions: Sequence[Mapping[str, Any]]) -> None:
    worksheet.append(
        [
            "decision_id",
            "review_item_ref",
            "categoria",
            "decision",
            "reviewed_by",
            "decided_at",
            "observacion",
            "snapshot",
        ]
    )
    for record in decisions:
        worksheet.append(
            [
                _safe_cell(record.get("decision_id")),
                _safe_cell(record.get("review_item_ref")),
                _safe_cell(record.get("review_category")),
                _safe_cell(record.get("decision")),
                _safe_cell(record.get("reviewed_by")),
                _safe_cell(record.get("decided_at")),
                _safe_cell(record.get("observation")),
                _safe_cell(record.get("review_item_snapshot")),
            ]
        )


def _write_pending(worksheet, pending: Sequence[Mapping[str, Any]]) -> None:
    worksheet.append(["review_item_ref", "categoria", "estado", "caso"])
    for item in pending:
        worksheet.append(
            [
                _safe_cell(item.get("review_item_ref")),
                _safe_cell(item.get("review_category")),
                _safe_cell(item.get("human_decision")),
                _safe_cell(item.get("review_item")),
            ]
        )


def _write_traceability(
    worksheet,
    *,
    reconciliation_packet: Mapping[str, Any],
    run: Mapping[str, Any],
    assisted: Mapping[str, Any],
) -> None:
    worksheet.append(["campo", "valor"])
    rows = [
        ("product_packet_status", reconciliation_packet.get("status")),
        ("reconciliation_run_schema", run.get("schema_version")),
        ("assisted_review_schema", assisted.get("schema_version")),
        ("reconciler_ref", assisted.get("reconciler_ref")),
        ("provenance", assisted.get("provenance")),
        ("source_data_modified", False),
        ("accounting_closure_authorized", False),
        ("runtime_authorized", False),
        ("llm_used", False),
    ]
    for key, value in rows:
        worksheet.append([key, _safe_cell(value)])


def _write_limits(worksheet) -> None:
    worksheet.append(["limitacion"])
    for value in _LIMITATIONS:
        worksheet.append([value])


def _summary_items(container: Mapping[str, Any], key: str) -> list[Any]:
    summary = container.get(key)
    if not isinstance(summary, Mapping):
        return []
    items = summary.get("items")
    return list(items) if isinstance(items, list) else []


def _list_value(container: Mapping[str, Any], key: str) -> list[Any]:
    value = container.get(key)
    return list(value) if isinstance(value, list) else []


def _safe_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        text = "true" if value else "false"
    elif isinstance(value, (int, float, str)):
        text = str(value)
    else:
        text = json.dumps(value, ensure_ascii=False, sort_keys=True, default=str)
    if text.startswith(("=", "+", "-", "@")):
        return "'" + text
    return text


def _safe_filename(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", value).strip("._")
    return cleaned[:80] or "reconciliacion"


def _workpaper_id(
    *,
    case_id: str,
    reconciliation_type: str,
    cases: Sequence[Mapping[str, Any]],
    decisions: Sequence[Mapping[str, Any]],
) -> str:
    payload = json.dumps(
        {
            "case_id": case_id,
            "reconciliation_type": reconciliation_type,
            "cases": list(cases),
            "decisions": list(decisions),
        },
        ensure_ascii=False,
        sort_keys=True,
        default=str,
    )
    digest = hashlib.sha256(payload.encode("utf-8")).hexdigest()[:16]
    return f"service_1_reconciliation_workpaper_xlsx_v1:{digest}"


def _text(value: object) -> str:
    return str(value or "").strip()


__all__ = [
    "SCHEMA_VERSION",
    "PACKET_TYPE",
    "SHEET_NAMES",
    "build_service_1_reconciliation_workpaper_xlsx_v1",
]

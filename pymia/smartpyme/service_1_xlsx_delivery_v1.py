from __future__ import annotations

import hashlib
import json
from datetime import datetime
from pathlib import Path
from typing import Final, NotRequired, TypedDict

from openpyxl import Workbook

DELIVERY_SCHEMA_VERSION: Final[str] = "1.0"
SERVICE_NAME: Final[str] = "SERVICE_1"
SHEET_NAMES: Final[tuple[str, ...]] = (
    "Resumen",
    "Datos usados",
    "Resultados",
    "Faltantes",
    "Limitaciones",
    "Claims prohibidos",
    "Notas técnicas",
)
_FIXED_WORKBOOK_TIMESTAMP: Final[datetime] = datetime(2000, 1, 1, 0, 0, 0)


class Service1XlsxDeliveryInputV1(TypedDict):
    service_name: str
    capability_ref: str
    status: str
    owner_summary: str
    inputs_used: dict[str, object]
    computed_results: dict[str, object]
    missing_inputs: list[str]
    limitations: list[str]
    forbidden_claims: list[str]
    technical_notes: list[str]
    runtime_authorized: bool
    summary_ref_label: NotRequired[str]


class Service1XlsxDeliveryV1(TypedDict):
    delivery_id: str
    schema_version: str
    service_name: str
    capability_ref: str
    output_path: str
    sheet_names: list[str]
    runtime_authorized: bool
    notes: list[str]


def build_service_1_xlsx_delivery_v1(
    *,
    delivery_input: Service1XlsxDeliveryInputV1,
    output_path: str | Path,
) -> Service1XlsxDeliveryV1:
    if delivery_input["runtime_authorized"]:
        raise ValueError("SERVICE_1_XLSX_DELIVERY_V1 does not accept runtime_authorized=True.")
    if delivery_input["service_name"] != SERVICE_NAME:
        raise ValueError("SERVICE_1_XLSX_DELIVERY_V1 only accepts SERVICE_1 inputs.")

    output_file = Path(output_path)
    parent_dir = output_file.parent
    if not parent_dir.exists():
        raise FileNotFoundError(f"Output directory does not exist: {parent_dir}")

    workbook = Workbook()
    _configure_workbook(workbook)
    _populate_workbook(workbook, delivery_input)
    workbook.save(output_file)

    return {
        "delivery_id": _build_delivery_id(delivery_input),
        "schema_version": DELIVERY_SCHEMA_VERSION,
        "service_name": SERVICE_NAME,
        "capability_ref": str(delivery_input["capability_ref"]),
        "output_path": str(output_file.resolve()),
        "sheet_names": list(SHEET_NAMES),
        "runtime_authorized": False,
        "notes": [
            "Deterministic XLSX delivery generated from Service1XlsxDeliveryInputV1 only.",
            "No formulas, macros, or runtime execution were used.",
        ],
    }


def _configure_workbook(workbook: Workbook) -> None:
    workbook.properties.creator = SERVICE_NAME
    workbook.properties.lastModifiedBy = SERVICE_NAME
    workbook.properties.created = _FIXED_WORKBOOK_TIMESTAMP
    workbook.properties.modified = _FIXED_WORKBOOK_TIMESTAMP


def _populate_workbook(workbook: Workbook, delivery_input: Service1XlsxDeliveryInputV1) -> None:
    summary_sheet = workbook.active
    summary_sheet.title = SHEET_NAMES[0]
    _write_key_value_sheet(
        summary_sheet,
        header=("field", "value"),
        rows=[
            ("service_name", delivery_input["service_name"]),
            (delivery_input.get("summary_ref_label", "capability_ref"), delivery_input["capability_ref"]),
            ("status", delivery_input["status"]),
            ("owner_summary", delivery_input["owner_summary"]),
            ("runtime_authorized", delivery_input["runtime_authorized"]),
        ],
    )

    _write_key_value_sheet(
        workbook.create_sheet(SHEET_NAMES[1]),
        header=("key", "value"),
        rows=list(delivery_input["inputs_used"].items()),
    )
    _write_key_value_sheet(
        workbook.create_sheet(SHEET_NAMES[2]),
        header=("key", "value"),
        rows=list(delivery_input["computed_results"].items()),
    )
    _write_single_column_sheet(
        workbook.create_sheet(SHEET_NAMES[3]),
        header="missing_input",
        rows=delivery_input["missing_inputs"],
    )
    _write_single_column_sheet(
        workbook.create_sheet(SHEET_NAMES[4]),
        header="limitation",
        rows=delivery_input["limitations"],
    )
    _write_single_column_sheet(
        workbook.create_sheet(SHEET_NAMES[5]),
        header="forbidden_claim",
        rows=delivery_input["forbidden_claims"],
    )
    _write_single_column_sheet(
        workbook.create_sheet(SHEET_NAMES[6]),
        header="technical_note",
        rows=delivery_input["technical_notes"],
    )


def _write_key_value_sheet(
    worksheet,
    *,
    header: tuple[str, str],
    rows: list[tuple[object, object]],
) -> None:
    worksheet.append([header[0], header[1]])
    for key, value in rows:
        worksheet.append([_to_safe_text(key), _to_safe_text(value)])


def _write_single_column_sheet(
    worksheet,
    *,
    header: str,
    rows: list[object],
) -> None:
    worksheet.append([header])
    for value in rows:
        worksheet.append([_to_safe_text(value)])


def _to_safe_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, str):
        if value.startswith(("=", "+", "-", "@")):
            return f"'{value}"
        return value

    text = json.dumps(value, ensure_ascii=False, sort_keys=True)
    if text.startswith(("=", "+", "-", "@")):
        return f"'{text}"
    return text


def _build_delivery_id(delivery_input: Service1XlsxDeliveryInputV1) -> str:
    canonical_payload = json.dumps(
        {
            "service_name": delivery_input["service_name"],
            "capability_ref": delivery_input["capability_ref"],
            "status": delivery_input["status"],
            "inputs_used": delivery_input["inputs_used"],
            "computed_results": delivery_input["computed_results"],
            "missing_inputs": delivery_input["missing_inputs"],
            "limitations": delivery_input["limitations"],
            "forbidden_claims": delivery_input["forbidden_claims"],
            "owner_summary": delivery_input["owner_summary"],
            "technical_notes": delivery_input["technical_notes"],
        },
        ensure_ascii=False,
        sort_keys=True,
    )
    payload_hash = hashlib.sha256(canonical_payload.encode("utf-8")).hexdigest()[:16]
    return f"service_1_xlsx_delivery_v1:{payload_hash}"

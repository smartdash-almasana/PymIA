from __future__ import annotations

from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from pymia.smartpyme.service_1_normalized_table_v1 import (
    NormalizedTableV1,
    build_normalized_table_v1,
)

_EMPTY_SHEET_ERROR = "XLSX sheet has no non-empty rows."


def read_xlsx_to_normalized_table_v1(
    xlsx_path: str | Path,
    *,
    sheet_name: str | None = None,
) -> NormalizedTableV1:
    """Read one worksheet through the canonical XLSX parser.

    With ``sheet_name`` omitted, the first non-empty worksheet is returned,
    preserving the historical single-sheet contract. The workbook is always
    closed before this function returns.
    """
    selected = (sheet_name,) if sheet_name is not None else None
    tables = read_xlsx_to_normalized_tables_v1(xlsx_path, sheet_names=selected)
    if not tables:
        return _blocked(str(Path(xlsx_path)), sheet_name, "XLSX workbook has no readable sheets.")
    return tables[0]


def read_xlsx_to_normalized_tables_v1(
    xlsx_path: str | Path,
    *,
    sheet_names: Iterable[str] | None = None,
) -> list[NormalizedTableV1]:
    """Read every selected non-empty worksheet using one canonical parser.

    ``sheet_names=None`` means all non-empty workbook sheets in workbook order.
    Explicit selection preserves the requested order and fails closed when a
    requested sheet is missing, duplicated or empty.
    """
    path = Path(xlsx_path)
    source_path = str(path)

    if not path.exists() or not path.is_file():
        return [_blocked(source_path, None, "File not found.")]
    if path.suffix.lower() != ".xlsx":
        return [_blocked(source_path, None, "Only .xlsx files are accepted.")]

    normalized_selection: tuple[str, ...] | None = None
    if sheet_names is not None:
        cleaned = tuple(str(name).strip() for name in sheet_names if str(name).strip())
        if not cleaned:
            return [_blocked(source_path, None, "At least one sheet name is required.")]
        if len(set(cleaned)) != len(cleaned):
            return [_blocked(source_path, None, "Duplicate sheet names are not allowed.")]
        normalized_selection = cleaned

    try:
        workbook = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        return [_blocked(source_path, None, str(exc))]

    try:
        if normalized_selection is not None:
            missing = [name for name in normalized_selection if name not in workbook.sheetnames]
            if missing:
                return [
                    _blocked(
                        source_path,
                        missing[0],
                        "Sheet not found: " + ", ".join(missing),
                    )
                ]
            worksheets = [workbook[name] for name in normalized_selection]
            return [
                _normalize_worksheet(source_path=source_path, worksheet=worksheet)
                for worksheet in worksheets
            ]

        tables: list[NormalizedTableV1] = []
        for worksheet in workbook.worksheets:
            table = _normalize_worksheet(source_path=source_path, worksheet=worksheet)
            if _EMPTY_SHEET_ERROR in table["blocking_errors"]:
                continue
            tables.append(table)
        if tables:
            return tables
        return [_blocked(source_path, None, "XLSX workbook has no non-empty rows.")]
    finally:
        workbook.close()


def _normalize_worksheet(*, source_path: str, worksheet: Any) -> NormalizedTableV1:
    selected_sheet_name = worksheet.title
    materialized_rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
    header_index = _first_non_empty_row_index(materialized_rows)

    if header_index is None:
        return _blocked(source_path, selected_sheet_name, _EMPTY_SHEET_ERROR)

    raw_headers = _trim_trailing_empty(materialized_rows[header_index])
    headers = [_clean(value) for value in raw_headers]
    if not headers or any(not header for header in headers):
        return build_normalized_table_v1(
            source_kind="xlsx",
            source_path=source_path,
            sheet_name=selected_sheet_name,
            headers=headers,
            rows=[],
            blocking_errors=["XLSX headers are missing or incomplete."],
        )

    rows: list[dict[str, Any]] = []
    source_row_numbers: list[int] = []
    warnings: list[str] = []
    width = len(headers)

    for row_number, raw_row in enumerate(
        materialized_rows[header_index + 1 :], start=header_index + 2
    ):
        if _is_empty_row(raw_row):
            continue
        if _last_non_empty_index(raw_row[:width]) < width - 1:
            warnings.append(f"Row {row_number} has fewer cells than headers.")
        if any(_clean(value) for value in raw_row[width:]):
            warnings.append(
                f"Row {row_number} has more cells than headers; extra cells ignored."
            )
        fitted = _fit_width(raw_row, width)
        rows.append({headers[index]: _clean(fitted[index]) for index in range(width)})
        source_row_numbers.append(row_number)

    return build_normalized_table_v1(
        source_kind="xlsx",
        source_path=source_path,
        sheet_name=selected_sheet_name,
        headers=headers,
        rows=rows,
        warnings=warnings,
        header_row_number=header_index + 1,
        source_row_numbers=source_row_numbers,
    )


def _blocked(source_path: str, sheet_name: str | None, message: str) -> NormalizedTableV1:
    return build_normalized_table_v1(
        source_kind="xlsx",
        source_path=source_path,
        sheet_name=sheet_name,
        headers=[],
        rows=[],
        blocking_errors=[message],
    )


def _first_non_empty_row_index(rows: list[list[Any]]) -> int | None:
    for index, row in enumerate(rows):
        if not _is_empty_row(row):
            return index
    return None


def _is_empty_row(row: list[Any]) -> bool:
    return all(not _clean(value) for value in row)


def _trim_trailing_empty(row: list[Any]) -> list[Any]:
    last_index = _last_non_empty_index(row)
    if last_index < 0:
        return []
    return row[: last_index + 1]


def _last_non_empty_index(row: list[Any]) -> int:
    for index in range(len(row) - 1, -1, -1):
        if _clean(row[index]):
            return index
    return -1


def _fit_width(row: list[Any], width: int) -> list[Any]:
    if len(row) < width:
        return row + [""] * (width - len(row))
    return row[:width]


def _clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


__all__ = [
    "read_xlsx_to_normalized_table_v1",
    "read_xlsx_to_normalized_tables_v1",
]

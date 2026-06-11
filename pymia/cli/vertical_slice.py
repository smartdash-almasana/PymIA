from __future__ import annotations

import argparse
import hashlib
from pathlib import Path

from openpyxl import load_workbook

from pymia.smartpyme.owner_facing_report import build_owner_facing_report


OPERATIONAL_TERMS = ("venta", "precio", "costo", "producto", "sku", "cantidad", "fecha")


def inspect_excel(path: Path) -> dict:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        headers = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        normalized_headers = [str(value).strip().lower() for value in headers if value]
        return {
            "sheet": worksheet.title,
            "rows": int(worksheet.max_row or 0),
            "columns": int(worksheet.max_column or 0),
            "headers": normalized_headers,
        }
    finally:
        workbook.close()


def has_operational_columns(headers: list[str]) -> bool:
    joined = " ".join(headers)
    return any(term in joined for term in OPERATIONAL_TERMS)


def calculate_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _write_jsonl_line(target: Path, payload: dict) -> Path:
    target.parent.mkdir(parents=True, exist_ok=True)
    import json

    with target.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(payload, sort_keys=True, ensure_ascii=False) + "\n")
    return target


def register_evidence_record(path: Path, tenant_id: str, intake_id: str, storage_dir: Path) -> dict:
    from pymia.smartpyme.evidence import (
        EVIDENCE_STATUS_REGISTERED,
        SOURCE_KIND_UPLOADED_FILE,
        create_evidence_record,
    )
    from pymia.smartpyme.storage import save_evidence_record

    record = create_evidence_record(
        tenant_id=tenant_id,
        intake_id=intake_id,
        evidence_type="xlsx_upload",
        source_kind=SOURCE_KIND_UPLOADED_FILE,
        source_ref=str(path),
        original_filename=path.name,
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        size_bytes=path.stat().st_size,
        content_hash=calculate_sha256(path),
        status=EVIDENCE_STATUS_REGISTERED,
        metadata={"registered_by": "vertical_slice_cli"},
    )
    save_evidence_record(tenant_id, record, base_dir=storage_dir)
    return record.to_dict()


def register_pipeline_run_record(
    *,
    tenant_id: str,
    intake_id: str,
    message: str,
    evidence_record: dict,
    structured_summary: dict,
    blocked: bool,
    storage_dir: Path,
) -> dict:
    from pymia.contracts.pipeline_run_v1 import build_pipeline_run_record

    output_payload = {
        "tenant_id": tenant_id,
        "intake_id": intake_id,
        "evidence_id": evidence_record["evidence_id"],
        "structured_evidence_status": structured_summary["status"],
        "blocked": blocked,
    }
    record = build_pipeline_run_record(
        tenant_id=tenant_id,
        intake_id=intake_id,
        message=message,
        evidence_ids=[evidence_record["evidence_id"]],
        status="BLOCKED" if blocked else "COMPLETED",
        output_payload=output_payload,
        steps_executed=[
            "evidence_record_registered",
            "structured_evidence_built",
            "evidence_sufficiency_checked",
            "owner_facing_report_built",
        ],
    )
    payload = record.model_dump(mode="json")
    _write_jsonl_line(storage_dir / tenant_id / "pipeline_runs.jsonl", payload)
    return payload


def build_structured_summary(
    path: Path,
    tenant_id: str,
    *,
    intake_id: str = "intake_cli_local",
    formula_ids: list[str] | None = None,
) -> dict:
    try:
        from pymia.smartpyme.structured_evidence_builder import build_structured_evidence_context

        formula_ids = formula_ids or []
        payload = build_structured_evidence_context(
            excel_path=path,
            tenant_id=tenant_id,
            intake_record={"evidence_requests": [{"formula_ids": formula_ids}]},
        )
        evidence = payload["structured_evidence"]
        computed = evidence.get("computed_variables") or {}
        tables = evidence.get("tables") or []
        summary = {
            "status": "available",
            "computed_variables_count": len(computed),
            "tables_count": len(tables),
            "case_id": intake_id,
            "sufficiency": [],
            "unsupported_formula_ids": [],
        }
        if formula_ids:
            from pymia.contracts.evidence_v1 import StructuredEvidence
            from pymia.contracts.formula_contract import SUPPORTED_FORMULAS
            from pymia.diagnostic_core.evidence_sufficiency import build_evidence_sufficiency_report_from_structured_evidence

            supported_formula_ids = [formula_id for formula_id in formula_ids if formula_id in SUPPORTED_FORMULAS]
            unsupported_formula_ids = [formula_id for formula_id in formula_ids if formula_id not in SUPPORTED_FORMULAS]
            summary["unsupported_formula_ids"] = unsupported_formula_ids
            if supported_formula_ids:
                sufficiency = build_evidence_sufficiency_report_from_structured_evidence(
                    StructuredEvidence.model_validate(evidence),
                    case_id=intake_id,
                    tenant_id=tenant_id,
                    formula_ids=supported_formula_ids,
                )
                summary["sufficiency"] = [item.model_dump(mode="json") for item in sufficiency]
        return summary
    except Exception as exc:
        return {
            "status": "unavailable",
            "reason": exc.__class__.__name__,
        }


def build_report(
    path: Path,
    message: str,
    profile: dict,
    *,
    tenant_id: str = "tenant_cli_local",
    intake_id: str = "intake_cli_local",
    formula_ids: list[str] | None = None,
    storage_dir: Path | None = None,
) -> dict:
    has_rows = profile["rows"] > 1 and profile["columns"] > 0
    has_columns = has_operational_columns(profile["headers"])
    actual_storage_dir = storage_dir or Path(".tmp/vertical_slice_storage")
    evidence_record = register_evidence_record(
        path,
        tenant_id,
        intake_id,
        actual_storage_dir,
    )
    structured_summary = build_structured_summary(
        path,
        tenant_id,
        intake_id=intake_id,
        formula_ids=formula_ids,
    )
    missing = []
    questions = []
    if not has_rows:
        missing.append("filas_de_datos")
        questions.append("Necesito al menos una fila de datos además de los encabezados.")
    if not has_columns:
        missing.append("columnas_operativas")
        questions.append("Necesito columnas como fecha, producto, ventas, precio, costo, cantidad o sku.")
    blocked = bool(missing)
    summary = "Falta evidencia mínima para avanzar." if blocked else "Planilla legible con señales operativas mínimas; resultado candidato, no diagnóstico final."
    report = build_owner_facing_report(
        operational_audit_result={"tenant_id": tenant_id, "status": "blocked" if blocked else "candidate", "evidence_used": ["excel_file_readable"], "missing_evidence": missing},
        render_contract={"tenant_id": tenant_id, "summary": summary, "blocked_message": summary if blocked else "", "next_questions": questions, "next_steps": ["Revisar con el dueño antes de diagnosticar."], "references": [str(path)], "forbidden_inferences": ["No inferir diagnóstico desde nombres de columnas."]},
        delivery_package={"tenant_id": tenant_id, "intake_id": intake_id, "status": "BLOCKED" if blocked else "DELIVERED", "summary": summary, "output_refs": ["stdout"], "warnings": ["Slice local; no es canal productivo."]},
    ).to_dict()
    pipeline_run_record = register_pipeline_run_record(
        tenant_id=tenant_id,
        intake_id=intake_id,
        message=message,
        evidence_record=evidence_record,
        structured_summary=structured_summary,
        blocked=blocked,
        storage_dir=actual_storage_dir,
    )
    report["evidence_record"] = evidence_record
    report["pipeline_run_record"] = pipeline_run_record
    report["structured_evidence_summary"] = structured_summary
    return report


def build_markdown(
    path: Path,
    message: str,
    profile: dict,
    *,
    tenant_id: str = "tenant_cli_local",
    intake_id: str = "intake_cli_local",
    formula_ids: list[str] | None = None,
    storage_dir: Path | None = None,
) -> str:
    report = build_report(
        path,
        message,
        profile,
        tenant_id=tenant_id,
        intake_id=intake_id,
        formula_ids=formula_ids,
        storage_dir=storage_dir,
    )
    return render_markdown_from_report(path, message, profile, report)


def render_markdown_from_report(path: Path, message: str, profile: dict, report: dict) -> str:
    lines = [
        "# Reporte owner-facing local",
        f"Estado: {report['status']}",
        f"Archivo: {path.name}",
        f"Mensaje: {message}",
        f"Tenant: {report['tenant_id']}",
        f"Intake: {report['intake_id']}",
        f"Evidence ID: {report['evidence_record']['evidence_id']}",
        f"Evidence SHA-256: {report['evidence_record']['content_hash']}",
        f"Run ID: {report['pipeline_run_record']['run_id']}",
        f"Hoja: {profile['sheet']}",
        f"Filas: {profile['rows']}",
        f"Columnas: {profile['columns']}",
        f"Resumen: {report['summary']}",
        "",
        "## Evidencia usada",
    ]
    for item in report["evidence_used"]:
        lines.append(f"- {item}")
    lines.append("")
    lines.append("## Evidencia faltante")
    if report["missing_evidence"]:
        for item in report["missing_evidence"]:
            lines.append(f"- {item}")
    else:
        lines.append("- Sin faltantes mínimos detectados en este slice.")
    lines.append("")
    lines.append("## Evidencia estructurada")
    structured_summary = report["structured_evidence_summary"]
    lines.append(f"- Estado: {structured_summary['status']}")
    if structured_summary["status"] == "available":
        lines.append(f"- Variables computables: {structured_summary['computed_variables_count']}")
        lines.append(f"- Tablas estructuradas: {structured_summary['tables_count']}")
        lines.append(f"- Case ID: {structured_summary['case_id']}")
        if structured_summary["sufficiency"] or structured_summary["unsupported_formula_ids"]:
            lines.append("")
            lines.append("## Suficiencia de evidencia")
            for item in structured_summary["sufficiency"]:
                lines.append(f"- {item['formula_id']}: {item['status']}")
                if item["missing_variables"]:
                    lines.append(f"  - Faltan: {', '.join(item['missing_variables'])}")
            for formula_id in structured_summary["unsupported_formula_ids"]:
                lines.append(f"- {formula_id}: UNSUPPORTED_FORMULA")
    else:
        lines.append(f"- Motivo: {structured_summary['reason']}")
    lines.append("")
    lines.append("## Próxima pregunta")
    if report["next_questions"]:
        for question in report["next_questions"]:
            lines.append(f"- {question}")
    else:
        lines.append("- Confirmar con el dueño si las columnas representan el proceso real.")
    lines.append("")
    lines.append("## Límites")
    for warning in report["limit_warnings"]:
        lines.append(f"- {warning}")
    lines.append("- No diagnostica sin evidencia suficiente ni confirmación del dueño.")
    return "\n".join(lines) + "\n"


def build_pipeline(
    path: Path,
    message: str,
    *,
    tenant_id: str = "tenant_cli_local",
    intake_id: str = "intake_cli_local",
    formula_ids: list[str] | None = None,
    storage_dir: Path | None = None,
) -> dict:
    profile = inspect_excel(path)
    report = build_report(
        path,
        message,
        profile,
        tenant_id=tenant_id,
        intake_id=intake_id,
        formula_ids=formula_ids,
        storage_dir=storage_dir,
    )
    markdown = render_markdown_from_report(path, message, profile, report)
    evidence_record = report["evidence_record"]
    pipeline_run_record = report["pipeline_run_record"]
    return {
        "status": report["status"],
        "profile": profile,
        "report": report,
        "markdown": markdown,
        "evidence_id": evidence_record["evidence_id"],
        "evidence_hash": evidence_record["content_hash"],
        "run_id": pipeline_run_record["run_id"],
        "output_hash": pipeline_run_record["output_hash"],
        "missing_evidence": report["missing_evidence"],
        "next_questions": report["next_questions"],
        "structured_summary": report["structured_evidence_summary"],
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", required=True)
    parser.add_argument("--message", required=True)
    parser.add_argument("--output")
    parser.add_argument("--tenant-id", default="tenant_cli_local")
    parser.add_argument("--intake-id", default="intake_cli_local")
    parser.add_argument("--formula-id", action="append", default=[])
    parser.add_argument("--storage-dir", default=".tmp/vertical_slice_storage")
    args = parser.parse_args(argv)
    path = Path(args.excel)
    if not path.exists():
        raise FileNotFoundError(path)
    pipeline = build_pipeline(
        path,
        args.message,
        tenant_id=args.tenant_id,
        intake_id=args.intake_id,
        formula_ids=args.formula_id,
        storage_dir=Path(args.storage_dir),
    )
    markdown = pipeline["markdown"]
    if args.output:
        output_path = Path(args.output)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(markdown, encoding="utf-8")
    else:
        print(markdown)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

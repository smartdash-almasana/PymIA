from __future__ import annotations

import inspect
from pathlib import Path

import pytest
from openpyxl import load_workbook

from pymia.smartpyme.first_aid_caja_diaria_triage_v1 import run_caja_diaria_triage_v1
from pymia.smartpyme.first_aid_precio_margen_basico_v1 import run_precio_margen_basico_v1
from pymia.smartpyme.first_aid_stock_alertas_basicas_v1 import run_stock_alertas_basicas_v1
from pymia.smartpyme.first_aid_xlsx_delivery_v1 import (
    DELIVERY_SCHEMA_VERSION,
    SERVICE_NAME,
    SHEET_NAMES,
    build_first_aid_xlsx_delivery_v1,
)


def test_generates_xlsx_for_precio_margen_ok_result(tmp_path: Path) -> None:
    tool_result = run_precio_margen_basico_v1(precio_venta=100, costo_unitario=60)
    output_path = tmp_path / "precio_margen.xlsx"

    delivery = build_first_aid_xlsx_delivery_v1(
        tool_result=tool_result,
        output_path=output_path,
    )

    assert output_path.exists()
    assert delivery["schema_version"] == DELIVERY_SCHEMA_VERSION
    assert delivery["service_name"] == SERVICE_NAME
    assert delivery["tool_ref"] == "precio_margen_basico"
    assert delivery["runtime_authorized"] is False


def test_generates_xlsx_for_missing_inputs_result(tmp_path: Path) -> None:
    tool_result = run_stock_alertas_basicas_v1(producto="SKU-1", stock_actual=5)
    output_path = tmp_path / "missing_inputs.xlsx"

    build_first_aid_xlsx_delivery_v1(
        tool_result=tool_result,
        output_path=output_path,
    )

    workbook = load_workbook(output_path)
    rows = list(workbook["Faltantes"].iter_rows(values_only=True))

    assert workbook["Resumen"]["B4"].value == "MISSING_INPUTS"
    assert ("stock_minimo",) in rows


def test_includes_minimum_expected_sheets(tmp_path: Path) -> None:
    tool_result = run_caja_diaria_triage_v1(saldo_inicial=100, ingresos=50, egresos=20)
    output_path = tmp_path / "caja.xlsx"

    build_first_aid_xlsx_delivery_v1(tool_result=tool_result, output_path=output_path)

    workbook = load_workbook(output_path)
    assert workbook.sheetnames == list(SHEET_NAMES)


def test_resumen_sheet_contains_required_fields(tmp_path: Path) -> None:
    tool_result = run_precio_margen_basico_v1(precio_venta=100, costo_unitario=60)
    output_path = tmp_path / "summary.xlsx"

    build_first_aid_xlsx_delivery_v1(tool_result=tool_result, output_path=output_path)

    workbook = load_workbook(output_path)
    summary_rows = list(workbook["Resumen"].iter_rows(values_only=True))

    assert summary_rows == [
        ("field", "value"),
        ("service_name", "SERVICE_1"),
        ("tool_ref", "precio_margen_basico"),
        ("status", "OK"),
        ("owner_summary", "Calculo preliminar sobre precio y costo declarados."),
        ("runtime_authorized", "false"),
    ]


def test_datos_usados_contains_inputs_used(tmp_path: Path) -> None:
    tool_result = run_caja_diaria_triage_v1(saldo_inicial=100, ingresos=50, egresos=20)
    output_path = tmp_path / "inputs.xlsx"

    build_first_aid_xlsx_delivery_v1(tool_result=tool_result, output_path=output_path)

    workbook = load_workbook(output_path)
    rows = list(workbook["Datos usados"].iter_rows(values_only=True))

    assert ("saldo_inicial", "100") in rows
    assert ("ingresos", "50") in rows
    assert ("egresos", "20") in rows


def test_resultados_contains_computed_results(tmp_path: Path) -> None:
    tool_result = run_precio_margen_basico_v1(precio_venta=100, costo_unitario=60)
    output_path = tmp_path / "results.xlsx"

    build_first_aid_xlsx_delivery_v1(tool_result=tool_result, output_path=output_path)

    workbook = load_workbook(output_path)
    rows = list(workbook["Resultados"].iter_rows(values_only=True))

    assert ("margen_bruto_pesos", "40.0") in rows
    assert ("margen_bruto_porcentaje", "0.4") in rows


def test_faltantes_contains_missing_inputs(tmp_path: Path) -> None:
    tool_result = run_precio_margen_basico_v1(precio_venta=100)
    output_path = tmp_path / "faltantes.xlsx"

    build_first_aid_xlsx_delivery_v1(tool_result=tool_result, output_path=output_path)

    workbook = load_workbook(output_path)
    rows = list(workbook["Faltantes"].iter_rows(values_only=True))

    assert ("costo_unitario",) in rows


def test_limitaciones_contains_limitations(tmp_path: Path) -> None:
    tool_result = run_stock_alertas_basicas_v1(producto="SKU-1", stock_actual=5, stock_minimo=10)
    output_path = tmp_path / "limitations.xlsx"

    build_first_aid_xlsx_delivery_v1(tool_result=tool_result, output_path=output_path)

    workbook = load_workbook(output_path)
    rows = list(workbook["Limitaciones"].iter_rows(values_only=True))

    assert ("No confirma stock fisico real.",) in rows


def test_claims_prohibidos_contains_forbidden_claims(tmp_path: Path) -> None:
    tool_result = run_caja_diaria_triage_v1(saldo_inicial=100, ingresos=50, egresos=20)
    output_path = tmp_path / "forbidden_claims.xlsx"

    build_first_aid_xlsx_delivery_v1(tool_result=tool_result, output_path=output_path)

    workbook = load_workbook(output_path)
    rows = list(workbook["Claims prohibidos"].iter_rows(values_only=True))

    assert ("No confirma saldo bancario real.",) in rows
    assert ("No confirma archivo normalizado.",) in rows


def test_notas_tecnicas_contains_technical_notes(tmp_path: Path) -> None:
    tool_result = run_caja_diaria_triage_v1(saldo_inicial=100, ingresos=50, egresos=20)
    output_path = tmp_path / "technical_notes.xlsx"

    build_first_aid_xlsx_delivery_v1(tool_result=tool_result, output_path=output_path)

    workbook = load_workbook(output_path)
    rows = list(workbook["Notas técnicas"].iter_rows(values_only=True))

    assert ("Tool scope is limited to deterministic math over explicit inputs.",) in rows


def test_module_does_not_import_concrete_tool_modules_or_execute_tools() -> None:
    import pymia.smartpyme.first_aid_xlsx_delivery_v1 as module

    source = inspect.getsource(module)

    assert "first_aid_precio_margen_basico_v1" not in source
    assert "first_aid_caja_diaria_triage_v1" not in source
    assert "first_aid_stock_alertas_basicas_v1" not in source
    assert "run_precio_margen_basico_v1" not in source
    assert "run_caja_diaria_triage_v1" not in source
    assert "run_stock_alertas_basicas_v1" not in source


def test_module_does_not_depend_on_pipeline_fsm_llm_chatbot_or_excelsystems() -> None:
    import pymia.smartpyme.first_aid_xlsx_delivery_v1 as module

    source = inspect.getsource(module)

    assert "vertical_pipeline" not in source
    assert "service_1_fsm_decision_patch_v1" not in source
    assert "service_1_boundary_chain_v1" not in source
    assert "document_ingestion" not in source
    assert "exceland" not in source.lower()
    assert "llm" not in source.lower()
    assert "chatbot" not in source.lower()


def test_rejects_runtime_authorized_tool_result(tmp_path: Path) -> None:
    tool_result = run_precio_margen_basico_v1(precio_venta=100, costo_unitario=60)
    tool_result["runtime_authorized"] = True

    with pytest.raises(ValueError, match="does not accept runtime_authorized=True"):
        build_first_aid_xlsx_delivery_v1(
            tool_result=tool_result,
            output_path=tmp_path / "runtime_blocked.xlsx",
        )


def test_xlsx_does_not_create_macros_or_formulas(tmp_path: Path) -> None:
    tool_result = run_precio_margen_basico_v1(precio_venta=100, costo_unitario=60)
    tool_result["owner_summary"] = "=DANGEROUS"
    output_path = tmp_path / "safe_text.xlsx"

    build_first_aid_xlsx_delivery_v1(tool_result=tool_result, output_path=output_path)

    workbook = load_workbook(output_path)
    assert workbook.vba_archive is None

    for sheet_name in workbook.sheetnames:
        for row in workbook[sheet_name].iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                assert cell.data_type != "f"

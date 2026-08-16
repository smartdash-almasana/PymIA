from __future__ import annotations

from datetime import datetime
from http.client import HTTPConnection
from pathlib import Path
from threading import Thread
from urllib.parse import urlencode

from openpyxl import Workbook

from pymia.smartpyme.service_1_assisted_web_v1 import create_assisted_web_server_v1
from pymia.smartpyme.service_1_radar_observation_policy_v1 import RadarObservationPolicyV1


class _RadarStore:
    def __init__(self) -> None:
        self.rows: dict[tuple[str, str], RadarObservationPolicyV1] = {}

    def save_policy(self, policy: RadarObservationPolicyV1) -> bool:
        key = (policy.tenant_id, policy.policy_ref)
        existing = self.rows.get(key)
        if existing is not None and existing.to_dict() != policy.to_dict():
            raise ValueError("conflicting policy")
        self.rows[key] = policy
        return True

    def list_policies(
        self, *, tenant_id: str, enabled_only: bool = False
    ) -> tuple[RadarObservationPolicyV1, ...]:
        rows = [
            policy
            for (row_tenant, _), policy in self.rows.items()
            if row_tenant == tenant_id and (not enabled_only or policy.enabled)
        ]
        return tuple(sorted(rows, key=lambda item: item.policy_ref))


def _xlsx_bytes(tmp_path: Path) -> bytes:
    path = tmp_path / "consorcio.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Expensas"
    sheet.append(["unidad_funcional", "saldo_anterior", "expensa_mes"])
    sheet.append(["UF-12", 200, 100])
    workbook.save(path)
    return path.read_bytes()


def _multipart(filename: str, content: bytes) -> tuple[bytes, dict[str, str]]:
    boundary = "Service1RadarWebBoundary"
    body = (
        f"--{boundary}\r\n"
        f'Content-Disposition: form-data; name="file"; filename="{filename}"\r\n'
        "Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n\r\n"
    ).encode("utf-8") + content + f"\r\n--{boundary}--\r\n".encode("utf-8")
    return body, {
        "Content-Type": f"multipart/form-data; boundary={boundary}",
        "Content-Length": str(len(body)),
        "Authorization": "Bearer test-token",
    }


def _request(server, method: str, path: str, *, body: bytes = b"", headers=None):
    connection = HTTPConnection("127.0.0.1", server.server_port, timeout=10)
    connection.request(method, path, body=body, headers=headers or {})
    response = connection.getresponse()
    payload = response.read().decode("utf-8")
    return response.status, response.getheaders(), payload


def _cookie(headers: list[tuple[str, str]]) -> str:
    for name, value in headers:
        if name.lower() == "set-cookie":
            return value.split(";", 1)[0]
    raise AssertionError("missing session cookie")


def test_authenticated_assisted_web_owner_can_open_radar_menu_and_persist_policy(
    tmp_path: Path,
) -> None:
    store = _RadarStore()

    def resolver(_handler):
        return {
            "tenant_id": "tenant-consorcio-web",
            "cliente_id": "cliente-consorcio-web",
            "owner_actor_id": "owner-consorcio-web",
            "owner_actor_role": "OWNER",
        }

    server = create_assisted_web_server_v1(
        host="127.0.0.1",
        port=0,
        output_dir=tmp_path / "outputs",
        tenant_identity_resolver=resolver,
        radar_policy_store=store,
    )
    thread = Thread(target=server.serve_forever, daemon=True)
    thread.start()
    try:
        upload_body, upload_headers = _multipart(
            "consorcio.xlsx", _xlsx_bytes(tmp_path)
        )
        status, headers, page = _request(
            server,
            "POST",
            "/upload",
            body=upload_body,
            headers=upload_headers,
        )
        assert status == 200
        assert "Esto entendí de tu Excel" in page or "¿Qué querés revisar?" in page
        cookie = _cookie(headers)

        status, _, radar_page = _request(
            server,
            "GET",
            "/radar",
            headers={"Cookie": cookie},
        )
        assert status == 200
        assert "Configurar RADAR del consorcio" in radar_page
        assert "Períodos equivalentes de deuda" in radar_page
        assert "Importe absoluto de movimientos bancarios sin imputar" in radar_page
        assert "RADAR no decide por vos" in radar_page
        assert 'name="communication_level"' in radar_page
        assert "Reporte a demanda" in radar_page
        assert "Notificación" in radar_page
        assert "Alerta" in radar_page
        assert "Urgencia" in radar_page

        form = urlencode(
            {
                "policy_ref": "owner-debt-two-periods-web",
                "observable_ref": "consorcios.debt_equivalent_periods",
                "enabled": "true",
                "operator": "GTE",
                "comparison_value": "2",
                "communication_level": "ALERT",
                "confirmed_by_owner": "true",
            }
        ).encode("utf-8")
        status, _, saved_page = _request(
            server,
            "POST",
            "/save-radar-policy",
            body=form,
            headers={
                "Content-Type": "application/x-www-form-urlencoded",
                "Content-Length": str(len(form)),
                "Cookie": cookie,
                "Authorization": "Bearer test-token",
            },
        )
        assert status == 200
        assert "Regla RADAR guardada" in saved_page
        assert "GTE 2" in saved_page
        assert "ALERT" in saved_page

        stored = store.rows[("tenant-consorcio-web", "owner-debt-two-periods-web")]
        assert stored.observable_ref == "consorcios.debt_equivalent_periods"
        assert stored.comparison_value == "2"
        assert stored.communication_level == "ALERT"
        assert stored.confirmed_by_owner is True
    finally:
        server.shutdown()
        server.server_close()
        thread.join(timeout=5)


def _table_xlsx(tmp_path: Path, name: str, headers: list[str], row: list[object]) -> bytes:
    path = tmp_path / name
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Datos"
    sheet.append(headers)
    sheet.append(row)
    workbook.save(path)
    return path.read_bytes()


def _multipart_files(files: dict[str, tuple[str, bytes]], cookie: str) -> tuple[bytes, dict[str, str]]:
    boundary = "Service1RadarReconciliationBoundary"
    chunks: list[bytes] = []
    for field_name, (filename, content) in files.items():
        chunks.append(
            (
                f"--{boundary}\r\n"
                f'Content-Disposition: form-data; name="{field_name}"; filename="{filename}"\r\n'
                "Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n\r\n"
            ).encode("utf-8")
            + content
            + b"\r\n"
        )
    chunks.append(f"--{boundary}--\r\n".encode("utf-8"))
    body = b"".join(chunks)
    return body, {
        "Content-Type": f"multipart/form-data; boundary={boundary}",
        "Content-Length": str(len(body)),
        "Cookie": cookie,
        "Authorization": "Bearer test-token",
    }


def _post_form(server, path: str, values: dict[str, str], cookie: str):
    body = urlencode(values).encode("utf-8")
    return _request(
        server,
        "POST",
        path,
        body=body,
        headers={
            "Content-Type": "application/x-www-form-urlencoded",
            "Content-Length": str(len(body)),
            "Cookie": cookie,
            "Authorization": "Bearer test-token",
        },
    )


def test_bank_reconciliation_result_page_presents_matching_owner_radar_event(
    tmp_path: Path,
) -> None:
    store = _RadarStore()

    def resolver(_handler):
        return {
            "tenant_id": "tenant-consorcio-event",
            "cliente_id": "cliente-consorcio-event",
            "owner_actor_id": "owner-consorcio-event",
            "owner_actor_role": "OWNER",
        }

    server = create_assisted_web_server_v1(
        host="127.0.0.1",
        port=0,
        output_dir=tmp_path / "outputs-event",
        tenant_identity_resolver=resolver,
        radar_policy_store=store,
    )
    thread = Thread(target=server.serve_forever, daemon=True)
    thread.start()
    try:
        upload_body, upload_headers = _multipart(
            "consorcio.xlsx", _xlsx_bytes(tmp_path)
        )
        status, headers, _ = _request(
            server,
            "POST",
            "/upload",
            body=upload_body,
            headers=upload_headers,
        )
        assert status == 200
        cookie = _cookie(headers)

        status, _, saved_page = _post_form(
            server,
            "/save-radar-policy",
            {
                "policy_ref": "owner-bank-unmatched-event",
                "observable_ref": "consorcios.bank_unmatched_amount",
                "enabled": "true",
                "operator": "GTE",
                "comparison_value": "1000",
                "communication_level": "ALERT",
                "confirmed_by_owner": "true",
            },
            cookie,
        )
        assert status == 200
        assert "Regla RADAR guardada" in saved_page

        status, _, _ = _post_form(
            server,
            "/start-reconciliation",
            {"reconciliation_type": "BANK_RECONCILIATION"},
            cookie,
        )
        assert status == 200

        bank = _table_xlsx(
            tmp_path,
            "bank_event.xlsx",
            ["Número de operación", "Fecha movimiento", "Importe acreditado", "Referencia bancaria"],
            ["BANK-1", datetime(2026, 8, 1), 1250, "BANK-REF-1"],
        )
        internal = _table_xlsx(
            tmp_path,
            "internal_event.xlsx",
            ["Número de factura", "Fecha prevista", "Importe esperado", "Referencia comercial"],
            ["INT-1", datetime(2026, 8, 8), 300, "INTERNAL-REF-9"],
        )
        body, headers = _multipart_files(
            {
                "source_bank": ("bank_event.xlsx", bank),
                "source_internal": ("internal_event.xlsx", internal),
            },
            cookie,
        )
        status, _, page = _request(
            server,
            "POST",
            "/upload-reconciliation",
            body=body,
            headers=headers,
        )
        assert status == 200
        assert "Esto entendí de tus archivos" in page

        status, _, result_page = _post_form(
            server,
            "/confirm-reconciliation-columns",
            {
                "bind_bank_id": "Número de operación",
                "bind_bank_fecha": "Fecha movimiento",
                "bind_bank_importe": "Importe acreditado",
                "bind_bank_referencia": "Referencia bancaria",
                "bind_internal_id": "Número de factura",
                "bind_internal_fecha": "Fecha prevista",
                "bind_internal_importe": "Importe esperado",
                "bind_internal_referencia": "Referencia comercial",
            },
            cookie,
        )
        assert status == 200
        assert "RADAR" in result_page
        assert "consorcios.bank_unmatched_amount" in result_page
        assert "ALERT" in result_page
        assert "1250.0" in result_page
        assert "GTE 1000" in result_page
        assert "no es una severidad asignada por PymIA" in result_page
        assert "HIGH" not in result_page
        assert "MODERATE" not in result_page
    finally:
        server.shutdown()
        server.server_close()
        thread.join(timeout=5)


def _consorcios_multi_xlsx_bytes(tmp_path: Path) -> bytes:
    path = tmp_path / "consorcio_multi.xlsx"
    workbook = Workbook()
    aging = workbook.active
    aging.title = "Expensas"
    aging.append(["unidad_funcional", "saldo_anterior", "expensa_mes"])
    aging.append(["UF-12", 250, 100])
    expenses = workbook.create_sheet("Gastos")
    expenses.append(["rubro", "importe"])
    expenses.append(["Limpieza", 150])
    budget = workbook.create_sheet("Presupuesto")
    budget.append(["rubro", "presupuesto_mensual", "promedio_historico"])
    budget.append(["Limpieza", 100, 100])
    workbook.save(path)
    return path.read_bytes()


def test_consorcios_radar_web_runs_collection_aging_and_expense_variance_with_owner_policies(
    tmp_path: Path,
) -> None:
    store = _RadarStore()
    store.save_policy(
        RadarObservationPolicyV1(
            tenant_id="tenant-consorcio-multi",
            policy_ref="aging-owner-policy",
            observable_ref="consorcios.debt_equivalent_periods",
            enabled=True,
            operator="GTE",
            comparison_value="2",
            communication_level="ALERT",
            confirmed_by_owner=True,
        )
    )
    store.save_policy(
        RadarObservationPolicyV1(
            tenant_id="tenant-consorcio-multi",
            policy_ref="expense-owner-policy",
            observable_ref="consorcios.expense_budget_deviation_pct",
            enabled=True,
            operator="GT",
            comparison_value="40",
            communication_level="NOTIFICATION",
            confirmed_by_owner=True,
        )
    )

    def resolver(_handler):
        return {
            "tenant_id": "tenant-consorcio-multi",
            "cliente_id": "cliente-consorcio-multi",
            "owner_actor_id": "owner-consorcio-multi",
            "owner_actor_role": "OWNER",
        }

    server = create_assisted_web_server_v1(
        host="127.0.0.1",
        port=0,
        output_dir=tmp_path / "outputs-multi",
        tenant_identity_resolver=resolver,
        radar_policy_store=store,
    )
    thread = Thread(target=server.serve_forever, daemon=True)
    thread.start()
    try:
        upload_body, upload_headers = _multipart(
            "consorcio_multi.xlsx", _consorcios_multi_xlsx_bytes(tmp_path)
        )
        status, headers, _ = _request(
            server, "POST", "/upload", body=upload_body, headers=upload_headers
        )
        assert status == 200
        cookie = _cookie(headers)

        status, _, page = _request(
            server,
            "GET",
            "/consorcios-radar-analysis",
            headers={"Cookie": cookie},
        )
        assert status == 200
        assert "Analizar Consorcio con RADAR" in page
        assert "Antigüedad de deuda" in page
        assert "Gastos contra presupuesto e histórico" in page

        aging_form = urlencode(
            {
                "sheet_name": "Expensas",
                "unidad_funcional": "unidad_funcional",
                "saldo_anterior": "saldo_anterior",
                "expensa_mes": "expensa_mes",
            }
        ).encode("utf-8")
        status, _, aging_page = _request(
            server,
            "POST",
            "/run-consorcios-collection-aging",
            body=aging_form,
            headers={
                "Content-Type": "application/x-www-form-urlencoded",
                "Content-Length": str(len(aging_form)),
                "Cookie": cookie,
                "Authorization": "Bearer test-token",
            },
        )
        assert status == 200
        assert "Antigüedad de deuda" in aging_page
        assert "UF-12" in aging_page
        assert "2.5" in aging_page
        assert "RADAR" in aging_page
        assert "ALERT" in aging_page
        assert "consorcios.debt_equivalent_periods" in aging_page
        assert "GTE 2" in aging_page

        expense_form = urlencode(
            {
                "expense_sheet": "Gastos",
                "expense_rubro": "rubro",
                "expense_importe": "importe",
                "budget_sheet": "Presupuesto",
                "budget_rubro": "rubro",
                "presupuesto_mensual": "presupuesto_mensual",
                "promedio_historico": "promedio_historico",
            }
        ).encode("utf-8")
        status, _, expense_page = _request(
            server,
            "POST",
            "/run-consorcios-expense-variance",
            body=expense_form,
            headers={
                "Content-Type": "application/x-www-form-urlencoded",
                "Content-Length": str(len(expense_form)),
                "Cookie": cookie,
                "Authorization": "Bearer test-token",
            },
        )
        assert status == 200
        assert "Gastos del consorcio" in expense_page
        assert "Limpieza" in expense_page
        assert "50.0" in expense_page
        assert "RADAR" in expense_page
        assert "NOTIFICATION" in expense_page
        assert "consorcios.expense_budget_deviation_pct" in expense_page
        assert "GT 40" in expense_page
        assert "HIGH" not in expense_page
        assert "MODERATE" not in expense_page
    finally:
        server.shutdown()
        server.server_close()
        thread.join(timeout=5)

from pathlib import Path

from openpyxl import load_workbook

from pymia.smartpyme.service_1_product_pipeline_v1 import run_service_1_product_pipeline_v1


def _rows(path: Path, sheet: str) -> list[dict[str, object]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    return [
        dict(zip(headers, row))
        for row in ws.iter_rows(min_row=2, values_only=True)
        if any(value is not None and str(value).strip() for value in row)
    ]


def _request(path: Path, case_id: str) -> dict[str, object]:
    expense_rows = _rows(path, "Gastos")
    budget_rows = _rows(path, "Presupuesto")
    approved = [
        "rubro_gasto",
        "importe_gasto",
        "rubro_presupuesto",
        "presupuesto_mes",
        "promedio_ultimos_6_meses",
    ]
    expense_rows = [
        {
            **row,
            "rubro_gasto": row["rubro"],
            "importe_gasto": row["importe"],
        }
        for row in expense_rows
    ]
    budget_rows = [
        {
            **row,
            "rubro_presupuesto": row["rubro"],
        }
        for row in budget_rows
    ]
    return {
        "case_id": case_id,
        "owner_requested": True,
        "expense_rows": expense_rows,
        "budget_rows": budget_rows,
        "expense_bindings": {"rubro": "rubro_gasto", "importe": "importe_gasto"},
        "budget_bindings": {
            "rubro": "rubro_presupuesto",
            "presupuesto_mensual": "presupuesto_mes",
            "promedio_historico": "promedio_ultimos_6_meses",
        },
        "governance": {
            "p5_status": "CONFIRMED",
            "p6_decisions": [{"column_ref": value, "status": "APPROVED"} for value in approved],
            "p7_status": "REQUIREMENT_MATCHED",
            "p8_status": "COMPUTABLE",
            "runtime_authorized": False,
            "tool_execution_authorized": False,
            "product_ready": False,
            "delivery_authorized": False,
            "diagnosis_generated": False,
        },
    }


def _run(path: Path, case_id: str) -> dict[str, object]:
    return run_service_1_product_pipeline_v1(
        ingestion_output=None,
        tool_requests=[],
        output_dir=Path(".tmp_consorcios_expense_variance"),
        expense_variance_request=_request(path, case_id),
    )


def test_cabildo_expense_variance_matches_pilot_oracle() -> None:
    path = Path("prueba_excels/PYMIA_CONSORCIO_CABILDO_2026_07.xlsx")
    packet = _run(path, "consorcio-cabildo-2026-07")
    assert packet["status"] == "EXPENSE_VARIANCE_REVIEW_READY", packet
    outcome = packet["bounded_outcome"]
    flagged = {row["rubro"]: row["classification"] for row in outcome["rows"]}
    assert flagged == {"Electricidad": "ALTO", "Limpieza": "ALTO"}


def test_rivadavia_expense_variance_matches_pilot_oracle() -> None:
    path = Path("prueba_excels/PYMIA_CONSORCIO_RIVADAVIA_2026_07.xlsx")
    packet = _run(path, "consorcio-rivadavia-2026-07")
    assert packet["status"] == "EXPENSE_VARIANCE_REVIEW_READY", packet
    outcome = packet["bounded_outcome"]
    flagged = {row["rubro"]: row["classification"] for row in outcome["rows"]}
    assert flagged == {"Mantenimiento": "ALTO", "Seguro": "MODERADO"}


def test_expense_variance_fails_closed_without_owner_confirmation() -> None:
    path = Path("prueba_excels/PYMIA_CONSORCIO_CABILDO_2026_07.xlsx")
    request = _request(path, "consorcio-cabildo-2026-07")
    request["governance"]["p5_status"] = "NEEDS_OWNER_CONFIRMATION"
    packet = run_service_1_product_pipeline_v1(
        ingestion_output=None,
        tool_requests=[],
        output_dir=Path(".tmp_consorcios_expense_variance"),
        expense_variance_request=request,
    )
    assert packet["status"] == "BLOCKED"
    assert packet["blocked_reason"] == "P5_CONFIRMATION_REQUIRED"

from __future__ import annotations

import html
import re
from http.client import HTTPConnection
from pathlib import Path
from threading import Thread
from urllib.parse import urlencode

import pytest
from openpyxl import Workbook

from pymia.smartpyme.service_1_assisted_web_semantic_reception_v1 import (
    create_semantic_reception_server_v1,
)
from tests.smartpyme.test_service_1_assisted_semantic_product_wiring_v1 import (
    _proposal_from_assignments,
)


@pytest.fixture()
def assisted_server(tmp_path: Path):
    def provider(payload: dict) -> dict:
        assignments = {
            "Ventas.fecha": "business_period",
            "Ventas.venta_total": "sold_amount",
            "Ventas.cobrado": "collected_amount",
            "Resumen.ventas_periodo": "sale_price",
            "Resumen.cmv_total": "costs",
            "Resumen.impuestos_periodo": "taxes",
            "CapitalTrabajo.saldo_inicial": "initial_balance",
            "CapitalTrabajo.cobros_esperados": "expected_collections",
            "CapitalTrabajo.pagos_esperados": "expected_payments",
            "CapitalTrabajo.cuentas_por_cobrar": "accounts_receivable",
            "CapitalTrabajo.ventas_periodo": "sold_amount",
            "CapitalTrabajo.dias_periodo": "days",
            "CapitalTrabajo.activo_corriente": "current_assets",
            "CapitalTrabajo.pasivo_corriente": "current_liabilities",
        }
        available = {
            str(item.get("column_ref") or "")
            for item in payload.get("workbook_profile", {}).get("columns", [])
            if isinstance(item, dict)
        }
        selected = {ref: role for ref, role in assignments.items() if ref in available}
        for ref in available:
            if ref in selected or "." not in ref:
                continue
            _, column_name = ref.split(".", 1)
            fallback_by_column = {
                "fecha": "business_period",
                "venta_total": "sold_amount",
                "cobrado": "collected_amount",
                "ventas_periodo": "sale_price",
                "cmv_total": "costs",
                "impuestos_periodo": "taxes",
                "saldo_inicial": "initial_balance",
                "cobros_esperados": "expected_collections",
                "pagos_esperados": "expected_payments",
                "cuentas_por_cobrar": "accounts_receivable",
                "dias_periodo": "days",
                "activo_corriente": "current_assets",
                "pasivo_corriente": "current_liabilities",
            }
            variable_name = fallback_by_column.get(column_name)
            if variable_name:
                selected[ref] = variable_name
        return _proposal_from_assignments(payload, selected)

    server = create_semantic_reception_server_v1(
        host="127.0.0.1",
        port=0,
        output_dir=tmp_path / "outputs",
        semantic_provider=provider,
    )
    thread = Thread(target=server.serve_forever, daemon=True)
    thread.start()
    try:
        yield server
    finally:
        server.shutdown()
        server.server_close()
        thread.join(timeout=5)


def _request_raw(server, method: str, path: str, body: bytes = b"", headers: dict[str, str] | None = None):
    connection = HTTPConnection("127.0.0.1", server.server_port, timeout=10)
    connection.request(method, path, body=body, headers=headers or {})
    response = connection.getresponse()
    return response.status, response.getheaders(), response.read()


def _request(server, method: str, path: str, body: bytes = b"", headers: dict[str, str] | None = None):
    status, response_headers, response_body = _request_raw(server, method, path, body, headers)
    return status, response_headers, response_body.decode("utf-8")


def _multipart(filename: str, content: bytes, *, launch_review: str | None = None) -> tuple[bytes, dict[str, str]]:
    boundary = "Service1AssistedWebBoundary"
    prefix = b""
    if launch_review is not None:
        prefix = (
            f"--{boundary}\r\n"
            'Content-Disposition: form-data; name="launch_review"\r\n\r\n'
            f"{launch_review}\r\n"
        ).encode("utf-8")
    body = prefix + (
        f"--{boundary}\r\n"
        f'Content-Disposition: form-data; name="file"; filename="{filename}"\r\n'
        "Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n\r\n"
    ).encode("utf-8") + content + f"\r\n--{boundary}--\r\n".encode("utf-8")
    return body, {"Content-Type": f"multipart/form-data; boundary={boundary}", "Content-Length": str(len(body))}


def _sales_xlsx(tmp_path: Path) -> bytes:
    path = tmp_path / "ventas.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Ventas"
    sheet.append(["fecha", "venta_total", "cobrado"])
    sheet.append(["2026-06-01", 1000, 800])
    sheet.append(["2026-06-02", 2000, 1500])
    workbook.save(path)
    return path.read_bytes()


def _margin_xlsx(tmp_path: Path) -> bytes:
    path = tmp_path / "margen.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Resumen"
    sheet.append(["ventas_periodo", "cmv_total", "impuestos_periodo"])
    sheet.append([100000, 60000, 10000])
    workbook.save(path)
    return path.read_bytes()


def _cookie(headers: list[tuple[str, str]]) -> str:
    for key, value in headers:
        if key.lower() == "set-cookie":
            return value.split(";", 1)[0]
    raise AssertionError("missing session cookie")


def _form(server, path: str, values: dict[str, str], cookie: str):
    body = urlencode(values).encode("utf-8")
    return _request(
        server,
        "POST",
        path,
        body,
        {"Content-Type": "application/x-www-form-urlencoded", "Content-Length": str(len(body)), "Cookie": cookie},
    )


class _BrowserAuthResolver:
    def sign_in_with_password(self, email: str, password: str) -> str:
        if email != "owner@example.test" or password != "secret":
            raise ValueError("invalid credentials")
        return "browser.jwt"

    def __call__(self, handler) -> dict[str, str]:
        cookie = str(handler.headers.get("Cookie") or "")
        if "service1_access_token=browser.jwt" not in cookie:
            raise ValueError("verified session required")
        return {
            "tenant_id": "tenant-browser",
            "cliente_id": "cliente-browser",
            "owner_actor_id": "owner-browser",
            "owner_actor_role": "OWNER",
        }


def _semantic_confirmation_answers(page: str) -> dict[str, str]:
    answers: dict[str, str] = {}
    for question_id, option_id in re.findall(
        r'name="answer_([^"]+)" value="([^"]+)"',
        page,
    ):
        if option_id not in {"OTHER", "IGNORE", "not_sure"}:
            answers.setdefault(f"answer_{question_id}", option_id)
    if answers:
        return answers
    for decision_id in re.findall(
        r'name="action_([^"]+)" value="ACCEPT"',
        page,
    ):
        clean = html.unescape(decision_id)
        answers[f"action_{clean}"] = "ACCEPT"
    assert answers
    return answers


def test_browser_login_cookie_allows_real_upload_without_manual_authorization(tmp_path: Path) -> None:
    resolver = _BrowserAuthResolver()
    server = create_semantic_reception_server_v1(
        host="127.0.0.1",
        port=0,
        output_dir=tmp_path / "browser-auth",
        tenant_identity_resolver=resolver,
    )
    thread = Thread(target=server.serve_forever, daemon=True)
    thread.start()
    try:
        status, response_headers, page = _request(server, "GET", "/")
        assert status == 200
        assert "Ingresá a PymIA" in page
        session_cookie = _cookie(response_headers)

        login_body = urlencode(
            {"email": "owner@example.test", "password": "secret"}
        ).encode("utf-8")
        status, login_headers, page = _request(
            server,
            "POST",
            "/login",
            login_body,
            {
                "Content-Type": "application/x-www-form-urlencoded",
                "Content-Length": str(len(login_body)),
                "Cookie": session_cookie,
            },
        )
        assert status == 200
        assert "Subí tu Excel" in page
        assert "Revisar archivo" in page
        cookies = "; ".join(
            value.split(";", 1)[0]
            for key, value in login_headers
            if key.lower() == "set-cookie"
        )
        assert "service1_access_token=browser.jwt" in cookies
        assert "service1_session=" in cookies

        body, headers = _multipart(
            "ventas.xlsx",
            _sales_xlsx(tmp_path),
            launch_review="sold_vs_collected_gap",
        )
        headers["Cookie"] = cookies
        status, _, page = _request(server, "POST", "/upload", body, headers)
        assert status == 200
        assert "Confirmá que entendimos bien" in page
    finally:
        server.shutdown()
        server.server_close()
        thread.join(timeout=5)


def test_http_assisted_flow_uploads_xlsx_confirms_and_evaluates(assisted_server, tmp_path: Path) -> None:
    status, _, page = _request(assisted_server, "GET", "/")
    assert status == 200
    assert "PymIA · Servicio 1" in page

    invalid_body, invalid_headers = _multipart("ventas.csv", b"not an xlsx")
    status, _, page = _request(assisted_server, "POST", "/upload", invalid_body, invalid_headers)
    assert status == 400
    assert "Solo se pueden subir archivos .xlsx." in page

    body, headers = _multipart("ventas.xlsx", _sales_xlsx(tmp_path))
    status, response_headers, page = _request(assisted_server, "POST", "/upload", body, headers)
    assert status == 200
    assert "Revisión del archivo" in page or "Revisión del archivo" in page
    assert "Confirmá que entendimos bien" in page
    cookie = _cookie(response_headers)

    semantic_questions_page = page
    status, _, page = _form(
        assisted_server,
        "/confirm-meanings",
        {"answer_cobrado": "A"},
        cookie,
    )
    assert status == 200
    assert "Elegí confirmar o no usar la interpretación visible." in page

    status, _, page = _form(
        assisted_server,
        "/confirm-meanings",
        _semantic_confirmation_answers(semantic_questions_page),
        cookie,
    )
    assert status == 200
    assert "Elegí qué querés revisar" in page
    assert "Elegí uno o varios análisis" in page
    assert 'name="review_sales_total"' in page

    status, _, page = _form(
        assisted_server,
        "/run-review",
        {"review_sales_total": "1"},
        cookie,
    )
    assert status == 200
    assert "Resumen del archivo" in page
    assert "Resumen de ventas" in page
    assert "Ventas" in page
    assert "$ 3.000,00" in page
    assert "Hojas utilizadas: Ventas" in page
    assert "¿Querés seguir analizando tu planilla Excel?" in page


def test_upload_first_flow_confirms_excel_then_offers_analysis_menu(assisted_server, tmp_path: Path) -> None:
    status, _, home = _request(assisted_server, "GET", "/")
    assert status == 200
    assert "Subí tu Excel" in home
    assert "Revisar archivo" in home
    assert "Ventas y cobranzas" not in home
    assert "Margen real" not in home
    assert "Flujo de caja" not in home

    body, headers = _multipart("ventas.xlsx", _sales_xlsx(tmp_path))
    status, response_headers, page = _request(assisted_server, "POST", "/upload", body, headers)
    assert status == 200
    assert "Revisión del archivo" in page
    cookie = _cookie(response_headers)

    status, _, page = _form(
        assisted_server,
        "/confirm-meanings",
        _semantic_confirmation_answers(page),
        cookie,
    )
    assert status == 200
    assert "Elegí qué querés revisar" in page
    assert "Ventas y cobranzas" not in page
    assert "Margen real" not in page
    assert "Flujo de caja" not in page
    assert 'type="checkbox" name="review_sales_total"' in page
    status, _, page = _form(
        assisted_server,
        "/run-review",
        {"review_sales_total": "1"},
        cookie,
    )
    assert status == 200
    assert "Resumen del archivo" in page
    assert "Resumen de ventas" in page
    assert "Ventas" in page
    assert "$ 3.000,00" in page


def test_one_excel_can_return_multiple_selected_analyses(assisted_server, tmp_path: Path) -> None:
    path = tmp_path / "ventas_y_caja.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Resumen"
    sheet.append([
        "fecha",
        "venta_total",
        "cobrado",
        "saldo_inicial",
        "cobros_esperados",
        "pagos_esperados",
    ])
    sheet.append(["2026-06-01", 3000, 2300, 1000, 2500, 1800])
    workbook.save(path)

    body, headers = _multipart("ventas_y_caja.xlsx", path.read_bytes())
    status, response_headers, page = _request(assisted_server, "POST", "/upload", body, headers)
    assert status == 200
    cookie = _cookie(response_headers)
    if "Revisión del archivo" in page:
        status, _, page = _form(
            assisted_server,
            "/confirm-meanings",
            _semantic_confirmation_answers(page),
            cookie,
        )
        assert status == 200

    assert "Elegí qué querés revisar" in page
    status, _, page = _form(
        assisted_server,
        "/run-review",
        {
            "review_sales_total": "1",
            "review_sales_series_day": "1",
        },
        cookie,
    )

    assert status == 200
    assert "Resumen del archivo" in page
    assert "Resumen de ventas" in page
    assert "Evolución diaria de ventas" in page
    assert "Ventas" in page


def test_launch_margin_real_flow_reaches_real_delivery(assisted_server, tmp_path: Path) -> None:
    body, headers = _multipart(
        "margen.xlsx",
        _margin_xlsx(tmp_path),
        launch_review="net_margin_real",
    )
    status, response_headers, page = _request(assisted_server, "POST", "/upload", body, headers)
    assert status == 200
    assert "Confirmá que entendimos bien" in page
    assert "Sí, está bien" in page
    cookie = _cookie(response_headers)

    status, _, page = _form(
        assisted_server,
        "/confirm-meanings",
        _semantic_confirmation_answers(page),
        cookie,
    )
    assert status == 200
    assert "Margen real" in page
    assert 'href="/download-net-margin"' in page

    status, download_headers, content = _request_raw(
        assisted_server,
        "GET",
        "/download-net-margin",
        headers={"Cookie": cookie},
    )
    assert status == 200
    assert content.startswith(b"PK")
    assert any(
        key.lower() == "content-disposition" and "service_1_ren_001_result.xlsx" in value
        for key, value in download_headers
    )


def test_completed_launch_control_appears_in_recent_cases_and_can_reopen(assisted_server, tmp_path: Path) -> None:
    body, headers = _multipart(
        "ventas.xlsx",
        _sales_xlsx(tmp_path),
        launch_review="sold_vs_collected_gap",
    )
    status, response_headers, page = _request(assisted_server, "POST", "/upload", body, headers)
    assert status == 200
    cookie = _cookie(response_headers)
    status, _, page = _form(
        assisted_server,
        "/confirm-meanings",
        _semantic_confirmation_answers(page),
        cookie,
    )
    assert status == 200
    assert "Ventas y cobranzas" in page

    status, _, cases = _request(
        assisted_server,
        "GET",
        "/cases",
        headers={"Cookie": cookie},
    )
    assert status == 200
    assert "Mis análisis" in cases
    assert "Ventas y cobranzas" in cases
    match = re.search(r'href="(/case\?case_ref=[^"]+)"', cases)
    assert match is not None

    status, _, reopened = _request(
        assisted_server,
        "GET",
        match.group(1),
        headers={"Cookie": cookie},
    )
    assert status == 200
    assert "Ventas y cobranzas" in reopened
    assert "Total vendido" in reopened
    assert "3.000,00" in reopened

    status, _, other_session_cases = _request(assisted_server, "GET", "/cases")
    assert status == 200
    assert "Todavía no hay análisis para mostrar" in other_session_cases
    assert "Ventas y cobranzas" not in other_session_cases


def _working_capital_xlsx(tmp_path: Path) -> bytes:
    path = tmp_path / "capital_trabajo.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "CapitalTrabajo"
    sheet.append([
        "saldo_inicial",
        "cobros_esperados",
        "pagos_esperados",
        "cuentas_por_cobrar",
        "ventas_periodo",
        "dias_periodo",
        "activo_corriente",
        "pasivo_corriente",
    ])
    sheet.append([1000, 2500, 1800, 3000, 9000, 30, 15000, 10000])
    workbook.save(path)
    return path.read_bytes()


def test_launch_working_capital_composes_three_governed_controls(assisted_server, tmp_path: Path) -> None:
    body, headers = _multipart(
        "capital_trabajo.xlsx",
        _working_capital_xlsx(tmp_path),
        launch_review="working_capital",
    )
    status, response_headers, page = _request(assisted_server, "POST", "/upload", body, headers)
    assert status == 200
    cookie = _cookie(response_headers)
    if "Confirmá que entendimos bien" in page:
        assert "Confirmá que entendimos bien" in page
        status, _, page = _form(
            assisted_server,
            "/confirm-meanings",
            _semantic_confirmation_answers(page),
            cookie,
        )
        assert status == 200
    assert "Flujo de caja" in page
    assert "Saldo de caja proyectado" in page
    assert "Tiempo promedio de cobro" in page
    assert "Cobertura de corto plazo" in page
    assert "1.700" in page or "1700" in page
    assert "10.0 días" in page or "10 días" in page
    assert "1.5" in page
    assert "no explica por sí solo la causa" in page.lower()


def test_working_capital_cash_only_is_presented_as_complete_cash_result(assisted_server, tmp_path: Path) -> None:
    path = tmp_path / "flujo_caja.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Proyeccion_Caja"
    sheet.append(["saldo_inicial", "cobros_esperados", "pagos_esperados"])
    sheet.append([1000, 2500, 1800])
    workbook.save(path)

    body, headers = _multipart(
        "flujo_caja.xlsx",
        path.read_bytes(),
        launch_review="working_capital",
    )
    status, response_headers, page = _request(assisted_server, "POST", "/upload", body, headers)
    assert status == 200
    cookie = _cookie(response_headers)
    if "Confirmá que entendimos bien" in page:
        status, _, page = _form(
            assisted_server,
            "/confirm-meanings",
            _semantic_confirmation_answers(page),
            cookie,
        )
        assert status == 200

    assert "Flujo de caja" in page
    assert "Resultados" in page
    assert "Saldo de caja proyectado" in page
    assert "1.700,00" in page or "1700" in page
    assert "Este sería el saldo al cierre" in page
    assert "Podés ampliar este análisis" in page
    assert "Tiempo promedio de cobro" in page
    assert "Cobertura de corto plazo" in page
    assert "No disponible" not in page
    assert "FALTA INFORMACIÓN" not in page


def test_not_sure_keeps_case_open_and_preserves_confirmed_choices(assisted_server, tmp_path: Path) -> None:
    body, headers = _multipart("ventas.xlsx", _sales_xlsx(tmp_path))
    status, response_headers, page = _request(assisted_server, "POST", "/upload", body, headers)
    assert status == 200
    cookie = _cookie(response_headers)
    answers = _semantic_confirmation_answers(page)
    unresolved_key = next(iter(answers))
    partial = dict(answers)
    partial[unresolved_key] = "not_sure"

    status, _, page = _form(assisted_server, "/confirm-meanings", partial, cookie)

    assert status == 200
    assert "Revisión del archivo" in page
    assert "Elegí confirmar o no usar la interpretación visible." in page
    assert 'value="ACCEPT"' in page

    status, _, page = _form(assisted_server, "/confirm-meanings", answers, cookie)
    assert status == 200
    assert "Elegí qué querés revisar" in page


def test_htmx_upload_returns_only_needed_semantic_questions_fragment(
    assisted_server,
    tmp_path: Path,
) -> None:
    body, headers = _multipart("ventas.xlsx", _sales_xlsx(tmp_path))
    headers["HX-Request"] = "true"
    status, _, fragment = _request(
        assisted_server,
        "POST",
        "/upload",
        body,
        headers,
    )

    assert status == 200
    assert fragment.lstrip().startswith("<main")
    assert "<!doctype" not in fragment.lower()
    assert "Revisión del archivo" in fragment
    assert "cobrado" in fragment
    assert "¿Qué representa la columna fecha?" not in fragment


def test_http_assisted_flow_rejects_missing_file_and_surfaces_blocked_result(assisted_server, tmp_path: Path) -> None:
    status, response_headers, page = _request(
        assisted_server,
        "POST",
        "/upload",
        b"",
        {"Content-Type": "multipart/form-data; boundary=empty", "Content-Length": "0"},
    )
    assert status == 400
    assert "Elegí un archivo" in page or "No se pudo leer el envío" in page

    corrupt_body, corrupt_headers = _multipart("incompleto.xlsx", b"not a workbook")
    status, _, page = _request(assisted_server, "POST", "/upload", corrupt_body, corrupt_headers)
    assert status == 400
    assert "No se pudo usar el archivo" in page

    body, headers = _multipart("ventas.xlsx", _sales_xlsx(tmp_path))
    _, response_headers, page = _request(assisted_server, "POST", "/upload", body, headers)
    cookie = _cookie(response_headers)
    status, _, _ = _form(
        assisted_server,
        "/confirm-meanings",
        _semantic_confirmation_answers(page),
        cookie,
    )
    assert status == 200

    status, _, page = _form(assisted_server, "/run-review", {"review": "payment_collection_gap"}, cookie)
    assert status == 200
    assert "Revisión necesaria" in page
    assert "Hay un dato pendiente" in page
    assert "La comprensión semántica no pudo iniciarse de forma segura." in page


def test_upload_first_menu_offers_margin_only_when_preflight_can_close_it(assisted_server, tmp_path: Path) -> None:
    body, headers = _multipart("margen.xlsx", _margin_xlsx(tmp_path))
    status, response_headers, page = _request(assisted_server, "POST", "/upload", body, headers)
    assert status == 200
    cookie = _cookie(response_headers)

    if "Revisión del archivo" in page:
        status, _, page = _form(
            assisted_server,
            "/confirm-meanings",
            _semantic_confirmation_answers(page),
            cookie,
        )
    assert status == 200

    assert "Elegí qué querés revisar" in page
    assert "Margen real" not in page
    assert 'name="review_net_margin_real"' not in page
    assert 'name="review_sold_vs_collected_gap"' not in page
    assert 'name="review_working_capital"' not in page
    assert "Análisis que necesitan más datos" in page

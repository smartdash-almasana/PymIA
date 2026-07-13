from __future__ import annotations

import inspect
from pathlib import Path

import pytest
from openpyxl import load_workbook

from pymia.smartpyme.service_1_manual_first_aid_smoke_case_v1 import (
    SERVICE_NAME,
    SMOKE_CASE_ID,
    run_service_1_manual_first_aid_smoke_case_v1,
)


def test_smoke_case_generates_three_xlsx_files(tmp_path: Path) -> None:
    result = run_service_1_manual_first_aid_smoke_case_v1(tmp_path)

    assert result["smoke_case_id"] == SMOKE_CASE_ID
    assert result["schema_version"] == "1.0"
    assert result["service_name"] == SERVICE_NAME
    assert result["runtime_authorized"] is False
    assert result["flow"]["delivery_count"] == 3
    assert len(result["generated_files"]) == 3

    for generated_file in result["generated_files"]:
        generated_path = Path(generated_file)
        assert generated_path.exists()
        assert generated_path.suffix == ".xlsx"
        assert generated_path.stat().st_size > 0


def test_smoke_case_uses_expected_first_aid_tools_in_order(tmp_path: Path) -> None:
    result = run_service_1_manual_first_aid_smoke_case_v1(tmp_path)

    assert result["flow"]["tool_refs"] == [
        "precio_margen_basico",
        "caja_diaria_triage",
        "stock_alertas_basicas",
    ]
    assert result["flow"]["statuses"] == ["OK", "OK", "OK"]


def test_smoke_case_filename_outputs_are_unique_and_ordered(tmp_path: Path) -> None:
    result = run_service_1_manual_first_aid_smoke_case_v1(tmp_path)
    filenames = [Path(path).name for path in result["generated_files"]]

    assert filenames == [
        "first_aid_001_precio_margen_basico.xlsx",
        "first_aid_002_caja_diaria_triage.xlsx",
        "first_aid_003_stock_alertas_basicas.xlsx",
    ]
    assert len(set(filenames)) == 3


def test_smoke_case_xlsx_files_are_readable(tmp_path: Path) -> None:
    result = run_service_1_manual_first_aid_smoke_case_v1(tmp_path)

    for generated_file in result["generated_files"]:
        workbook = load_workbook(generated_file)
        assert "Resumen" in workbook.sheetnames
        assert "Resultados" in workbook.sheetnames
        assert workbook["Resumen"]["B2"].value == "SERVICE_1"


def test_smoke_case_summary_is_owner_facing_and_conservative(tmp_path: Path) -> None:
    result = run_service_1_manual_first_aid_smoke_case_v1(tmp_path)
    summary_text = result["flow"]["summary_text"]

    assert "Resultados procesados: 3" in summary_text
    assert "precio_margen_basico: OK" in summary_text
    assert "caja_diaria_triage: OK" in summary_text
    assert "stock_alertas_basicas: OK" in summary_text
    assert "Entrega preliminar basada en datos declarados." in summary_text
    assert "diagnostico integral" in summary_text.lower()
    assert "saldo bancario real" in summary_text.lower()
    assert "stock fisico real" in summary_text.lower()


def test_smoke_case_operator_runbook_has_delivery_steps() -> None:
    with pytest.raises(FileNotFoundError):
        run_service_1_manual_first_aid_smoke_case_v1("/nonexistent/manual/smoke/path")


def test_smoke_case_operator_runbook_is_present(tmp_path: Path) -> None:
    result = run_service_1_manual_first_aid_smoke_case_v1(tmp_path)
    runbook = " ".join(result["operator_runbook"]).lower()

    assert "declared inputs" in runbook
    assert "generated xlsx" in runbook
    assert "owner-facing summary" in runbook
    assert "full diagnosis" in runbook
    assert "real stock" in runbook
    assert "real bank balance" in runbook


def test_smoke_case_does_not_depend_on_pipeline_fsm_llm_chatbot_document_ingestion_or_excelsystems() -> None:
    import pymia.smartpyme.service_1_manual_first_aid_smoke_case_v1 as module

    source = inspect.getsource(module)

    assert "vertical_pipeline" not in source
    assert "fsm" not in source.lower()
    assert "llm" not in source.lower()
    assert "chatbot" not in source.lower()
    assert "document_ingestion" not in source
    assert "exceland" not in source.lower()


def test_smoke_case_uses_manual_delivery_flow_not_runtime_pipeline() -> None:
    import pymia.smartpyme.service_1_manual_first_aid_smoke_case_v1 as module

    source = inspect.getsource(module)

    assert "build_service_1_manual_first_aid_delivery_flow_v1" in source
    assert "run_precio_margen_basico_v1" in source
    assert "run_caja_diaria_triage_v1" in source
    assert "run_stock_alertas_basicas_v1" in source
    assert "runtime_authorized" in source

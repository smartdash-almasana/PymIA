"""
Audit tests for SERVICE_1_WEB_COLUMN_CONFIRMATION_INTAKE_BOUNDARY_V1.

Scope: boundary intake only (uploaded/local XLSX -> canonical reader ->
owner question packet). These tests do NOT authorize runtime/product/delivery,
do NOT execute tools, do NOT create delivery, and rely on real XLSX fixtures.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from pymia.smartpyme.service_1_web_column_confirmation_intake_boundary_v1 import (
    BLOCK_DUAL_SOURCE,
    BLOCK_INVALID_EXTENSION,
    BLOCK_MISSING_FILENAME,
    BLOCK_NO_SOURCE,
    BLOCK_READER_FAILED,
    BLOCK_RUNTIME_FLAG_FORBIDDEN,
    BLOCK_SHEET_SELECTION_CONFLICT,
    build_service_1_web_column_confirmation_intake_boundary_v1 as build_intake,
)

# --- Fixture resolution ---------------------------------------------------

_REPO_ROOT = Path(__file__).resolve().parents[2]          # PymIA/
_PARENT_ROOT = _REPO_ROOT.parent                          # PymIA/

_LOCAL_FIXTURE_CANDIDATES = [
    _REPO_ROOT / "prueba_excels" / "cafeteria_abc.xlsx",
    _PARENT_ROOT / "prueba_excels" / "cafeteria_abc.xlsx",
]
_CASE_001_CANDIDATES = [
    _PARENT_ROOT / "prueba_excels" / "CASE_001_ventas_junio_2026_margin_leak.xlsx",
    _REPO_ROOT / "prueba_excels" / "CASE_001_ventas_junio_2026_margin_leak.xlsx",
]


def _first_existing(candidates: list[Path]) -> Path:
    for candidate in candidates:
        if candidate.exists():
            return candidate
    pytest.skip(f"No fixture found among: {[str(c) for c in candidates]}")


@pytest.fixture()
def local_xlsx() -> Path:
    return _first_existing(_LOCAL_FIXTURE_CANDIDATES)


@pytest.fixture()
def case_001_xlsx() -> Path:
    return _first_existing(_CASE_001_CANDIDATES)


# --- 1. local real XLSX path produces owner question packet ---------------

def test_local_real_xlsx_produces_owner_question_packet(local_xlsx: Path) -> None:
    packet = build_intake(local_xlsx_path=str(local_xlsx))

    assert packet["status"] == "NEEDS_OWNER_CONFIRMATION"
    assert packet["blocked_reason"] is None
    assert packet["source_kind"] == "local_path"
    assert packet["filename"] == local_xlsx.name
    assert packet["columns"], "expected detected columns"
    assert packet["owner_questions"], "expected owner questions"
    assert packet["question_count"] == len(packet["owner_questions"])
    assert packet["case_id"]
    assert packet["source_artifact_ref"].startswith("xlsx:sha256:")
    assert packet["workbook_ref"]
    assert packet["workbook_ref"] != packet["filename"]
    assert packet["ingestion_scope"] == "first_non_empty_sheet"
    assert packet["sheet_ref"].startswith("sheet:sha256:")
    assert packet["source_system_ref"] == "local_path"
    assert packet["source_context_ref"] is None


# --- 2. uploaded XLSX bytes + filename produce owner question packet ------

def test_uploaded_bytes_produce_owner_question_packet(local_xlsx: Path) -> None:
    data = local_xlsx.read_bytes()
    packet = build_intake(
        uploaded_xlsx_bytes=data,
        uploaded_filename=local_xlsx.name,
    )

    assert packet["status"] == "NEEDS_OWNER_CONFIRMATION"
    assert packet["source_kind"] == "uploaded_bytes"
    assert packet["filename"] == local_xlsx.name
    assert packet["source_artifact_ref"].startswith("xlsx:sha256:")
    assert packet["workbook_ref"]
    assert packet["workbook_ref"] != packet["filename"]
    assert packet["source_system_ref"] == "uploaded_bytes"
    assert packet["source_context_ref"] is None
    assert packet["columns"]
    assert packet["question_count"] == len(packet["owner_questions"])


# --- 3. no source blocked -------------------------------------------------

def test_no_source_blocked() -> None:
    packet = build_intake()
    assert packet["status"] == "BLOCKED"
    assert packet["blocked_reason"] == BLOCK_NO_SOURCE


# --- 4. dual source blocked -----------------------------------------------

def test_dual_source_blocked(local_xlsx: Path) -> None:
    packet = build_intake(
        local_xlsx_path=str(local_xlsx),
        uploaded_xlsx_bytes=b"x",
    )
    assert packet["status"] == "BLOCKED"
    assert packet["blocked_reason"] == BLOCK_DUAL_SOURCE


# --- 5. invalid extension blocked -----------------------------------------

def test_invalid_extension_blocked() -> None:
    packet = build_intake(local_xlsx_path="ventas.csv")
    assert packet["status"] == "BLOCKED"
    assert packet["blocked_reason"] == BLOCK_INVALID_EXTENSION


# --- 6. missing uploaded filename blocked ---------------------------------

def test_missing_uploaded_filename_blocked() -> None:
    packet = build_intake(uploaded_xlsx_bytes=b"x")
    assert packet["status"] == "BLOCKED"
    assert packet["blocked_reason"] == BLOCK_MISSING_FILENAME


# --- 7. canonical reader fails -> blocked ---------------------------------

def test_canonical_reader_failure_blocked(tmp_path: Path) -> None:
    # A .xlsx-named file with non-xlsx content forces the canonical reader
    # to fail (openpyxl cannot open it), which must surface as a block.
    corrupt = tmp_path / "corrupt.xlsx"
    corrupt.write_bytes(b"this is not a real xlsx workbook")

    packet = build_intake(local_xlsx_path=str(corrupt))
    assert packet["status"] == "BLOCKED"
    assert packet["blocked_reason"] == BLOCK_READER_FAILED


# --- 8. CASE_001 question-count lock --------------------------------------

def test_case_001_question_count_lock(case_001_xlsx: Path) -> None:
    """Lock the deterministic question count for the canonical CASE_001 fixture.

    Decision (confirmed): the lock is the fixture reality, not a magic number.
    'CASE_001_ventas_junio_2026_margin_leak.xlsx' has 10 detected columns, so
    the deterministic packet yields exactly 10 questions. The earlier spec
    value of 12 was wrong and is not used. This test locks the strong
    invariant: question_count == len(columns) == len(owner_questions) == 10.
    """
    packet = build_intake(local_xlsx_path=str(case_001_xlsx))
    assert packet["status"] == "NEEDS_OWNER_CONFIRMATION"
    # Strong invariant: the lock is the fixture reality, not a magic number.
    assert (
        packet["question_count"]
        == len(packet["columns"])
        == len(packet["owner_questions"])
        == 10
    )


# --- 9/10/11. authorization flags always False in a valid packet ----------

def test_valid_packet_flags_are_false(local_xlsx: Path) -> None:
    packet = build_intake(local_xlsx_path=str(local_xlsx))
    assert packet["runtime_authorized"] is False
    assert packet["product_ready"] is False
    assert packet["delivery_authorized"] is False


# --- 12. passing any authorization flag True is itself blocking -----------

@pytest.mark.parametrize(
    "flag",
    ["runtime_authorized", "product_ready", "delivery_authorized"],
)
def test_authorization_flag_true_blocks(local_xlsx: Path, flag: str) -> None:
    packet = build_intake(**{"local_xlsx_path": str(local_xlsx), flag: True})
    assert packet["status"] == "BLOCKED"
    assert packet["blocked_reason"] == BLOCK_RUNTIME_FLAG_FORBIDDEN
    # Even when blocked, output flags must remain False.
    assert packet["runtime_authorized"] is False
    assert packet["product_ready"] is False
    assert packet["delivery_authorized"] is False


def test_sheet_selection_modes_are_mutually_exclusive(local_xlsx: Path) -> None:
    packet = build_intake(
        local_xlsx_path=local_xlsx,
        sheet_name="Ventas",
        include_all_sheets=True,
    )

    assert packet["status"] == "BLOCKED"
    assert packet["blocked_reason"] == BLOCK_SHEET_SELECTION_CONFLICT
    assert packet["runtime_authorized"] is False
    assert packet["product_ready"] is False
    assert packet["delivery_authorized"] is False


def _two_sheet_xlsx_bytes() -> bytes:
    from io import BytesIO

    workbook = Workbook()
    first = workbook.active
    first.title = "Ventas"
    first.append(["id", "importe"])
    first.append(["A-1", "10"])
    second = workbook.create_sheet("Productos")
    second.append(["sku", "descripcion"])
    second.append(["P-1", "Producto"])
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _one_sheet_xlsx_bytes(value: str) -> bytes:
    from io import BytesIO

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Ventas"
    worksheet.append(["id", "importe"])
    worksheet.append(["A-1", value])
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def test_source_artifact_identity_uses_bytes_not_basename(tmp_path: Path) -> None:
    first_dir = tmp_path / "one"
    second_dir = tmp_path / "two"
    first_dir.mkdir()
    second_dir.mkdir()
    first = first_dir / "same.xlsx"
    second = second_dir / "same.xlsx"
    first.write_bytes(_one_sheet_xlsx_bytes("10"))
    second.write_bytes(_one_sheet_xlsx_bytes("20"))

    first_packet = build_intake(local_xlsx_path=first)
    second_packet = build_intake(local_xlsx_path=second)

    assert first_packet["source_artifact_ref"] != second_packet["source_artifact_ref"]


def test_source_artifact_identity_ignores_filename_for_same_bytes() -> None:
    content = _two_sheet_xlsx_bytes()
    first = build_intake(uploaded_xlsx_bytes=content, uploaded_filename="enero.xlsx")
    second = build_intake(uploaded_xlsx_bytes=content, uploaded_filename="renamed.xlsx")

    assert first["source_artifact_ref"] == second["source_artifact_ref"]
    assert first["workbook_ref"] == second["workbook_ref"]
    assert first["sheet_ref"] == second["sheet_ref"]
    assert first["column_refs"][0]["sheet_ref"] == first["sheet_ref"]
    assert first["owner_questions"][0]["sheet_ref"] == first["sheet_ref"]


def test_local_path_artifact_hash_matches_uploaded_bytes(tmp_path: Path) -> None:
    content = _two_sheet_xlsx_bytes()
    source = tmp_path / "local.xlsx"
    source.write_bytes(content)

    local = build_intake(local_xlsx_path=source)
    uploaded = build_intake(uploaded_xlsx_bytes=content, uploaded_filename="upload.xlsx")

    assert local["source_artifact_ref"] == uploaded["source_artifact_ref"]


def test_workbook_ref_changes_with_ingestion_scope() -> None:
    content = _two_sheet_xlsx_bytes()
    one_sheet = build_intake(
        uploaded_xlsx_bytes=content,
        uploaded_filename="workbook.xlsx",
        sheet_name="Ventas",
    )
    all_sheets = build_intake(
        uploaded_xlsx_bytes=content,
        uploaded_filename="workbook.xlsx",
        include_all_sheets=True,
    )

    assert one_sheet["source_artifact_ref"] == all_sheets["source_artifact_ref"]
    assert one_sheet["workbook_ref"] != all_sheets["workbook_ref"]


def test_missing_requested_sheet_fails_closed(tmp_path: Path) -> None:
    source = tmp_path / "known.xlsx"
    source.write_bytes(_two_sheet_xlsx_bytes())

    packet = build_intake(local_xlsx_path=source, sheet_name="Missing")

    assert packet["status"] == "BLOCKED"
    assert packet["blocked_reason"] == BLOCK_READER_FAILED
    assert packet["sheet_names"] == []

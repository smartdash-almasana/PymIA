from pathlib import Path

from openpyxl import load_workbook

from pymia.smartpyme.service_1_product_pipeline_v1 import run_service_1_product_pipeline_v1


def _rows(path: Path) -> list[dict[str, object]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["Expensas"]
    values = list(ws.iter_rows(values_only=True))
    headers = [str(value) for value in values[0]]
    return [dict(zip(headers, row)) for row in values[1:]]


def _request(path: Path, case_id: str) -> dict[str, object]:
    columns = ["unidad_funcional", "saldo_anterior", "expensa_mes"]
    return {
        "case_id": case_id,
        "owner_requested": True,
        "sheet_name": "Expensas",
        "rows": _rows(path),
        "field_bindings": {name: name for name in columns},
        "governance": {
            "p5_status": "CONFIRMED",
            "p6_decisions": [
                {"column_ref": name, "status": "APPROVED"} for name in columns
            ],
            "p7_status": "REQUIREMENT_MATCHED",
            "p8_status": "COMPUTABLE",
            "runtime_authorized": False,
            "tool_execution_authorized": False,
            "product_ready": False,
            "delivery_authorized": False,
            "diagnosis_generated": False,
        },
    }


def _run(path: Path, case_id: str) -> dict[str, object]:
    return run_service_1_product_pipeline_v1(
        ingestion_output=None,
        tool_requests=[],
        output_dir=path.parent,
        collection_aging_request=_request(path, case_id),
    )


def test_cabildo_collection_aging_physical_fixture() -> None:
    path = Path("prueba_excels/PYMIA_CONSORCIO_CABILDO_2026_07.xlsx")
    packet = _run(path, "consorcio-cabildo-2026-07")
    assert packet["status"] == "AGING_REVIEW_READY"
    result = packet["computation_result"]
    assert result["status"] == "EVALUATED"
    assert result["summary"] == {
        "total_units": 28,
        "current": 18,
        "one_period": 4,
        "two_periods": 3,
        "three_plus_periods": 3,
    }
    three_plus = {
        row["unidad_funcional"]
        for row in result["rows"]
        if row["aging_bucket"] == "THREE_PLUS_PERIODS_EQUIVALENT"
    }
    assert three_plus == {"3C", "5B", "7A"}
    assert packet["runtime_authorized"] is False
    assert packet["delivery_authorized"] is False


def test_rivadavia_collection_aging_physical_fixture() -> None:
    path = Path("prueba_excels/PYMIA_CONSORCIO_RIVADAVIA_2026_07.xlsx")
    packet = _run(path, "consorcio-rivadavia-2026-07")
    assert packet["status"] == "AGING_REVIEW_READY"
    result = packet["computation_result"]
    assert result["status"] == "EVALUATED"
    buckets = {row["unidad_funcional"]: row["aging_bucket"] for row in result["rows"]}
    assert buckets["4A"] == "TWO_PERIODS_EQUIVALENT"
    assert buckets["6C"] == "TWO_PERIODS_EQUIVALENT"
    assert buckets["8B"] == "THREE_PLUS_PERIODS_EQUIVALENT"
    assert result["summary"]["total_units"] == 36


def test_collection_aging_fails_closed_without_owner_confirmation() -> None:
    path = Path("prueba_excels/PYMIA_CONSORCIO_CABILDO_2026_07.xlsx")
    request = _request(path, "consorcio-cabildo-2026-07")
    request["governance"]["p5_status"] = "NEEDS_OWNER_CONFIRMATION"
    packet = run_service_1_product_pipeline_v1(
        ingestion_output=None,
        tool_requests=[],
        output_dir=path.parent,
        collection_aging_request=request,
    )
    assert packet["status"] == "BLOCKED"
    assert packet["blocked_reason"] == "P5_CONFIRMATION_REQUIRED"

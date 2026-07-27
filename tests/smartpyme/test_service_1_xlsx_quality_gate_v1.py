from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from pymia.smartpyme.service_1_xlsx_quality_gate_v1 import (
    STATUS_FAIL,
    STATUS_PASS,
    evaluate_service_1_xlsx_quality_v1,
)


def test_quality_gate_passes_valid_workbook_with_expected_sheets(tmp_path: Path) -> None:
    path = tmp_path / "valid.xlsx"
    workbook = Workbook()
    workbook.active.title = "Resumen"
    workbook.create_sheet("Resultados")
    workbook.save(path)

    result = evaluate_service_1_xlsx_quality_v1(
        xlsx_path=path,
        expected_sheet_names=("Resumen", "Resultados"),
    )

    assert result.verdict == STATUS_PASS
    assert result.zip_integrity is True
    assert result.workbook_readable is True
    assert result.missing_expected_sheets == ()
    assert result.excel_error_cells == ()
    assert result.runtime_authorized is False
    assert result.delivery_authorized is False


def test_quality_gate_fails_corrupt_xlsx_package(tmp_path: Path) -> None:
    path = tmp_path / "corrupt.xlsx"
    path.write_bytes(b"not-an-xlsx")

    result = evaluate_service_1_xlsx_quality_v1(xlsx_path=path)

    assert result.verdict == STATUS_FAIL
    assert result.zip_integrity is False
    assert "INVALID_XLSX_ZIP_PACKAGE" in result.failures


def test_quality_gate_fails_when_expected_sheet_is_missing(tmp_path: Path) -> None:
    path = tmp_path / "missing.xlsx"
    workbook = Workbook()
    workbook.active.title = "Resumen"
    workbook.save(path)

    result = evaluate_service_1_xlsx_quality_v1(
        xlsx_path=path,
        expected_sheet_names=("Resumen", "Resultados"),
    )

    assert result.verdict == STATUS_FAIL
    assert result.missing_expected_sheets == ("Resultados",)
    assert "EXPECTED_SHEET_MISSING" in result.failures


def test_quality_gate_detects_excel_error_cells(tmp_path: Path) -> None:
    path = tmp_path / "error.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Resultados"
    sheet["A1"] = "#REF!"
    sheet["A1"].data_type = "e"
    workbook.save(path)

    result = evaluate_service_1_xlsx_quality_v1(xlsx_path=path)

    assert result.verdict == STATUS_FAIL
    assert result.excel_error_cells == ("Resultados!A1:#REF!",)
    assert "EXCEL_ERROR_CELLS_PRESENT" in result.failures


def test_quality_gate_reports_formula_cells_without_treating_them_as_errors(tmp_path: Path) -> None:
    path = tmp_path / "formula.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = 2
    sheet["A2"] = 3
    sheet["A3"] = "=SUM(A1:A2)"
    workbook.save(path)

    result = evaluate_service_1_xlsx_quality_v1(xlsx_path=path)

    assert result.verdict == STATUS_PASS
    assert result.formula_cell_count == 1
    assert result.excel_error_cells == ()

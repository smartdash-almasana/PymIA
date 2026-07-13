from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from pymia.smartpyme.excel_treatment_lab_completion_slice_v1 import (
    CAPABILITY_REF,
    REVIEW_PACKET_CAPABILITY_REF,
    SERVICE_NAME,
    SYNTHETIC_CASE_ID,
    run_excel_treatment_lab_completion_slice_v1,
)


def test_completion_slice_builds_ready_excel_treatment_lab_packet(tmp_path: Path) -> None:
    result = run_excel_treatment_lab_completion_slice_v1(tmp_path)

    assert result["schema_version"] == "1.0"
    assert result["service_name"] == SERVICE_NAME
    assert result["capability_ref"] == CAPABILITY_REF
    assert result["case_id"] == SYNTHETIC_CASE_ID
    assert result["synthetic_data"] is True
    assert result["real_client_data"] is False
    assert result["runtime_authorized"] is False
    assert result["production_allowed"] is False
    assert result["final_status"] == "READY"


def test_completion_slice_component_chain_is_ready_and_safe(tmp_path: Path) -> None:
    result = run_excel_treatment_lab_completion_slice_v1(tmp_path)

    assert result["lab_result"]["status"] == "OK"
    assert result["lab_result"]["runtime_authorized"] is False
    assert result["lab_result"]["computed_results"]["detected_columns_count"] == 5
    assert result["lab_result"]["computed_results"]["confirmed_columns_count"] == 5
    assert result["lab_result"]["computed_results"]["pending_confirmation_columns"] == []
    assert result["exceland_bridge"]["status"] == "OK"
    assert result["exceland_bridge"]["runtime_authorized"] is False
    assert result["xlsx_delivery"]["runtime_authorized"] is False


def test_completion_slice_generates_three_reviewable_output_files(tmp_path: Path) -> None:
    result = run_excel_treatment_lab_completion_slice_v1(tmp_path)

    assert len(result["output_files"]) == 3
    for output_file in result["output_files"]:
        assert Path(output_file).exists()
        assert Path(output_file).stat().st_size > 0
        assert output_file in result["output_hashes"]
        assert len(result["output_hashes"][output_file]) == 64


def test_completion_slice_generates_readable_xlsx_delivery(tmp_path: Path) -> None:
    result = run_excel_treatment_lab_completion_slice_v1(tmp_path)
    workbook = load_workbook(result["xlsx_delivery"]["output_path"])

    assert workbook["Resumen"]["B2"].value == "SERVICE_1"
    assert workbook["Resumen"]["B3"].value == REVIEW_PACKET_CAPABILITY_REF
    assert workbook["Resumen"]["B4"].value == "READY"
    assert ("detected_columns_count", "5") in list(workbook["Resultados"].iter_rows(values_only=True))
    assert ("confirmed_columns_count", "5") in list(workbook["Resultados"].iter_rows(values_only=True))
    assert "Datos usados" in workbook.sheetnames
    assert "Resultados" in workbook.sheetnames
    assert "Limitaciones" in workbook.sheetnames
    assert "Claims prohibidos" in workbook.sheetnames


def test_completion_slice_owner_summary_is_conservative(tmp_path: Path) -> None:
    result = run_excel_treatment_lab_completion_slice_v1(tmp_path)
    owner_summary = Path(result["owner_summary_path"]).read_text(encoding="utf-8")

    assert "Paquete de revisión Laboratorio Excel" in owner_summary
    assert "Columnas detectadas: 5" in owner_summary
    assert "Columnas confirmadas: 5" in owner_summary
    assert "Estado bridge Exceland: OK" in owner_summary
    assert "No procesa archivos reales." in owner_summary
    assert "No confirma normalización final del archivo del cliente." in owner_summary
    assert "No ejecuta factoría Exceland real." in owner_summary
    assert "No ejecuta fórmulas ni cálculos de negocio sobre datos reales." in owner_summary


def test_completion_slice_operator_notes_explain_limits(tmp_path: Path) -> None:
    result = run_excel_treatment_lab_completion_slice_v1(tmp_path)
    notes = Path(result["operator_notes_path"]).read_text(encoding="utf-8")

    assert "lab_status=OK" in notes
    assert "exceland_bridge_status=OK" in notes
    assert "detected_columns_count=5" in notes
    assert "confirmed_columns_count=5" in notes
    assert "rows_processed=25" in notes
    assert "Use as synthetic review packet only" in notes
    assert "No source workbook was read or modified." in notes
    assert "No external factory or formula execution was run." in notes


def test_completion_slice_rejects_missing_output_dir() -> None:
    missing_dir = Path("/nonexistent/service1/excel-treatment-lab-completion")

    try:
        run_excel_treatment_lab_completion_slice_v1(missing_dir)
    except FileNotFoundError as exc:
        assert "Output directory does not exist" in str(exc)
    else:
        raise AssertionError("Expected FileNotFoundError")


def test_completion_slice_source_does_not_open_forbidden_external_layers() -> None:
    module_path = (
        Path(__file__).parents[2]
        / "pymia"
        / "smartpyme"
        / "excel_treatment_lab_completion_slice_v1.py"
    )
    source = module_path.read_text(encoding="utf-8").lower()

    forbidden_fragments = (
        "openai",
        "langchain",
        "requests",
        "httpx",
        "vertical_pipeline",
        "servicio_2",
        "ocr_runtime",
        "parser_runtime",
        "api_call",
        "read_excel",
        "load_workbook",
    )
    for fragment in forbidden_fragments:
        assert fragment not in source

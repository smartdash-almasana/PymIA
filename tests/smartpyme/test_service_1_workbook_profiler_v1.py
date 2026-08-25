from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook

from pymia.smartpyme.service_1_owner_confirmation_to_canonical_ingestion_output_v1 import (
    STATUS_UNCONFIRMED_READY,
    build_service_1_unconfirmed_canonical_ingestion_output_v1,
)
from pymia.smartpyme.service_1_web_column_confirmation_intake_boundary_v1 import (
    build_service_1_web_column_confirmation_intake_boundary_v1,
)
from pymia.smartpyme.service_1_workbook_profiler_v1 import (
    BLOCK_COLUMN_REF_NOT_FOUND,
    BLOCK_INPUT_AUTHORITY_FORBIDDEN,
    BLOCK_SOURCE_TABLES_MISSING,
    STATUS_BLOCKED,
    STATUS_READY,
    build_service_1_workbook_profile_v1,
)


def _ingestion_output() -> dict:
    return {
        "case_id": "case-cafeteria",
        "filename": "cafeteria.xlsx",
        "source_file_ref": "cafeteria.xlsx",
        "workbook_context": {
            "case_id": "case-cafeteria",
            "source_artifact_ref": "artifact:cafeteria",
            "workbook_ref": "workbook:cafeteria",
            "ingestion_scope": "all_sheets",
            "canonical_reader_schema_version": "SERVICE_1_XLSX_TO_NORMALIZED_TABLE_V1",
        },
        "provenance": {
            "source_kind": "xlsx",
            "source_artifact_ref": "artifact:cafeteria",
            "source_file_ref": "workbook:cafeteria",
            "workbook_ref": "workbook:cafeteria",
            "filename": "cafeteria.xlsx",
            "sheet_names": ["Ventas", "Productos", "Sucursales"],
            "sheet_refs": [],
        },
        "column_refs": [
            {
                "question_id": "col_confirm_001",
                "field_id": "col_confirm_001",
                "sheet_name": "Ventas",
                "column_name": "ProductoID",
                "normalized_column_name": "productoid",
            },
            {
                "question_id": "col_confirm_002",
                "field_id": "col_confirm_002",
                "sheet_name": "Ventas",
                "column_name": "Cantidad",
                "normalized_column_name": "cantidad",
            },
            {
                "question_id": "col_confirm_003",
                "field_id": "col_confirm_003",
                "sheet_name": "Productos",
                "column_name": "ProductoID",
                "normalized_column_name": "productoid",
            },
            {
                "question_id": "col_confirm_004",
                "field_id": "col_confirm_004",
                "sheet_name": "Productos",
                "column_name": "Costo",
                "normalized_column_name": "costo",
            },
        ],
        "normalized_tables": [
            {
                "schema_version": "1.0",
                "service_name": "SERVICE_1",
                "status": "OK",
                "source_kind": "xlsx",
                "source_path": "cafeteria.xlsx",
                "sheet_name": "Ventas",
                "headers": ["ProductoID", "Cantidad"],
                "normalized_headers": ["productoid", "cantidad"],
                "rows": [
                    {"productoid": "P001", "cantidad": "1"},
                    {"productoid": "P002", "cantidad": "2"},
                    {"productoid": "P001", "cantidad": "1"},
                ],
                "header_row_number": 1,
                "source_row_numbers": [2, 3, 4],
                "row_count": 3,
                "column_count": 2,
                "warnings": [],
                "blocking_errors": [],
                "runtime_authorized": False,
            },
            {
                "schema_version": "1.0",
                "service_name": "SERVICE_1",
                "status": "OK",
                "source_kind": "xlsx",
                "source_path": "cafeteria.xlsx",
                "sheet_name": "Productos",
                "headers": ["ProductoID", "Costo"],
                "normalized_headers": ["productoid", "costo"],
                "rows": [
                    {"productoid": "P001", "costo": "10"},
                    {"productoid": "P002", "costo": "15"},
                ],
                "header_row_number": 1,
                "source_row_numbers": [2, 3],
                "row_count": 2,
                "column_count": 2,
                "warnings": [],
                "blocking_errors": [],
                "runtime_authorized": False,
            },
        ],
        "runtime_authorized": False,
    }


def test_workbook_profiler_detects_cross_sheet_many_to_one_relationship() -> None:
    profile = build_service_1_workbook_profile_v1(ingestion_output=_ingestion_output())

    assert profile["status"] == STATUS_READY
    columns = {item["column_ref"]: item for item in profile["columns"]}
    assert columns["Productos.ProductoID"]["candidate_primary_key"] is True
    assert columns["Productos.ProductoID"]["uniqueness_class"] == "UNIQUE"
    assert columns["Ventas.ProductoID"]["uniqueness_class"] == "NON_UNIQUE"

    relationships = {
        (item["left_column_ref"], item["right_column_ref"]): item
        for item in profile["relationships"]
    }
    relation = relationships[("Ventas.ProductoID", "Productos.ProductoID")]
    assert ("Productos.ProductoID", "Ventas.ProductoID") not in relationships
    assert relation["relationship_kind"] == "MANY_TO_ONE"
    assert relation["candidate_foreign_key"] is True
    assert relation["left_value_coverage"] == 1.0
    assert relation["right_value_coverage"] == 1.0

    evidence = profile["evidence_registry"]
    assert "ev:column:Productos.ProductoID:uniqueness" in evidence
    assert "ev:relationship:Ventas.ProductoID->Productos.ProductoID:overlap" in evidence
    assert all(profile[flag] is False for flag in (
        "runtime_authorized",
        "tool_execution_authorized",
        "product_ready",
        "delivery_authorized",
        "diagnosis_generated",
    ))


def test_workbook_profiler_computes_null_cardinality_ranges_and_samples() -> None:
    output = _ingestion_output()
    output["normalized_tables"][0]["rows"] = [
        {"productoid": "P001", "cantidad": "1"},
        {"productoid": "P002", "cantidad": "2"},
        {"productoid": "", "cantidad": "3"},
        {"productoid": "P001", "cantidad": "4"},
    ]
    profile = build_service_1_workbook_profile_v1(ingestion_output=output)

    assert profile["status"] == STATUS_READY
    columns = {item["column_ref"]: item for item in profile["columns"]}
    product = columns["Ventas.ProductoID"]
    quantity = columns["Ventas.Cantidad"]
    assert product["row_count"] == 4
    assert product["non_null_count"] == 3
    assert product["null_count"] == 1
    assert product["null_ratio"] == 0.25
    assert product["cardinality"] == 2
    assert product["sample_values"] == ["P001", "P002"]
    assert quantity["inferred_type"] == "number"
    assert quantity["numeric_range"] == {"min": 1.0, "max": 4.0}


def test_workbook_profiler_fails_closed_without_canonical_tables() -> None:
    result = build_service_1_workbook_profile_v1(
        ingestion_output={"case_id": "case-1", "column_refs": [], "runtime_authorized": False}
    )
    assert result["status"] == STATUS_BLOCKED
    assert result["blocked_reason"] == BLOCK_SOURCE_TABLES_MISSING
    assert result["evidence_registry"] == {}


def test_workbook_profiler_rejects_authority_flags() -> None:
    output = _ingestion_output()
    output["runtime_authorized"] = True
    result = build_service_1_workbook_profile_v1(ingestion_output=output)
    assert result["status"] == STATUS_BLOCKED
    assert result["blocked_reason"] == BLOCK_INPUT_AUTHORITY_FORBIDDEN


def test_workbook_profiler_rejects_column_ref_not_present_in_canonical_table() -> None:
    output = _ingestion_output()
    output["column_refs"][0]["column_name"] = "Inventada"
    result = build_service_1_workbook_profile_v1(ingestion_output=output)
    assert result["status"] == STATUS_BLOCKED
    assert result["blocked_reason"] == BLOCK_COLUMN_REF_NOT_FOUND


def _cafeteria_xlsx_bytes() -> bytes:
    stream = BytesIO()
    workbook = Workbook()

    ventas = workbook.active
    ventas.title = "Ventas"
    ventas.append(["VentaID", "ProductoID", "Cantidad", "PrecioUnitario"])
    ventas.append(["V001", "P001", 1, 60])
    ventas.append(["V002", "P002", 2, 45])
    ventas.append(["V003", "P001", 1, 60])

    productos = workbook.create_sheet("Productos")
    productos.append(["ProductoID", "Producto", "Costo"])
    productos.append(["P001", "Latte", 28])
    productos.append(["P002", "Americano", 18])

    sucursales = workbook.create_sheet("Sucursales")
    sucursales.append(["SucursalID", "Sucursal"])
    sucursales.append(["S001", "Centro"])
    sucursales.append(["S002", "Norte"])

    workbook.save(stream)
    return stream.getvalue()


def test_workbook_profiler_vertical_from_real_canonical_xlsx_intake() -> None:
    intake = build_service_1_web_column_confirmation_intake_boundary_v1(
        uploaded_xlsx_bytes=_cafeteria_xlsx_bytes(),
        uploaded_filename="cafeteria_vertical.xlsx",
        include_all_sheets=True,
    )
    assert intake["status"] == "NEEDS_OWNER_CONFIRMATION"
    assert intake["sheet_names"] == ["Ventas", "Productos", "Sucursales"]
    assert len(intake["normalized_tables"]) == 3

    canonical = build_service_1_unconfirmed_canonical_ingestion_output_v1(
        owner_question_packet=intake
    )
    assert canonical["status"] == STATUS_UNCONFIRMED_READY
    ingestion_output = canonical["ingestion_output"]
    assert len(ingestion_output["normalized_tables"]) == 3

    profile = build_service_1_workbook_profile_v1(
        ingestion_output=ingestion_output
    )
    assert profile["status"] == STATUS_READY

    columns = {item["column_ref"]: item for item in profile["columns"]}
    assert columns["Productos.ProductoID"]["candidate_primary_key"] is True
    assert columns["Productos.ProductoID"]["uniqueness_class"] == "UNIQUE"
    assert columns["Ventas.ProductoID"]["candidate_primary_key"] is False

    relations = {
        (item["left_column_ref"], item["right_column_ref"]): item
        for item in profile["relationships"]
    }
    relation = relations[("Ventas.ProductoID", "Productos.ProductoID")]
    assert relation["relationship_kind"] == "MANY_TO_ONE"
    assert relation["candidate_foreign_key"] is True
    assert relation["left_value_coverage"] == 1.0
    assert relation["right_value_coverage"] == 1.0

    evidence = profile["evidence_registry"]
    assert "ev:column:Productos.ProductoID:uniqueness" in evidence
    assert "ev:relationship:Ventas.ProductoID->Productos.ProductoID:overlap" in evidence

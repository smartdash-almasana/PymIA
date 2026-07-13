from __future__ import annotations

import inspect
from pathlib import Path

from openpyxl import load_workbook

from pymia.smartpyme.service_1_accounting_contracts_v1 import (
    CAPABILITY_REF,
    build_service_1_accounting_contract_v1,
)
from pymia.smartpyme.service_1_xlsx_delivery_v1 import build_service_1_xlsx_delivery_v1


def _contract_input() -> dict[str, object]:
    return {
        "contract_ref": "bank_reconciliation_basic",
        "owner_requested_output": "contract_review_report",
        "source_files_required": ["extracto_banco", "archivo_contable"],
        "source_files_received": ["extracto_banco", "archivo_contable"],
        "required_fields": ["fecha", "importe", "referencia"],
        "received_fields": ["fecha", "importe", "referencia"],
    }


def test_returns_ready_for_review_with_supported_contract_and_complete_inputs() -> None:
    result = build_service_1_accounting_contract_v1(contract_input=_contract_input())

    assert result["status"] == "READY_FOR_REVIEW"
    assert result["family"] == "bank_reconciliation"
    assert result["runtime_authorized"] is False
    assert result["missing_sources"] == []
    assert result["missing_fields"] == []


def test_returns_missing_sources_when_required_source_is_missing() -> None:
    contract_input = _contract_input()
    contract_input["source_files_received"] = ["extracto_banco"]

    result = build_service_1_accounting_contract_v1(contract_input=contract_input)

    assert result["status"] == "MISSING_SOURCES"
    assert result["missing_sources"] == ["archivo_contable"]


def test_returns_missing_fields_when_required_field_is_missing() -> None:
    contract_input = _contract_input()
    contract_input["received_fields"] = ["fecha", "importe"]

    result = build_service_1_accounting_contract_v1(contract_input=contract_input)

    assert result["status"] == "MISSING_FIELDS"
    assert result["missing_fields"] == ["referencia"]


def test_returns_unsupported_contract_when_contract_ref_is_not_supported() -> None:
    contract_input = _contract_input()
    contract_input["contract_ref"] = "erp_auto_posting_basic"

    result = build_service_1_accounting_contract_v1(contract_input=contract_input)

    assert result["status"] == "UNSUPPORTED_CONTRACT"
    assert "allowlist mínima actual" in result["delivery_input"]["owner_summary"]


def test_returns_invalid_input_when_required_fields_is_not_list_of_strings() -> None:
    contract_input = _contract_input()
    contract_input["required_fields"] = {"fecha": True}

    result = build_service_1_accounting_contract_v1(contract_input=contract_input)  # type: ignore[arg-type]

    assert result["status"] == "INVALID_INPUT"
    assert result["next_allowed_action"] == "fix_invalid_contract_input"


def test_returns_invalid_input_when_received_fields_is_not_list_of_strings() -> None:
    contract_input = _contract_input()
    contract_input["received_fields"] = {"fecha": True}

    result = build_service_1_accounting_contract_v1(contract_input=contract_input)  # type: ignore[arg-type]

    assert result["status"] == "INVALID_INPUT"
    assert result["delivery_input"]["status"] == "INVALID_INPUT"


def test_output_is_compatible_with_service_1_xlsx_delivery_input_v1() -> None:
    result = build_service_1_accounting_contract_v1(contract_input=_contract_input())
    delivery_input = result["delivery_input"]

    assert delivery_input["service_name"] == "SERVICE_1"
    assert delivery_input["capability_ref"] == CAPABILITY_REF
    assert delivery_input["computed_results"]["next_allowed_action"] == "manual_accounting_review"


def test_delivery_input_integrates_with_generic_xlsx_delivery(tmp_path: Path) -> None:
    result = build_service_1_accounting_contract_v1(contract_input=_contract_input())
    output_path = tmp_path / "accounting_contract.xlsx"

    delivery = build_service_1_xlsx_delivery_v1(
        delivery_input=result["delivery_input"],
        output_path=output_path,
    )

    workbook = load_workbook(output_path)

    assert output_path.exists()
    assert delivery["capability_ref"] == CAPABILITY_REF
    assert workbook["Resumen"]["B3"].value == CAPABILITY_REF
    assert ("contract_ref", "bank_reconciliation_basic") in list(workbook["Datos usados"].iter_rows(values_only=True))


def test_module_has_no_io_openpyxl_or_forbidden_runtime_dependencies() -> None:
    import pymia.smartpyme.service_1_accounting_contracts_v1 as module

    source = inspect.getsource(module)

    assert "openpyxl" not in source
    assert "first_aid" not in source
    assert "exceland_runtime" not in source
    assert "vertical_pipeline" not in source
    assert "service_1_pipeline" not in source
    assert "fsm" not in source.lower()
    assert "llm" not in source.lower()
    assert "chatbot" not in source.lower()
    assert "open(" not in source
    assert ".save(" not in source
    assert "read_text(" not in source
    assert "read_bytes(" not in source


def test_forbidden_accounting_claims_are_present() -> None:
    result = build_service_1_accounting_contract_v1(contract_input=_contract_input())

    assert "No confirma conciliación bancaria cerrada." in result["forbidden_claims"]
    assert "No confirma auditoría contable certificada." in result["forbidden_claims"]
    assert "No genera asientos contables automáticos." in result["forbidden_claims"]

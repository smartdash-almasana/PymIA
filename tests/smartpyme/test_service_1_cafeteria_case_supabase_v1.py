from __future__ import annotations

import os
import re
import uuid
from io import BytesIO

import pytest
from openpyxl import Workbook

from pymia.smartpyme.service_1_assisted_web_v1 import AssistedWebApplicationV1
from pymia.smartpyme.service_1_supabase_persistence_v1 import (
    OWNER_CONFIRMATIONS_TABLE,
    SEMANTIC_CONTRACTS_TABLE,
    Service1SupabasePersistenceAdapterV1,
)

_REQUIRED_ENV = ("PYMIA_SUPABASE_URL", "PYMIA_SUPABASE_SERVICE_ROLE_KEY")


def _supabase_ready() -> bool:
    return all(str(os.environ.get(name) or "").strip() for name in _REQUIRED_ENV)


def _cafeteria_xlsx_bytes() -> bytes:
    stream = BytesIO()
    workbook = Workbook()
    ventas = workbook.active
    ventas.title = "Ventas"
    ventas.append(["VentaID", "Fecha", "Hora", "SucursalID", "ProductoID", "Cantidad", "PrecioUnitario", "MetodoPago", "CanalVenta", "Descuento", "Empleado"])
    ventas.append(["V0001", "2026-01-01", "07:15", "S001", "P008", 1, 60, "Tarjeta de Crédito", "Llevar", 0, "Carlos Pérez"])
    ventas.append(["V0002", "2026-01-01", "07:37", "S004", "P008", 1, 60, "Tarjeta de Débito", "Llevar", 0, "Fernanda Ruiz"])
    sucursales = workbook.create_sheet("Sucursales")
    sucursales.append(["SucursalID", "Sucursal", "Ciudad"])
    sucursales.append(["S001", "Centro", "Querétaro"])
    sucursales.append(["S004", "Roma Norte", "CDMX"])
    productos = workbook.create_sheet("Productos")
    productos.append(["ProductoID", "Producto", "Categoria", "Costo", "Precio"])
    productos.append(["P001", "Latte", "Café", 28, 65])
    productos.append(["P008", "Brownie", "Panadería", 25, 60])
    workbook.save(stream)
    return stream.getvalue()


def _answers(page: str) -> dict[str, str]:
    by_question: dict[str, list[str]] = {}
    for question_id, option_id in re.findall(r'name="answer_([^"]+)" value="([^"]+)"', page):
        by_question.setdefault(question_id, []).append(option_id)
    answers: dict[str, str] = {}
    for question_id, options in by_question.items():
        preferred = next((option for option in options if option not in {"OTHER", "IGNORE", "not_sure"}), None)
        answers[f"answer_{question_id}"] = preferred or "IGNORE"
    assert answers
    return answers


@pytest.mark.skipif(not _supabase_ready(), reason="Supabase semantic persistence environment is not configured")
def test_cafeteria_case_persists_semantics_and_stays_actionable_when_margin_evidence_is_incomplete(tmp_path) -> None:
    store = Service1SupabasePersistenceAdapterV1.from_environment()
    tenant_id = f"cafeteria-test-{uuid.uuid4()}"
    cliente_id = f"cafeteria-client-{uuid.uuid4()}"
    owner_id = f"cafeteria-owner-{uuid.uuid4()}"
    filename = "cafeteria_abc_case.xlsx"
    app = AssistedWebApplicationV1(
        output_dir=tmp_path / "outputs",
        persist_tenant_confirmation=store,
        load_tenant_memory=store.list_owner_confirmation_memory,
        load_prior_semantic_contract=store.load_current_semantic_contract,
        require_tenant_persistence=True,
    )
    app.bind_tenant_identity(
        session_id="cafeteria-session",
        tenant_id=tenant_id,
        cliente_id=cliente_id,
        owner_actor_id=owner_id,
        owner_actor_role="OWNER",
    )
    try:
        status, page = app.receive_xlsx(
            session_id="cafeteria-session",
            filename=filename,
            content=_cafeteria_xlsx_bytes(),
            selected_launch_review="net_margin_real",
        )
        assert status == 200
        assert "Esto entendí de tu Excel" in page
        status, page = app.confirm_meanings(session_id="cafeteria-session", fields=_answers(page))
        assert status == 200
        assert "FALTA INFORMACIÓN" in page
        assert "Caso guardado" in page
        state = app.session("cafeteria-session")
        case_id = str(
            (state.ingestion_output.get("workbook_context") or {}).get("case_id") or ""
        )
        assert case_id
        owner_rows = store._client.table(OWNER_CONFIRMATIONS_TABLE).select("confirmation_event_ref,tenant_id,case_id,workbook_ref").eq("tenant_id", tenant_id).eq("case_id", case_id).execute().data
        contract_rows = store._client.table(SEMANTIC_CONTRACTS_TABLE).select("contract_id,tenant_id,case_id,workbook_ref").eq("tenant_id", tenant_id).eq("case_id", case_id).execute().data
        assert owner_rows
        assert contract_rows
        assert all(row["workbook_ref"] == filename for row in owner_rows)
        assert all(row["workbook_ref"] == filename for row in contract_rows)
    finally:
        store._client.table(SEMANTIC_CONTRACTS_TABLE).delete().eq("tenant_id", tenant_id).execute()
        store._client.table(OWNER_CONFIRMATIONS_TABLE).delete().eq("tenant_id", tenant_id).execute()

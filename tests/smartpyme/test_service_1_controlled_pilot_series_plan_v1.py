from __future__ import annotations
import json
from pathlib import Path
import openpyxl

def _root() -> Path:
    return Path(__file__).resolve().parents[2]

def _plan() -> dict:
    return json.loads((_root()/"docs"/"service_1_controlled_pilot_series_plan.v1.json").read_text(encoding="utf-8"))

def test_controlled_pilot_series_is_active_and_excel_based() -> None:
    plan=_plan()
    assert plan["schema_version"]=="SERVICE_1_CONTROLLED_PILOT_SERIES_PLAN_V1"
    assert plan["status"]=="ACTIVE"
    assert plan["source_folder"]=="prueba_excels"
    assert len(plan["pilot_sequence"])==7
    assert plan["next_execution_order"]==["S1-PILOT-008", "S1-PILOT-005"]

def test_controlled_pilot_execution_status_is_current() -> None:
    status=_plan()["execution_status"]
    assert status=={
        "S1-PILOT-001": "PASS",
        "S1-PILOT-003": "PASS",
        "S1-PILOT-004": "PASS",
        "S1-PILOT-005": "PLANNED_AFTER_008",
        "S1-PILOT-006": "PASS",
        "S1-PILOT-007": "PASS",
        "S1-PILOT-008": "NEXT",
    }

def test_all_active_pilot_files_exist_and_have_primary_headers() -> None:
    root=_root()
    for pilot in _plan()["pilot_sequence"]:
        path=root/pilot["file"]
        assert path.exists(), pilot["file"]
        wb=openpyxl.load_workbook(path,read_only=True,data_only=True)
        assert pilot["primary_sheet"] in wb.sheetnames
        ws=wb[pilot["primary_sheet"]]
        headers=[str(c).strip() for c in next(ws.iter_rows(min_row=1,max_row=1,values_only=True)) if c is not None and str(c).strip()]
        assert headers, pilot["file"]
        for expected in pilot["observed_headers"]:
            assert expected in headers
        wb.close()

def test_quarantined_inputs_are_not_active_pilots() -> None:
    plan=_plan()
    active={p["file"] for p in plan["pilot_sequence"]}
    quarantined={q["file"] for q in plan["quarantined_inputs"]}
    assert active.isdisjoint(quarantined)
    assert "prueba_excels/cobros_marzo_2026.xlsx" in quarantined
    assert "prueba_excels/ventas_marzo_2026.xlsx" in quarantined
    assert "prueba_excels/simple_bem_test.xlsx" in quarantined

def test_plan_preserves_product_boundaries() -> None:
    rules="\n".join(_plan()["policy"]["rules"])
    assert "selección automática de tool" in rules
    assert "tool_requests explícitos" in rules
    assert "capacidades formulaicas" in rules

def test_current_readme_lists_controlled_pilot_series_doc() -> None:
    root=_root()
    readme=(root/"docs"/"current"/"README.md").read_text(encoding="utf-8")
    assert "SERVICE_1_CONTROLLED_PILOT_SERIES_PLAN.md" in readme
    assert (root/"docs"/"current"/"SERVICE_1_CONTROLLED_PILOT_SERIES_PLAN.md").exists()

def test_headers_do_not_grant_productive_pilot_authority() -> None:
    plan=_plan()
    guard=plan["policy"]["anti_drift_guard"]
    assert "columnas observadas" in guard["principle"] or "sheet names" in guard["principle"]
    disqualified={item["file"]: item for item in guard["known_disqualified_inputs"]}
    assert "prueba_excels/simple_bem_test.xlsx" in disqualified
    assert disqualified["prueba_excels/simple_bem_test.xlsx"]["trap"]=="perfil_superficial_de_columnas"

def test_disqualified_bem_fixture_cannot_be_next_or_active_even_with_headers() -> None:
    root=_root()
    plan=_plan()
    active={p["file"] for p in plan["pilot_sequence"]}
    next_ids=set(plan["next_execution_order"])
    assert "prueba_excels/simple_bem_test.xlsx" not in active
    assert "S1-PILOT-002" not in next_ids
    path=root/"prueba_excels/simple_bem_test.xlsx"
    wb=openpyxl.load_workbook(path,read_only=True,data_only=True)
    ws=wb["Sheet1"]
    headers=[str(c).strip() for c in next(ws.iter_rows(min_row=1,max_row=1,values_only=True)) if c is not None and str(c).strip()]
    wb.close()
    assert {"fecha","producto","cantidad","precio_unitario","venta_total"}.issubset(set(headers))
    quarantined={q["file"] for q in plan["quarantined_inputs"]}
    assert "prueba_excels/simple_bem_test.xlsx" in quarantined

def test_markdown_explains_bem_trap_and_forbids_superficial_promotion() -> None:
    root=_root()
    doc=(root/"docs"/"current"/"SERVICE_1_CONTROLLED_PILOT_SERIES_PLAN.md").read_text(encoding="utf-8")
    assert "Guarda anti-deriva" in doc
    assert "Tener headers válidos no autoriza un piloto" in doc
    assert "simple_bem_test.xlsx tiene columnas válidas" in doc
    assert "fixture BEM descartado" in doc

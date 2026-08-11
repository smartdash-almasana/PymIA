from __future__ import annotations

import os
import uuid
from http.client import HTTPConnection
from pathlib import Path
from threading import Thread
from urllib.parse import urlencode

import pytest
from openpyxl import Workbook

from pymia.smartpyme.service_1_assisted_web_v1 import create_assisted_web_server_v1
from pymia.smartpyme.service_1_radar_supabase_persistence_v1 import (
    RADAR_POLICIES_TABLE,
    Service1RadarSupabasePersistenceAdapterV1,
)
from pymia.smartpyme.service_1_supabase_identity_resolver_v1 import (
    Service1SupabaseIdentityResolverV1,
    create_service_1_supabase_identity_client_v1,
    load_service_1_supabase_identity_config_v1,
)

_TEST_EMAIL = "neoalmasana@gmail.com"
_REQUIRED_ENV = (
    "PYMIA_SUPABASE_URL",
    "PYMIA_SUPABASE_SERVICE_ROLE_KEY",
    "PYMIA_SUPABASE_PUBLISHABLE_KEY",
    "SUPABASE_TEST_PASSWORD",
)


def _supabase_ready() -> bool:
    return all(str(os.environ.get(name) or "").strip() for name in _REQUIRED_ENV)


def _xlsx_bytes(tmp_path: Path) -> bytes:
    path = tmp_path / "consorcio_summary.xlsx"
    workbook = Workbook()
    expensas = workbook.active
    expensas.title = "Expensas"
    expensas.append(["unidad", "saldo", "expensa"])
    expensas.append(["UF-12", 250, 100])
    gastos = workbook.create_sheet("Gastos")
    gastos.append(["rubro_gasto", "importe_gasto"])
    gastos.append(["Limpieza", 150])
    presupuesto = workbook.create_sheet("Presupuesto")
    presupuesto.append(["rubro_presupuesto", "presupuesto", "historico"])
    presupuesto.append(["Limpieza", 100, 100])
    workbook.save(path)
    return path.read_bytes()


def _multipart_with_context(
    filename: str, content: bytes, token: str, *, period: str, cookie: str = ""
) -> tuple[bytes, dict[str, str]]:
    boundary = "Service1CaseSummarySupabaseBoundary"
    chunks: list[bytes] = []
    for name, value in (
        ("consorcio_id", "rivadavia-1200"),
        ("consorcio_name", "Rivadavia 1200"),
        ("period", period),
    ):
        chunks.append(
            (
                f"--{boundary}\r\n"
                f'Content-Disposition: form-data; name="{name}"\r\n\r\n'
                f"{value}\r\n"
            ).encode("utf-8")
        )
    chunks.append(
        (
            f"--{boundary}\r\n"
            f'Content-Disposition: form-data; name="file"; filename="{filename}"\r\n'
            "Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n\r\n"
        ).encode("utf-8")
        + content
        + b"\r\n"
    )
    chunks.append(f"--{boundary}--\r\n".encode("utf-8"))
    body = b"".join(chunks)
    headers = {
        "Content-Type": f"multipart/form-data; boundary={boundary}",
        "Content-Length": str(len(body)),
        "Authorization": f"Bearer {token}",
    }
    if cookie:
        headers["Cookie"] = cookie
    return body, headers


def _post_form(server, path: str, cookie: str, token: str, fields: dict[str, str]):
    body = urlencode(fields).encode("utf-8")
    return _request(
        server,
        "POST",
        path,
        body=body,
        headers={
            "Content-Type": "application/x-www-form-urlencoded",
            "Content-Length": str(len(body)),
            "Cookie": cookie,
            "Authorization": f"Bearer {token}",
        },
    )


def _request(server, method: str, path: str, *, body: bytes = b"", headers=None):
    connection = HTTPConnection("127.0.0.1", server.server_port, timeout=20)
    connection.request(method, path, body=body, headers=headers or {})
    response = connection.getresponse()
    payload = response.read().decode("utf-8")
    return response.status, response.getheaders(), payload


def _cookie(headers: list[tuple[str, str]]) -> str:
    for name, value in headers:
        if name.lower() == "set-cookie":
            return value.split(";", 1)[0]
    raise AssertionError("missing session cookie")


def _assert_no_radar_severity_semantics(page: str) -> None:
    assert "HIGH" not in page
    assert "MODERATE" not in page
    assert "LOW" not in page
    assert "severity" not in page
    assert "risk" not in page


@pytest.mark.skipif(not _supabase_ready(), reason="Supabase RADAR environment is not configured")
def test_consorcios_case_summary_physical_supabase_http(tmp_path: Path) -> None:
    store = Service1RadarSupabasePersistenceAdapterV1.from_environment()
    identity_config = load_service_1_supabase_identity_config_v1()
    identity_client = create_service_1_supabase_identity_client_v1(identity_config)

    password = str(os.environ["SUPABASE_TEST_PASSWORD"]).strip()
    login = identity_client.auth.sign_in_with_password(
        {"email": _TEST_EMAIL, "password": password}
    )
    token = login.session.access_token
    assert token

    resolver = Service1SupabaseIdentityResolverV1(identity_client)
    identity = resolver.__call__(
        type("Handler", (), {"headers": {"Authorization": f"Bearer {token}"}})()
    )
    tenant_id = identity["tenant_id"]
    suffix = uuid.uuid4().hex[:12]
    aging_policy_ref = f"owner-radar-summary-aging-{suffix}"
    expense_policy_ref = f"owner-radar-summary-expense-{suffix}"

    server = create_assisted_web_server_v1(
        host="127.0.0.1",
        port=0,
        output_dir=tmp_path / "outputs-summary",
        tenant_identity_resolver=resolver,
        radar_policy_store=store,
    )
    thread = Thread(target=server.serve_forever, daemon=True)
    thread.start()
    try:
        upload_body, upload_headers = _multipart_with_context(
            "agosto.xlsx", _xlsx_bytes(tmp_path), token, period="2026-08"
        )
        status, headers, _ = _request(
            server, "POST", "/upload", body=upload_body, headers=upload_headers
        )
        assert status == 200
        cookie = _cookie(headers)

        status, _, saved_page = _post_form(
            server,
            "/save-radar-policy",
            cookie,
            token,
            {
                "policy_ref": aging_policy_ref,
                "observable_ref": "consorcios.debt_equivalent_periods",
                "enabled": "true",
                "operator": "GTE",
                "comparison_value": "2",
                "communication_level": "ALERT",
                "confirmed_by_owner": "true",
            },
        )
        assert status == 200
        assert "Regla RADAR guardada" in saved_page

        status, _, saved_page = _post_form(
            server,
            "/save-radar-policy",
            cookie,
            token,
            {
                "policy_ref": expense_policy_ref,
                "observable_ref": "consorcios.expense_budget_deviation_pct",
                "enabled": "true",
                "operator": "GT",
                "comparison_value": "40",
                "communication_level": "NOTIFICATION",
                "confirmed_by_owner": "true",
            },
        )
        assert status == 200
        assert "Regla RADAR guardada" in saved_page

        status, _, result = _post_form(
            server,
            "/run-consorcios-collection-aging",
            cookie,
            token,
            {
                "sheet_name": "Expensas",
                "unidad_funcional": "unidad",
                "saldo_anterior": "saldo",
                "expensa_mes": "expensa",
            },
        )
        assert status == 200
        assert "UF-12" in result
        assert "2.5" in result

        status, _, result = _post_form(
            server,
            "/run-consorcios-expense-variance",
            cookie,
            token,
            {
                "expense_sheet": "Gastos",
                "expense_rubro": "rubro_gasto",
                "expense_importe": "importe_gasto",
                "budget_sheet": "Presupuesto",
                "budget_rubro": "rubro_presupuesto",
                "presupuesto_mensual": "presupuesto",
                "promedio_historico": "historico",
            },
        )
        assert status == 200
        assert "Limpieza" in result
        assert "50" in result

        status, _, summary = _request(
            server,
            "GET",
            "/consorcios-case-summary",
            headers={"Cookie": cookie, "Authorization": f"Bearer {token}"},
        )
        assert status == 200
        assert "Resumen del período" in summary
        assert "Rivadavia 1200" in summary
        assert "2026-08" in summary
        assert "Cobranzas y deuda" in summary
        assert "Gastos" in summary
        assert "Banco" in summary
        assert "RADAR" in summary
        assert "Pendientes de revisión" in summary
        assert "Descargas" in summary
        assert "Realizado · 1 unidad(es) revisada(s)" in summary
        assert "Realizado · 1 rubro(s) revisado(s)" in summary
        assert "ALERT" in summary
        assert "NOTIFICATION" in summary
        assert "PymIA no asigna severidad" in summary
        assert "Realizado · 1 unidad(es) revisada(s)" in summary
        assert "Realizado · 1 rubro(s) revisado(s)" in summary
        _assert_no_radar_severity_semantics(summary)

        upload_body, upload_headers = _multipart_with_context(
            "septiembre.xlsx", _xlsx_bytes(tmp_path), token, period="2026-09", cookie=cookie
        )
        status, _, _ = _request(
            server, "POST", "/upload", body=upload_body, headers=upload_headers
        )
        assert status == 200

        status, _, new_summary = _request(
            server,
            "GET",
            "/consorcios-case-summary",
            headers={"Cookie": cookie, "Authorization": f"Bearer {token}"},
        )
        assert status == 200
        assert "2026-09" in new_summary
        assert new_summary.count("Pendiente") >= 3
        assert "Sin eventos RADAR" in new_summary
        assert "0 caso(s) bancario(s)" in new_summary
        assert "ALERT" not in new_summary
        assert "NOTIFICATION" not in new_summary
        _assert_no_radar_severity_semantics(new_summary)
    finally:
        server.shutdown()
        server.server_close()
        thread.join(timeout=5)

    for policy_ref in (aging_policy_ref, expense_policy_ref):
        store._client.table(RADAR_POLICIES_TABLE).delete().eq("tenant_id", tenant_id).eq(
            "policy_ref", policy_ref
        ).execute()
    for policy_ref in (aging_policy_ref, expense_policy_ref):
        residual = (
            store._client.table(RADAR_POLICIES_TABLE)
            .select("policy_ref")
            .eq("tenant_id", tenant_id)
            .eq("policy_ref", policy_ref)
            .execute()
        )
        assert residual.data == []

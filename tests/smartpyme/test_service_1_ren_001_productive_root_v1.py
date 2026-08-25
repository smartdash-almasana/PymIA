from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from pymia.smartpyme import service_1_product_pipeline_v1 as product
from pymia.smartpyme.service_1_ren_001_normalized_evidence_v1 import (
    STATUS_EVIDENCE_BLOCKED,
    evaluate_ren_001_from_normalized_tables_v1,
)
from pymia.smartpyme.service_1_ren_001_outcome_v1 import (
    STATUS_BLOCKED,
    STATUS_READY,
    build_ren_001_outcome_v1,
    deliver_ren_001_outcome_xlsx_v1,
)
from tests.smartpyme.service_1_p8_test_support import computable_decision_from_legacy_fixture


def _plan() -> dict[str, object]:
    governed = {
        "schema_version": "SERVICE_1_GOVERNED_COMPUTATION_INPUT_V1",
        "case_id": "case_ren_001_productive_root",
        "requested_capability": "net_margin_real",
        "family_id": "SALES_MARGIN",
        "pathology_code": "REN_001",
        "formula_id": "REN_001_margen_neto_real",
        "formula_expression": "sale_price - costs - taxes",
        "required_variables": ["sale_price", "costs", "taxes"],
        "required_evidence": [],
        "source_bindings": {"sale_price": "ventas", "costs": "costos", "taxes": "impuestos"},
        "grain": {"structural_scope": "REGION", "business_entity_grain": "NONE", "temporal_grain": "NONE", "aggregation_grain": "ATOMIC"},
        "catalog_versions": {},
        "provenance": {"source": "TEST_P8"},
        "runtime_authorized": False,
        "tool_execution_authorized": False,
        "product_ready": False,
        "delivery_authorized": False,
        "diagnosis_generated": False,
    }
    return {
        "schema_version": "SERVICE_1_COMPUTATION_PLAN_V1",
        "status": "READY_FOR_COMPUTATION",
        "requested_capability": "net_margin_real",
        "pathology_code": "REN_001",
        "formula_id": "REN_001_margen_neto_real",
        "required_variables": ["sale_price", "costs", "taxes"],
        "source_bindings": {"sale_price": "ventas", "costs": "costos", "taxes": "impuestos"},
        "governed_computation_input": governed,
        "computation_candidate_ready": True,
        "runtime_authorized": False,
        "tool_execution_authorized": False,
        "product_ready": False,
        "delivery_authorized": False,
        "diagnosis_generated": False,
    }


def _evidence() -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    tables = [
        {
            "sheet_name": "ventas",
            "rows": [
                {"ventas": 1000, "costos": 600, "impuestos": 100},
                {"ventas": 500, "costos": 350, "impuestos": 50},
            ],
        }
    ]
    refs = [
        {"sheet_name": "ventas", "column_name": "ventas", "normalized_column_name": "ventas"},
        {"sheet_name": "ventas", "column_name": "costos", "normalized_column_name": "costos"},
        {"sheet_name": "ventas", "column_name": "impuestos", "normalized_column_name": "impuestos"},
    ]
    return tables, refs


def test_ren_001_aggregates_confirmed_normalized_rows() -> None:
    tables, refs = _evidence()
    result = evaluate_ren_001_from_normalized_tables_v1(
        computation_plan=_plan(), normalized_tables=tables, column_refs=refs
    )

    assert result["status"] == "EVALUATED"
    assert result["classification"] == "POSITIVE_MARGIN"
    assert result["inputs"] == {"sale_price": 1500.0, "costs": 950.0, "taxes": 150.0}
    assert result["computed"]["net_margin_amount"] == 400.0
    assert result["aggregation"]["sample_based"] is False


def test_ren_001_blocks_ambiguous_column_resolution() -> None:
    tables, refs = _evidence()
    refs.append(dict(refs[0]))
    result = evaluate_ren_001_from_normalized_tables_v1(
        computation_plan=_plan(), normalized_tables=tables, column_refs=refs
    )

    assert result["status"] == STATUS_EVIDENCE_BLOCKED
    assert result["diagnosis_generated"] is False
    assert any("must resolve exactly once" in error for error in result["errors"])


def test_ren_001_outcome_is_bounded_and_non_causal() -> None:
    tables, refs = _evidence()
    evaluation = evaluate_ren_001_from_normalized_tables_v1(
        computation_plan=_plan(), normalized_tables=tables, column_refs=refs
    )
    outcome = build_ren_001_outcome_v1(computation_result=evaluation)

    assert outcome["status"] == STATUS_READY
    assert outcome["bounded_finding_generated"] is True
    assert outcome["causal_diagnosis_generated"] is False
    assert outcome["runtime_authorized"] is False
    assert outcome["delivery_authorized"] is False
    assert outcome["treatment_actions"]
    assert outcome["forbidden_claims"]




def test_ren_001_delivery_writes_deterministic_seven_sheet_workbook(tmp_path: Path) -> None:
    tables, refs = _evidence()
    evaluation = evaluate_ren_001_from_normalized_tables_v1(
        computation_plan=_plan(), normalized_tables=tables, column_refs=refs
    )
    outcome = build_ren_001_outcome_v1(computation_result=evaluation)

    delivered = deliver_ren_001_outcome_xlsx_v1(outcome=outcome, output_dir=tmp_path)

    assert delivered["status"] == "DELIVERED"
    assert delivered["causal_diagnosis_generated"] is False
    assert delivered["runtime_authorized"] is False
    assert delivered["delivery_authorized"] is False
    output = Path(delivered["delivery"]["output_path"])
    assert output.exists()
    workbook = load_workbook(output, data_only=False)
    assert workbook.sheetnames == [
        "Resumen",
        "Datos usados",
        "Resultados",
        "Faltantes",
        "Limitaciones",
        "Claims prohibidos",
        "Notas técnicas",
    ]
    summary = {row[0].value: row[1].value for row in workbook["Resumen"].iter_rows(min_row=2)}
    assert summary["status"] == "POSITIVE_MARGIN"
    assert "margen neto real positivo" in summary["owner_summary"].lower()
    inputs = {row[0].value: row[1].value for row in workbook["Datos usados"].iter_rows(min_row=2)}
    assert inputs == {"sale_price": "1500.0", "costs": "950.0", "taxes": "150.0"}
    results = {row[0].value: row[1].value for row in workbook["Resultados"].iter_rows(min_row=2)}
    assert results["net_margin_amount"] == "400.0"
    assert "Conservar el cálculo" in results["treatment_actions"]
    limitations = [row[0].value for row in workbook["Limitaciones"].iter_rows(min_row=2)]
    claims = [row[0].value for row in workbook["Claims prohibidos"].iter_rows(min_row=2)]
    assert outcome["limitations"][0] in limitations
    assert outcome["forbidden_claims"][0] in claims


def test_ren_001_delivery_blocks_non_ready_outcome(tmp_path: Path) -> None:
    delivered = deliver_ren_001_outcome_xlsx_v1(
        outcome={"status": "OUTCOME_BLOCKED"},
        output_dir=tmp_path,
    )

    assert delivered["status"] == STATUS_BLOCKED
    assert delivered["bounded_finding_generated"] is False
    assert delivered["causal_diagnosis_generated"] is False
    assert not list(tmp_path.glob("*.xlsx"))

from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import load_workbook

from pymia.smartpyme.service_1_reconciliation_request_gate_v1 import BANK_RECONCILIATION
from pymia.smartpyme.service_1_reconciliation_workpaper_xlsx_v1 import (
    build_service_1_reconciliation_workpaper_xlsx_v1,
)


def _packet() -> dict[str, object]:
    return {
        "status": "RECONCILIATION_REVIEW_READY",
        "reconciliation_run": {
            "schema_version": "SERVICE_1_RECONCILIATION_PRODUCT_REQUEST_V1",
            "case_id": "CASE-WORKPAPER-1",
            "reconciliation_type": BANK_RECONCILIATION,
            "assisted_review": {
                "schema_version": "SERVICE_1_RECONCILIATION_ASSISTED_REVIEW_V1",
                "reconciler_ref": "bank_reconciliation_contract_v1",
                "provenance": {"source": "test"},
                "review_result": {
                    "exact_matches_summary": {
                        "items": [{"banco_id": "B-1", "interno_id": "C-1"}]
                    },
                    "probable_matches_summary": {"items": []},
                    "ambiguous_matches_summary": {"items": []},
                    "amount_differences_summary": {"items": []},
                    "date_differences_summary": {"items": []},
                    "bank_pending_summary": {
                        "items": [{"banco_id": "B-2", "importe": 500}]
                    },
                    "internal_pending_summary": {"items": []},
                    "missing_evidence_summary": {"items": []},
                },
            },
        },
    }


def _decision() -> dict[str, object]:
    return {
        "decision_id": "DEC-1",
        "case_id": "CASE-WORKPAPER-1",
        "reconciliation_type": BANK_RECONCILIATION,
        "review_item_ref": "exact:1",
        "review_category": "exact",
        "decision": "CONFIRM",
        "reviewed_by": "María Administración",
        "decided_at": "2026-07-30T15:00:00+00:00",
        "observation": "Comprobante verificado",
        "review_item_snapshot": {"banco_id": "B-1", "interno_id": "C-1"},
        "human_decision": True,
        "source_data_modified": False,
    }


def test_builds_downloadable_workpaper_with_decisions_and_pending_cases() -> None:
    result = build_service_1_reconciliation_workpaper_xlsx_v1(
        reconciliation_packet=_packet(),
        human_decisions=[_decision()],
    )

    assert result["case_count"] == 2
    assert result["confirmed_count"] == 1
    assert result["rejected_count"] == 0
    assert result["pending_count"] == 1
    assert result["source_data_modified"] is False
    assert result["accounting_closure_authorized"] is False
    assert result["runtime_authorized"] is False
    assert result["llm_used"] is False

    workbook = load_workbook(BytesIO(result["content"]), data_only=True)
    assert workbook.sheetnames == [
        "Resumen",
        "Casos",
        "Decisiones",
        "Pendientes",
        "Trazabilidad",
        "Limites",
    ]

    cases = list(workbook["Casos"].values)
    assert any("CONFIRM" in row for row in cases[1:])
    assert any("PENDING_REVIEW" in row for row in cases[1:])

    decisions = list(workbook["Decisiones"].values)
    assert any("María Administración" in row for row in decisions[1:])

    pending = list(workbook["Pendientes"].values)
    assert len(pending) == 2
    assert "bank_pending:1" in pending[1]

    limits = [row[0] for row in list(workbook["Limites"].values)[1:]]
    assert any("no es certificación" in str(value) for value in limits)
    assert any("No modifica" in str(value) for value in limits)


def test_rejects_decision_from_another_case() -> None:
    decision = _decision()
    decision["case_id"] = "OTHER-CASE"

    with pytest.raises(ValueError, match="another case"):
        build_service_1_reconciliation_workpaper_xlsx_v1(
            reconciliation_packet=_packet(),
            human_decisions=[decision],
        )

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from pymia.smartpyme.service_1_synthetic_real_case_pilot_v1 import (
    SERVICE_NAME,
    SYNTHETIC_REAL_CASE_ID,
    run_service_1_synthetic_real_case_pilot_v1,
)


def test_synthetic_real_case_pilot_runs_full_safe_delivery_chain(tmp_path: Path) -> None:
    result = run_service_1_synthetic_real_case_pilot_v1(tmp_path)

    assert result["schema_version"] == "1.0"
    assert result["service_name"] == SERVICE_NAME
    assert result["case_id"] == SYNTHETIC_REAL_CASE_ID
    assert result["case_type"] == "synthetic_real_case_pilot"
    assert result["synthetic_data"] is True
    assert result["real_client_data"] is False
    assert result["runtime_authorized"] is False
    assert result["final_delivery_allowed"] is True


def test_synthetic_real_case_pilot_activation_manifest_audit_and_release_gate_are_coherent(tmp_path: Path) -> None:
    result = run_service_1_synthetic_real_case_pilot_v1(tmp_path)

    assert result["activation"]["status"] == "ACTIVATION_ALLOWED"
    assert result["activation"]["activation_allowed"] is True
    assert result["case_manifest"]["status"] == "READY_FOR_QA"
    assert result["case_manifest"]["delivery_allowed"] is True
    assert result["delivery_audit"]["status"] == "PASS_WITH_WARNINGS_REQUIRES_HUMAN_REVIEW"
    assert result["delivery_audit"]["delivery_allowed"] is True
    assert result["owner_release_action_gate"]["status"] == "READY_FOR_OPERATIONAL_DRAFT_DELIVERY"
    assert result["owner_release_action_gate"]["delivery_allowed"] is True


def test_synthetic_real_case_pilot_generates_owner_delivery_package(tmp_path: Path) -> None:
    result = run_service_1_synthetic_real_case_pilot_v1(tmp_path)
    package = result["delivery_package"]

    assert package["runtime_authorized"] is False
    assert package["file_count"] >= 8
    assert Path(package["package_dir"]).exists()
    assert Path(package["readme_path"]).exists()
    assert Path(package["manifest_path"]).exists()

    manifest = json.loads(Path(package["manifest_path"]).read_text(encoding="utf-8"))
    assert manifest["service_name"] == "SERVICE_1"
    assert manifest["runtime_authorized"] is False
    assert "No confirma conciliacion cerrada." in manifest["limitations"]


def test_synthetic_real_case_pilot_xlsx_outputs_are_readable_and_limited(tmp_path: Path) -> None:
    result = run_service_1_synthetic_real_case_pilot_v1(tmp_path)
    generated_files = result["harness_run"]["generated_files"]

    assert len(generated_files) == 5
    for generated_file in generated_files:
        workbook = load_workbook(generated_file)
        assert workbook["Resumen"]["B2"].value == "SERVICE_1"
        assert "Resultados" in workbook.sheetnames
        assert "Limitaciones" in workbook.sheetnames
        assert "Claims prohibidos" in workbook.sheetnames


def test_synthetic_real_case_pilot_summary_keeps_owner_claims_conservative(tmp_path: Path) -> None:
    result = run_service_1_synthetic_real_case_pilot_v1(tmp_path)
    summary_text = result["harness_run"]["summary_text"]

    assert "Entrega preliminar basada en datos declarados." in summary_text
    assert "No es un diagnostico integral" in summary_text
    assert "No confirma saldo bancario real" in summary_text
    assert "No confirma stock fisico real" in summary_text


def test_synthetic_real_case_pilot_readme_keeps_delivery_claims_conservative(tmp_path: Path) -> None:
    result = run_service_1_synthetic_real_case_pilot_v1(tmp_path)
    readme = Path(result["delivery_package"]["readme_path"]).read_text(encoding="utf-8")

    assert "Entrega preliminar basada en datos declarados." in readme
    assert "No es un diagnostico integral" in readme
    assert "No confirma rentabilidad real." in readme
    assert "No confirma conciliacion cerrada." in readme


def test_synthetic_real_case_pilot_rejects_missing_output_root() -> None:
    missing_dir = Path("/nonexistent/service1/synthetic-real-case")

    try:
        run_service_1_synthetic_real_case_pilot_v1(missing_dir)
    except FileNotFoundError as exc:
        assert "Output root does not exist" in str(exc)
    else:
        raise AssertionError("Expected FileNotFoundError")


def test_synthetic_real_case_pilot_source_does_not_open_forbidden_layers() -> None:
    module_path = Path(__file__).parents[2] / "pymia" / "smartpyme" / "service_1_synthetic_real_case_pilot_v1.py"
    source = module_path.read_text(encoding="utf-8")

    forbidden_fragments = (
        "openai",
        "langchain",
        "requests",
        "httpx",
        "document_ingestion",
        "vertical_pipeline",
        "servicio_2",
        "chatbot",
        "ocr",
        "parser",
    )
    for fragment in forbidden_fragments:
        assert fragment not in source.lower()

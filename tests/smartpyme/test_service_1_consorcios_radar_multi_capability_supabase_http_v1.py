from __future__ import annotations

import os
import uuid
from http.client import HTTPConnection
from pathlib import Path
from threading import Thread
from urllib.parse import urlencode

import pytest
from openpyxl import Workbook

from pymia.smartpyme.service_1_assisted_web_v1 import create_assisted_web_server_v1
from pymia.smartpyme.service_1_radar_observation_policy_v1 import (
    RadarObservationPolicyV1,
)
from pymia.smartpyme.service_1_radar_supabase_persistence_v1 import (
    RADAR_POLICIES_TABLE,
    Service1RadarSupabasePersistenceAdapterV1,
)
from pymia.smartpyme.service_1_supabase_identity_resolver_v1 import (
    Service1SupabaseIdentityResolverV1,
    create_service_1_supabase_identity_client_v1,
    load_service_1_supabase_identity_config_v1,
)

_TEST_EMAIL = "neoalmasana@gmail.com"
_REQUIRED_ENV = (
    "PYMIA_SUPABASE_URL",
    "PYMIA_SUPABASE_SERVICE_ROLE_KEY",
    "PYMIA_SUPABASE_PUBLISHABLE_KEY",
    "SUPABASE_TEST_PASSWORD",
)


def _supabase_ready() -> bool:
    return all(str(os.environ.get(name) or "").strip() for name in _REQUIRED_ENV)


def _consorcios_multi_xlsx_bytes(tmp_path: Path) -> bytes:
    path = tmp_path / "consorcio_multi.xlsx"
    workbook = Workbook()
    aging = workbook.active
    aging.title = "Expensas"
    aging.append(["unidad_funcional", "saldo_anterior", "expensa_mes"])
    aging.append(["UF-12", 250, 100])
    expenses = workbook.create_sheet("Gastos")
    expenses.append(["rubro", "importe"])
    expenses.append(["Limpieza", 150])
    budget = workbook.create_sheet("Presupuesto")
    budget.append(["rubro", "presupuesto_mensual", "promedio_historico"])
    budget.append(["Limpieza", 100, 100])
    workbook.save(path)
    return path.read_bytes()


def _multipart(filename: str, content: bytes, token: str) -> tuple[bytes, dict[str, str]]:
    boundary = "Service1RadarMultiSupabaseBoundary"
    body = (
        f"--{boundary}\r\n"
        f'Content-Disposition: form-data; name="file"; filename="{filename}"\r\n'
        "Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n\r\n"
    ).encode("utf-8") + content + f"\r\n--{boundary}--\r\n".encode("utf-8")
    return body, {
        "Content-Type": f"multipart/form-data; boundary={boundary}",
        "Content-Length": str(len(body)),
        "Authorization": f"Bearer {token}",
    }


def _post_form(
    server, path: str, values: dict[str, str], cookie: str, token: str
):
    body = urlencode(values).encode("utf-8")
    return _request(
        server,
        "POST",
        path,
        body=body,
        headers={
            "Content-Type": "application/x-www-form-urlencoded",
            "Content-Length": str(len(body)),
            "Cookie": cookie,
            "Authorization": f"Bearer {token}",
        },
    )


def _request(server, method: str, path: str, *, body: bytes = b"", headers=None):
    connection = HTTPConnection("127.0.0.1", server.server_port, timeout=20)
    connection.request(method, path, body=body, headers=headers or {})
    response = connection.getresponse()
    payload = response.read().decode("utf-8")
    return response.status, response.getheaders(), payload


def _cookie(headers: list[tuple[str, str]]) -> str:
    for name, value in headers:
        if name.lower() == "set-cookie":
            return value.split(";", 1)[0]
    raise AssertionError("missing session cookie")


def _assert_no_radar_severity_semantics(page: str) -> None:
    assert "HIGH" not in page
    assert "MODERATE" not in page
    assert "LOW" not in page
    assert "severity" not in page
    assert "risk" not in page


@pytest.mark.skipif(not _supabase_ready(), reason="Supabase RADAR environment is not configured")
def test_consorcios_radar_multi_capability_physical_supabase_http_roundtrip(
    tmp_path: Path,
) -> None:
    store = Service1RadarSupabasePersistenceAdapterV1.from_environment()
    identity_config = load_service_1_supabase_identity_config_v1()
    identity_client = create_service_1_supabase_identity_client_v1(identity_config)

    password = str(os.environ["SUPABASE_TEST_PASSWORD"]).strip()
    login = identity_client.auth.sign_in_with_password(
        {"email": _TEST_EMAIL, "password": password}
    )
    token = login.session.access_token
    assert token

    resolver = Service1SupabaseIdentityResolverV1(identity_client)
    identity = resolver.__call__(
        type("Handler", (), {"headers": {"Authorization": f"Bearer {token}"}})()
    )
    tenant_id = identity["tenant_id"]
    suffix = uuid.uuid4().hex[:12]
    aging_policy_ref = f"owner-radar-aging-{suffix}"
    expense_policy_ref = f"owner-radar-expense-{suffix}"
    isolation_tenant = f"tenant_radar_isolation_{suffix}"
    isolation_policy_ref = f"owner-radar-isolation-{suffix}"

    server = create_assisted_web_server_v1(
        host="127.0.0.1",
        port=0,
        output_dir=tmp_path / "outputs-multi",
        tenant_identity_resolver=resolver,
        radar_policy_store=store,
    )
    thread = Thread(target=server.serve_forever, daemon=True)
    thread.start()
    try:
        upload_body, upload_headers = _multipart(
            "consorcio_multi.xlsx", _consorcios_multi_xlsx_bytes(tmp_path), token
        )
        status, headers, page = _request(
            server,
            "POST",
            "/upload",
            body=upload_body,
            headers=upload_headers,
        )
        assert status == 200
        assert "Confirmar qué significa cada dato" in page or "¿Qué querés revisar?" in page
        cookie = _cookie(headers)

        status, _, radar_page = _request(
            server,
            "GET",
            "/consorcios-radar-analysis",
            headers={"Cookie": cookie, "Authorization": f"Bearer {token}"},
        )
        assert status == 200
        assert "Analizar Consorcio con RADAR" in radar_page
        assert "Antigüedad de deuda" in radar_page
        assert "Gastos contra presupuesto e histórico" in radar_page

        status, _, saved_page = _post_form(
            server,
            "/save-radar-policy",
            {
                "policy_ref": aging_policy_ref,
                "observable_ref": "consorcios.debt_equivalent_periods",
                "enabled": "true",
                "operator": "GTE",
                "comparison_value": "2",
                "communication_level": "ALERT",
                "confirmed_by_owner": "true",
            },
            cookie,
            token,
        )
        assert status == 200
        assert "Regla RADAR guardada" in saved_page

        status, _, saved_page = _post_form(
            server,
            "/save-radar-policy",
            {
                "policy_ref": expense_policy_ref,
                "observable_ref": "consorcios.expense_budget_deviation_pct",
                "enabled": "true",
                "operator": "GT",
                "comparison_value": "40",
                "communication_level": "NOTIFICATION",
                "confirmed_by_owner": "true",
            },
            cookie,
            token,
        )
        assert status == 200
        assert "Regla RADAR guardada" in saved_page

        aging_form = urlencode(
            {
                "sheet_name": "Expensas",
                "unidad_funcional": "unidad_funcional",
                "saldo_anterior": "saldo_anterior",
                "expensa_mes": "expensa_mes",
            }
        ).encode("utf-8")
        status, _, aging_page = _request(
            server,
            "POST",
            "/run-consorcios-collection-aging",
            body=aging_form,
            headers={
                "Content-Type": "application/x-www-form-urlencoded",
                "Content-Length": str(len(aging_form)),
                "Cookie": cookie,
                "Authorization": f"Bearer {token}",
            },
        )
        assert status == 200
        assert "Antigüedad de deuda" in aging_page
        assert "UF-12" in aging_page
        assert "2.5" in aging_page
        assert "RADAR" in aging_page
        assert "ALERT" in aging_page
        assert "consorcios.debt_equivalent_periods" in aging_page
        assert "GTE 2" in aging_page
        _assert_no_radar_severity_semantics(aging_page)

        expense_form = urlencode(
            {
                "expense_sheet": "Gastos",
                "expense_rubro": "rubro",
                "expense_importe": "importe",
                "budget_sheet": "Presupuesto",
                "budget_rubro": "rubro",
                "presupuesto_mensual": "presupuesto_mensual",
                "promedio_historico": "promedio_historico",
            }
        ).encode("utf-8")
        status, _, expense_page = _request(
            server,
            "POST",
            "/run-consorcios-expense-variance",
            body=expense_form,
            headers={
                "Content-Type": "application/x-www-form-urlencoded",
                "Content-Length": str(len(expense_form)),
                "Cookie": cookie,
                "Authorization": f"Bearer {token}",
            },
        )
        assert status == 200
        assert "Gastos del consorcio" in expense_page
        assert "Limpieza" in expense_page
        assert "50.0" in expense_page
        assert "RADAR" in expense_page
        assert "NOTIFICATION" in expense_page
        assert "consorcios.expense_budget_deviation_pct" in expense_page
        assert "GT 40" in expense_page
        _assert_no_radar_severity_semantics(expense_page)
    finally:
        server.shutdown()
        server.server_close()
        thread.join(timeout=5)

    loaded_aging = store.load_policy(tenant_id=tenant_id, policy_ref=aging_policy_ref)
    assert loaded_aging is not None
    assert loaded_aging.tenant_id == tenant_id
    assert loaded_aging.observable_ref == "consorcios.debt_equivalent_periods"
    assert loaded_aging.operator == "GTE"
    assert loaded_aging.comparison_value == "2"
    assert loaded_aging.communication_level == "ALERT"
    assert loaded_aging.confirmed_by_owner is True

    loaded_expense = store.load_policy(tenant_id=tenant_id, policy_ref=expense_policy_ref)
    assert loaded_expense is not None
    assert loaded_expense.tenant_id == tenant_id
    assert loaded_expense.observable_ref == "consorcios.expense_budget_deviation_pct"
    assert loaded_expense.operator == "GT"
    assert loaded_expense.comparison_value == "40"
    assert loaded_expense.communication_level == "NOTIFICATION"
    assert loaded_expense.confirmed_by_owner is True

    isolation_policy = RadarObservationPolicyV1(
        tenant_id=isolation_tenant,
        policy_ref=isolation_policy_ref,
        observable_ref="consorcios.debt_equivalent_periods",
        enabled=True,
        operator="GTE",
        comparison_value="2",
        communication_level="ALERT",
        confirmed_by_owner=True,
    )
    assert store.save_policy(isolation_policy)
    try:
        assert (
            store.load_policy(
                tenant_id=tenant_id, policy_ref=isolation_policy_ref
            )
            is None
        )
        tenant_policies = store.list_policies(tenant_id=tenant_id)
        refs = {policy.policy_ref for policy in tenant_policies}
        assert isolation_policy_ref not in refs
        owned = store.load_policy(
            tenant_id=isolation_tenant, policy_ref=isolation_policy_ref
        )
        assert owned is not None
        assert owned.tenant_id == isolation_tenant
    finally:
        store._client.table(RADAR_POLICIES_TABLE).delete().eq(
            "tenant_id", isolation_tenant
        ).eq("policy_ref", isolation_policy_ref).execute()

    for policy_ref in (aging_policy_ref, expense_policy_ref):
        store._client.table(RADAR_POLICIES_TABLE).delete().eq("tenant_id", tenant_id).eq(
            "policy_ref", policy_ref
        ).execute()

    residual_refs = [aging_policy_ref, expense_policy_ref]
    for policy_ref in residual_refs:
        residual = (
            store._client.table(RADAR_POLICIES_TABLE)
            .select("policy_ref")
            .eq("tenant_id", tenant_id)
            .eq("policy_ref", policy_ref)
            .execute()
        )
        assert residual.data == []
    isolation_residual = (
        store._client.table(RADAR_POLICIES_TABLE)
        .select("policy_ref")
        .eq("tenant_id", isolation_tenant)
        .eq("policy_ref", isolation_policy_ref)
        .execute()
    )
    assert isolation_residual.data == []

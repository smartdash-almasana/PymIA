from __future__ import annotations

import html
import re
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook

from pymia.smartpyme.service_1_assisted_web_v1 import AssistedWebApplicationV1
from pymia.smartpyme.service_1_deterministic_semantic_proposal_provider_v1 import (
    build_service_1_deterministic_semantic_proposal_v1,
)
from pymia.smartpyme.service_1_owner_unit_confirmation_event_v1 import UNIT_DISCOUNT_FRACTION


def _cafeteria_xlsx_bytes(*, include_taxes: bool = False) -> bytes:
    stream = BytesIO()
    workbook = Workbook()

    ventas = workbook.active
    ventas.title = "Ventas"
    ventas.append([
        "VentaID", "Fecha", "Hora", "SucursalID", "ProductoID", "Cantidad",
        "PrecioUnitario", "MetodoPago", "CanalVenta", "Descuento", "Empleado",
    ])
    ventas.append([
        "V0001", "2026-01-01", "07:15:44", "S001", "P008", 1,
        60, "Tarjeta de Crédito", "Llevar", 0, "Carlos Pérez",
    ])
    ventas.append([
        "V0002", "2026-01-01", "07:37:24", "S004", "P008", 2,
        60, "Tarjeta de Débito", "Llevar", 0.1, "Fernanda Ruiz",
    ])

    sucursales = workbook.create_sheet("Sucursales")
    sucursales.append(["SucursalID", "Sucursal", "Ciudad"])
    sucursales.append(["S001", "Centro", "Querétaro"])
    sucursales.append(["S004", "Roma Norte", "CDMX"])

    productos = workbook.create_sheet("Productos")
    productos.append(["ProductoID", "Producto", "Categoria", "Costo", "Precio"])
    productos.append(["P008", "Latte", "Café", 28, 60])
    productos.append(["P013", "Americano", "Café", 18, 45])

    if include_taxes:
        resumen = workbook.create_sheet("Resumen")
        resumen.append(["impuestos_periodo"])
        resumen.append([20])

    workbook.save(stream)
    return stream.getvalue()


def _answers(page: str) -> dict[str, str]:
    answers: dict[str, str] = {}
    for question_id, option_id in re.findall(
        r'name="answer_([^"]+)" value="([^"]+)"', page
    ):
        if option_id not in {"OTHER", "IGNORE", "not_sure"}:
            answers.setdefault(f"answer_{question_id}", option_id)
    if answers:
        return answers
    for decision_id in re.findall(r'name="action_([^"]+)" value="ACCEPT"', page):
        clean = html.unescape(decision_id)
        answers[f"action_{clean}"] = "ACCEPT"
    assert answers
    return answers


def _unit_answers(page: str) -> dict[str, str]:
    question_ids = [
        html.unescape(item)
        for item in re.findall(r'name="unit_([^"]+)" value="DISCOUNT_FRACTION_0_1"', page)
    ]
    assert question_ids
    return {f"unit_{question_id}": UNIT_DISCOUNT_FRACTION for question_id in question_ids}


def test_cafeteria_margin_asks_only_relevant_columns_and_keeps_case_actionable(tmp_path: Path) -> None:
    persisted = []

    def persist(event, contract):
        persisted.append((event, contract))
        return True

    app = AssistedWebApplicationV1(
        output_dir=tmp_path / "outputs",
        persist_tenant_confirmation=persist,
        require_tenant_persistence=True,
    )
    app.bind_tenant_identity(
        session_id="cafeteria-session",
        tenant_id="cafeteria-abc",
        cliente_id="cafeteria-abc",
        owner_actor_id="owner-cafeteria",
        owner_actor_role="OWNER",
    )

    status, page = app.receive_xlsx(
        session_id="cafeteria-session",
        filename="cafeteria_abc.xlsx",
        content=_cafeteria_xlsx_bytes(),
        selected_launch_review="net_margin_real",
    )

    assert status == 200
    assert "Revisión del archivo" in page
    assert "Sí, está bien" in page
    assert 1 <= page.count('class="understanding-card semantic-transaction"') < 10

    for relevant in ("ProductoID", "Cantidad", "PrecioUnitario", "Descuento", "Costo"):
        assert relevant in page
    for irrelevant in ("Hora", "SucursalID", "MetodoPago", "CanalVenta", "Empleado", "Ciudad"):
        assert f">{irrelevant}<" not in page

    status, unit_page = app.confirm_meanings(
        session_id="cafeteria-session",
        fields=_answers(page),
    )

    assert status == 200
    assert "Esto entendí de tu Excel" in unit_page
    assert "Confirmemos el descuento" in unit_page
    assert "0,10 significa 10%" in unit_page
    assert "Valores que encontré en esta columna" in unit_page
    assert "<code>0</code>" in unit_page
    assert "<code>0.1</code>" in unit_page
    assert "No lo puedo confirmar ahora" in unit_page
    assert persisted

    status, result_page = app.confirm_meanings(
        session_id="cafeteria-session",
        fields=_unit_answers(unit_page),
    )
    assert status == 200
    packet = app.session("cafeteria-session").last_review_result or {}
    assert "Análisis pendiente" in result_page
    assert "Margen calculado" not in result_page
    assert packet.get("computation_executed") is False
    assert len(app.session("cafeteria-session").owner_unit_confirmation_events) == 1

    persisted_columns = {event.column_ref for event, _contract in persisted}
    assert persisted_columns.isdisjoint(
        {"Hora", "SucursalID", "MetodoPago", "CanalVenta", "Empleado", "Sucursal", "Ciudad", "Categoria"}
    )
    assert {"Cantidad", "PrecioUnitario", "Costo"}.issubset(persisted_columns)



def test_cafeteria_margin_can_defer_discount_unit_without_creating_evidence_or_calculating(tmp_path: Path) -> None:
    app = AssistedWebApplicationV1(output_dir=tmp_path / "outputs")

    status, semantic_page = app.receive_xlsx(
        session_id="cafeteria-defer-unit",
        filename="cafeteria_defer.xlsx",
        content=_cafeteria_xlsx_bytes(include_taxes=True),
        selected_launch_review="net_margin_real",
    )
    assert status == 200

    status, unit_page = app.confirm_meanings(
        session_id="cafeteria-defer-unit",
        fields=_answers(semantic_page),
    )
    assert status == 200
    assert "No lo puedo confirmar ahora" in unit_page
    question_ids = [
        html.unescape(item)
        for item in re.findall(r'name="unit_([^"]+)" value="not_sure"', unit_page)
    ]
    assert question_ids

    status, deferred_page = app.confirm_meanings(
        session_id="cafeteria-defer-unit",
        fields={f"unit_{question_id}": "not_sure" for question_id in question_ids},
    )

    assert status == 200
    assert "Necesito una confirmación para continuar" in deferred_page
    assert "Análisis pendiente" in deferred_page
    assert "No generé ningún cálculo con este dato pendiente" in deferred_page
    state = app.session("cafeteria-defer-unit")
    assert state.owner_unit_confirmation_events == []
    packet = state.last_review_result or {}
    assert packet.get("computation_executed") is False


def test_cafeteria_margin_confirms_semantics_then_discount_unit_and_executes_kernel_once_semantics_are_fixed(tmp_path: Path) -> None:
    provider_calls = 0

    def provider(payload: dict) -> dict:
        nonlocal provider_calls
        provider_calls += 1
        return build_service_1_deterministic_semantic_proposal_v1(payload)

    app = AssistedWebApplicationV1(
        output_dir=tmp_path / "outputs",
        semantic_provider=provider,
    )

    status, semantic_page = app.receive_xlsx(
        session_id="cafeteria-full",
        filename="cafeteria_full.xlsx",
        content=_cafeteria_xlsx_bytes(include_taxes=True),
        selected_launch_review="net_margin_real",
    )
    assert status == 200
    assert provider_calls == 1
    assert "Revisión del archivo" in semantic_page

    status, unit_page = app.confirm_meanings(
        session_id="cafeteria-full",
        fields=_answers(semantic_page),
    )
    assert status == 200
    assert provider_calls == 1
    assert "Esto entendí de tu Excel" in unit_page
    assert "Confirmemos el descuento" in unit_page

    status, result_page = app.confirm_meanings(
        session_id="cafeteria-full",
        fields=_unit_answers(unit_page),
    )
    assert status == 200
    assert provider_calls == 1, "unit reentry must not recall semantic provider"
    packet = app.session("cafeteria-full").last_review_result or {}
    assert packet["status"] == "COMPUTATION_PLAN_READY"
    assert packet["computation_executed"] is True
    assert packet["computation_result"]["status"] == "EVALUATED"
    assert packet["computation_result"]["inputs"] == {
        "sale_price": 168.0,
        "costs": 84.0,
        "taxes": 20.0,
    }
    assert packet["computation_result"]["computed"]["net_margin_amount"] == 64.0
    assert round(packet["computation_result"]["computed"]["net_margin_percentage"], 6) == round(64 / 168 * 100, 6)
    assert packet["runtime_authorized"] is False
    assert packet["delivery_authorized"] is False
    assert "Margen" in result_page
    assert 'href="/download-net-margin"' in result_page
    assert len(app.session("cafeteria-full").owner_unit_confirmation_events) == 1

from __future__ import annotations

import inspect
from pathlib import Path

import openpyxl
import pytest

import pymia.smartpyme.exceland_runtime_v1 as runtime_module
from pymia.smartpyme.exceland_runtime_v1 import run_exceland_runtime_v1


@pytest.fixture(autouse=True)
def _portable_fake_factory(monkeypatch: pytest.MonkeyPatch):
    def fake_build_product(product_ref: str, *, output_path: Path) -> None:
        workbook = openpyxl.Workbook()
        workbook.active.title = "BIENVENIDA"
        workbook.active["A1"] = product_ref
        workbook.save(output_path)

    monkeypatch.setattr(runtime_module, "_load_build_product", lambda: fake_build_product)


def test_success_path_generates_xlsx_with_sheets(tmp_path: Path) -> None:
    result = run_exceland_runtime_v1(product_ref="precio_margen", output_dir=tmp_path)
    assert result["status"] == "OK"
    assert result["artifact_exists"] is True
    assert result["runtime_authorized"] is False
    output_path = Path(result["output_path"])
    assert output_path.exists() and output_path.stat().st_size > 0
    assert openpyxl.load_workbook(output_path).sheetnames


def test_success_path_custom_filename(tmp_path: Path) -> None:
    result = run_exceland_runtime_v1(
        product_ref="caja_diaria", output_dir=tmp_path, output_filename="mi_caja.xlsx"
    )
    assert result["status"] == "OK"
    assert Path(result["output_path"]).name == "mi_caja.xlsx"


def test_missing_product_ref(tmp_path: Path) -> None:
    result = run_exceland_runtime_v1(product_ref=None, output_dir=tmp_path)
    assert result["status"] == "MISSING_PRODUCT_REF"
    assert result["runtime_authorized"] is False


def test_empty_product_ref(tmp_path: Path) -> None:
    assert run_exceland_runtime_v1(product_ref="", output_dir=tmp_path)["status"] == "MISSING_PRODUCT_REF"


def test_unknown_product_ref(tmp_path: Path) -> None:
    result = run_exceland_runtime_v1(product_ref="producto_inexistente", output_dir=tmp_path)
    assert result["status"] == "UNKNOWN_PRODUCT"
    assert result["artifact_exists"] is False


def test_invalid_output_dir(tmp_path: Path) -> None:
    result = run_exceland_runtime_v1(product_ref="caja_diaria", output_dir=tmp_path / "no_existe")
    assert result["status"] == "INVALID_OUTPUT_DIR"
    assert result["artifact_exists"] is False


def test_factory_unavailable_is_fail_closed(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(runtime_module, "_load_build_product", lambda: None)
    result = run_exceland_runtime_v1(product_ref="caja_diaria", output_dir=tmp_path)
    assert result["status"] == "FACTORY_UNAVAILABLE"
    assert result["artifact_exists"] is False
    assert result["runtime_authorized"] is False


def test_runtime_authorized_always_false(tmp_path: Path) -> None:
    for product_ref in ("precio_margen", "caja_diaria"):
        assert run_exceland_runtime_v1(product_ref=product_ref, output_dir=tmp_path)["runtime_authorized"] is False


def test_module_has_no_forbidden_imports() -> None:
    source = inspect.getsource(runtime_module)
    assert "service_1_pipeline" not in source
    assert "diagnostic_core" not in source
    assert "fsm" not in source.lower()
    assert "llm" not in source.lower()
    assert "chatbot" not in source.lower()


def test_allowed_product_refs_are_callable(tmp_path: Path) -> None:
    for product_ref in runtime_module._EXCELAND_PRODUCT_REFS:
        result = run_exceland_runtime_v1(product_ref=product_ref, output_dir=tmp_path)
        assert result["status"] == "OK"
        assert result["artifact_exists"] is True

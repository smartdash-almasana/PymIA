from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from pymia.smartpyme.service_1_reconciliation_candidate_to_assisted_review_v1 import (
    STATUS_PARTIAL as REVIEW_PARTIAL,
    STATUS_READY as REVIEW_READY,
    build_service_1_reconciliation_assisted_review_v1,
)
from pymia.smartpyme.service_1_reconciliation_request_gate_v1 import (
    BANK_RECONCILIATION,
    MERCADO_PAGO_BANK_RECONCILIATION,
    STATUS_READY as GATE_READY,
    build_service_1_reconciliation_request_gate_v1,
)


REPO_ROOT = Path(__file__).resolve().parents[2]


def _rows_from_sheet(path: Path, sheet_name: str) -> list[dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        headers = [str(value) if value is not None else "" for value in next(rows)]
        return [
            dict(zip(headers, row))
            for row in rows
            if any(value is not None for value in row)
        ]
    finally:
        workbook.close()


def _excel_date(value: object) -> object:
    if isinstance(value, (date, datetime)):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return from_excel(value)
    return value


def _governance(columns: tuple[str, ...]) -> dict[str, object]:
    return {
        "p5_status": "CONFIRMED",
        "p6_decisions": [
            {
                "column_ref": column,
                "status": "APPROVED",
                "approved_role": f"role:{column}",
            }
            for column in columns
        ],
        "p7_status": "REQUIREMENT_MATCHED",
        "p8_status": "COMPUTABLE",
        "runtime_authorized": False,
        "tool_execution_authorized": False,
        "product_ready": False,
        "delivery_authorized": False,
        "diagnosis_generated": False,
    }


def _bank_source() -> dict[str, object]:
    columns = ("mov_id", "fecha", "importe", "referencia")
    return {
        "source_kind": "bank",
        "source_ref": "BANCO",
        "rows": [
            {
                "mov_id": "B-100",
                "fecha": "2026-07-01",
                "importe": 1250.0,
                "referencia": "COBRO-100",
            }
        ],
        "field_bindings": {
            "id": "mov_id",
            "fecha": "fecha",
            "importe": "importe",
            "referencia": "referencia",
        },
        "governance": _governance(columns),
    }


def _internal_source() -> dict[str, object]:
    columns = ("cobro_id", "fecha", "importe", "referencia")
    return {
        "source_kind": "internal",
        "source_ref": "COBROS",
        "rows": [
            {
                "cobro_id": "C-100",
                "fecha": "2026-07-01",
                "importe": 1250.0,
                "referencia": "COBRO-100",
            }
        ],
        "field_bindings": {
            "id": "cobro_id",
            "fecha": "fecha",
            "importe": "importe",
            "referencia": "referencia",
        },
        "governance": _governance(columns),
    }


def _mercado_pago_source() -> dict[str, object]:
    columns = (
        "op_id",
        "fecha",
        "bruto",
        "comision",
        "retencion",
        "neto",
        "lote",
        "referencia",
    )
    return {
        "source_kind": "mercado_pago",
        "source_ref": "MERCADO_PAGO",
        "rows": [
            {
                "op_id": "MP-100",
                "fecha": "2026-07-02",
                "bruto": 2000.0,
                "comision": 80.0,
                "retencion": 20.0,
                "neto": 1900.0,
                "lote": "LOTE-100",
                "referencia": "MPREF-100",
            }
        ],
        "field_bindings": {
            "operacion_mp_id": "op_id",
            "fecha_operacion": "fecha",
            "importe_bruto": "bruto",
            "comision": "comision",
            "retencion": "retencion",
            "importe_neto": "neto",
            "lote_id": "lote",
            "referencia": "referencia",
        },
        "governance": _governance(columns),
    }


def _mercado_pago_bank_source() -> dict[str, object]:
    columns = ("mov_id", "fecha", "importe", "lote", "referencia")
    return {
        "source_kind": "bank",
        "source_ref": "BANCO",
        "rows": [
            {
                "mov_id": "B-MP-100",
                "fecha": "2026-07-02",
                "importe": 1900.0,
                "lote": "LOTE-100",
                "referencia": "MPREF-100",
            }
        ],
        "field_bindings": {
            "movimiento_banco_id": "mov_id",
            "fecha": "fecha",
            "importe": "importe",
            "lote_id": "lote",
            "referencia": "referencia",
        },
        "governance": _governance(columns),
    }


def _assert_safe_review(result: dict[str, object]) -> None:
    assert result["requires_human_review"] is True
    assert result["runtime_authorized"] is False
    assert result["tool_execution_authorized"] is False
    assert result["product_ready"] is False
    assert result["delivery_authorized"] is False
    assert result["diagnosis_generated"] is False
    assert result["io_performed"] is False
    assert result["files_created"] == []
    assert result["api_used"] is False
    assert result["llm_used"] is False


def test_governed_bank_request_reaches_assisted_review_without_shortcuts() -> None:
    gate = build_service_1_reconciliation_request_gate_v1(
        case_id="CASE-BANK-100",
        owner_requested=True,
        reconciliation_type=BANK_RECONCILIATION,
        source_packets=[_bank_source(), _internal_source()],
    )
    assert gate["status"] == GATE_READY

    review = build_service_1_reconciliation_assisted_review_v1(gate_packet=gate)

    assert review["status"] == REVIEW_READY
    assert review["case_id"] == "CASE-BANK-100"
    assert review["reconciliation_type"] == BANK_RECONCILIATION
    assert review["review_summary"]["confirmed_candidates"] == 1
    assert review["review_summary"]["ambiguous_groups"] == 0
    assert review["next_allowed_action"] == "human_reconciliation_review"
    assert review["provenance"]["candidate_source"]["source"] == (
        "SERVICE_1_RECONCILIATION_REQUEST_GATE_V1"
    )
    _assert_safe_review(review)


def test_governed_mercado_pago_request_reaches_assisted_review() -> None:
    gate = build_service_1_reconciliation_request_gate_v1(
        case_id="CASE-MP-100",
        owner_requested=True,
        reconciliation_type=MERCADO_PAGO_BANK_RECONCILIATION,
        source_packets=[_mercado_pago_source(), _mercado_pago_bank_source()],
    )
    assert gate["status"] == GATE_READY

    review = build_service_1_reconciliation_assisted_review_v1(gate_packet=gate)

    assert review["status"] == REVIEW_READY
    assert review["case_id"] == "CASE-MP-100"
    assert review["reconciliation_type"] == MERCADO_PAGO_BANK_RECONCILIATION
    assert review["review_summary"]["confirmed_candidates"] == 1
    assert review["review_summary"]["calculation_inconsistencies"] == 0
    assert review["review_result"]["conciliaciones"][0]["importe_neto_esperado"] == 1900.0
    assert review["next_allowed_action"] == "human_reconciliation_review"
    _assert_safe_review(review)



def _physical_bank_sources() -> list[dict[str, object]]:
    path = REPO_ROOT / "prueba_excels" / "conciliacion_pyme_argentina_corregida.xlsx"
    bank_rows = _rows_from_sheet(path, "BANCO")
    collection_rows = _rows_from_sheet(path, "COBROS")
    for row in bank_rows:
        row["fecha"] = _excel_date(row.get("fecha"))
    for row in collection_rows:
        row["fecha_cobro"] = _excel_date(row.get("fecha_cobro"))

    bank_columns = (
        "movimiento_id",
        "fecha",
        "importe",
        "referencia",
        "descripcion",
    )
    internal_columns = (
        "cobro_id",
        "fecha_cobro",
        "importe_cobrado",
        "referencia",
        "cliente",
    )
    return [
        {
            "source_kind": "bank",
            "source_ref": "BANCO",
            "rows": [
                row
                for row in bank_rows
                if isinstance(row.get("movimiento_id"), str)
            ],
            "field_bindings": {
                "id": "movimiento_id",
                "fecha": "fecha",
                "importe": "importe",
                "referencia": "referencia",
                "descripcion": "descripcion",
            },
            "governance": _governance(bank_columns),
        },
        {
            "source_kind": "internal",
            "source_ref": "COBROS",
            "rows": [
                row
                for row in collection_rows
                if isinstance(row.get("cobro_id"), str)
            ],
            "field_bindings": {
                "id": "cobro_id",
                "fecha": "fecha_cobro",
                "importe": "importe_cobrado",
                "referencia": "referencia",
                "descripcion": "cliente",
            },
            "governance": _governance(internal_columns),
        },
    ]


def _physical_mercado_pago_sources() -> list[dict[str, object]]:
    path = (
        REPO_ROOT
        / "prueba_excels"
        / "conciliacion_mercado_pago_banco_corregida.xlsx"
    )
    mp_rows = _rows_from_sheet(path, "MERCADO_PAGO")
    bank_rows = _rows_from_sheet(path, "BANCO")
    for row in mp_rows:
        row["fecha_operacion"] = _excel_date(row.get("fecha_operacion"))
    for row in bank_rows:
        row["fecha"] = _excel_date(row.get("fecha"))

    mp_columns = (
        "operacion_mp_id",
        "fecha_operacion",
        "importe_bruto",
        "comision",
        "retencion",
        "importe_neto",
        "lote_id",
        "referencia",
        "estado",
    )
    bank_columns = (
        "movimiento_banco_id",
        "fecha",
        "importe",
        "lote_id",
        "referencia",
        "descripcion",
    )
    return [
        {
            "source_kind": "mercado_pago",
            "source_ref": "MERCADO_PAGO",
            "rows": [
                row
                for row in mp_rows
                if isinstance(row.get("operacion_mp_id"), str)
            ],
            "field_bindings": {
                "operacion_mp_id": "operacion_mp_id",
                "fecha_operacion": "fecha_operacion",
                "importe_bruto": "importe_bruto",
                "comision": "comision",
                "retencion": "retencion",
                "importe_neto": "importe_neto",
                "lote_id": "lote_id",
                "referencia": "referencia",
                "estado": "estado",
            },
            "governance": _governance(mp_columns),
        },
        {
            "source_kind": "bank",
            "source_ref": "BANCO",
            "rows": [
                row
                for row in bank_rows
                if isinstance(row.get("movimiento_banco_id"), str)
            ],
            "field_bindings": {
                "movimiento_banco_id": "movimiento_banco_id",
                "fecha": "fecha",
                "importe": "importe",
                "lote_id": "lote_id",
                "referencia": "referencia",
                "descripcion": "descripcion",
            },
            "governance": _governance(bank_columns),
        },
    ]


def test_physical_bank_workbook_crosses_gate_and_adapter() -> None:
    gate = build_service_1_reconciliation_request_gate_v1(
        case_id="PHYSICAL-BANK-001",
        owner_requested=True,
        reconciliation_type=BANK_RECONCILIATION,
        source_packets=_physical_bank_sources(),
    )
    assert gate["status"] == GATE_READY

    review = build_service_1_reconciliation_assisted_review_v1(gate_packet=gate)

    assert review["status"] == REVIEW_PARTIAL
    assert review["review_summary"]["confirmed_candidates"] > 0
    assert review["review_summary"]["ambiguous_groups"] == 3
    assert review["review_summary"]["amount_differences"] == 1
    assert review["next_allowed_action"] == "human_reconciliation_review"
    _assert_safe_review(review)


def test_physical_mercado_pago_workbook_crosses_gate_and_adapter() -> None:
    gate = build_service_1_reconciliation_request_gate_v1(
        case_id="PHYSICAL-MP-001",
        owner_requested=True,
        reconciliation_type=MERCADO_PAGO_BANK_RECONCILIATION,
        source_packets=_physical_mercado_pago_sources(),
    )
    assert gate["status"] == GATE_READY

    review = build_service_1_reconciliation_assisted_review_v1(gate_packet=gate)

    assert review["status"] == REVIEW_PARTIAL
    assert review["review_summary"] == {
        "confirmed_candidates": 10,
        "probable_candidates": 0,
        "ambiguous_groups": 2,
        "amount_differences": 1,
        "date_differences": 0,
        "bank_pending": 1,
        "internal_pending": 1,
        "missing_evidence": 0,
        "calculation_inconsistencies": 0,
    }
    assert review["next_allowed_action"] == "human_reconciliation_review"
    _assert_safe_review(review)

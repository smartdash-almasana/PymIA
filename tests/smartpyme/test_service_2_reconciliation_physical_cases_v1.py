from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from pymia.smartpyme.service_2_reconciliation_match_candidates_v1 import (
    build_reconciliation_match_candidates_v1,
)


REPO_ROOT = Path(__file__).resolve().parents[2]


def _rows_from_sheet(path: Path, sheet_name: str) -> list[dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        headers = [str(value) if value is not None else "" for value in next(rows)]
        return [dict(zip(headers, row)) for row in rows]
    finally:
        workbook.close()


def _sales_movements() -> list[dict[str, object]]:
    rows = _rows_from_sheet(REPO_ROOT / "prueba_excels" / "ventas_marzo_2026.xlsx", "Ventas_Marzo_2026")
    movements: list[dict[str, object]] = []
    for row in rows:
        ticket_id = row.get("ticket_id")
        if not isinstance(ticket_id, str) or not ticket_id.startswith("T-"):
            continue
        movements.append(
            {
                "id": ticket_id,
                "fecha": row["fecha"],
                "importe": row["total_venta"],
                "referencia": ticket_id,
                "descripcion": row.get("producto"),
            }
        )
    return movements


def _collection_movements(*, mercado_pago_only: bool = False) -> list[dict[str, object]]:
    rows = _rows_from_sheet(REPO_ROOT / "prueba_excels" / "cobros_marzo_2026.xlsx", "Cobros_Marzo_2026")
    movements: list[dict[str, object]] = []
    for row in rows:
        collection_id = row.get("cobro_id")
        if not isinstance(collection_id, str) or not collection_id.startswith("C-"):
            continue
        if mercado_pago_only and row.get("medio_de_cobro") != "Mercado Pago":
            continue
        movements.append(
            {
                "id": collection_id,
                "fecha": row["fecha"],
                "importe": row["importe_cobrado"],
                "referencia": row.get("ticket_relacionado"),
                "descripcion": row.get("referencia"),
            }
        )
    return movements


def _mercado_pago_sales() -> list[dict[str, object]]:
    rows = _rows_from_sheet(REPO_ROOT / "prueba_excels" / "ventas_marzo_2026.xlsx", "Ventas_Marzo_2026")
    movements: list[dict[str, object]] = []
    for row in rows:
        ticket_id = row.get("ticket_id")
        if not isinstance(ticket_id, str) or not ticket_id.startswith("T-"):
            continue
        if row.get("medio_de_cobro_declarado") != "Mercado Pago":
            continue
        movements.append(
            {
                "id": ticket_id,
                "fecha": row["fecha"],
                "importe": row["total_venta"],
                "referencia": ticket_id,
                "descripcion": row.get("producto"),
            }
        )
    return movements


def _cash_bank_movements() -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    rows = _rows_from_sheet(
        REPO_ROOT / "prueba_excels" / "first_aid_pilot_004_cash_bank_reconciliation_demo.xlsx",
        "Caja_Banco",
    )
    bank: list[dict[str, object]] = []
    internal: list[dict[str, object]] = []
    for row in rows:
        movement_id = row.get("ID movimiento")
        if not isinstance(movement_id, str) or not movement_id.startswith("MOV-"):
            continue
        common = {
            "fecha": row["Fecha"],
            "referencia": row.get("Referencia externa"),
            "descripcion": row.get("Descripción"),
        }
        bank_amount = row.get("Importe banco")
        if isinstance(bank_amount, (int, float)) and not isinstance(bank_amount, bool) and float(bank_amount) != 0.0:
            bank.append({"id": f"bank:{movement_id}", "importe": bank_amount, **common})
        internal_amount = row.get("Importe caja/POS")
        if isinstance(internal_amount, (int, float)) and not isinstance(internal_amount, bool) and float(internal_amount) != 0.0:
            internal.append({"id": f"internal:{movement_id}", "importe": internal_amount, **common})
    return bank, internal


def test_physical_sales_and_collections_cross_source_case() -> None:
    result = build_reconciliation_match_candidates_v1(
        bank_movements=_collection_movements(),
        internal_movements=_sales_movements(),
    )

    assert len(result["matches_exactos"]) == 16
    assert all(match["tipo_match"] == "MATCH_REFERENCE_EXACT" for match in result["matches_exactos"])
    assert result["matches_ambiguos"] == []
    assert result["matches_probables"] == []
    assert {
        (str(item["banco_id"]), str(item["interno_id"]))
        for item in result["diferencias_importe"]
    } == {
        ("C-0001", "T-0001"),
        ("C-0012", "T-0013"),
        ("C-0019", "T-0020"),
    }
    assert {item["id"] for item in result["banco_sin_imputar"]} == {"C-0001", "C-0012", "C-0019", "C-0020"}
    assert {item["id"] for item in result["interno_sin_banco"]} == {"T-0001", "T-0003", "T-0013", "T-0020"}
    assert result["requires_human_review"] is True


def test_physical_mercado_pago_subset_preserves_amount_difference() -> None:
    result = build_reconciliation_match_candidates_v1(
        bank_movements=_collection_movements(mercado_pago_only=True),
        internal_movements=_mercado_pago_sales(),
    )

    assert len(result["matches_exactos"]) == 4
    assert {match["interno_id"] for match in result["matches_exactos"]} == {"T-0002", "T-0005", "T-0010", "T-0017"}
    assert len(result["diferencias_importe"]) == 1
    assert result["diferencias_importe"][0]["banco_id"] == "C-0012"
    assert result["diferencias_importe"][0]["interno_id"] == "T-0013"
    assert {item["id"] for item in result["banco_sin_imputar"]} == {"C-0012"}
    assert {item["id"] for item in result["interno_sin_banco"]} == {"T-0013"}


def test_physical_cash_bank_case_surfaces_duplicate_as_nm_ambiguity() -> None:
    bank_movements, internal_movements = _cash_bank_movements()
    result = build_reconciliation_match_candidates_v1(
        bank_movements=bank_movements,
        internal_movements=internal_movements,
    )

    assert len(result["matches_ambiguos"]) == 1
    ambiguous = result["matches_ambiguos"][0]
    assert ambiguous["tipo"] == "AMBIGUOUS"
    assert ambiguous["cardinalidad"] == "N:M"
    assert ambiguous["candidate_count"] == 4
    assert set(ambiguous["banco_ids"]) == {"bank:MOV-010", "bank:MOV-011"}
    assert set(ambiguous["interno_ids"]) == {"internal:MOV-010", "internal:MOV-011"}
    assert any(item["banco_id"] == "bank:MOV-002" for item in result["diferencias_importe"])
    assert {item["id"] for item in result["banco_sin_imputar"]} >= {"bank:MOV-002", "bank:MOV-010", "bank:MOV-011"}
    assert {item["id"] for item in result["interno_sin_banco"]} >= {"internal:MOV-002", "internal:MOV-010", "internal:MOV-011"}


def _split_ground_truth_ids(value: object) -> list[str]:
    if not isinstance(value, str):
        return []
    return [part.strip() for part in value.split(",") if part.strip()]


def _excel_date_value(value: object) -> object:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return from_excel(value)
    return value


def _corrected_reconciliation_case() -> tuple[
    list[dict[str, object]],
    list[dict[str, object]],
    list[dict[str, object]],
]:
    path = REPO_ROOT / "prueba_excels" / "conciliacion_pyme_argentina_corregida.xlsx"
    bank_rows = _rows_from_sheet(path, "BANCO")
    collection_rows = _rows_from_sheet(path, "COBROS")
    expected_rows = _rows_from_sheet(path, "CASO_ESPERADO")

    bank_movements = [
        {
            "id": row["movimiento_id"],
            "fecha": _excel_date_value(row["fecha"]),
            "importe": row["importe"],
            "referencia": row.get("referencia"),
            "descripcion": row.get("descripcion"),
        }
        for row in bank_rows
        if isinstance(row.get("movimiento_id"), str)
    ]
    internal_movements = [
        {
            "id": row["cobro_id"],
            "fecha": _excel_date_value(row["fecha_cobro"]),
            "importe": row["importe_cobrado"],
            "referencia": row.get("referencia"),
            "descripcion": row.get("cliente"),
        }
        for row in collection_rows
        if isinstance(row.get("cobro_id"), str)
    ]
    return bank_movements, internal_movements, expected_rows


def test_corrected_reconciliation_workbook_matches_declared_ground_truth() -> None:
    bank_movements, internal_movements, expected_rows = _corrected_reconciliation_case()
    result = build_reconciliation_match_candidates_v1(bank_movements, internal_movements)

    expected_pairs: dict[str, set[tuple[str, str]]] = {
        "MATCH_REFERENCE_EXACT": set(),
        "MATCH_ATTRIBUTES_EXACT": set(),
        "MATCH_PROBABLE_DATE": set(),
    }
    expected_ambiguities: set[tuple[str, frozenset[str], frozenset[str], int]] = set()
    expected_unmatched_bank: set[str] = set()
    expected_unmatched_internal: set[str] = set()

    for row in expected_rows:
        expected_result = row.get("resultado_esperado")
        bank_ids = _split_ground_truth_ids(row.get("movimiento_banco_ids"))
        internal_ids = _split_ground_truth_ids(row.get("cobro_ids"))

        if expected_result in expected_pairs:
            assert len(bank_ids) == len(internal_ids), row.get("caso_id")
            expected_pairs[str(expected_result)].update(zip(bank_ids, internal_ids))
            continue

        if expected_result == "AMBIGUOUS":
            cardinality = str(row.get("cardinalidad"))
            expected_ambiguities.add(
                (
                    cardinality,
                    frozenset(bank_ids),
                    frozenset(internal_ids),
                    len(bank_ids) * len(internal_ids),
                )
            )
            expected_unmatched_bank.update(bank_ids)
            expected_unmatched_internal.update(internal_ids)
            continue

        if expected_result == "NO_MATCH":
            expected_unmatched_bank.update(bank_ids)
            expected_unmatched_internal.update(internal_ids)

    actual_pairs = {
        "MATCH_REFERENCE_EXACT": {
            (str(item["banco_id"]), str(item["interno_id"]))
            for item in result["matches_exactos"]
            if item["tipo_match"] == "MATCH_REFERENCE_EXACT"
        },
        "MATCH_ATTRIBUTES_EXACT": {
            (str(item["banco_id"]), str(item["interno_id"]))
            for item in result["matches_exactos"]
            if item["tipo_match"] == "MATCH_ATTRIBUTES_EXACT"
        },
        "MATCH_PROBABLE_DATE": {
            (str(item["banco_id"]), str(item["interno_id"]))
            for item in result["matches_probables"]
            if item["tipo_match"] == "MATCH_PROBABLE_DATE"
        },
    }
    actual_ambiguities = {
        (
            str(item["cardinalidad"]),
            frozenset(str(value) for value in item["banco_ids"]),
            frozenset(str(value) for value in item["interno_ids"]),
            int(item["candidate_count"]),
        )
        for item in result["matches_ambiguos"]
    }

    assert result["status"] == "PARTIAL_MATCHES_FOUND"
    assert actual_pairs == expected_pairs
    assert actual_ambiguities == expected_ambiguities
    assert {str(item["id"]) for item in result["banco_sin_imputar"]} == expected_unmatched_bank
    assert {str(item["id"]) for item in result["interno_sin_banco"]} == expected_unmatched_internal
    assert {
        (str(item["banco_id"]), str(item["interno_id"]))
        for item in result["diferencias_importe"]
    } == {("M-3029", "C-2027")}
    assert result["requires_human_review"] is True

    def keys_in(value: object) -> list[str]:
        if isinstance(value, dict):
            keys = [str(key) for key in value]
            for nested in value.values():
                keys.extend(keys_in(nested))
            return keys
        if isinstance(value, list):
            keys: list[str] = []
            for nested in value:
                keys.extend(keys_in(nested))
            return keys
        return []

    assert not [key for key in keys_in(result) if "confianza" in key.lower() or "confidence" in key.lower()]

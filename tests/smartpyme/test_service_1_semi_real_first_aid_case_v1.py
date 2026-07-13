from __future__ import annotations

import inspect
from pathlib import Path

from openpyxl import load_workbook

from pymia.smartpyme.service_1_semi_real_first_aid_case_v1 import (
    SEMI_REAL_CASE_ID,
    SERVICE_NAME,
    run_service_1_semi_real_first_aid_case_v1,
)


def test_semi_real_case_generates_three_first_aid_deliveries(tmp_path: Path) -> None:
    result = run_service_1_semi_real_first_aid_case_v1(tmp_path)

    assert result["case_id"] == SEMI_REAL_CASE_ID
    assert result["schema_version"] == "1.0"
    assert result["service_name"] == SERVICE_NAME
    assert result["runtime_authorized"] is False
    assert result["flow"]["delivery_count"] == 3
    assert len(result["generated_files"]) == 3


def test_semi_real_case_business_profile_is_plausible(tmp_path: Path) -> None:
    result = run_service_1_semi_real_first_aid_case_v1(tmp_path)
    profile = result["business_profile"]

    assert profile["business_type"] == "comercio_minorista_alimentos"
    assert profile["business_size"] == "micro_pyme"
    assert "revisión rápida" in str(profile["operator_scenario"])


def test_semi_real_case_contains_business_context(tmp_path: Path) -> None:
    result = run_service_1_semi_real_first_aid_case_v1(tmp_path)

    assert result["business_profile"]["business_type"] == "comercio_minorista_alimentos"
    assert result["business_profile"]["business_size"] == "micro_pyme"
    assert "datos declarados" in str(result["business_profile"]["data_source"])
    assert "comercio minorista" in result["owner_context"]


def test_semi_real_case_declared_inputs_are_preserved(tmp_path: Path) -> None:
    result = run_service_1_semi_real_first_aid_case_v1(tmp_path)
    declared_inputs = result["declared_inputs"]

    assert declared_inputs["precio_margen_basico"] == {
        "precio_venta": 2500,
        "costo_unitario": 1625,
    }
    assert declared_inputs["caja_diaria_triage"] == {
        "saldo_inicial": 180000,
        "ingresos": 324500,
        "egresos": 286750,
    }
    assert declared_inputs["stock_alertas_basicas"] == {
        "producto": "Pack yerba 1kg",
        "stock_actual": 8,
        "stock_minimo": 15,
        "ventas_diarias_promedio": 3,
    }


def test_semi_real_case_uses_expected_tools_and_statuses(tmp_path: Path) -> None:
    result = run_service_1_semi_real_first_aid_case_v1(tmp_path)

    assert result["flow"]["tool_refs"] == [
        "precio_margen_basico",
        "caja_diaria_triage",
        "stock_alertas_basicas",
    ]
    assert result["flow"]["statuses"] == ["OK", "OK", "OK"]


def test_semi_real_case_generates_readable_xlsx_files(tmp_path: Path) -> None:
    result = run_service_1_semi_real_first_aid_case_v1(tmp_path)

    for generated_file in result["generated_files"]:
        workbook = load_workbook(generated_file)
        assert workbook["Resumen"]["B2"].value == "SERVICE_1"
        assert "Resultados" in workbook.sheetnames
        assert "Limitaciones" in workbook.sheetnames
        assert "Claims prohibidos" in workbook.sheetnames


def test_semi_real_case_summary_contains_owner_relevant_content(tmp_path: Path) -> None:
    result = run_service_1_semi_real_first_aid_case_v1(tmp_path)
    summary_text = result["flow"]["summary_text"]

    assert "Resultados procesados: 3" in summary_text
    assert "precio_margen_basico: OK" in summary_text
    assert "caja_diaria_triage: OK" in summary_text
    assert "stock_alertas_basicas: OK" in summary_text
    assert "Entrega preliminar basada en datos declarados." in summary_text
    assert "No es un diagnostico integral" in summary_text
    assert "No confirma saldo bancario real" in summary_text
    assert "No confirma stock fisico real" in summary_text


def test_semi_real_case_operator_notes_are_conservative(tmp_path: Path) -> None:
    result = run_service_1_semi_real_first_aid_case_v1(tmp_path)
    notes = " ".join(result["operator_review_notes"]).lower()

    assert "preliminar" in notes
    assert "no como certificación" in notes
    assert "evidencia adicional" in notes
    assert "rentabilidad real" in notes


def test_semi_real_case_rejects_missing_output_dir() -> None:
    missing_dir = Path("/nonexistent/service1/semi-real/case")

    try:
        run_service_1_semi_real_first_aid_case_v1(missing_dir)
    except FileNotFoundError as exc:
        assert "Output directory does not exist" in str(exc)
    else:
        raise AssertionError("Expected FileNotFoundError")


def test_semi_real_case_does_not_depend_on_forbidden_product_layers() -> None:
    import pymia.smartpyme.service_1_semi_real_first_aid_case_v1 as module

    source = inspect.getsource(module)

    assert "vertical_pipeline" not in source
    assert "fsm" not in source.lower()
    assert "llm" not in source.lower()
    assert "chatbot" not in source.lower()
    assert "document_ingestion" not in source
    assert "exceland" not in source.lower()

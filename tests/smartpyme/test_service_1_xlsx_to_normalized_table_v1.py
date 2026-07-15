from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from pymia.smartpyme.service_1_xlsx_to_normalized_table_v1 import (
    read_xlsx_to_normalized_table_v1,
    read_xlsx_to_normalized_tables_v1,
)


def _save_workbook(path: Path, rows_by_sheet: dict[str, list[list[object | None]]]) -> None:
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)
    for sheet_name, rows in rows_by_sheet.items():
        worksheet = workbook.create_sheet(sheet_name)
        for row in rows:
            worksheet.append(row)
    workbook.save(path)
    workbook.close()


def test_file_not_found_is_blocked(tmp_path: Path) -> None:
    result = read_xlsx_to_normalized_table_v1(tmp_path / "missing.xlsx")

    assert result["status"] == "BLOCKED"
    assert "File not found." in result["blocking_errors"]
    assert result["runtime_authorized"] is False


def test_non_xlsx_extension_is_blocked(tmp_path: Path) -> None:
    path = tmp_path / "ventas.csv"
    path.write_text("Fecha,Importe\n2026-06-01,100\n", encoding="utf-8")

    result = read_xlsx_to_normalized_table_v1(path)

    assert result["status"] == "BLOCKED"
    assert "Only .xlsx files are accepted." in result["blocking_errors"]
    assert result["runtime_authorized"] is False


def test_empty_workbook_is_blocked(tmp_path: Path) -> None:
    path = tmp_path / "empty.xlsx"
    Workbook().save(path)

    result = read_xlsx_to_normalized_table_v1(path)

    assert result["status"] == "BLOCKED"
    assert "XLSX workbook has no non-empty rows." in result["blocking_errors"]
    assert result["runtime_authorized"] is False


def test_missing_headers_is_blocked(tmp_path: Path) -> None:
    path = tmp_path / "bad.xlsx"
    _save_workbook(path, {"Datos": [["Fecha", None, "Importe"], ["2026-06-01", "Ana", 100]]})

    result = read_xlsx_to_normalized_table_v1(path)

    assert result["status"] == "BLOCKED"
    assert "XLSX headers are missing or incomplete." in result["blocking_errors"]
    assert "headers_required" in result["blocking_errors"]
    assert result["runtime_authorized"] is False


def test_duplicate_headers_is_blocked(tmp_path: Path) -> None:
    path = tmp_path / "dup.xlsx"
    _save_workbook(path, {"Datos": [["Cliente", "cliente", "Importe"], ["Ana", "Ana SA", 100]]})

    result = read_xlsx_to_normalized_table_v1(path)

    assert result["status"] == "BLOCKED"
    assert any(error.startswith("duplicate_headers") for error in result["blocking_errors"])
    assert result["rows"] == []
    assert result["runtime_authorized"] is False


def test_valid_single_sheet_is_ok(tmp_path: Path) -> None:
    path = tmp_path / "ventas.xlsx"
    _save_workbook(
        path,
        {"Ventas": [["Fecha", "Cliente", "Importe"], ["2026-06-01", "Ana", 1200], ["2026-06-02", "Beto", 900]]},
    )

    result = read_xlsx_to_normalized_table_v1(path)

    assert result["status"] == "OK"
    assert result["source_kind"] == "xlsx"
    assert result["sheet_name"] == "Ventas"
    assert result["headers"] == ["Fecha", "Cliente", "Importe"]
    assert result["normalized_headers"] == ["fecha", "cliente", "importe"]
    assert result["row_count"] == 2
    assert result["rows"][0] == {"fecha": "2026-06-01", "cliente": "Ana", "importe": "1200"}
    assert result["runtime_authorized"] is False


def test_valid_with_sheet_name_is_ok(tmp_path: Path) -> None:
    path = tmp_path / "multi.xlsx"
    _save_workbook(
        path,
        {
            "Resumen": [[None]],
            "Inventario": [["Producto", "Stock"], ["Yerba", 50]],
        },
    )

    result = read_xlsx_to_normalized_table_v1(path, sheet_name="Inventario")

    assert result["status"] == "OK"
    assert result["sheet_name"] == "Inventario"
    assert result["rows"] == [{"producto": "Yerba", "stock": "50"}]


def test_missing_sheet_name_is_blocked(tmp_path: Path) -> None:
    path = tmp_path / "ventas.xlsx"
    _save_workbook(path, {"Ventas": [["Fecha"], ["2026-06-01"]]})

    result = read_xlsx_to_normalized_table_v1(path, sheet_name="NoExiste")

    assert result["status"] == "BLOCKED"
    assert result["sheet_name"] == "NoExiste"
    assert "Sheet not found: NoExiste" in result["blocking_errors"]


def test_multiple_sheets_defaults_to_first_data_sheet(tmp_path: Path) -> None:
    path = tmp_path / "multi.xlsx"
    _save_workbook(
        path,
        {
            "Vacia": [[None]],
            "Ventas": [["Fecha", "Importe"], ["2026-06-01", 100]],
            "Stock": [["Producto", "Stock"], ["Yerba", 50]],
        },
    )

    result = read_xlsx_to_normalized_table_v1(path)

    assert result["status"] == "OK"
    assert result["sheet_name"] == "Ventas"
    assert result["rows"] == [{"fecha": "2026-06-01", "importe": "100"}]


def test_warns_on_short_rows(tmp_path: Path) -> None:
    path = tmp_path / "short.xlsx"
    _save_workbook(path, {"Datos": [["Fecha", "Cliente", "Importe"], ["2026-06-01", "Ana"]]})

    result = read_xlsx_to_normalized_table_v1(path)

    assert result["status"] == "OK"
    assert result["rows"] == [{"fecha": "2026-06-01", "cliente": "Ana", "importe": ""}]
    assert "Row 2 has fewer cells than headers." in result["warnings"]


def test_warns_on_extra_cells(tmp_path: Path) -> None:
    path = tmp_path / "extra.xlsx"
    _save_workbook(path, {"Datos": [["Fecha", "Importe"], ["2026-06-01", 100, "extra"]]})

    result = read_xlsx_to_normalized_table_v1(path)

    assert result["status"] == "OK"
    assert result["rows"] == [{"fecha": "2026-06-01", "importe": "100"}]
    assert "Row 2 has more cells than headers; extra cells ignored." in result["warnings"]


def test_runtime_authorized_always_false(tmp_path: Path) -> None:
    path = tmp_path / "ventas.xlsx"
    _save_workbook(path, {"Ventas": [["Fecha"], ["2026-06-01"]]})

    ok_result = read_xlsx_to_normalized_table_v1(path)
    blocked_result = read_xlsx_to_normalized_table_v1(tmp_path / "missing.xlsx")

    assert ok_result["runtime_authorized"] is False
    assert blocked_result["runtime_authorized"] is False


def test_normalizes_accents(tmp_path: Path) -> None:
    path = tmp_path / "acentos.xlsx"
    _save_workbook(path, {"Datos": [["Fecha de emisión", "Código Ñ"], ["2026-06-01", "A1"]]})

    result = read_xlsx_to_normalized_table_v1(path)

    assert result["status"] == "OK"
    assert result["normalized_headers"] == ["fecha_de_emision", "codigo_n"]


def test_reads_formula_cached_values_without_executing_formulas(tmp_path: Path) -> None:
    path = tmp_path / "formula.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Datos"
    worksheet.append(["Formula"])
    worksheet["A2"] = "=1+1"
    workbook.save(path)

    result = read_xlsx_to_normalized_table_v1(path)

    assert result["status"] == "OK"
    assert result["rows"] == []
    assert result["runtime_authorized"] is False
    assert path.exists()
    assert load_workbook(path, data_only=False).active["A2"].value == "=1+1"



def test_reads_all_non_empty_sheets_in_workbook_order(tmp_path: Path) -> None:
    path = tmp_path / "multi_all.xlsx"
    _save_workbook(
        path,
        {
            "Vacia": [[None]],
            "Ventas": [["Fecha", "Importe"], ["2026-07-01", 100]],
            "Cobros": [["Fecha", "Cobrado"], ["2026-07-02", 80]],
        },
    )

    tables = read_xlsx_to_normalized_tables_v1(path)

    assert [table["sheet_name"] for table in tables] == ["Ventas", "Cobros"]
    assert all(table["status"] == "OK" for table in tables)
    assert tables[0]["rows"] == [{"fecha": "2026-07-01", "importe": "100"}]
    assert tables[1]["rows"] == [{"fecha": "2026-07-02", "cobrado": "80"}]


def test_reads_explicit_sheet_subset_in_requested_order(tmp_path: Path) -> None:
    path = tmp_path / "multi_selected.xlsx"
    _save_workbook(
        path,
        {
            "Ventas": [["Fecha"], ["2026-07-01"]],
            "Cobros": [["Fecha"], ["2026-07-02"]],
            "Stock": [["Producto"], ["Yerba"]],
        },
    )

    tables = read_xlsx_to_normalized_tables_v1(
        path,
        sheet_names=("Stock", "Ventas"),
    )

    assert [table["sheet_name"] for table in tables] == ["Stock", "Ventas"]
    assert all(table["status"] == "OK" for table in tables)


def test_reader_releases_workbook_handle_on_windows(tmp_path: Path) -> None:
    path = tmp_path / "release.xlsx"
    _save_workbook(path, {"Ventas": [["Fecha"], ["2026-07-01"]]})

    result = read_xlsx_to_normalized_table_v1(path)
    assert result["status"] == "OK"

    renamed = tmp_path / "released.xlsx"
    path.rename(renamed)
    assert renamed.exists()

from __future__ import annotations

import inspect
from pathlib import Path

from openpyxl import load_workbook

from pymia.smartpyme.accounting_sandbox_release_gate_v1 import evaluate_accounting_sandbox_release_gate_v1
from pymia.smartpyme.bank_reconciliation_contract_v1 import build_bank_reconciliation_contract_v1
from pymia.smartpyme.bank_reconciliation_sandbox_contract_v1 import (
    CAPABILITY_REF,
    build_bank_reconciliation_sandbox_contract_v1,
)
from pymia.smartpyme.service_1_xlsx_delivery_v1 import build_service_1_xlsx_delivery_v1


def _bank_contract() -> dict[str, object]:
    return build_bank_reconciliation_contract_v1(
        contract_input={
            "owner_requested_output": "bank_reconciliation_scope_report",
            "source_files_received": ["extracto_banco", "archivo_contable"],
            "received_fields": ["fecha", "importe", "referencia"],
        }
    )


def _sandbox_release_gate() -> dict[str, object]:
    return evaluate_accounting_sandbox_release_gate_v1(
        gate_input={
            "capability_ref": "bank_reconciliation_basic",
            "responsible_role": "owner_or_accountant",
            "decision": "APPROVED",
            "scope_ok": True,
            "evidence_ok": True,
            "forbidden_claims": [],
            "live_use": False,
        }
    )


def _sandbox_input() -> dict[str, object]:
    return {
        "bank_contract": _bank_contract(),
        "sandbox_release_gate": _sandbox_release_gate(),
        "fixture_refs": ["bank_statement_fixture", "internal_ledger_fixture"],
        "live_use_requested": False,
    }


def test_ready_for_sandbox_contract_when_contract_gate_and_fixtures_are_ready() -> None:
    result = build_bank_reconciliation_sandbox_contract_v1(sandbox_input=_sandbox_input())

    assert result["status"] == "READY_FOR_SANDBOX_CONTRACT"
    assert result["capability_ref"] == CAPABILITY_REF
    assert result["sandbox_candidate_allowed"] is True
    assert result["runtime_authorized"] is False
    assert result["production_allowed"] is False
    assert result["next_allowed_action"] == "prepare_fixture_sandbox"


def test_blocks_when_bank_contract_is_not_ready() -> None:
    sandbox_input = _sandbox_input()
    bank_contract = _bank_contract()
    bank_contract["status"] = "MISSING_BANK_STATEMENT"
    sandbox_input["bank_contract"] = bank_contract

    result = build_bank_reconciliation_sandbox_contract_v1(sandbox_input=sandbox_input)  # type: ignore[arg-type]

    assert result["status"] == "BLOCKED_BANK_CONTRACT_NOT_READY"
    assert result["sandbox_candidate_allowed"] is False
    assert result["runtime_authorized"] is False


def test_blocks_when_sandbox_release_gate_did_not_pass() -> None:
    sandbox_input = _sandbox_input()
    sandbox_release_gate = _sandbox_release_gate()
    sandbox_release_gate["status"] = "PENDING"
    sandbox_input["sandbox_release_gate"] = sandbox_release_gate

    result = build_bank_reconciliation_sandbox_contract_v1(sandbox_input=sandbox_input)  # type: ignore[arg-type]

    assert result["status"] == "BLOCKED_SANDBOX_RELEASE_GATE_NOT_PASSED"
    assert result["next_allowed_action"] == "complete_sandbox_release_gate_first"


def test_blocks_live_use_request() -> None:
    sandbox_input = _sandbox_input()
    sandbox_input["live_use_requested"] = True

    result = build_bank_reconciliation_sandbox_contract_v1(sandbox_input=sandbox_input)  # type: ignore[arg-type]

    assert result["status"] == "BLOCKED_LIVE_USE"
    assert result["sandbox_candidate_allowed"] is False
    assert result["production_allowed"] is False


def test_missing_fixtures_blocks_sandbox_candidate() -> None:
    sandbox_input = _sandbox_input()
    sandbox_input["fixture_refs"] = ["bank_statement_fixture"]

    result = build_bank_reconciliation_sandbox_contract_v1(sandbox_input=sandbox_input)  # type: ignore[arg-type]

    assert result["status"] == "MISSING_FIXTURES"
    assert result["missing_fixture_refs"] == ["internal_ledger_fixture"]
    assert result["sandbox_candidate_allowed"] is False


def test_invalid_fixture_refs_shape_returns_invalid_input() -> None:
    sandbox_input = _sandbox_input()
    sandbox_input["fixture_refs"] = {"bad": True}

    result = build_bank_reconciliation_sandbox_contract_v1(sandbox_input=sandbox_input)  # type: ignore[arg-type]

    assert result["status"] == "INVALID_INPUT"
    assert result["runtime_authorized"] is False


def test_delivery_input_is_compatible_with_generic_xlsx_delivery(tmp_path: Path) -> None:
    result = build_bank_reconciliation_sandbox_contract_v1(sandbox_input=_sandbox_input())
    output_path = tmp_path / "bank_sandbox_contract.xlsx"

    delivery = build_service_1_xlsx_delivery_v1(
        delivery_input=result["delivery_input"],
        output_path=output_path,
    )
    workbook = load_workbook(output_path)

    assert output_path.exists()
    assert delivery["capability_ref"] == CAPABILITY_REF
    assert workbook["Resumen"]["B3"].value == CAPABILITY_REF
    assert ("sandbox_candidate_allowed", "true") in list(workbook["Resultados"].iter_rows(values_only=True))


def test_forbidden_claims_prevent_productive_accounting_claims() -> None:
    result = build_bank_reconciliation_sandbox_contract_v1(sandbox_input=_sandbox_input())
    forbidden_claims = result["delivery_input"]["forbidden_claims"]

    assert "No confirma saldo conciliado." in forbidden_claims
    assert "No genera asientos contables." in forbidden_claims
    assert "No certifica exactitud contable o fiscal." in forbidden_claims


def test_module_has_no_io_or_forbidden_dependencies() -> None:
    import pymia.smartpyme.bank_reconciliation_sandbox_contract_v1 as module

    source = inspect.getsource(module)

    assert "openpyxl" not in source
    assert "requests" not in source
    assert "httpx" not in source
    assert "api" not in source.lower()
    assert "mercado_pago" not in source.lower()
    assert "vertical_pipeline" not in source
    assert "service_1_pipeline" not in source
    assert "fsm" not in source.lower()
    assert "llm" not in source.lower()
    assert "chatbot" not in source.lower()
    assert "open(" not in source
    assert ".save(" not in source
    assert "read_text(" not in source
    assert "read_bytes(" not in source

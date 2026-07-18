from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from pymia.smartpyme.service_1_liq_001_outcome_v1 import (
    STATUS_BLOCKED,
    STATUS_READY,
    build_liq_001_outcome_v1,
    deliver_liq_001_outcome_xlsx_v1,
)


def _evaluated(classification: str = "SALES_PENDING_COLLECTION") -> dict:
    return {
        "status": "EVALUATED",
        "capability_ref": "sold_vs_collected_gap",
        "classification": classification,
        "inputs": {"sold_amount": 3000.0, "collected_amount": 2300.0},
        "computed": {
            "gap_amount": 700.0,
            "collection_ratio": 2300 / 3000,
            "gap_ratio": 700 / 3000,
        },
    }


def test_positive_gap_builds_bounded_finding_and_treatment() -> None:
    outcome = build_liq_001_outcome_v1(computation_result=_evaluated())

    assert outcome["status"] == STATUS_READY
    assert outcome["classification"] == "SALES_PENDING_COLLECTION"
    assert "vendido supera" in outcome["finding"]
    assert len(outcome["treatment_actions"]) == 3
    assert outcome["bounded_finding_generated"] is True
    assert outcome["causal_diagnosis_generated"] is False
    assert outcome["runtime_authorized"] is False
    assert outcome["delivery_authorized"] is False


def test_outcome_explicitly_forbids_unsupported_causal_claims() -> None:
    outcome = build_liq_001_outcome_v1(computation_result=_evaluated())

    rendered = " ".join(outcome["forbidden_claims"]).lower()
    assert "morosidad" in rendered
    assert "incobrable" in rendered
    assert "fraude" in rendered


def test_non_evaluated_result_is_blocked() -> None:
    result = build_liq_001_outcome_v1(
        computation_result={"status": "EVIDENCE_BLOCKED", "capability_ref": "sold_vs_collected_gap"}
    )
    assert result["status"] == STATUS_BLOCKED
    assert result["bounded_finding_generated"] is False


def test_delivery_writes_deterministic_owner_facing_workbook(tmp_path: Path) -> None:
    outcome = build_liq_001_outcome_v1(computation_result=_evaluated())
    delivered = deliver_liq_001_outcome_xlsx_v1(outcome=outcome, output_dir=tmp_path)

    assert delivered["status"] == "DELIVERED"
    assert delivered["bounded_finding_generated"] is True
    assert delivered["causal_diagnosis_generated"] is False
    assert delivered["delivery_authorized"] is False
    output = Path(delivered["delivery"]["output_path"])
    assert output.exists()
    workbook = load_workbook(output, data_only=False)
    assert workbook.sheetnames == [
        "Resumen",
        "Datos usados",
        "Resultados",
        "Faltantes",
        "Limitaciones",
        "Claims prohibidos",
        "Notas técnicas",
    ]
    summary = {row[0].value: row[1].value for row in workbook["Resumen"].iter_rows(min_row=2)}
    assert summary["status"] == "SALES_PENDING_COLLECTION"
    assert "vendido supera" in summary["owner_summary"]
    results = {row[0].value: row[1].value for row in workbook["Resultados"].iter_rows(min_row=2)}
    assert results["gap_amount"] == "700.0"
    assert "Identificar las operaciones" in results["treatment_actions"]

from __future__ import annotations

import inspect
from pathlib import Path

from openpyxl import Workbook

from pymia.smartpyme import service_1_web_column_confirmation_intake_boundary_v1 as intake
from pymia.smartpyme.service_1_canonical_ingestion_output_to_semantic_bridge_v1 import (
    build_service_1_semantic_bridge_from_canonical_ingestion_output_v1,
)
from pymia.smartpyme.service_1_assisted_semantic_product_wiring_v1 import (
    STATUS_CONFIRMED as SEM8_CONFIRMED,
    STATUS_OWNER_DIALOGUE_REQUIRED as SEM8_OWNER_REQUIRED,
    run_service_1_assisted_semantic_initial_v1,
    run_service_1_assisted_semantic_reentry_v1,
)
from pymia.smartpyme.service_1_owner_confirmation_to_canonical_ingestion_output_v1 import (
    BLOCK_MISSING_ANSWERS,
    BLOCK_UNKNOWN_COLUMNS,
    STATUS_UNCONFIRMED_READY,
    build_service_1_canonical_ingestion_output_from_owner_confirmation_v1,
    build_service_1_unconfirmed_canonical_ingestion_output_v1,
)
from pymia.smartpyme.service_1_xlsx_to_normalized_table_v1 import (
    read_xlsx_to_normalized_tables_v1,
)


def _write_multisheet_xlsx(path: Path) -> None:
    workbook = Workbook()
    ventas = workbook.active
    ventas.title = "Ventas"
    ventas.append(["fecha", "importe", "canal"])
    ventas.append(["2026-07-01", 1000, "Mostrador"])
    ventas.append(["2026-07-02", 1500, "Online"])

    cobros = workbook.create_sheet("Cobros")
    cobros.append(["fecha", "importe", "medio_pago"])
    cobros.append(["2026-07-03", 800, "Transferencia"])
    cobros.append(["2026-07-04", 700, "Efectivo"])
    workbook.save(path)
    workbook.close()


def _multisheet_packet(tmp_path: Path) -> dict:
    xlsx = tmp_path / "multisheet.xlsx"
    _write_multisheet_xlsx(xlsx)
    packet = intake.build_service_1_web_column_confirmation_intake_boundary_v1(
        local_xlsx_path=xlsx,
        include_all_sheets=True,
    )
    assert packet["status"] == "NEEDS_OWNER_CONFIRMATION"
    return packet


def _answers_by_question_id(packet: dict) -> dict[str, str]:
    return {
        ref["question_id"]: f"significado de {ref['sheet_name']}.{ref['column_name']}"
        for ref in packet["column_refs"]
    }


def test_canonical_intake_exposes_single_and_multiple_sheet_selection() -> None:
    parameters = inspect.signature(
        intake.build_service_1_web_column_confirmation_intake_boundary_v1
    ).parameters

    assert "sheet_name" in parameters
    assert "sheet_names" in parameters
    assert "include_all_sheets" in parameters


def test_canonical_reader_returns_all_non_empty_sheets_and_releases_file(
    tmp_path: Path,
) -> None:
    xlsx = tmp_path / "multisheet.xlsx"
    _write_multisheet_xlsx(xlsx)

    tables = read_xlsx_to_normalized_tables_v1(xlsx)

    assert [table["sheet_name"] for table in tables] == ["Ventas", "Cobros"]
    assert all(table["status"] == "OK" for table in tables)
    assert tables[0]["headers"] == ["fecha", "importe", "canal"]
    assert tables[1]["headers"] == ["fecha", "importe", "medio_pago"]

    replacement = tmp_path / "renamed.xlsx"
    xlsx.rename(replacement)
    assert replacement.exists()


def test_canonical_owner_questions_are_sheet_qualified_and_collision_safe(
    tmp_path: Path,
) -> None:
    packet = _multisheet_packet(tmp_path)

    assert packet["sheet_names"] == ["Ventas", "Cobros"]
    assert packet["normalized_table"] is None
    assert len(packet["normalized_tables"]) == 2
    assert packet["question_count"] == len(packet["column_refs"]) == 6
    assert len({ref["question_id"] for ref in packet["column_refs"]}) == 6
    assert [ref["column_name"] for ref in packet["column_refs"]].count("fecha") == 2
    assert [ref["column_name"] for ref in packet["column_refs"]].count("importe") == 2
    assert all(question["sheet_name"] in question["question"] for question in packet["owner_questions"])
    assert all(question["field_id"] == question["question_id"] for question in packet["owner_questions"])


def test_multisheet_connector_requires_question_ids_and_preserves_evidence(
    tmp_path: Path,
) -> None:
    packet = _multisheet_packet(tmp_path)

    unconfirmed = build_service_1_unconfirmed_canonical_ingestion_output_v1(
        owner_question_packet=packet,
    )
    assert unconfirmed["status"] == STATUS_UNCONFIRMED_READY
    unconfirmed_ingestion = unconfirmed["ingestion_output"]
    assert unconfirmed_ingestion["provenance"]["sheet_names"] == ["Ventas", "Cobros"]
    assert unconfirmed_ingestion["column_refs"] == packet["column_refs"]
    assert unconfirmed_ingestion["normalized_tables"] == packet["normalized_tables"]
    assert unconfirmed_ingestion["physical_lineage"]

    legacy_answers = {column: f"significado de {column}" for column in packet["columns"]}
    blocked = build_service_1_canonical_ingestion_output_from_owner_confirmation_v1(
        owner_question_packet=packet,
        owner_answers=legacy_answers,
    )
    assert blocked["status"] == "BLOCKED"
    assert blocked["blocked_reason"] in {BLOCK_UNKNOWN_COLUMNS, BLOCK_MISSING_ANSWERS}

    result = build_service_1_canonical_ingestion_output_from_owner_confirmation_v1(
        owner_question_packet=packet,
        owner_answers=_answers_by_question_id(packet),
    )

    assert result["status"] == "INGESTION_OUTPUT_READY"
    ingestion = result["ingestion_output"]
    assert "sheet_name" not in ingestion
    assert ingestion["provenance"]["sheet_names"] == ["Ventas", "Cobros"]
    assert len(ingestion["column_refs"]) == 6
    assert {ref["field_id"] for ref in ingestion["column_refs"]} == {
        ref["question_id"] for ref in packet["column_refs"]
    }
    assert all(ref.get("owner_meaning") for ref in ingestion["column_refs"])
    assert ingestion["normalized_tables"] == packet["normalized_tables"]


def test_semantic_bridge_preserves_sheet_identity_for_duplicate_headers(
    tmp_path: Path,
) -> None:
    packet = _multisheet_packet(tmp_path)
    connector = build_service_1_unconfirmed_canonical_ingestion_output_v1(
        owner_question_packet=packet,
    )

    bridge = build_service_1_semantic_bridge_from_canonical_ingestion_output_v1(
        ingestion_output=connector["ingestion_output"]
    )

    assert bridge["status"] == "SEMANTIC_CANDIDATES_READY"
    entries = bridge["confirmation_matrix"].entries
    assert len(entries) == 6
    assert {(entry.sheet_name, entry.original_column_name) for entry in entries} == {
        ("Ventas", "fecha"),
        ("Ventas", "importe"),
        ("Ventas", "canal"),
        ("Cobros", "fecha"),
        ("Cobros", "importe"),
        ("Cobros", "medio_pago"),
    }
    assert all(entry.sample_values for entry in entries)
    assert all(entry.owner_confirmed_role is None for entry in entries)
    assert {entry.inferred_type for entry in entries} >= {"date", "number", "text"}
    assert {candidate.sheet_name for candidate in bridge["column_candidates"]} == {
        "Ventas",
        "Cobros",
    }


def test_explicit_single_sheet_selection_preserves_canonical_answer_input(
    tmp_path: Path,
) -> None:
    xlsx = tmp_path / "multisheet.xlsx"
    _write_multisheet_xlsx(xlsx)
    packet = intake.build_service_1_web_column_confirmation_intake_boundary_v1(
        local_xlsx_path=xlsx,
        sheet_name="Cobros",
    )

    assert packet["sheet_names"] == ["Cobros"]
    assert packet["columns"] == ["fecha", "importe", "medio_pago"]
    assert all(ref["field_id"] == ref["column_name"] for ref in packet["column_refs"])

    answers = {column: f"significado de {column}" for column in packet["columns"]}
    result = build_service_1_canonical_ingestion_output_from_owner_confirmation_v1(
        owner_question_packet=packet,
        owner_answers=answers,
    )
    assert result["status"] == "INGESTION_OUTPUT_READY"
    assert all(
        ref["owner_meaning"] == answers[ref["field_id"]]
        for ref in result["ingestion_output"]["column_refs"]
    )
    assert "owner_answers" not in result
    assert result["ingestion_output"]["provenance"]["sheet_names"] == ["Cobros"]


def test_default_single_sheet_and_explicit_all_sheets_are_distinct_scopes(
    tmp_path: Path,
) -> None:
    xlsx = tmp_path / "multisheet.xlsx"
    _write_multisheet_xlsx(xlsx)

    default_packet = intake.build_service_1_web_column_confirmation_intake_boundary_v1(
        local_xlsx_path=xlsx
    )
    all_packet = intake.build_service_1_web_column_confirmation_intake_boundary_v1(
        local_xlsx_path=xlsx,
        include_all_sheets=True,
    )

    assert default_packet["sheet_names"] == ["Ventas"]
    assert default_packet["question_count"] == 3
    assert all_packet["sheet_names"] == ["Ventas", "Cobros"]
    assert all_packet["question_count"] == 6
    assert default_packet["case_id"] != all_packet["case_id"]


def test_duplicate_ambiguous_headers_reenter_semantic_loop_by_question_id(
    tmp_path: Path,
) -> None:
    xlsx = tmp_path / "ambiguous_multisheet.xlsx"
    workbook = Workbook()
    ventas = workbook.active
    ventas.title = "Ventas"
    ventas.append(["monto"])
    ventas.append([1000])
    cobros = workbook.create_sheet("Cobros")
    cobros.append(["monto"])
    cobros.append([800])
    workbook.save(xlsx)
    workbook.close()

    packet = intake.build_service_1_web_column_confirmation_intake_boundary_v1(
        local_xlsx_path=xlsx,
        include_all_sheets=True,
    )
    connector = build_service_1_canonical_ingestion_output_from_owner_confirmation_v1(
        owner_question_packet=packet,
        owner_answers={
            ref["question_id"]: f"importe de {ref['sheet_name']}"
            for ref in packet["column_refs"]
        },
    )

    def provider(_payload: dict) -> dict:
        return {
            "schema_version": "SERVICE_1_LLM_SEMANTIC_PROPOSAL_V1",
            "concept_proposals": [
                {
                    "proposal_id": "p1",
                    "target_column_refs": ["Ventas.monto"],
                    "semantic_role": "sales_amount",
                    "variable_name": "sold_amount",
                    "confidence": 0.95,
                    "rationale": "explicit test proposal",
                    "evidence_refs": ["ev:column:Ventas.monto:type"],
                },
                {
                    "proposal_id": "p2",
                    "target_column_refs": ["Cobros.monto"],
                    "semantic_role": "unit_sale_price",
                    "variable_name": "sale_price",
                    "confidence": 0.95,
                    "rationale": "explicit test proposal",
                    "evidence_refs": ["ev:column:Cobros.monto:type"],
                },
            ],
            "relationship_proposals": [],
            "duplicate_semantics": [],
            "irrelevant_refs": [],
            "material_ambiguities": [],
        }

    initial = run_service_1_assisted_semantic_initial_v1(
        ingestion_output=connector["ingestion_output"],
        requested_capability="sold_vs_collected_gap",
        provider=provider,
    )

    assert initial["status"] == SEM8_OWNER_REQUIRED
    questions = initial["owner_questions"]
    assert questions
    assert any("Ventas.monto" in question["column_refs"] for question in questions)
    assert len({question["decision_id"] for question in questions}) == len(questions)
    semantic_answers = {
        question["decision_id"]: {"action": "ACCEPT"} for question in questions
    }
    final = run_service_1_assisted_semantic_reentry_v1(
        previous_state=initial,
        owner_responses=[
            {"decision_id": decision_id, **response}
            for decision_id, response in semantic_answers.items()
        ],
        owner_actor_id="owner-multisheet",
        owner_actor_role="OWNER",
        file_ref=connector["ingestion_output"]["provenance"].get("source_file_ref"),
    )

    assert final["status"] == SEM8_CONFIRMED
    reinjected = final["sem6_packet"]["reentry_packet"]["column_candidates"]
    assert len(reinjected) == 2
    assert all(candidate.owner_confirmation_required is False for candidate in reinjected)
    assert {
        (candidate.sheet_name, candidate.source_column_name) for candidate in reinjected
    } == {("Ventas", "monto"), ("Cobros", "monto")}


def test_uploaded_bytes_support_explicit_all_sheets(tmp_path: Path) -> None:
    xlsx = tmp_path / "uploaded_multisheet.xlsx"
    _write_multisheet_xlsx(xlsx)

    packet = intake.build_service_1_web_column_confirmation_intake_boundary_v1(
        uploaded_xlsx_bytes=xlsx.read_bytes(),
        uploaded_filename="uploaded_multisheet.xlsx",
        include_all_sheets=True,
    )

    assert packet["status"] == "NEEDS_OWNER_CONFIRMATION"
    assert packet["source_kind"] == "uploaded_bytes"
    assert packet["sheet_names"] == ["Ventas", "Cobros"]
    assert packet["question_count"] == 6
    assert len({ref["question_id"] for ref in packet["column_refs"]}) == 6

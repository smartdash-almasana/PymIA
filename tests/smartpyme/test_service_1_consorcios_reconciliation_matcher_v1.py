from __future__ import annotations

import io
from pathlib import Path

from openpyxl import Workbook, load_workbook

from pymia.smartpyme.service_1_assisted_web_v1 import (
    AssistedWebApplicationV1,
    BANK_RECONCILIATION,
)


def _sheet_bytes(path: Path, sheet_name: str) -> bytes:
    source = load_workbook(path, data_only=True, read_only=True)
    sheet = source[sheet_name]
    output = Workbook()
    target = output.active
    target.title = sheet_name
    for row in sheet.iter_rows(values_only=True):
        target.append(list(row))
    buffer = io.BytesIO()
    output.save(buffer)
    return buffer.getvalue()


def _run_case(path: Path, *, label: str) -> dict[str, object]:
    app = AssistedWebApplicationV1(output_dir=Path(".tmp_consorcios_outputs"))
    session_id = f"consorcios-matcher-{label}"
    app.start_reconciliation(
        session_id=session_id,
        reconciliation_type=BANK_RECONCILIATION,
    )
    status_upload, _ = app.receive_reconciliation_sources(
        session_id=session_id,
        files={
            "source_bank": (f"{label}_Banco.xlsx", _sheet_bytes(path, "Banco")),
            "source_internal": (
                f"{label}_Cobranzas.xlsx",
                _sheet_bytes(path, "Cobranzas"),
            ),
        },
    )
    assert int(status_upload) == 200

    status_confirm, _ = app.confirm_reconciliation_columns(
        session_id=session_id,
        fields={
            "bind_bank_id": "movimiento_id",
            "bind_bank_fecha": "fecha",
            "bind_bank_importe": "importe",
            "bind_bank_referencia": "referencia",
            "bind_internal_id": "unidad_funcional",
            "bind_internal_fecha": "fecha_pago",
            "bind_internal_importe": "importe",
            "bind_internal_referencia": "referencia",
        },
    )
    assert int(status_confirm) == 200

    packet = app.session(session_id).reconciliation_result
    assert isinstance(packet, dict)
    assisted = packet["reconciliation_run"]["assisted_review"]
    return assisted["review_result"]["source_result"]


def test_consorcios_generic_references_remain_ambiguous_and_stronger_amount_anchor_wins() -> None:
    base = Path("prueba_excels")
    cabildo = _run_case(
        base / "PYMIA_CONSORCIO_CABILDO_2026_07.xlsx",
        label="cabildo",
    )
    rivadavia = _run_case(
        base / "PYMIA_CONSORCIO_RIVADAVIA_2026_07.xlsx",
        label="rivadavia",
    )

    assert any(
        "MOV-0006" in group.get("banco_ids", [])
        for group in cabildo["matches_ambiguos"]
    )
    assert "MOV-0006" not in {
        item["banco_id"] for item in cabildo["matches_exactos"]
    }

    assert any(
        "MOV-0104" in group.get("banco_ids", [])
        for group in rivadavia["matches_ambiguos"]
    )
    assert "MOV-0104" not in {
        item["banco_id"] for item in rivadavia["matches_exactos"]
    }
    assert any(
        item["banco_id"] == "MOV-0115" and item["interno_id"] == "3A"
        for item in rivadavia["diferencias_importe"]
    )
    assert any(
        item["id"] == "MOV-0109"
        for item in rivadavia["banco_sin_imputar"]
    )

from __future__ import annotations

import inspect
from pathlib import Path

from openpyxl import load_workbook

from pymia.smartpyme.accounting_workpaper_contract_v1 import (
    ACCOUNTING_WORKPAPER_CAPABILITY_REF,
    ACCOUNTING_WORKPAPER_CONTRACT_REF,
    build_accounting_workpaper_contract_v1,
)
from pymia.smartpyme.service_1_xlsx_delivery_v1 import build_service_1_xlsx_delivery_v1


def _contract_input() -> dict[str, object]:
    return {
        "owner_requested_output": "accounting_workpaper_scope_report",
        "source_files_received": ["evidencia_soporte", "plantilla_papel_trabajo"],
        "received_fields": ["periodo", "cliente", "area_revision", "responsable"],
    }


def test_ready_for_review_with_supporting_evidence_template_and_required_fields() -> None:
    result = build_accounting_workpaper_contract_v1(contract_input=_contract_input())

    assert result["status"] == "READY_FOR_REVIEW"
    assert result["contract_ref"] == ACCOUNTING_WORKPAPER_CONTRACT_REF
    assert result["runtime_authorized"] is False
    assert result["missing_sources"] == []
    assert result["missing_fields"] == []
    assert result["next_allowed_action"] == "manual_accounting_workpaper_review"


def test_missing_supporting_evidence_when_evidencia_soporte_is_absent() -> None:
    contract_input = _contract_input()
    contract_input["source_files_received"] = ["plantilla_papel_trabajo"]

    result = build_accounting_workpaper_contract_v1(contract_input=contract_input)

    assert result["status"] == "MISSING_SUPPORTING_EVIDENCE"
    assert result["missing_sources"] == ["evidencia_soporte"]
    assert result["next_allowed_action"] == "request_supporting_evidence"


def test_missing_workpaper_template_when_plantilla_papel_trabajo_is_absent() -> None:
    contract_input = _contract_input()
    contract_input["source_files_received"] = ["evidencia_soporte"]

    result = build_accounting_workpaper_contract_v1(contract_input=contract_input)

    assert result["status"] == "MISSING_WORKPAPER_TEMPLATE"
    assert result["missing_sources"] == ["plantilla_papel_trabajo"]
    assert result["next_allowed_action"] == "request_workpaper_template"


def test_missing_fields_when_required_field_is_absent() -> None:
    contract_input = _contract_input()
    contract_input["received_fields"] = ["periodo", "cliente", "responsable"]

    result = build_accounting_workpaper_contract_v1(contract_input=contract_input)

    assert result["status"] == "MISSING_FIELDS"
    assert result["missing_fields"] == ["area_revision"]
    assert result["next_allowed_action"] == "request_missing_accounting_workpaper_fields"


def test_invalid_input_when_sources_are_not_list_of_strings() -> None:
    contract_input = _contract_input()
    contract_input["source_files_received"] = {"evidencia_soporte": True}

    result = build_accounting_workpaper_contract_v1(contract_input=contract_input)  # type: ignore[arg-type]

    assert result["status"] == "INVALID_INPUT"
    assert result["next_allowed_action"] == "fix_invalid_accounting_workpaper_contract_input"


def test_invalid_input_when_fields_are_not_list_of_strings() -> None:
    contract_input = _contract_input()
    contract_input["received_fields"] = {"periodo": True}

    result = build_accounting_workpaper_contract_v1(contract_input=contract_input)  # type: ignore[arg-type]

    assert result["status"] == "INVALID_INPUT"
    assert result["delivery_input"]["status"] == "INVALID_INPUT"


def test_wraps_base_accounting_contract_without_runtime_workpaper_generation() -> None:
    result = build_accounting_workpaper_contract_v1(contract_input=_contract_input())

    accounting_contract = result["accounting_contract"]
    assert accounting_contract["contract_ref"] == "accounting_workpaper_basic"
    assert accounting_contract["family"] == "accounting_workpaper"
    assert accounting_contract["runtime_authorized"] is False
    assert accounting_contract["next_allowed_action"] == "manual_accounting_review"


def test_delivery_input_is_compatible_with_generic_xlsx_delivery(tmp_path: Path) -> None:
    result = build_accounting_workpaper_contract_v1(contract_input=_contract_input())
    output_path = tmp_path / "accounting_workpaper_contract.xlsx"

    delivery = build_service_1_xlsx_delivery_v1(
        delivery_input=result["delivery_input"],
        output_path=output_path,
    )

    workbook = load_workbook(output_path)

    assert output_path.exists()
    assert delivery["capability_ref"] == ACCOUNTING_WORKPAPER_CAPABILITY_REF
    assert workbook["Resumen"]["B3"].value == ACCOUNTING_WORKPAPER_CAPABILITY_REF
    assert ("accounting_workpaper_status", "READY_FOR_REVIEW") in list(
        workbook["Resultados"].iter_rows(values_only=True)
    )


def test_forbidden_claims_prevent_certified_workpaper_or_accounting_runtime() -> None:
    result = build_accounting_workpaper_contract_v1(contract_input=_contract_input())
    forbidden_claims = result["delivery_input"]["forbidden_claims"]

    assert "No confirma auditoría contable certificada." in forbidden_claims
    assert "No confirma exactitud fiscal." in forbidden_claims
    assert "No genera asientos contables automáticos." in forbidden_claims


def test_module_has_no_io_openpyxl_or_forbidden_runtime_dependencies() -> None:
    import pymia.smartpyme.accounting_workpaper_contract_v1 as module

    source = inspect.getsource(module)

    assert "openpyxl" not in source
    assert "first_aid" not in source
    assert "exceland" not in source.lower()
    assert "vertical_pipeline" not in source
    assert "service_1_pipeline" not in source
    assert "fsm" not in source.lower()
    assert "llm" not in source.lower()
    assert "chatbot" not in source.lower()
    assert "mercado_pago" not in source.lower()
    assert "open(" not in source
    assert ".save(" not in source
    assert "read_text(" not in source
    assert "read_bytes(" not in source

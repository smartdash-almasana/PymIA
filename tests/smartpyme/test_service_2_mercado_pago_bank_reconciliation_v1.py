from __future__ import annotations

import inspect
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from pymia.smartpyme.service_2_mercado_pago_bank_reconciliation_v1 import (
    AMBIGUO,
    COINCIDENCIA_FECHA_CERCANA,
    COINCIDENCIA_IMPORTE_NETO,
    COINCIDENCIA_LOTE,
    COINCIDENCIA_REFERENCIA_EXACTA,
    DIFERENCIA_IMPORTE,
    build_mercado_pago_bank_reconciliation_v1,
)


REPO_ROOT = Path(__file__).resolve().parents[2]
FIXTURE_PATH = REPO_ROOT / "prueba_excels" / "conciliacion_mercado_pago_banco_corregida.xlsx"


def _rows_from_sheet(path: Path, sheet_name: str) -> list[dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        headers = [str(value) if value is not None else "" for value in next(rows)]
        return [dict(zip(headers, row)) for row in rows if any(value is not None for value in row)]
    finally:
        workbook.close()


def _excel_date(value: object) -> object:
    if isinstance(value, (date, datetime)):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return from_excel(value)
    return value


def _physical_inputs() -> tuple[list[dict[str, object]], list[dict[str, object]], list[dict[str, object]]]:
    mp_rows = _rows_from_sheet(FIXTURE_PATH, "MERCADO_PAGO")
    bank_rows = _rows_from_sheet(FIXTURE_PATH, "BANCO")
    expected_rows = _rows_from_sheet(FIXTURE_PATH, "CASO_ESPERADO")
    for row in mp_rows:
        row["fecha_operacion"] = _excel_date(row.get("fecha_operacion"))
    for row in bank_rows:
        row["fecha"] = _excel_date(row.get("fecha"))
    return mp_rows, bank_rows, expected_rows


def _split_ids(value: object) -> frozenset[str]:
    if value is None:
        return frozenset()
    text = str(value).strip()
    if not text or text == "NINGUNO":
        return frozenset()
    return frozenset(part.strip() for part in text.split(",") if part.strip())


def _result_index(result: dict[str, object]) -> dict[tuple[frozenset[str], frozenset[str]], dict[str, object]]:
    records = [
        *result["conciliaciones"],  # type: ignore[index]
        *result["ambiguos"],  # type: ignore[index]
        *result["diferencias_importe"],  # type: ignore[index]
    ]
    return {
        (
            frozenset(str(value) for value in record["operaciones_mp_ids"]),
            frozenset(str(value) for value in record["movimientos_banco_ids"]),
        ): record
        for record in records
    }


def test_net_formula_is_calculated_before_matching() -> None:
    result = build_mercado_pago_bank_reconciliation_v1(
        mercado_pago_operations=[
            {
                "operacion_mp_id": "MP-1",
                "fecha_operacion": "2026-06-01",
                "importe_bruto": 100000,
                "comision": 4000,
                "retencion": 1500,
                "importe_neto": 94500,
                "lote_id": "LOTE-1",
                "referencia": "REF-1",
            }
        ],
        bank_movements=[
            {
                "movimiento_banco_id": "BAN-1",
                "fecha": "2026-06-01",
                "importe": 94500,
                "lote_id": "LOTE-1",
                "referencia": "REF-1",
            }
        ],
    )

    assert result["status"] == "READY_FOR_HUMAN_REVIEW"
    assert result["inconsistencias_calculo"] == []
    assert result["conciliaciones"][0]["resultado"] == COINCIDENCIA_REFERENCIA_EXACTA  # type: ignore[index]
    assert result["conciliaciones"][0]["importe_neto_esperado"] == 94500.0  # type: ignore[index]


def test_multiple_operations_and_split_bank_movements_are_reconciled_by_lot_sum() -> None:
    result = build_mercado_pago_bank_reconciliation_v1(
        mercado_pago_operations=[
            {
                "operacion_mp_id": "MP-1",
                "fecha_operacion": "2026-06-01",
                "importe_bruto": 60000,
                "comision": 2400,
                "retencion": 900,
                "importe_neto": 56700,
                "lote_id": "LOTE-1",
                "referencia": "REF-A",
            },
            {
                "operacion_mp_id": "MP-2",
                "fecha_operacion": "2026-06-01",
                "importe_bruto": 40000,
                "comision": 1600,
                "retencion": 600,
                "importe_neto": 37800,
                "lote_id": "LOTE-1",
                "referencia": "REF-B",
            },
        ],
        bank_movements=[
            {
                "movimiento_banco_id": "BAN-1",
                "fecha": "2026-06-01",
                "importe": 50000,
                "lote_id": "LOTE-1",
                "referencia": "REF-LOTE",
            },
            {
                "movimiento_banco_id": "BAN-2",
                "fecha": "2026-06-01",
                "importe": 44500,
                "lote_id": "LOTE-1",
                "referencia": "REF-LOTE",
            },
        ],
    )

    match = result["conciliaciones"][0]  # type: ignore[index]
    assert match["resultado"] == COINCIDENCIA_LOTE
    assert match["cardinalidad"] == "N:M"
    assert match["operaciones_mp_ids"] == ["MP-1", "MP-2"]
    assert match["movimientos_banco_ids"] == ["BAN-1", "BAN-2"]
    assert match["importe_neto_esperado"] == 94500.0
    assert match["importe_banco_total"] == 94500.0


def test_duplicate_bank_accreditation_remains_ambiguous() -> None:
    result = build_mercado_pago_bank_reconciliation_v1(
        mercado_pago_operations=[
            {
                "operacion_mp_id": "MP-1",
                "fecha_operacion": "2026-06-01",
                "importe_bruto": 100000,
                "comision": 4000,
                "retencion": 1500,
                "importe_neto": 94500,
                "lote_id": "LOTE-1",
                "referencia": "REF-1",
            }
        ],
        bank_movements=[
            {
                "movimiento_banco_id": "BAN-1",
                "fecha": "2026-06-01",
                "importe": 94500,
                "lote_id": "LOTE-1",
                "referencia": "REF-1",
            },
            {
                "movimiento_banco_id": "BAN-2",
                "fecha": "2026-06-01",
                "importe": 94500,
                "lote_id": "LOTE-1",
                "referencia": "REF-1",
            },
        ],
    )

    assert result["conciliaciones"] == []
    assert result["ambiguos"][0]["resultado"] == AMBIGUO  # type: ignore[index]
    assert result["ambiguos"][0]["cardinalidad"] == "1:N"  # type: ignore[index]
    assert result["operaciones_mp_sin_acreditacion"] == []
    assert result["movimientos_banco_sin_operacion_mp"] == []


def test_inconsistent_declared_net_is_not_silently_reconciled() -> None:
    result = build_mercado_pago_bank_reconciliation_v1(
        mercado_pago_operations=[
            {
                "operacion_mp_id": "MP-1",
                "fecha_operacion": "2026-06-01",
                "importe_bruto": 100000,
                "comision": 4000,
                "retencion": 1500,
                "importe_neto": 95000,
                "lote_id": "LOTE-1",
            }
        ],
        bank_movements=[
            {
                "movimiento_banco_id": "BAN-1",
                "fecha": "2026-06-01",
                "importe": 95000,
                "lote_id": "LOTE-1",
            }
        ],
    )

    assert result["status"] == "PARTIAL_MATCHES_FOUND"
    assert result["conciliaciones"] == []
    assert result["inconsistencias_calculo"] == [
        {
            "operacion_mp_id": "MP-1",
            "importe_neto_declarado": 95000.0,
            "importe_neto_calculado": 94500.0,
            "diferencia": 500.0,
        }
    ]
    assert result["movimientos_banco_sin_operacion_mp"][0]["movimiento_banco_id"] == "BAN-1"  # type: ignore[index]


def test_physical_mercado_pago_workbook_is_processed_by_pymia() -> None:
    mp_rows, bank_rows, expected_rows = _physical_inputs()
    result = build_mercado_pago_bank_reconciliation_v1(mp_rows, bank_rows)
    actual = _result_index(result)

    assert result["status"] == "PARTIAL_MATCHES_FOUND"
    assert result["inconsistencias_calculo"] == []
    assert len(result["conciliaciones"]) == 10
    assert len(result["diferencias_importe"]) == 1
    assert len(result["ambiguos"]) == 2

    expected_types = {
        "CASO-01": COINCIDENCIA_REFERENCIA_EXACTA,
        "CASO-02": COINCIDENCIA_IMPORTE_NETO,
        "CASO-03": COINCIDENCIA_LOTE,
        "CASO-04": COINCIDENCIA_FECHA_CERCANA,
        "CASO-05": DIFERENCIA_IMPORTE,
        "CASO-07": COINCIDENCIA_LOTE,
        "CASO-08": COINCIDENCIA_LOTE,
        "CASO-09": COINCIDENCIA_IMPORTE_NETO,
        "CASO-11": AMBIGUO,
        "CASO-12": AMBIGUO,
        "CASO-15": COINCIDENCIA_LOTE,
    }

    for row in expected_rows:
        case_id = str(row["caso_id"])
        operation_ids = _split_ids(row.get("operaciones_mp_ids"))
        bank_ids = _split_ids(row.get("movimientos_banco_ids"))
        if not operation_ids or not bank_ids:
            continue
        record = actual[(operation_ids, bank_ids)]
        assert record["cardinalidad"] == row["cardinalidad"]
        if case_id in expected_types:
            assert record["resultado"] == expected_types[case_id]

    # CASO-06 y CASO-10 son indistinguibles por estructura: ambos tienen dos
    # operaciones con la misma referencia y un banco del mismo lote. PymIA no
    # inventa una diferencia semantica y clasifica ambos por evidencia de lote.
    assert actual[(_split_ids("MP-2011A, MP-2011B"), _split_ids("BAN-8006"))]["resultado"] == COINCIDENCIA_LOTE
    assert actual[(_split_ids("MP-2015A, MP-2015B"), _split_ids("BAN-8010"))]["resultado"] == COINCIDENCIA_LOTE

    difference = actual[(_split_ids("MP-2008"), _split_ids("BAN-8005"))]
    assert difference["diferencia"] == -1500.0
    assert {item["operacion_mp_id"] for item in result["operaciones_mp_sin_acreditacion"]} == {"MP-2017"}  # type: ignore[index]
    assert {item["movimiento_banco_id"] for item in result["movimientos_banco_sin_operacion_mp"]} == {"BAN-8015"}  # type: ignore[index]


def test_module_is_pure_deterministic_and_reuses_general_matcher() -> None:
    import pymia.smartpyme.service_2_mercado_pago_bank_reconciliation_v1 as module

    source = inspect.getsource(module)

    assert "build_reconciliation_match_candidates_v1" in source
    assert "openpyxl" not in source
    assert "pandas" not in source
    assert "llm" not in source.lower()
    assert "open(" not in source
    assert ".save(" not in source

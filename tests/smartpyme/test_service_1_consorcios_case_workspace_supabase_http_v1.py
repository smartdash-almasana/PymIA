from __future__ import annotations

import os
from http.client import HTTPConnection
from pathlib import Path
from threading import Thread
from urllib.parse import urlencode

import pytest
from openpyxl import Workbook

from pymia.smartpyme.service_1_assisted_web_v1 import create_assisted_web_server_v1
from pymia.smartpyme.service_1_supabase_identity_resolver_v1 import (
    Service1SupabaseIdentityResolverV1,
    create_service_1_supabase_identity_client_v1,
    load_service_1_supabase_identity_config_v1,
)

_TEST_EMAIL = "neoalmasana@gmail.com"
_REQUIRED_ENV = (
    "PYMIA_SUPABASE_URL",
    "PYMIA_SUPABASE_SERVICE_ROLE_KEY",
    "PYMIA_SUPABASE_PUBLISHABLE_KEY",
    "SUPABASE_TEST_PASSWORD",
)


def _supabase_ready() -> bool:
    return all(str(os.environ.get(name) or "").strip() for name in _REQUIRED_ENV)


def _xlsx_bytes(tmp_path: Path) -> bytes:
    path = tmp_path / "consorcio_workspace.xlsx"
    workbook = Workbook()
    expensas = workbook.active
    expensas.title = "Expensas"
    expensas.append(["unidad", "saldo", "expensa"])
    expensas.append(["UF-12", 250, 100])
    gastos = workbook.create_sheet("Gastos")
    gastos.append(["rubro_gasto", "importe_gasto"])
    gastos.append(["Limpieza", 150])
    presupuesto = workbook.create_sheet("Presupuesto")
    presupuesto.append(["rubro_presupuesto", "presupuesto", "historico"])
    presupuesto.append(["Limpieza", 100, 100])
    workbook.save(path)
    return path.read_bytes()


def _multipart_with_context(
    filename: str, content: bytes, token: str, *, period: str, cookie: str = ""
) -> tuple[bytes, dict[str, str]]:
    boundary = "Service1ConsorciosWorkspaceSupabaseBoundary"
    chunks: list[bytes] = []
    for name, value in (
        ("consorcio_id", "rivadavia-1200"),
        ("consorcio_name", "Rivadavia 1200"),
        ("period", period),
    ):
        chunks.append(
            (
                f"--{boundary}\r\n"
                f'Content-Disposition: form-data; name="{name}"\r\n\r\n'
                f"{value}\r\n"
            ).encode("utf-8")
        )
    chunks.append(
        (
            f"--{boundary}\r\n"
            f'Content-Disposition: form-data; name="file"; filename="{filename}"\r\n'
            "Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n\r\n"
        ).encode("utf-8")
        + content
        + b"\r\n"
    )
    chunks.append(f"--{boundary}--\r\n".encode("utf-8"))
    body = b"".join(chunks)
    headers = {
        "Content-Type": f"multipart/form-data; boundary={boundary}",
        "Content-Length": str(len(body)),
        "Authorization": f"Bearer {token}",
    }
    if cookie:
        headers["Cookie"] = cookie
    return body, headers


def _post_form(server, path: str, cookie: str, token: str, fields: dict[str, str]):
    body = urlencode(fields).encode("utf-8")
    return _request(
        server,
        "POST",
        path,
        body=body,
        headers={
            "Content-Type": "application/x-www-form-urlencoded",
            "Content-Length": str(len(body)),
            "Cookie": cookie,
            "Authorization": f"Bearer {token}",
        },
    )


def _request(server, method: str, path: str, *, body: bytes = b"", headers=None):
    connection = HTTPConnection("127.0.0.1", server.server_port, timeout=20)
    connection.request(method, path, body=body, headers=headers or {})
    response = connection.getresponse()
    payload = response.read().decode("utf-8")
    return response.status, response.getheaders(), payload


def _cookie(headers: list[tuple[str, str]]) -> str:
    for name, value in headers:
        if name.lower() == "set-cookie":
            return value.split(";", 1)[0]
    raise AssertionError("missing session cookie")


@pytest.mark.skipif(not _supabase_ready(), reason="Supabase RADAR environment is not configured")
def test_consorcios_case_workspace_physical_supabase_http(tmp_path: Path) -> None:
    identity_config = load_service_1_supabase_identity_config_v1()
    identity_client = create_service_1_supabase_identity_client_v1(identity_config)

    password = str(os.environ["SUPABASE_TEST_PASSWORD"]).strip()
    login = identity_client.auth.sign_in_with_password(
        {"email": _TEST_EMAIL, "password": password}
    )
    token = login.session.access_token
    assert token

    resolver = Service1SupabaseIdentityResolverV1(identity_client)
    identity = resolver.__call__(
        type("Handler", (), {"headers": {"Authorization": f"Bearer {token}"}})()
    )
    assert identity["tenant_id"]

    server = create_assisted_web_server_v1(
        host="127.0.0.1",
        port=0,
        output_dir=tmp_path / "outputs-workspace",
        tenant_identity_resolver=resolver,
    )
    thread = Thread(target=server.serve_forever, daemon=True)
    thread.start()
    try:
        upload_body, upload_headers = _multipart_with_context(
            "agosto.xlsx", _xlsx_bytes(tmp_path), token, period="2026-08"
        )
        status, headers, _ = _request(
            server, "POST", "/upload", body=upload_body, headers=upload_headers
        )
        assert status == 200
        cookie = _cookie(headers)

        status, _, workspace = _request(
            server, "GET", "/consorcios-case", headers={"Cookie": cookie}
        )
        assert status == 200
        assert "Rivadavia 1200" in workspace
        assert "2026-08" in workspace
        assert "agosto.xlsx" in workspace
        assert "Cobranzas y deuda" in workspace
        assert "Gastos" in workspace
        assert "Banco" in workspace
        assert "RADAR" in workspace
        assert 'name="unidad_funcional"' in workspace

        status, _, result = _post_form(
            server,
            "/run-consorcios-collection-aging",
            cookie,
            token,
            {
                "sheet_name": "Expensas",
                "unidad_funcional": "unidad",
                "saldo_anterior": "saldo",
                "expensa_mes": "expensa",
            },
        )
        assert status == 200
        assert "UF-12" in result
        assert "2.5" in result
        assert "/consorcios-case" in result

        status, _, workspace = _request(
            server, "GET", "/consorcios-case", headers={"Cookie": cookie}
        )
        assert status == 200
        assert "Las columnas de este control ya fueron confirmadas para este caso." in workspace

        status, _, rerun = _post_form(
            server, "/run-consorcios-collection-aging", cookie, token, {}
        )
        assert status == 200
        assert "UF-12" in rerun
        assert "2.5" in rerun

        status, _, result = _post_form(
            server,
            "/run-consorcios-expense-variance",
            cookie,
            token,
            {
                "expense_sheet": "Gastos",
                "expense_rubro": "rubro_gasto",
                "expense_importe": "importe_gasto",
                "budget_sheet": "Presupuesto",
                "budget_rubro": "rubro_presupuesto",
                "presupuesto_mensual": "presupuesto",
                "promedio_historico": "historico",
            },
        )
        assert status == 200
        assert "Limpieza" in result
        assert "50" in result
        assert "/consorcios-case" in result

        status, _, workspace = _request(
            server, "GET", "/consorcios-case", headers={"Cookie": cookie}
        )
        assert status == 200
        assert "Las columnas de gastos y presupuesto ya fueron confirmadas para este caso." in workspace

        status, _, rerun = _post_form(
            server, "/run-consorcios-expense-variance", cookie, token, {}
        )
        assert status == 200
        assert "Limpieza" in rerun
        assert "50" in rerun

        upload_body, upload_headers = _multipart_with_context(
            "septiembre.xlsx", _xlsx_bytes(tmp_path), token, period="2026-09", cookie=cookie
        )
        status, headers, _ = _request(
            server, "POST", "/upload", body=upload_body, headers=upload_headers
        )
        assert status == 200

        status, _, workspace = _request(
            server, "GET", "/consorcios-case", headers={"Cookie": cookie}
        )
        assert status == 200
        assert "2026-09" in workspace
        assert 'name="unidad_funcional"' in workspace
        assert 'name="expense_rubro"' in workspace
        assert "ya fueron confirmadas" not in workspace

        status, _, blocked = _post_form(
            server, "/run-consorcios-collection-aging", cookie, token, {}
        )
        assert status == 400
        assert "Elegí una hoja disponible" in blocked

        status, _, blocked = _post_form(
            server, "/run-consorcios-expense-variance", cookie, token, {}
        )
        assert status == 400
        assert "Elegí una hoja disponible" in blocked

        other_body, other_headers = _multipart_with_context(
            "octubre.xlsx", _xlsx_bytes(tmp_path), token, period="2026-10", cookie=""
        )
        other_status, other_headers, _ = _request(
            server, "POST", "/upload", body=other_body, headers=other_headers
        )
        assert other_status == 200
        other_cookie = _cookie(other_headers)
        assert other_cookie != cookie
        status, _, other_workspace = _request(
            server, "GET", "/consorcios-case", headers={"Cookie": other_cookie}
        )
        assert status == 200
        assert "2026-10" in other_workspace
        assert 'name="unidad_funcional"' in other_workspace
        assert "ya fueron confirmadas" not in other_workspace
        status, _, other_blocked = _post_form(
            server, "/run-consorcios-collection-aging", other_cookie, token, {}
        )
        assert status == 400
        assert "Elegí una hoja disponible" in other_blocked
    finally:
        server.shutdown()
        server.server_close()
        thread.join(timeout=5)

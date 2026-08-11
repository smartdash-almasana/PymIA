from __future__ import annotations

from http.client import HTTPConnection
from pathlib import Path
from threading import Thread
from urllib.parse import urlencode

from openpyxl import Workbook

from pymia.smartpyme.service_1_assisted_web_v1 import create_assisted_web_server_v1


def _xlsx_bytes(tmp_path: Path) -> bytes:
    path = tmp_path / "consorcio_workspace.xlsx"
    workbook = Workbook()
    expensas = workbook.active
    expensas.title = "Expensas"
    expensas.append(["unidad", "saldo", "expensa"])
    expensas.append(["UF-12", 250, 100])
    gastos = workbook.create_sheet("Gastos")
    gastos.append(["rubro_gasto", "importe_gasto"])
    gastos.append(["Limpieza", 150])
    presupuesto = workbook.create_sheet("Presupuesto")
    presupuesto.append(["rubro_presupuesto", "presupuesto", "historico"])
    presupuesto.append(["Limpieza", 100, 100])
    workbook.save(path)
    return path.read_bytes()


def _multipart_with_context(filename: str, content: bytes) -> tuple[bytes, dict[str, str]]:
    boundary = "Service1ConsorciosWorkspaceBoundary"
    chunks: list[bytes] = []
    for name, value in (
        ("consorcio_id", "rivadavia-1200"),
        ("consorcio_name", "Rivadavia 1200"),
        ("period", "2026-08"),
    ):
        chunks.append(
            (
                f"--{boundary}\r\n"
                f'Content-Disposition: form-data; name="{name}"\r\n\r\n'
                f"{value}\r\n"
            ).encode("utf-8")
        )
    chunks.append(
        (
            f"--{boundary}\r\n"
            f'Content-Disposition: form-data; name="file"; filename="{filename}"\r\n'
            "Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n\r\n"
        ).encode("utf-8")
        + content
        + b"\r\n"
    )
    chunks.append(f"--{boundary}--\r\n".encode("utf-8"))
    body = b"".join(chunks)
    return body, {
        "Content-Type": f"multipart/form-data; boundary={boundary}",
        "Content-Length": str(len(body)),
        "Authorization": "Bearer test-token",
    }


def _request(server, method: str, path: str, *, body: bytes = b"", headers=None):
    connection = HTTPConnection("127.0.0.1", server.server_port, timeout=10)
    connection.request(method, path, body=body, headers=headers or {})
    response = connection.getresponse()
    payload = response.read().decode("utf-8")
    return response.status, response.getheaders(), payload


def _cookie(headers: list[tuple[str, str]]) -> str:
    for name, value in headers:
        if name.lower() == "set-cookie":
            return value.split(";", 1)[0]
    raise AssertionError("missing session cookie")


def _post_form(server, path: str, cookie: str, fields: dict[str, str]):
    body = urlencode(fields).encode("utf-8")
    return _request(
        server,
        "POST",
        path,
        body=body,
        headers={
            "Content-Type": "application/x-www-form-urlencoded",
            "Content-Length": str(len(body)),
            "Cookie": cookie,
            "Authorization": "Bearer test-token",
        },
    )


def test_http_case_workspace_reuses_confirmed_control_bindings_without_cross_case_inference(tmp_path: Path) -> None:
    def resolver(_handler):
        return {
            "tenant_id": "tenant-consorcios-workspace",
            "cliente_id": "cliente-consorcios-workspace",
            "owner_actor_id": "owner-consorcios-workspace",
            "owner_actor_role": "OWNER",
        }

    server = create_assisted_web_server_v1(
        host="127.0.0.1",
        port=0,
        output_dir=tmp_path / "outputs",
        tenant_identity_resolver=resolver,
    )
    thread = Thread(target=server.serve_forever, daemon=True)
    thread.start()
    try:
        upload_body, upload_headers = _multipart_with_context(
            "agosto.xlsx", _xlsx_bytes(tmp_path)
        )
        status, headers, _ = _request(
            server, "POST", "/upload", body=upload_body, headers=upload_headers
        )
        assert status == 200
        cookie = _cookie(headers)

        status, _, workspace = _request(
            server, "GET", "/consorcios-case", headers={"Cookie": cookie}
        )
        assert status == 200
        assert "Rivadavia 1200" in workspace
        assert "2026-08" in workspace
        assert "Cobranzas y deuda" in workspace
        assert "Gastos" in workspace
        assert "Banco" in workspace
        assert "RADAR" in workspace
        assert 'name="unidad_funcional"' in workspace

        status, _, result = _post_form(
            server,
            "/run-consorcios-collection-aging",
            cookie,
            {
                "sheet_name": "Expensas",
                "unidad_funcional": "unidad",
                "saldo_anterior": "saldo",
                "expensa_mes": "expensa",
            },
        )
        assert status == 200
        assert "UF-12" in result
        assert "2.5" in result
        assert "/consorcios-case" in result

        status, _, workspace = _request(
            server, "GET", "/consorcios-case", headers={"Cookie": cookie}
        )
        assert status == 200
        assert "Las columnas de este control ya fueron confirmadas para este caso." in workspace

        status, _, rerun = _post_form(
            server, "/run-consorcios-collection-aging", cookie, {}
        )
        assert status == 200
        assert "UF-12" in rerun
        assert "2.5" in rerun

        status, _, summary = _request(
            server, "GET", "/consorcios-case-summary", headers={"Cookie": cookie}
        )
        assert status == 200
        assert "Resumen del período" in summary
        assert "Rivadavia 1200" in summary
        assert "2026-08" in summary
        assert "Realizado · 1 unidad(es) revisada(s)" in summary
        assert "Gastos" in summary
        assert "Pendiente" in summary
        assert "Volver al caso" in summary
    finally:
        server.shutdown()
        server.server_close()
        thread.join(timeout=5)

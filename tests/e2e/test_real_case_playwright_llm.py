from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest
from playwright.sync_api import Page

from landing.build_service1_excel_ingestion_chat_web import (
    build_service1_excel_ingestion_chat_web,
)


def _build_file_context(xlsx_path: Path) -> str:
    """Extract structure from a real XLSX to give LLM context."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    lines: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        total = len(rows)
        headers = [str(c) for c in (rows[0] or []) if c is not None] if total else []
        samples = []
        for r in rows[1:4]:
            samples.append([str(c) if c is not None else "" for c in (r or [])])
        lines.append(
            f"Hoja '{sheet_name}': {total} filas, "
            f"encabezados={headers[:10]}, "
            f"ejemplos={samples}"
        )
    wb.close()
    return "\n".join(lines)


def _call_llm(client, system: str, prompt: str) -> str:
    """Call Anthropic (or DummyLLM) to generate an owner response."""
    response = client.messages.create(
        model="claude-sonnet-4-20250514",
        system=system,
        max_tokens=300,
        messages=[{"role": "user", "content": prompt}],
    )
    return response.content[0].text


_SYSTEM_PROMPT = (
    "Sos el dueño de una PyME argentina. "
    "Estás revisando un archivo Excel con información de tu negocio. "
    "Respondé las preguntas de PymIA de forma concisa, natural y realista, "
    "como lo haría el dueño de la PyME. "
    "No inventes datos que no existan en el contexto del archivo. "
    "Si no sabés, decí 'no sé'."
)


@pytest.mark.parametrize(
    "xlsx_name",
    [
        "cafeteria_abc.xlsx",
        "la_textil_cosida_srl_mar_abr_may_2026.xlsx",
    ],
)
def test_real_case_playwright_llm_chat(
    page: Page,
    llm_client,
    real_xlsx_dir: Path,
    xlsx_name: str,
    tmp_path: Path,
) -> None:
    html_path = build_service1_excel_ingestion_chat_web(
        tmp_path / "servicio1-excel-ingestion-chat.html"
    )
    xlsx_path = real_xlsx_dir / xlsx_name
    if not xlsx_path.exists():
        pytest.skip(f"XLSX not found: {xlsx_path}")

    page.goto(html_path.as_uri())

    page.wait_for_function("!document.getElementById('selectFileBtn').disabled")

    page.set_input_files("#fileInput", str(xlsx_path))

    page.wait_for_selector("#chatLog .msg.pymia", timeout=5000)

    file_context = _build_file_context(xlsx_path)

    max_iterations = 30
    for _ in range(max_iterations):
        if page.is_disabled("#ownerInput"):
            break

        prev_pymia_count = page.locator("#chatLog .msg.pymia").count()
        last_question = page.locator("#chatLog .msg.pymia").last.text_content() or ""

        user_prompt = (
            f"Contexto del archivo:\n{file_context}\n\n"
            f"Pregunta de PymIA:\n{last_question}\n\n"
            f"Respondé como el dueño:"
        )
        answer = _call_llm(llm_client, _SYSTEM_PROMPT, user_prompt)

        page.fill("#ownerInput", answer)
        page.click("#sendBtn")

        page.wait_for_function(
            "(prev) => document.querySelectorAll('#chatLog .msg.pymia').length > prev"
            " || document.getElementById('ownerInput').disabled",
            arg=prev_pymia_count,
            timeout=15000,
        )
    else:
        pytest.fail("Reached max iterations without completing all questions")

    assert page.is_disabled("#ownerInput"), "Chat did not reach end state"
    assert "Terminé las preguntas iniciales" in page.text_content("#chatLog")

    with page.expect_download() as download_info:
        page.click("#exportBtn")
    download = download_info.value
    out = tmp_path / download.suggested_filename
    download.save_as(out)

    content = out.read_text(encoding="utf-8")
    assert "PYMIA_SERVICE_1_EXCEL_INGESTION_CHAT_V1" in content
    assert "file_profile" in content
    assert "answers" in content
    assert '"runtime_authorized": false' in content
    assert '"production_allowed": false' in content
    assert '"final_diagnosis": false' in content
    assert '"final_accounting_result": false' in content
    assert xlsx_name in content


@pytest.mark.parametrize(
    "xlsx_name",
    ["cafeteria_abc.xlsx"],
)
def test_real_case_playwright_flow_without_llm(
    page: Page,
    real_xlsx_dir: Path,
    xlsx_name: str,
    tmp_path: Path,
) -> None:
    """Quick smoke test of the upload + chat flow using manual 'no sé' answers."""
    html_path = build_service1_excel_ingestion_chat_web(
        tmp_path / "servicio1-excel-ingestion-chat.html"
    )
    xlsx_path = real_xlsx_dir / xlsx_name
    if not xlsx_path.exists():
        pytest.skip(f"XLSX not found: {xlsx_path}")

    page.goto(html_path.as_uri())
    page.wait_for_function("!document.getElementById('selectFileBtn').disabled")
    page.set_input_files("#fileInput", str(xlsx_path))
    page.wait_for_selector("#chatLog .msg.pymia", timeout=5000)

    for _ in range(30):
        if page.is_disabled("#ownerInput"):
            break
        page.fill("#ownerInput", "No sé")
        page.click("#sendBtn")
        page.wait_for_timeout(300)
    else:
        pytest.fail("Did not finish questions with manual answers")

    assert page.is_disabled("#ownerInput")
    assert "Terminé las preguntas iniciales" in page.text_content("#chatLog")

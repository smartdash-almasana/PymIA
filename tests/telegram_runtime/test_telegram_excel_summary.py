from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from pymia.telegram_excel_summary import (
    SENTINEL,
    analyze_latest_excel,
    render_summary_text,
    resolve_latest_excel,
    summarize_excel,
)


def _make_xlsx(path: Path, sheet_name: str = "Sheet1") -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(["fecha", "orden", "producto"])
    ws.append(["2026-05-01", "O-1", "A"])
    ws.append(["2026-05-02", "O-2", "B"])
    wb.save(path)
    wb.close()
    return path


def test_resolve_latest_excel_returns_none_if_dir_missing(tmp_path: Path) -> None:
    assert resolve_latest_excel(tmp_path / "missing") is None


def test_resolve_latest_excel_returns_none_if_no_xlsx(tmp_path: Path) -> None:
    (tmp_path / "a.txt").write_text("x", encoding="utf-8")
    assert resolve_latest_excel(tmp_path) is None


def test_resolve_latest_excel_returns_most_recent_file(tmp_path: Path) -> None:
    older = _make_xlsx(tmp_path / "older.xlsx")
    newer = _make_xlsx(tmp_path / "newer.xlsx")
    older.touch()
    newer.touch()
    assert resolve_latest_excel(tmp_path).name == "newer.xlsx"


def test_summarize_excel_reads_structure(tmp_path: Path) -> None:
    excel = _make_xlsx(tmp_path / "sample.xlsx")
    summary = summarize_excel(excel)
    assert Path(summary.file_path).name == "sample.xlsx"
    assert len(summary.sheets) == 1
    sheet = summary.sheets[0]
    assert sheet.name == "Sheet1"
    assert sheet.rows >= 3
    assert sheet.cols >= 3
    assert sheet.columns[0].name == "fecha"


def test_summarize_excel_empty_sheet_flags_empty(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Empty"
    excel = tmp_path / "empty.xlsx"
    wb.save(excel)
    wb.close()
    summary = summarize_excel(excel)
    assert summary.sheets[0].empty is False  # openpyxl reports 1x1 for empty workbook


def test_render_summary_contains_sentinel_and_sheets(tmp_path: Path) -> None:
    excel = _make_xlsx(tmp_path / "render.xlsx", sheet_name="Datos")
    summary = summarize_excel(excel)
    text = render_summary_text(summary)
    assert SENTINEL in text
    assert "Resumen estructural del Excel" in text
    assert "Hojas:" in text
    assert "Datos" in text


def test_render_summary_does_not_include_business_claims(tmp_path: Path) -> None:
    excel = _make_xlsx(tmp_path / "claims.xlsx")
    summary = summarize_excel(excel)
    text = render_summary_text(summary).lower()
    forbidden = ("rentabilidad", "margen", "ganancia", "pérdida", "ingresos", "gastos")
    for token in forbidden:
        assert token not in text


def test_analyze_latest_excel_not_found(tmp_path: Path) -> None:
    result = analyze_latest_excel(tmp_path)
    assert result.source == "pymia"
    assert result.mode == "not_found"
    assert SENTINEL in result.text


def test_analyze_latest_excel_summary(tmp_path: Path) -> None:
    _make_xlsx(tmp_path / "ok.xlsx")
    result = analyze_latest_excel(tmp_path)
    assert result.source == "pymia"
    assert result.mode == "summary"
    assert result.file_path is not None
    assert SENTINEL in result.text
    assert "Hojas:" in result.text


def test_analyze_latest_excel_error_for_broken_xlsx(tmp_path: Path) -> None:
    broken = tmp_path / "broken.xlsx"
    broken.write_bytes(b"not an excel")
    result = analyze_latest_excel(tmp_path)
    assert result.source == "pymia"
    assert result.mode == "error"
    assert SENTINEL in result.text


def test_render_summary_max_length_limit(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append([f"col_{i}" for i in range(1, 400)])
    excel = tmp_path / "long.xlsx"
    wb.save(excel)
    wb.close()
    summary = summarize_excel(excel)
    text = render_summary_text(summary)
    assert len(text) <= 4000

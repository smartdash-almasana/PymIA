from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from pymia.cli.service_1_product import run_service_1_product_entrypoint_v1


ROOT = Path(__file__).resolve().parents[1]
FIXTURE = ROOT / "prueba_excels" / "fabrica_industrial_compleja.xlsx"
SHEET_NAME = "PRODUCCION"
WORK_DIR = ROOT / ".tmp" / "service_1_pilot_005_fabrica_industrial"
OUTPUT_DIR = WORK_DIR / "output"
OBSERVATION_PATH = WORK_DIR / "observed_run.json"
TOOL_REQUESTS = [
    {
        "tool_ref": "precio_margen_basico",
        "inputs": {"precio_venta": 1200, "costo_unitario": 800},
    }
]


def _headers() -> list[str]:
    workbook = load_workbook(FIXTURE, read_only=True, data_only=True)
    try:
        worksheet = workbook[SHEET_NAME]
        first_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        return [
            str(value).strip()
            for value in first_row
            if value is not None and str(value).strip()
        ]
    finally:
        workbook.close()


def _owner_column_answers(headers: list[str]) -> dict[str, str]:
    return {
        header: (
            f"La columna {header} pertenece al registro operativo de producción "
            "de la fábrica industrial."
        )
        for header in headers
    }


def _canonical_semantic_answers(
    first_result: dict[str, object],
) -> tuple[dict[str, str], dict[str, list[str]]]:
    product = first_result["product_pipeline"]
    questions = product["owner_questions"]
    selected: dict[str, str] = {}
    allowed_sources: dict[str, list[str]] = {}
    for question in questions:
        column_name = str(question["column_name"])
        allowed = [str(value) for value in question["allowed_option_ids"]]
        canonical = [
            value
            for value in allowed
            if value not in {"OTHER", "IGNORE", "IGNORED_NOT_RELEVANT"}
        ]
        if not canonical:
            raise RuntimeError(
                f"No canonical semantic option available for {column_name!r}."
            )
        selected[column_name] = canonical[0]
        allowed_sources[column_name] = allowed
    return selected, allowed_sources


def main() -> int:
    if not FIXTURE.exists():
        raise FileNotFoundError(FIXTURE)

    WORK_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    headers = _headers()
    owner_answers = _owner_column_answers(headers)

    first = run_service_1_product_entrypoint_v1(
        xlsx_path=FIXTURE,
        owner_column_answers=owner_answers,
        semantic_owner_answers=None,
        tool_requests=TOOL_REQUESTS,
        output_dir=OUTPUT_DIR,
        sheet_name=SHEET_NAME,
    )

    if first["status"] != "NEEDS_OWNER_CONFIRMATION":
        raise RuntimeError(f"Unexpected first-pass status: {first['status']!r}")
    if first["product_pipeline"]["tools_executed"] is not False:
        raise RuntimeError("Tool execution occurred before semantic owner confirmation.")

    semantic_answers, allowed_sources = _canonical_semantic_answers(first)

    final = run_service_1_product_entrypoint_v1(
        xlsx_path=FIXTURE,
        owner_column_answers=owner_answers,
        semantic_owner_answers=semantic_answers,
        tool_requests=TOOL_REQUESTS,
        output_dir=OUTPUT_DIR,
        sheet_name=SHEET_NAME,
    )

    product = final["product_pipeline"]
    observation = {
        "schema_version": "SERVICE_1_PILOT_005_FABRICA_INDUSTRIAL_OBSERVATION_V1",
        "cycle": "CYCLE_038_RUN_S1_PILOT_005_FABRICA_INDUSTRIAL",
        "status": "OBSERVED_NOT_YET_CERTIFIED",
        "fixture": "prueba_excels/fabrica_industrial_compleja.xlsx",
        "sheet_name": SHEET_NAME,
        "headers": headers,
        "first_pass": {
            "status": first["status"],
            "owner_questions_count": len(
                first["product_pipeline"]["owner_questions"]
            ),
            "tools_executed": first["product_pipeline"]["tools_executed"],
        },
        "semantic_owner_answers": {
            "source_rule": (
                "selected from product_pipeline.owner_questions[].allowed_option_ids; "
                "no free text"
            ),
            "selected": semantic_answers,
            "allowed_option_sources": allowed_sources,
        },
        "final_pass": {
            "status": final["status"],
            "blocked_reason": product.get("blocked_reason"),
            "semantic_bindings_confirmed": product.get(
                "semantic_bindings_confirmed"
            ),
            "tools_executed": product.get("tools_executed"),
            "executed_tool_refs": product.get("physical_run", {}).get(
                "executed_tool_refs", []
            ),
            "xlsx_outputs": sorted(path.name for path in OUTPUT_DIR.glob("*.xlsx")),
        },
        "limits": [
            "Run uses an explicit tool request; it does not claim automatic tool selection.",
            "Run validates the canonical product path on the PRODUCCION sheet of an industrial workbook, not a full industrial diagnosis.",
            "Scrap, OEE, machine efficiency, downtime and production-loss analysis remain unsupported unless added by a later governed cycle.",
            "No new formula, pathology, semantic capability, runtime authority or delivery authority is added by this observation.",
        ],
    }

    OBSERVATION_PATH.write_text(
        json.dumps(observation, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(json.dumps(observation, ensure_ascii=False, indent=2))
    print(f"\nObservation written to: {OBSERVATION_PATH}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

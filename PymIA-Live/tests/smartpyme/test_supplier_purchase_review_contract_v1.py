from __future__ import annotations

import inspect
from pathlib import Path

from openpyxl import load_workbook

from pymia.smartpyme.service_1_xlsx_delivery_v1 import build_service_1_xlsx_delivery_v1
from pymia.smartpyme.supplier_purchase_review_contract_v1 import (
    SUPPLIER_PURCHASE_REVIEW_CAPABILITY_REF,
    SUPPLIER_PURCHASE_REVIEW_CONTRACT_REF,
    build_supplier_purchase_review_contract_v1,
)


def _contract_input() -> dict[str, object]:
    return {
        "owner_requested_output": "supplier_purchase_review_scope_report",
        "source_files_received": ["registro_proveedores", "registro_compras"],
        "received_fields": ["fecha", "proveedor", "numero_comprobante", "importe"],
    }


def test_ready_for_review_with_supplier_and_purchase_registers_and_required_fields() -> None:
    result = build_supplier_purchase_review_contract_v1(contract_input=_contract_input())

    assert result["status"] == "READY_FOR_REVIEW"
    assert result["contract_ref"] == SUPPLIER_PURCHASE_REVIEW_CONTRACT_REF
    assert result["runtime_authorized"] is False
    assert result["missing_sources"] == []
    assert result["missing_fields"] == []
    assert result["next_allowed_action"] == "manual_supplier_purchase_review"


def test_missing_supplier_register_when_registro_proveedores_is_absent() -> None:
    contract_input = _contract_input()
    contract_input["source_files_received"] = ["registro_compras"]

    result = build_supplier_purchase_review_contract_v1(contract_input=contract_input)

    assert result["status"] == "MISSING_SUPPLIER_REGISTER"
    assert result["missing_sources"] == ["registro_proveedores"]
    assert result["next_allowed_action"] == "request_supplier_register"


def test_missing_purchase_register_when_registro_compras_is_absent() -> None:
    contract_input = _contract_input()
    contract_input["source_files_received"] = ["registro_proveedores"]

    result = build_supplier_purchase_review_contract_v1(contract_input=contract_input)

    assert result["status"] == "MISSING_PURCHASE_REGISTER"
    assert result["missing_sources"] == ["registro_compras"]
    assert result["next_allowed_action"] == "request_purchase_register"


def test_missing_fields_when_required_field_is_absent() -> None:
    contract_input = _contract_input()
    contract_input["received_fields"] = ["fecha", "proveedor", "importe"]

    result = build_supplier_purchase_review_contract_v1(contract_input=contract_input)

    assert result["status"] == "MISSING_FIELDS"
    assert result["missing_fields"] == ["numero_comprobante"]
    assert result["next_allowed_action"] == "request_missing_supplier_purchase_review_fields"


def test_invalid_input_when_sources_are_not_list_of_strings() -> None:
    contract_input = _contract_input()
    contract_input["source_files_received"] = {"registro_proveedores": True}

    result = build_supplier_purchase_review_contract_v1(contract_input=contract_input)  # type: ignore[arg-type]

    assert result["status"] == "INVALID_INPUT"
    assert result["next_allowed_action"] == "fix_invalid_supplier_purchase_review_contract_input"


def test_invalid_input_when_fields_are_not_list_of_strings() -> None:
    contract_input = _contract_input()
    contract_input["received_fields"] = {"fecha": True}

    result = build_supplier_purchase_review_contract_v1(contract_input=contract_input)  # type: ignore[arg-type]

    assert result["status"] == "INVALID_INPUT"
    assert result["delivery_input"]["status"] == "INVALID_INPUT"


def test_wraps_base_accounting_contract_without_runtime_review() -> None:
    result = build_supplier_purchase_review_contract_v1(contract_input=_contract_input())

    accounting_contract = result["accounting_contract"]
    assert accounting_contract["contract_ref"] == "supplier_purchase_review_basic"
    assert accounting_contract["family"] == "supplier_purchase_review"
    assert accounting_contract["runtime_authorized"] is False
    assert accounting_contract["next_allowed_action"] == "manual_accounting_review"


def test_delivery_input_is_compatible_with_generic_xlsx_delivery(tmp_path: Path) -> None:
    result = build_supplier_purchase_review_contract_v1(contract_input=_contract_input())
    output_path = tmp_path / "supplier_purchase_review_contract.xlsx"

    delivery = build_service_1_xlsx_delivery_v1(
        delivery_input=result["delivery_input"],
        output_path=output_path,
    )

    workbook = load_workbook(output_path)

    assert output_path.exists()
    assert delivery["capability_ref"] == SUPPLIER_PURCHASE_REVIEW_CAPABILITY_REF
    assert workbook["Resumen"]["B3"].value == SUPPLIER_PURCHASE_REVIEW_CAPABILITY_REF
    assert ("supplier_purchase_review_status", "READY_FOR_REVIEW") in list(
        workbook["Resultados"].iter_rows(values_only=True)
    )


def test_forbidden_claims_prevent_certified_review_or_accounting_runtime() -> None:
    result = build_supplier_purchase_review_contract_v1(contract_input=_contract_input())
    forbidden_claims = result["delivery_input"]["forbidden_claims"]

    assert "No confirma auditoría contable certificada." in forbidden_claims
    assert "No confirma exactitud fiscal." in forbidden_claims
    assert "No genera asientos contables automáticos." in forbidden_claims


def test_module_has_no_io_openpyxl_or_forbidden_runtime_dependencies() -> None:
    import pymia.smartpyme.supplier_purchase_review_contract_v1 as module

    source = inspect.getsource(module)

    assert "openpyxl" not in source
    assert "first_aid" not in source
    assert "exceland" not in source.lower()
    assert "vertical_pipeline" not in source
    assert "service_1_pipeline" not in source
    assert "fsm" not in source.lower()
    assert "llm" not in source.lower()
    assert "chatbot" not in source.lower()
    assert "mercado_pago" not in source.lower()
    assert "open(" not in source
    assert ".save(" not in source
    assert "read_text(" not in source
    assert "read_bytes(" not in source

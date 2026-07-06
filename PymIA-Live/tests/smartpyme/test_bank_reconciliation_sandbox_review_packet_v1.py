from __future__ import annotations

import inspect
from pathlib import Path

from openpyxl import load_workbook

from pymia.smartpyme.accounting_sandbox_release_gate_v1 import evaluate_accounting_sandbox_release_gate_v1
from pymia.smartpyme.bank_reconciliation_contract_v1 import build_bank_reconciliation_contract_v1
from pymia.smartpyme.bank_reconciliation_sandbox_contract_v1 import build_bank_reconciliation_sandbox_contract_v1
from pymia.smartpyme.bank_reconciliation_sandbox_fixture_handoff_v1 import build_bank_reconciliation_sandbox_fixture_handoff_v1
from pymia.smartpyme.bank_reconciliation_sandbox_fixture_model_v1 import build_bank_reconciliation_sandbox_fixture_model_v1
from pymia.smartpyme.bank_reconciliation_sandbox_review_packet_v1 import (
    CAPABILITY_REF,
    build_bank_reconciliation_sandbox_review_packet_v1,
)
from pymia.smartpyme.service_1_xlsx_delivery_v1 import build_service_1_xlsx_delivery_v1


def _movement(ref: str) -> dict[str, str]:
    return {"movement_ref": ref, "date": "2026-06-01", "amount": "100.00", "description": "sample"}


def _fixture_model() -> dict[str, object]:
    return build_bank_reconciliation_sandbox_fixture_model_v1(
        bundle_input={
            "bank_statement_fixture": {
                "fixture_id": "bank-fixture-001",
                "source_ref": "sample-bank",
                "period_ref": "2026-06",
                "currency": "ARS",
                "movements": [_movement("bank-001")],
                "live_source": False,
            },
            "internal_ledger_fixture": {
                "fixture_id": "ledger-fixture-001",
                "source_ref": "sample-ledger",
                "period_ref": "2026-06",
                "currency": "ARS",
                "movements": [_movement("ledger-001")],
                "live_source": False,
            },
        }
    )


def _base_contract() -> dict[str, object]:
    return build_bank_reconciliation_contract_v1(
        contract_input={
            "owner_requested_output": "bank_reconciliation_scope_report",
            "source_files_received": ["extracto_banco", "archivo_contable"],
            "received_fields": ["fecha", "importe", "referencia"],
        }
    )


def _human_gate() -> dict[str, object]:
    return evaluate_accounting_sandbox_release_gate_v1(
        gate_input={
            "capability_ref": "bank_reconciliation_basic",
            "reviewer_role": "operator",
            "decision": "APPROVED",
            "scope_ok": True,
            "evidence_ok": True,
            "forbidden_claims": [],
            "live_use": False,
        }
    )


def _packet_input() -> dict[str, object]:
    fixture_model = _fixture_model()
    handoff = build_bank_reconciliation_sandbox_fixture_handoff_v1(
        handoff_input={
            "fixture_model_result": fixture_model,
            "base_contract": _base_contract(),
            "sandbox_release_gate": _human_gate(),
            "live_use_requested": False,
        }
    )
    sandbox = build_bank_reconciliation_sandbox_contract_v1(sandbox_input=handoff["sandbox_input"])  # type: ignore[arg-type]
    return {
        "fixture_model_result": fixture_model,
        "fixture_handoff_result": handoff,
        "sandbox_contract_result": sandbox,
    }


def test_review_packet_ready_when_chain_is_ready() -> None:
    result = build_bank_reconciliation_sandbox_review_packet_v1(packet_input=_packet_input())

    assert result["status"] == "READY"
    assert result["capability_ref"] == CAPABILITY_REF
    assert result["runtime_authorized"] is False
    assert result["production_allowed"] is False
    assert result["blocked_reasons"] == []
    assert result["readiness_flags"]["fixture_model_valid"] is True
    assert result["readiness_flags"]["handoff_ready"] is True
    assert result["readiness_flags"]["sandbox_contract_ready"] is True
    assert result["next_allowed_action"] == "prepare_owner_operator_review"


def test_review_packet_blocks_when_fixture_model_is_not_valid() -> None:
    packet_input = _packet_input()
    fixture_model = dict(packet_input["fixture_model_result"])
    fixture_model["status"] = "INVALID_MOVEMENT"
    packet_input["fixture_model_result"] = fixture_model

    result = build_bank_reconciliation_sandbox_review_packet_v1(packet_input=packet_input)  # type: ignore[arg-type]

    assert result["status"] == "BLOCKED"
    assert "fixture_model_not_valid" in result["blocked_reasons"]
    assert result["runtime_authorized"] is False


def test_review_packet_blocks_when_handoff_is_not_ready() -> None:
    packet_input = _packet_input()
    handoff = dict(packet_input["fixture_handoff_result"])
    handoff["status"] = "BLOCKED_HUMAN_GATE"
    packet_input["fixture_handoff_result"] = handoff

    result = build_bank_reconciliation_sandbox_review_packet_v1(packet_input=packet_input)  # type: ignore[arg-type]

    assert result["status"] == "BLOCKED"
    assert "fixture_handoff_not_ready" in result["blocked_reasons"]


def test_review_packet_blocks_when_sandbox_contract_is_not_ready() -> None:
    packet_input = _packet_input()
    sandbox = dict(packet_input["sandbox_contract_result"])
    sandbox["status"] = "MISSING_FIXTURES"
    packet_input["sandbox_contract_result"] = sandbox

    result = build_bank_reconciliation_sandbox_review_packet_v1(packet_input=packet_input)  # type: ignore[arg-type]

    assert result["status"] == "BLOCKED"
    assert "sandbox_contract_not_ready" in result["blocked_reasons"]


def test_invalid_packet_components_return_invalid_input() -> None:
    result = build_bank_reconciliation_sandbox_review_packet_v1(
        packet_input={
            "fixture_model_result": [],
            "fixture_handoff_result": {},
            "sandbox_contract_result": {},
        }
    )

    assert result["status"] == "INVALID_INPUT"
    assert result["production_allowed"] is False


def test_delivery_input_is_compatible_with_generic_xlsx_delivery(tmp_path: Path) -> None:
    result = build_bank_reconciliation_sandbox_review_packet_v1(packet_input=_packet_input())
    output_path = tmp_path / "review_packet.xlsx"

    delivery = build_service_1_xlsx_delivery_v1(
        delivery_input=result["delivery_input"],
        output_path=output_path,
    )
    workbook = load_workbook(output_path)

    assert output_path.exists()
    assert delivery["capability_ref"] == CAPABILITY_REF
    assert workbook["Resumen"]["B3"].value == CAPABILITY_REF
    assert ("packet_sections", '["owner_summary", "operator_summary", "readiness_flags", "fixture_refs", "limitations", "forbidden_claims", "next_allowed_action"]') in list(workbook["Resultados"].iter_rows(values_only=True))


def test_forbidden_claims_prevent_productive_claims() -> None:
    result = build_bank_reconciliation_sandbox_review_packet_v1(packet_input=_packet_input())

    assert "No confirma saldo conciliado." in result["forbidden_claims"]
    assert "No genera asientos contables." in result["forbidden_claims"]
    assert "No certifica exactitud contable o fiscal." in result["forbidden_claims"]


def test_module_has_no_io_or_forbidden_dependencies() -> None:
    import pymia.smartpyme.bank_reconciliation_sandbox_review_packet_v1 as module

    source = inspect.getsource(module)

    assert "openpyxl" not in source
    assert "requests" not in source
    assert "httpx" not in source
    assert "api" not in source.lower()
    assert "mercado_pago" not in source.lower()
    assert "vertical_pipeline" not in source
    assert "service_1_pipeline" not in source
    assert "fsm" not in source.lower()
    assert "llm" not in source.lower()
    assert "chatbot" not in source.lower()
    assert "open(" not in source
    assert ".save(" not in source
    assert "read_text(" not in source
    assert "read_bytes(" not in source
    assert "score" not in source.lower()
    assert "match" not in source.lower()

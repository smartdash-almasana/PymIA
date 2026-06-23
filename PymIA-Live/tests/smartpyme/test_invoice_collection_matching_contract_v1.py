from __future__ import annotations

import inspect
from pathlib import Path

from openpyxl import load_workbook

from pymia.smartpyme.invoice_collection_matching_contract_v1 import (
    INVOICE_COLLECTION_MATCHING_CAPABILITY_REF,
    INVOICE_COLLECTION_MATCHING_CONTRACT_REF,
    build_invoice_collection_matching_contract_v1,
)
from pymia.smartpyme.service_1_xlsx_delivery_v1 import build_service_1_xlsx_delivery_v1


def _contract_input() -> dict[str, object]:
    return {
        "owner_requested_output": "invoice_collection_matching_scope_report",
        "source_files_received": ["registro_facturas", "registro_cobros"],
        "received_fields": ["fecha", "cliente", "numero_factura", "importe"],
    }


def test_ready_for_review_with_invoice_and_collection_registers_and_required_fields() -> None:
    result = build_invoice_collection_matching_contract_v1(contract_input=_contract_input())

    assert result["status"] == "READY_FOR_REVIEW"
    assert result["contract_ref"] == INVOICE_COLLECTION_MATCHING_CONTRACT_REF
    assert result["runtime_authorized"] is False
    assert result["missing_sources"] == []
    assert result["missing_fields"] == []
    assert result["next_allowed_action"] == "manual_invoice_collection_matching_review"


def test_missing_invoice_register_when_registro_facturas_is_absent() -> None:
    contract_input = _contract_input()
    contract_input["source_files_received"] = ["registro_cobros"]

    result = build_invoice_collection_matching_contract_v1(contract_input=contract_input)

    assert result["status"] == "MISSING_INVOICE_REGISTER"
    assert result["missing_sources"] == ["registro_facturas"]
    assert result["next_allowed_action"] == "request_invoice_register"


def test_missing_collection_register_when_registro_cobros_is_absent() -> None:
    contract_input = _contract_input()
    contract_input["source_files_received"] = ["registro_facturas"]

    result = build_invoice_collection_matching_contract_v1(contract_input=contract_input)

    assert result["status"] == "MISSING_COLLECTION_REGISTER"
    assert result["missing_sources"] == ["registro_cobros"]
    assert result["next_allowed_action"] == "request_collection_register"


def test_missing_fields_when_required_field_is_absent() -> None:
    contract_input = _contract_input()
    contract_input["received_fields"] = ["fecha", "cliente", "importe"]

    result = build_invoice_collection_matching_contract_v1(contract_input=contract_input)

    assert result["status"] == "MISSING_FIELDS"
    assert result["missing_fields"] == ["numero_factura"]
    assert result["next_allowed_action"] == "request_missing_invoice_collection_matching_fields"


def test_invalid_input_when_sources_are_not_list_of_strings() -> None:
    contract_input = _contract_input()
    contract_input["source_files_received"] = {"registro_facturas": True}

    result = build_invoice_collection_matching_contract_v1(contract_input=contract_input)  # type: ignore[arg-type]

    assert result["status"] == "INVALID_INPUT"
    assert result["next_allowed_action"] == "fix_invalid_invoice_collection_matching_contract_input"


def test_invalid_input_when_fields_are_not_list_of_strings() -> None:
    contract_input = _contract_input()
    contract_input["received_fields"] = {"fecha": True}

    result = build_invoice_collection_matching_contract_v1(contract_input=contract_input)  # type: ignore[arg-type]

    assert result["status"] == "INVALID_INPUT"
    assert result["delivery_input"]["status"] == "INVALID_INPUT"


def test_wraps_base_accounting_contract_without_runtime_matching() -> None:
    result = build_invoice_collection_matching_contract_v1(contract_input=_contract_input())

    accounting_contract = result["accounting_contract"]
    assert accounting_contract["contract_ref"] == "invoice_collection_matching_basic"
    assert accounting_contract["family"] == "invoice_collection_matching"
    assert accounting_contract["runtime_authorized"] is False
    assert accounting_contract["next_allowed_action"] == "manual_accounting_review"


def test_delivery_input_is_compatible_with_generic_xlsx_delivery(tmp_path: Path) -> None:
    result = build_invoice_collection_matching_contract_v1(contract_input=_contract_input())
    output_path = tmp_path / "invoice_collection_matching_contract.xlsx"

    delivery = build_service_1_xlsx_delivery_v1(
        delivery_input=result["delivery_input"],
        output_path=output_path,
    )

    workbook = load_workbook(output_path)

    assert output_path.exists()
    assert delivery["capability_ref"] == INVOICE_COLLECTION_MATCHING_CAPABILITY_REF
    assert workbook["Resumen"]["B3"].value == INVOICE_COLLECTION_MATCHING_CAPABILITY_REF
    assert ("invoice_collection_matching_status", "READY_FOR_REVIEW") in list(
        workbook["Resultados"].iter_rows(values_only=True)
    )


def test_forbidden_claims_prevent_closed_matching_or_accounting_runtime() -> None:
    result = build_invoice_collection_matching_contract_v1(contract_input=_contract_input())
    forbidden_claims = result["delivery_input"]["forbidden_claims"]

    assert "No confirma conciliación bancaria cerrada." in forbidden_claims
    assert "No confirma auditoría contable certificada." in forbidden_claims
    assert "No genera asientos contables automáticos." in forbidden_claims


def test_module_has_no_io_openpyxl_or_forbidden_runtime_dependencies() -> None:
    import pymia.smartpyme.invoice_collection_matching_contract_v1 as module

    source = inspect.getsource(module)

    assert "openpyxl" not in source
    assert "first_aid" not in source
    assert "exceland" not in source.lower()
    assert "vertical_pipeline" not in source
    assert "service_1_pipeline" not in source
    assert "fsm" not in source.lower()
    assert "llm" not in source.lower()
    assert "chatbot" not in source.lower()
    assert "mercado_pago" not in source.lower()
    assert "open(" not in source
    assert ".save(" not in source
    assert "read_text(" not in source
    assert "read_bytes(" not in source

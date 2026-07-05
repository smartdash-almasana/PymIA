from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from pymia.smartpyme.service_1_operator_delivery_package_v1 import (
    build_service_1_operator_delivery_package_v1,
)
from pymia.smartpyme.service_1_controlled_delivery_demo_harness_v1 import (
    build_service_1_controlled_delivery_demo_sample_case_v1,
    run_service_1_controlled_delivery_demo_harness_v1,
)


def _build_delivery_package(tmp_path: Path):
    harness_root = tmp_path / "harness"
    package_root = tmp_path / "packages"
    harness_root.mkdir()
    package_root.mkdir()

    harness_run = run_service_1_controlled_delivery_demo_harness_v1(
        case=build_service_1_controlled_delivery_demo_sample_case_v1(),
        output_root=harness_root,
    )
    package = build_service_1_operator_delivery_package_v1(
        harness_run=harness_run,
        package_root=package_root,
    )
    return harness_run, package


def test_delivery_folder_smoke_generates_complete_delivery_package(tmp_path: Path) -> None:
    harness_run, package = _build_delivery_package(tmp_path)

    package_dir = Path(package["package_dir"])
    assert package_dir.exists()
    assert package_dir.is_dir()
    assert package["runtime_authorized"] is False
    assert harness_run["runtime_authorized"] is False

    inventory = sorted(path.name for path in package_dir.iterdir() if path.is_file())
    assert inventory == [
        "README_ENTREGA.md",
        "delivery_report.txt",
        "first_aid_001_precio_margen_basico.xlsx",
        "first_aid_002_caja_diaria_triage.xlsx",
        "first_aid_003_stock_alertas_basicas.xlsx",
        "first_aid_004_gastos_triage.xlsx",
        "first_aid_005_proveedores_precio_variacion_triage.xlsx",
        "manifest.json",
        "summary.txt",
    ]


def test_delivery_folder_smoke_validates_xlsx_and_text_artifacts(tmp_path: Path) -> None:
    _, package = _build_delivery_package(tmp_path)

    package_dir = Path(package["package_dir"])
    readme_text = (package_dir / "README_ENTREGA.md").read_text(encoding="utf-8")
    manifest = json.loads((package_dir / "manifest.json").read_text(encoding="utf-8"))
    summary_text = (package_dir / "summary.txt").read_text(encoding="utf-8")
    delivery_report = (package_dir / "delivery_report.txt").read_text(encoding="utf-8")

    assert "Entrega preliminar basada en datos declarados." in readme_text
    assert manifest["runtime_authorized"] is False
    assert "Resultados procesados: 5" in summary_text
    assert "Tools ejecutadas: 5" in delivery_report

    xlsx_files = sorted(package_dir.glob("*.xlsx"))
    assert len(xlsx_files) == 5
    for xlsx_file in xlsx_files:
        workbook = load_workbook(xlsx_file)
        assert workbook["Resumen"]["B2"].value == "SERVICE_1"
        assert "Claims prohibidos" in workbook.sheetnames

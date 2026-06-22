from __future__ import annotations

import inspect
from pathlib import Path

import pytest
from openpyxl import load_workbook

from pymia.smartpyme.service_1_pipeline_v1 import (
    SERVICE_NAME,
    run_service_1_pipeline_v1,
)


def _three_tool_requests() -> list[dict[str, object]]:
    return [
        {
            "tool_ref": "precio_margen_basico",
            "inputs": {"precio_venta": 2500, "costo_unitario": 1625},
        },
        {
            "tool_ref": "caja_diaria_triage",
            "inputs": {"saldo_inicial": 180000, "ingresos": 324500, "egresos": 286750},
        },
        {
            "tool_ref": "stock_alertas_basicas",
            "inputs": {
                "producto": "Pack yerba 1kg",
                "stock_actual": 8,
                "stock_minimo": 15,
                "ventas_diarias_promedio": 3,
            },
        },
    ]


def test_pipeline_executes_explicit_three_tool_request_and_generates_deliveries(tmp_path: Path) -> None:
    result = run_service_1_pipeline_v1(
        tool_requests=_three_tool_requests(),
        output_dir=tmp_path,
    )

    assert result["schema_version"] == "1.0"
    assert result["service_name"] == SERVICE_NAME
    assert result["requested_tool_count"] == 3
    assert result["runtime_authorized"] is False
    assert result["executed_tool_refs"] == [
        "precio_margen_basico",
        "caja_diaria_triage",
        "stock_alertas_basicas",
    ]
    assert result["delivery_flow"]["delivery_count"] == 3
    assert len(result["delivery_flow"]["deliveries"]) == 3


def _five_tool_requests() -> list[dict[str, object]]:
    return [
        *_three_tool_requests(),
        {
            "tool_ref": "gastos_triage",
            "inputs": {
                "concepto": ["alquiler", "luz", "insumos"],
                "importe": [1000, 200, 300],
                "categoria": ["fijo", "fijo", "variable"],
            },
        },
        {
            "tool_ref": "proveedores_precio_variacion_triage",
            "inputs": {
                "proveedor": ["Proveedor A", "Proveedor B"],
                "producto_o_insumo": ["Harina", "Harina"],
                "precio_o_costo": [1000, 1250],
            },
        },
    ]


def test_pipeline_executes_full_five_tool_first_aid_family_and_generates_deliveries(tmp_path: Path) -> None:
    result = run_service_1_pipeline_v1(
        tool_requests=_five_tool_requests(),
        output_dir=tmp_path,
    )

    assert result["requested_tool_count"] == 5
    assert result["executed_tool_refs"] == [
        "precio_margen_basico",
        "caja_diaria_triage",
        "stock_alertas_basicas",
        "gastos_triage",
        "proveedores_precio_variacion_triage",
    ]
    assert result["delivery_flow"]["delivery_count"] == 5
    assert len(result["delivery_flow"]["deliveries"]) == 5


def test_pipeline_generates_readable_xlsx_outputs(tmp_path: Path) -> None:
    result = run_service_1_pipeline_v1(
        tool_requests=_five_tool_requests(),
        output_dir=tmp_path,
    )

    for delivery in result["delivery_flow"]["deliveries"]:
        output_path = Path(delivery["output_path"])
        assert output_path.exists()
        workbook = load_workbook(output_path)
        assert workbook["Resumen"]["B2"].value == "SERVICE_1"
        assert "Resultados" in workbook.sheetnames
        assert "Claims prohibidos" in workbook.sheetnames


def test_pipeline_preserves_request_order(tmp_path: Path) -> None:
    requests = [
        {
            "tool_ref": "stock_alertas_basicas",
            "inputs": {"producto": "SKU", "stock_actual": 2, "stock_minimo": 5},
        },
        {
            "tool_ref": "precio_margen_basico",
            "inputs": {"precio_venta": 100, "costo_unitario": 60},
        },
    ]

    result = run_service_1_pipeline_v1(tool_requests=requests, output_dir=tmp_path)

    assert result["executed_tool_refs"] == ["stock_alertas_basicas", "precio_margen_basico"]
    assert result["delivery_flow"]["tool_refs"] == ["stock_alertas_basicas", "precio_margen_basico"]


def test_pipeline_returns_missing_inputs_without_guessing(tmp_path: Path) -> None:
    result = run_service_1_pipeline_v1(
        tool_requests=[{"tool_ref": "precio_margen_basico", "inputs": {"precio_venta": 100}}],
        output_dir=tmp_path,
    )

    assert result["tool_results"][0]["status"] == "MISSING_INPUTS"
    assert result["tool_results"][0]["missing_inputs"] == ["costo_unitario"]
    assert "Faltantes detectados" in result["delivery_flow"]["summary_text"]


def test_pipeline_returns_invalid_input_without_correction(tmp_path: Path) -> None:
    result = run_service_1_pipeline_v1(
        tool_requests=[
            {
                "tool_ref": "caja_diaria_triage",
                "inputs": {"saldo_inicial": 10, "ingresos": -1, "egresos": 2},
            }
        ],
        output_dir=tmp_path,
    )

    assert result["tool_results"][0]["status"] == "INVALID_INPUT"
    assert result["delivery_flow"]["statuses"] == ["INVALID_INPUT"]


def test_pipeline_rejects_empty_requests(tmp_path: Path) -> None:
    with pytest.raises(ValueError, match="requires at least one tool request"):
        run_service_1_pipeline_v1(tool_requests=[], output_dir=tmp_path)


def test_pipeline_rejects_unsupported_tool_ref(tmp_path: Path) -> None:
    with pytest.raises(ValueError, match="Unsupported SERVICE_1_PIPELINE_V1 tool_ref"):
        run_service_1_pipeline_v1(
            tool_requests=[{"tool_ref": "unsupported_tool", "inputs": {}}],  # type: ignore[list-item]
            output_dir=tmp_path,
        )


def test_pipeline_rejects_missing_output_dir() -> None:
    with pytest.raises(FileNotFoundError, match="Output directory does not exist"):
        run_service_1_pipeline_v1(
            tool_requests=_three_tool_requests(),
            output_dir="/nonexistent/service1/pipeline/path",
        )


def test_pipeline_does_not_depend_on_forbidden_product_layers() -> None:
    import pymia.smartpyme.service_1_pipeline_v1 as module

    source = inspect.getsource(module)

    assert "vertical_pipeline" not in source
    assert "service_1_fsm_decision_patch_v1" not in source
    assert "service_1_boundary_chain_v1" not in source
    assert "document_ingestion" not in source
    assert "exceland" not in source.lower()


def test_pipeline_has_no_llm_or_chatbot_dependency() -> None:
    import pymia.smartpyme.service_1_pipeline_v1 as module

    source = inspect.getsource(module)

    assert "openai" not in source.lower()
    assert "chatbot" not in source.lower()


def test_pipeline_is_allowlist_based_not_inference_based() -> None:
    import pymia.smartpyme.service_1_pipeline_v1 as module

    source = inspect.getsource(module)

    assert "_ALLOWED_TOOL_REFS" in source
    assert "candidate_tool" not in source
    assert "select" not in source.lower()
    assert "infer" not in source.lower()

from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from pymia.smartpyme.service_1_operator_delivery_package_v1 import (
    SERVICE_NAME,
    build_service_1_operator_delivery_package_v1,
)
from pymia.smartpyme.service_1_operator_harness_v1 import (
    build_service_1_operator_harness_sample_case_v1,
    run_service_1_operator_harness_v1,
)


def _build_harness_run(tmp_path: Path):
    source_root = tmp_path / "source"
    source_root.mkdir()
    return run_service_1_operator_harness_v1(
        case=build_service_1_operator_harness_sample_case_v1(),
        output_root=source_root,
    )


def test_delivery_package_creates_presentable_folder(tmp_path: Path) -> None:
    harness_run = _build_harness_run(tmp_path)
    package_root = tmp_path / "packages"
    package_root.mkdir()

    package = build_service_1_operator_delivery_package_v1(
        harness_run=harness_run,
        package_root=package_root,
    )

    package_dir = Path(package["package_dir"])
    assert package["schema_version"] == "1.0"
    assert package["service_name"] == SERVICE_NAME
    assert package["case_id"] == harness_run["case_id"]
    assert package["runtime_authorized"] is False
    assert package_dir.exists()
    assert package_dir.is_dir()


def test_delivery_package_inventory(tmp_path: Path) -> None:
    harness_run = _build_harness_run(tmp_path)
    package_root = tmp_path / "packages"
    package_root.mkdir()

    package = build_service_1_operator_delivery_package_v1(
        harness_run=harness_run,
        package_root=package_root,
    )

    files = sorted(path.name for path in Path(package["package_dir"]).iterdir() if path.is_file())
    assert files == [
        "README_ENTREGA.md",
        "first_aid_001_precio_margen_basico.xlsx",
        "first_aid_002_caja_diaria_triage.xlsx",
        "first_aid_003_stock_alertas_basicas.xlsx",
        "manifest.json",
        "operator_report.txt",
        "summary.txt",
    ]
    assert package["file_count"] == 7


def test_delivery_package_manifest_matches_artifacts(tmp_path: Path) -> None:
    harness_run = _build_harness_run(tmp_path)
    package_root = tmp_path / "packages"
    package_root.mkdir()

    package = build_service_1_operator_delivery_package_v1(
        harness_run=harness_run,
        package_root=package_root,
    )

    manifest = json.loads(Path(package["manifest_path"]).read_text(encoding="utf-8"))
    assert manifest["schema_version"] == "1.0"
    assert manifest["service_name"] == "SERVICE_1"
    assert manifest["case_id"] == harness_run["case_id"]
    assert manifest["runtime_authorized"] is False
    assert manifest["artifact_count"] == 6
    assert len(manifest["artifacts"]) == 6
    assert {artifact["filename"] for artifact in manifest["artifacts"]} == {
        "first_aid_001_precio_margen_basico.xlsx",
        "first_aid_002_caja_diaria_triage.xlsx",
        "first_aid_003_stock_alertas_basicas.xlsx",
        "README_ENTREGA.md",
        "operator_report.txt",
        "summary.txt",
    }


def test_delivery_package_files_include_hashes_and_sizes(tmp_path: Path) -> None:
    harness_run = _build_harness_run(tmp_path)
    package_root = tmp_path / "packages"
    package_root.mkdir()

    package = build_service_1_operator_delivery_package_v1(
        harness_run=harness_run,
        package_root=package_root,
    )

    for file_record in package["files"]:
        package_path = Path(file_record["package_path"])
        assert package_path.exists()
        assert len(file_record["sha256"]) == 64
        assert file_record["bytes"] == package_path.stat().st_size
        assert file_record["bytes"] > 0


def test_delivery_package_readme_is_client_and_operator_readable(tmp_path: Path) -> None:
    harness_run = _build_harness_run(tmp_path)
    package_root = tmp_path / "packages"
    package_root.mkdir()

    package = build_service_1_operator_delivery_package_v1(
        harness_run=harness_run,
        package_root=package_root,
    )

    readme = Path(package["readme_path"]).read_text(encoding="utf-8")
    assert "# Entrega preliminar — Servicio 1 First Aid" in readme
    assert "## Archivos incluidos" in readme
    assert "summary.txt" in readme
    assert "operator_report.txt" in readme
    assert "manifest.json" in readme
    assert "Entrega preliminar basada en datos declarados." in readme
    assert "No es un diagnostico integral" in readme
    assert "No confirma saldo bancario real" in readme
    assert "No confirma stock fisico real" in readme


def test_delivery_package_copied_xlsx_files_are_readable(tmp_path: Path) -> None:
    harness_run = _build_harness_run(tmp_path)
    package_root = tmp_path / "packages"
    package_root.mkdir()

    package = build_service_1_operator_delivery_package_v1(
        harness_run=harness_run,
        package_root=package_root,
    )

    xlsx_paths = sorted(Path(package["package_dir"]).glob("*.xlsx"))
    assert len(xlsx_paths) == 3
    for xlsx_path in xlsx_paths:
        workbook = load_workbook(xlsx_path)
        assert workbook["Resumen"]["B2"].value == "SERVICE_1"
        assert "Claims prohibidos" in workbook.sheetnames


def test_delivery_package_rejects_missing_package_root(tmp_path: Path) -> None:
    harness_run = _build_harness_run(tmp_path)

    with pytest.raises(FileNotFoundError, match="Package root does not exist"):
        build_service_1_operator_delivery_package_v1(
            harness_run=harness_run,
            package_root=tmp_path / "missing",
        )


def test_delivery_package_rejects_runtime_authorized_true(tmp_path: Path) -> None:
    harness_run = _build_harness_run(tmp_path)
    harness_run["runtime_authorized"] = True
    package_root = tmp_path / "packages"
    package_root.mkdir()

    with pytest.raises(ValueError, match="rejects runtime_authorized=True"):
        build_service_1_operator_delivery_package_v1(
            harness_run=harness_run,
            package_root=package_root,
        )


def test_delivery_package_rejects_missing_artifact(tmp_path: Path) -> None:
    harness_run = _build_harness_run(tmp_path)
    Path(harness_run["summary_path"]).unlink()
    package_root = tmp_path / "packages"
    package_root.mkdir()

    with pytest.raises(FileNotFoundError, match="Delivery artifact does not exist"):
        build_service_1_operator_delivery_package_v1(
            harness_run=harness_run,
            package_root=package_root,
        )

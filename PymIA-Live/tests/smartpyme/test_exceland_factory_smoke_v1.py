from __future__ import annotations

import os
import sys
from pathlib import Path

import pytest

try:
    from exceland_factory.factory import build_product
except ImportError:
    exceland_factory = None
else:
    exceland_factory = build_product


def _locate_exceland2_src() -> Path | None:
    test_dir = Path(__file__).resolve().parent
    for candidate in (
        test_dir.parents[3] / ".." / ".." / ".." / "exeland2" / "src",
        test_dir.parents[4] / ".." / ".." / "exeland2" / "src",
    ):
        resolved = candidate.resolve()
        if resolved.is_dir():
            return resolved
    return None


def _ensure_exceland_importable() -> bool:
    if exceland_factory is not None:
        return True
    src = _locate_exceland2_src()
    if src is not None and str(src) not in sys.path:
        sys.path.insert(0, str(src))
    try:
        from exceland_factory.factory import build_product  # noqa: F811

        globals()["exceland_factory"] = build_product
        return True
    except ImportError:
        return False


def _find_spec_path(slug: str) -> Path | None:
    spec_dir_candidates = [
        Path(__file__).resolve().parents[0] / ".." / ".." / ".." / ".." / ".." / "exeland2" / "specs",
        Path(__file__).resolve().parents[0] / ".." / ".." / ".." / ".." / "exeland2" / "specs",
    ]
    for candidate in spec_dir_candidates:
        spec_file = (candidate / f"{slug}.yaml").resolve()
        if spec_file.exists():
            return spec_file
    return None


@pytest.mark.skipif(
    not _ensure_exceland_importable(),
    reason="exceland-factory is not importable; run pip install -e '.[exceland]' or pip install -e ../../../exeland2",
)
class TestExcelandFactorySmoke:
    def test_build_product_creates_xlsx_with_sheets(self, tmp_path: Path) -> None:
        output_path = tmp_path / "smoke_caja_diaria.xlsx"

        build_product("caja_diaria", output_path=output_path)

        assert output_path.exists(), (
            f"build_product did not create {output_path}"
        )
        assert output_path.stat().st_size > 0, (
            f"Generated XLSX is empty: {output_path}"
        )

        import openpyxl

        wb = openpyxl.load_workbook(output_path)
        sheet_names = wb.sheetnames
        assert len(sheet_names) >= 1, (
            f"XLSX has no sheets: {output_path}"
        )
        assert "BIENVENIDA" in sheet_names or any(
            "welcome" in s.lower() or "bienvenida" in s.lower() for s in sheet_names
        ), f"Expected a welcome sheet in {sheet_names}"

    def test_build_product_respects_runtime_authorized_false(self, tmp_path: Path) -> None:
        output_path = tmp_path / "smoke_precio_margen.xlsx"

        build_product("precio_margen", output_path=output_path)

        import openpyxl

        wb = openpyxl.load_workbook(output_path)
        assert "BIENVENIDA" in wb.sheetnames or "DATOS" in wb.sheetnames
        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and "runtime_authorized" in cell.value.lower():
                        assert "true" not in cell.value.lower(), (
                            f"runtime_authorized=true found in sheet '{sheet.title}' cell {cell.coordinate}"
                        )

    def test_build_product_with_minimal_spec(self, tmp_path: Path) -> None:
        spec_path = _find_spec_path("precio_margen")
        if spec_path is None:
            pytest.skip("precio_margen spec not found relative to test location")

        output_path = tmp_path / "smoke_precio_margen_minimal.xlsx"
        build_product(str(spec_path), output_path=output_path)

        assert output_path.exists()
        assert output_path.stat().st_size > 0

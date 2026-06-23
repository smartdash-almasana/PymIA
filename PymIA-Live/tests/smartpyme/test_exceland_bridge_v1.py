from __future__ import annotations

import inspect
from pathlib import Path

from openpyxl import load_workbook

from pymia.smartpyme.exceland_bridge_v1 import (
    CAPABILITY_REF,
    SOURCE_SYSTEM,
    build_exceland_bridge_v1,
)
from pymia.smartpyme.service_1_xlsx_delivery_v1 import build_service_1_xlsx_delivery_v1


def _bridge_input() -> dict[str, object]:
    return {
        "requested_template_ref": "precio_margen_basico_template",
        "requested_formula_refs": ["margen_bruto", "markup"],
        "input_fields_required": ["precio_venta", "costo_unitario"],
        "input_fields_received": {
            "precio_venta": 120,
            "costo_unitario": 80,
        },
        "warnings": ["Template request is logical only."],
        "limitations": ["No ejecuta spec YAML real todavía."],
        "technical_notes": ["Allowlist-only bridge contract."],
    }


def test_returns_ok_with_supported_template_formula_refs_and_complete_fields() -> None:
    result = build_exceland_bridge_v1(bridge_input=_bridge_input())

    assert result["status"] == "OK"
    assert result["source_system"] == SOURCE_SYSTEM
    assert result["capability_ref"] == CAPABILITY_REF
    assert result["runtime_authorized"] is False
    assert result["missing_inputs"] == []
    assert result["delivery_input"]["status"] == "OK"


def test_returns_missing_inputs_when_requested_template_ref_is_missing() -> None:
    bridge_input = _bridge_input()
    bridge_input["requested_template_ref"] = None

    result = build_exceland_bridge_v1(bridge_input=bridge_input)

    assert result["status"] == "MISSING_INPUTS"
    assert "requested_template_ref" in result["missing_inputs"]


def test_returns_unknown_template_when_template_is_outside_allowlist() -> None:
    bridge_input = _bridge_input()
    bridge_input["requested_template_ref"] = "rentabilidad_extrema_template"

    result = build_exceland_bridge_v1(bridge_input=bridge_input)

    assert result["status"] == "UNKNOWN_TEMPLATE"
    assert "allowlist mínima actual del bridge" in result["owner_summary"]


def test_returns_unsupported_formula_when_formula_is_outside_allowlist() -> None:
    bridge_input = _bridge_input()
    bridge_input["requested_formula_refs"] = ["margen_bruto", "resultado_neto"]

    result = build_exceland_bridge_v1(bridge_input=bridge_input)

    assert result["status"] == "UNSUPPORTED_FORMULA"
    assert result["delivery_input"]["computed_results"]["unsupported_formula_refs"] == ["resultado_neto"]


def test_returns_invalid_input_when_input_fields_required_is_not_list_of_strings() -> None:
    bridge_input = _bridge_input()
    bridge_input["input_fields_required"] = {"precio_venta": True}

    result = build_exceland_bridge_v1(bridge_input=bridge_input)  # type: ignore[arg-type]

    assert result["status"] == "INVALID_INPUT"
    assert result["delivery_input"]["status"] == "INVALID_INPUT"


def test_returns_missing_inputs_when_required_fields_are_missing() -> None:
    bridge_input = _bridge_input()
    bridge_input["input_fields_received"] = {
        "precio_venta": 120,
    }

    result = build_exceland_bridge_v1(bridge_input=bridge_input)

    assert result["status"] == "MISSING_INPUTS"
    assert "costo_unitario" in result["missing_inputs"]


def test_output_includes_delivery_input_compatible_with_service_1_xlsx_delivery() -> None:
    result = build_exceland_bridge_v1(bridge_input=_bridge_input())
    delivery_input = result["delivery_input"]

    assert delivery_input["service_name"] == "SERVICE_1"
    assert delivery_input["capability_ref"] == CAPABILITY_REF
    assert delivery_input["inputs_used"]["source_system"] == SOURCE_SYSTEM


def test_delivery_input_integrates_with_generic_xlsx_delivery(tmp_path: Path) -> None:
    result = build_exceland_bridge_v1(bridge_input=_bridge_input())
    output_path = tmp_path / "exceland_bridge.xlsx"

    delivery = build_service_1_xlsx_delivery_v1(
        delivery_input=result["delivery_input"],
        output_path=output_path,
    )

    workbook = load_workbook(output_path)

    assert output_path.exists()
    assert delivery["capability_ref"] == CAPABILITY_REF
    assert workbook["Resumen"]["B3"].value == CAPABILITY_REF
    assert ("source_system", SOURCE_SYSTEM) in list(workbook["Datos usados"].iter_rows(values_only=True))


def test_module_has_no_io_openpyxl_or_forbidden_runtime_dependencies() -> None:
    import pymia.smartpyme.exceland_bridge_v1 as module

    source = inspect.getsource(module)

    assert "openpyxl" not in source
    assert "first_aid" not in source
    assert "vertical_pipeline" not in source
    assert "service_1_pipeline" not in source
    assert "fsm" not in source.lower()
    assert "llm" not in source.lower()
    assert "chatbot" not in source.lower()
    assert "document_ingestion" not in source
    assert "open(" not in source
    assert ".save(" not in source
    assert "read_text(" not in source
    assert "read_bytes(" not in source

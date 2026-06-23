from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from pymia.smartpyme.invoice_collection_matching_sandbox_completion_slice_v1 import (
    CAPABILITY_REF,
    REVIEW_PACKET_CAPABILITY_REF,
    SERVICE_NAME,
    SYNTHETIC_CASE_ID,
    run_invoice_collection_matching_sandbox_completion_slice_v1,
)


def test_completion_slice_builds_ready_invoice_collection_sandbox_packet(tmp_path: Path) -> None:
    result = run_invoice_collection_matching_sandbox_completion_slice_v1(tmp_path)

    assert result["schema_version"] == "1.0"
    assert result["service_name"] == SERVICE_NAME
    assert result["capability_ref"] == CAPABILITY_REF
    assert result["case_id"] == SYNTHETIC_CASE_ID
    assert result["synthetic_data"] is True
    assert result["real_client_data"] is False
    assert result["runtime_authorized"] is False
    assert result["production_allowed"] is False
    assert result["final_status"] == "READY"


def test_completion_slice_component_chain_is_ready_and_safe(tmp_path: Path) -> None:
    result = run_invoice_collection_matching_sandbox_completion_slice_v1(tmp_path)

    assert result["base_contract"]["status"] == "READY_FOR_REVIEW"
    assert result["base_contract"]["runtime_authorized"] is False
    assert result["base_contract"]["next_allowed_action"] == "manual_invoice_collection_matching_review"
    assert result["human_review_gate"]["status"] == "PASS"
    assert result["human_review_gate"]["runtime_authorized"] is False
    assert result["xlsx_delivery"]["runtime_authorized"] is False


def test_completion_slice_generates_expected_matching_statuses(tmp_path: Path) -> None:
    result = run_invoice_collection_matching_sandbox_completion_slice_v1(tmp_path)

    assert result["status_counts"] == {
        "MATCHED_BY_INVOICE_NUMBER": 1,
        "PENDING_COLLECTION": 2,
        "UNMATCHED_COLLECTION": 1,
        "AMOUNT_DIFFERENCE_REVIEW": 1,
    }
    statuses = {row["status"] for row in result["matching_rows"]}
    assert statuses == {
        "MATCHED_BY_INVOICE_NUMBER",
        "PENDING_COLLECTION",
        "UNMATCHED_COLLECTION",
        "AMOUNT_DIFFERENCE_REVIEW",
    }


def test_completion_slice_generates_three_reviewable_output_files(tmp_path: Path) -> None:
    result = run_invoice_collection_matching_sandbox_completion_slice_v1(tmp_path)

    assert len(result["output_files"]) == 3
    for output_file in result["output_files"]:
        assert Path(output_file).exists()
        assert Path(output_file).stat().st_size > 0
        assert output_file in result["output_hashes"]
        assert len(result["output_hashes"][output_file]) == 64


def test_completion_slice_generates_readable_xlsx_delivery(tmp_path: Path) -> None:
    result = run_invoice_collection_matching_sandbox_completion_slice_v1(tmp_path)
    workbook = load_workbook(result["xlsx_delivery"]["output_path"])

    assert workbook["Resumen"]["B2"].value == "SERVICE_1"
    assert workbook["Resumen"]["B3"].value == REVIEW_PACKET_CAPABILITY_REF
    assert workbook["Resumen"]["B4"].value == "READY"
    assert ("matching_rows_count", "5") in list(workbook["Resultados"].iter_rows(values_only=True))
    assert ("matched_by_invoice_number", "1") in list(workbook["Resultados"].iter_rows(values_only=True))
    assert "Datos usados" in workbook.sheetnames
    assert "Resultados" in workbook.sheetnames
    assert "Limitaciones" in workbook.sheetnames
    assert "Claims prohibidos" in workbook.sheetnames


def test_completion_slice_owner_summary_is_conservative(tmp_path: Path) -> None:
    result = run_invoice_collection_matching_sandbox_completion_slice_v1(tmp_path)
    owner_summary = Path(result["owner_summary_path"]).read_text(encoding="utf-8")

    assert "Paquete sandbox de revisión facturas-cobros" in owner_summary
    assert "Coincidencias por número de factura: 1" in owner_summary
    assert "Facturas pendientes de cobro: 2" in owner_summary
    assert "Cobros sin factura asociada: 1" in owner_summary
    assert "Diferencias de importe para revisar: 1" in owner_summary
    assert "No confirma deuda final." in owner_summary
    assert "No confirma cobranza aplicada definitiva." in owner_summary
    assert "No certifica saldo de cliente." in owner_summary
    assert "No genera asientos contables." in owner_summary
    assert "No usa API ni archivos reales." in owner_summary


def test_completion_slice_operator_notes_explain_limits(tmp_path: Path) -> None:
    result = run_invoice_collection_matching_sandbox_completion_slice_v1(tmp_path)
    notes = Path(result["operator_notes_path"]).read_text(encoding="utf-8")

    assert "base_contract_status=READY_FOR_REVIEW" in notes
    assert "human_review_gate_status=PASS" in notes
    assert "matching_rows_count=5" in notes
    assert "matched_by_invoice_number=1" in notes
    assert "pending_collection=2" in notes
    assert "unmatched_collection=1" in notes
    assert "amount_difference_review=1" in notes
    assert "No API was called." in notes
    assert "No source files were read or parsed" in notes
    assert "No Mercado Pago logic is included." in notes


def test_completion_slice_rejects_missing_output_dir() -> None:
    missing_dir = Path("/nonexistent/service1/invoice-collection-matching-sandbox-completion")

    try:
        run_invoice_collection_matching_sandbox_completion_slice_v1(missing_dir)
    except FileNotFoundError as exc:
        assert "Output directory does not exist" in str(exc)
    else:
        raise AssertionError("Expected FileNotFoundError")


def test_completion_slice_source_does_not_open_forbidden_external_layers() -> None:
    module_path = (
        Path(__file__).parents[2]
        / "pymia"
        / "smartpyme"
        / "invoice_collection_matching_sandbox_completion_slice_v1.py"
    )
    source = module_path.read_text(encoding="utf-8").lower()

    forbidden_fragments = (
        "openai",
        "langchain",
        "requests",
        "httpx",
        "vertical_pipeline",
        "servicio_2",
        "chatbot",
        "ocr",
        "parser_runtime",
        "api_call",
        "mercado_pago_api",
    )
    for fragment in forbidden_fragments:
        assert fragment not in source

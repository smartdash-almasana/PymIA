from __future__ import annotations

import inspect

from openpyxl import load_workbook

from pymia.smartpyme.accounting_workpaper_manifest_model_v1 import (
    CAPABILITY_REF,
    build_accounting_workpaper_manifest_model_v1,
)
from pymia.smartpyme.service_1_xlsx_delivery_v1 import build_service_1_xlsx_delivery_v1


def _evidence_item(ref: str = "evidence-001") -> dict[str, object]:
    return {
        "evidence_ref": ref,
        "source_name": "Libro ventas enero",
        "source_kind": "xlsx_declared",
        "period_ref": "2026-01",
        "owner_supplied": True,
        "operator_notes": "Declarado por operador; archivo no parseado.",
        "sensitive_data_present": True,
    }


def _evidence_manifest() -> dict[str, object]:
    return {
        "manifest_id": "evidence-manifest-001",
        "period_ref": "2026-01",
        "evidence_items": [_evidence_item()],
        "live_source": False,
    }


def _template_manifest() -> dict[str, object]:
    return {
        "template_ref": "template-ventas-001",
        "template_name": "Papel de trabajo ventas mensual",
        "area_revision": "ventas",
        "required_sections": ["alcance", "evidencia_recibida", "faltantes"],
        "optional_sections": ["notas_operador"],
        "review_owner": "contador",
        "template_runtime_requested": False,
    }


def _bundle_input() -> dict[str, object]:
    return {"evidence_manifest": _evidence_manifest(), "template_manifest": _template_manifest()}


def test_valid_manifest_bundle_is_ready_for_draft_packet_handoff_only() -> None:
    result = build_accounting_workpaper_manifest_model_v1(bundle_input=_bundle_input())

    assert result["status"] == "VALID"
    assert result["capability_ref"] == CAPABILITY_REF
    assert result["runtime_authorized"] is False
    assert result["production_allowed"] is False
    assert result["valid_for_draft_packet"] is True
    assert result["evidence_manifest_id"] == "evidence-manifest-001"
    assert result["template_ref"] == "template-ventas-001"
    assert result["period_ref"] == "2026-01"
    assert result["area_revision"] == "ventas"
    assert result["evidence_count"] == 1
    assert result["required_section_count"] == 3
    assert result["optional_section_count"] == 1
    assert result["handoff_refs"] == ["evidence_manifest", "template_manifest"]


def test_missing_evidence_manifest_blocks_handoff() -> None:
    bundle_input = _bundle_input()
    bundle_input["evidence_manifest"] = None

    result = build_accounting_workpaper_manifest_model_v1(bundle_input=bundle_input)  # type: ignore[arg-type]

    assert result["status"] == "MISSING_EVIDENCE_MANIFEST"
    assert result["valid_for_draft_packet"] is False
    assert result["missing_inputs"] == ["evidence_manifest"]
    assert result["handoff_refs"] == ["template_manifest"]


def test_missing_template_manifest_blocks_handoff() -> None:
    bundle_input = _bundle_input()
    bundle_input["template_manifest"] = None

    result = build_accounting_workpaper_manifest_model_v1(bundle_input=bundle_input)  # type: ignore[arg-type]

    assert result["status"] == "MISSING_TEMPLATE_MANIFEST"
    assert result["valid_for_draft_packet"] is False
    assert result["missing_inputs"] == ["template_manifest"]
    assert result["handoff_refs"] == ["evidence_manifest"]


def test_invalid_input_blocks_handoff() -> None:
    result = build_accounting_workpaper_manifest_model_v1(bundle_input=[])  # type: ignore[arg-type]

    assert result["status"] == "INVALID_INPUT"
    assert result["missing_inputs"] == ["bundle_input"]
    assert result["runtime_authorized"] is False


def test_blocks_live_source_or_template_runtime_request() -> None:
    bundle_input = _bundle_input()
    evidence_manifest = dict(_evidence_manifest())
    evidence_manifest["live_source"] = True
    bundle_input["evidence_manifest"] = evidence_manifest

    result = build_accounting_workpaper_manifest_model_v1(bundle_input=bundle_input)  # type: ignore[arg-type]

    assert result["status"] == "BLOCKED_LIVE_SOURCE"
    assert result["valid_for_draft_packet"] is False
    assert "No source file was parsed." in result["forbidden_claims"]


def test_blocks_template_runtime_request() -> None:
    bundle_input = _bundle_input()
    template_manifest = dict(_template_manifest())
    template_manifest["template_runtime_requested"] = True
    bundle_input["template_manifest"] = template_manifest

    result = build_accounting_workpaper_manifest_model_v1(bundle_input=bundle_input)  # type: ignore[arg-type]

    assert result["status"] == "BLOCKED_LIVE_SOURCE"
    assert result["valid_for_draft_packet"] is False
    assert "No template was executed as runtime." in result["forbidden_claims"]


def test_invalid_evidence_item_when_required_field_is_missing() -> None:
    bundle_input = _bundle_input()
    evidence_manifest = _evidence_manifest()
    evidence_manifest["evidence_items"] = [{**_evidence_item(), "source_name": ""}]
    bundle_input["evidence_manifest"] = evidence_manifest

    result = build_accounting_workpaper_manifest_model_v1(bundle_input=bundle_input)  # type: ignore[arg-type]

    assert result["status"] == "INVALID_EVIDENCE_ITEM"
    assert result["valid_for_draft_packet"] is False
    assert result["reasons"] == ["Evidence item 0 missing source_name."]


def test_invalid_evidence_item_when_boolean_field_is_not_boolean() -> None:
    bundle_input = _bundle_input()
    evidence_manifest = _evidence_manifest()
    evidence_manifest["evidence_items"] = [{**_evidence_item(), "owner_supplied": "yes"}]
    bundle_input["evidence_manifest"] = evidence_manifest

    result = build_accounting_workpaper_manifest_model_v1(bundle_input=bundle_input)  # type: ignore[arg-type]

    assert result["status"] == "INVALID_EVIDENCE_ITEM"
    assert result["reasons"] == ["Evidence item 0 owner_supplied must be boolean."]


def test_invalid_template_manifest_when_required_sections_empty() -> None:
    bundle_input = _bundle_input()
    template_manifest = dict(_template_manifest())
    template_manifest["required_sections"] = []
    bundle_input["template_manifest"] = template_manifest

    result = build_accounting_workpaper_manifest_model_v1(bundle_input=bundle_input)  # type: ignore[arg-type]

    assert result["status"] == "INVALID_TEMPLATE_SECTION"
    assert result["reasons"] == ["Template manifest required_sections must be a non-empty list of strings."]


def test_duplicate_evidence_ref_blocks_handoff() -> None:
    bundle_input = _bundle_input()
    evidence_manifest = _evidence_manifest()
    evidence_manifest["evidence_items"] = [_evidence_item("dup"), _evidence_item("dup")]
    bundle_input["evidence_manifest"] = evidence_manifest

    result = build_accounting_workpaper_manifest_model_v1(bundle_input=bundle_input)  # type: ignore[arg-type]

    assert result["status"] == "DUPLICATE_EVIDENCE_REF"
    assert result["reasons"] == ["Duplicate evidence_ref: dup."]


def test_duplicate_template_section_blocks_handoff() -> None:
    bundle_input = _bundle_input()
    template_manifest = dict(_template_manifest())
    template_manifest["optional_sections"] = ["alcance"]
    bundle_input["template_manifest"] = template_manifest

    result = build_accounting_workpaper_manifest_model_v1(bundle_input=bundle_input)  # type: ignore[arg-type]

    assert result["status"] == "DUPLICATE_TEMPLATE_SECTION"
    assert result["reasons"] == ["Duplicate template section: alcance."]


def test_delivery_input_is_compatible_with_xlsx_delivery(tmp_path) -> None:
    result = build_accounting_workpaper_manifest_model_v1(bundle_input=_bundle_input())
    output_path = tmp_path / "workpaper_manifest.xlsx"

    delivery = build_service_1_xlsx_delivery_v1(
        delivery_input=result["delivery_input"],
        output_path=output_path,
    )

    workbook = load_workbook(output_path)
    assert delivery["capability_ref"] == CAPABILITY_REF
    assert workbook["Resumen"]["B3"].value == CAPABILITY_REF
    assert workbook["Resumen"]["B4"].value == "VALID"
    assert ("valid_for_draft_packet", "true") in list(workbook["Resultados"].iter_rows(values_only=True))


def test_model_does_not_import_runtime_io_or_excel_dependencies() -> None:
    import pymia.smartpyme.accounting_workpaper_manifest_model_v1 as module

    source = inspect.getsource(module)

    forbidden_tokens = [
        "openpyxl",
        "pandas",
        "Path(",
        "open(",
        "requests",
        "httpx",
        "subprocess",
        "vertical_slice",
    ]
    for token in forbidden_tokens:
        assert token not in source

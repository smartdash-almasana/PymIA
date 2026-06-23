from __future__ import annotations

import inspect
from pathlib import Path

from openpyxl import load_workbook

from pymia.smartpyme.accounting_human_review_gate_v1 import evaluate_accounting_human_review_gate_v1
from pymia.smartpyme.accounting_workpaper_contract_v1 import build_accounting_workpaper_contract_v1
from pymia.smartpyme.accounting_workpaper_draft_packet_v1 import (
    CAPABILITY_REF,
    build_accounting_workpaper_draft_packet_v1,
)
from pymia.smartpyme.accounting_workpaper_manifest_model_v1 import build_accounting_workpaper_manifest_model_v1
from pymia.smartpyme.service_1_xlsx_delivery_v1 import build_service_1_xlsx_delivery_v1


def _contract_result() -> dict[str, object]:
    return build_accounting_workpaper_contract_v1(
        contract_input={
            "owner_requested_output": "accounting_workpaper_scope_report",
            "source_files_received": ["evidencia_soporte", "plantilla_papel_trabajo"],
            "received_fields": ["periodo", "cliente", "area_revision", "responsable"],
        }
    )


def _manifest_result() -> dict[str, object]:
    return build_accounting_workpaper_manifest_model_v1(
        bundle_input={
            "evidence_manifest": {
                "manifest_id": "evidence-manifest-001",
                "period_ref": "2026-06",
                "evidence_items": [
                    {
                        "evidence_ref": "evidence-001",
                        "source_name": "Balance soporte junio",
                        "source_kind": "xlsx_declared",
                        "period_ref": "2026-06",
                        "owner_supplied": True,
                        "operator_notes": "Declarado por operador; no parseado.",
                        "sensitive_data_present": True,
                    }
                ],
                "live_source": False,
            },
            "template_manifest": {
                "template_ref": "workpaper-template-001",
                "template_name": "Borrador revisión mensual",
                "area_revision": "caja",
                "required_sections": ["alcance", "evidencia", "faltantes"],
                "optional_sections": ["notas_operador"],
                "review_owner": "operator",
                "template_runtime_requested": False,
            },
        }
    )


def _human_gate_result() -> dict[str, object]:
    return evaluate_accounting_human_review_gate_v1(
        gate_input={
            "capability_ref": "accounting_workpaper_basic",
            "reviewer_role": "operator",
            "decision": "APPROVED",
            "scope_ok": True,
            "evidence_ok": True,
            "forbidden_claims": [],
            "live_use": False,
        }
    )


def _packet_input() -> dict[str, object]:
    return {
        "workpaper_contract_result": _contract_result(),
        "manifest_model_result": _manifest_result(),
        "human_review_gate_result": _human_gate_result(),
    }


def test_ready_when_contract_manifest_and_human_gate_are_ready() -> None:
    result = build_accounting_workpaper_draft_packet_v1(packet_input=_packet_input())

    assert result["status"] == "READY"
    assert result["capability_ref"] == CAPABILITY_REF
    assert result["runtime_authorized"] is False
    assert result["production_allowed"] is False
    assert result["blocked_reasons"] == []
    assert result["readiness_flags"]["workpaper_contract_ready"] is True
    assert result["readiness_flags"]["manifest_model_valid"] is True
    assert result["readiness_flags"]["human_review_gate_passed"] is True
    assert result["readiness_flags"]["runtime_authorization_clear"] is True
    assert result["readiness_flags"]["production_use_clear"] is True


def test_invalid_input_if_packet_input_is_not_dict() -> None:
    result = build_accounting_workpaper_draft_packet_v1(packet_input=[])  # type: ignore[arg-type]

    assert result["status"] == "INVALID_INPUT"
    assert result["blocked_reasons"] == ["invalid_packet_input"]
    assert result["production_allowed"] is False


def test_invalid_input_if_components_are_missing() -> None:
    result = build_accounting_workpaper_draft_packet_v1(
        packet_input={"workpaper_contract_result": _contract_result()}  # type: ignore[arg-type]
    )

    assert result["status"] == "INVALID_INPUT"
    assert result["blocked_reasons"] == ["invalid_packet_components"]


def test_blocked_if_contract_is_not_ready_for_review() -> None:
    packet_input = _packet_input()
    packet_input["workpaper_contract_result"] = {
        **packet_input["workpaper_contract_result"],
        "status": "MISSING_FIELDS",
    }

    result = build_accounting_workpaper_draft_packet_v1(packet_input=packet_input)  # type: ignore[arg-type]

    assert result["status"] == "BLOCKED"
    assert "workpaper_contract_not_ready" in result["blocked_reasons"]


def test_blocked_if_manifest_is_not_valid() -> None:
    packet_input = _packet_input()
    packet_input["manifest_model_result"] = {
        **packet_input["manifest_model_result"],
        "status": "INVALID_TEMPLATE_SECTION",
    }

    result = build_accounting_workpaper_draft_packet_v1(packet_input=packet_input)  # type: ignore[arg-type]

    assert result["status"] == "BLOCKED"
    assert "manifest_model_not_valid" in result["blocked_reasons"]


def test_blocked_if_human_gate_is_not_pass() -> None:
    packet_input = _packet_input()
    packet_input["human_review_gate_result"] = {
        **packet_input["human_review_gate_result"],
        "status": "PENDING",
    }

    result = build_accounting_workpaper_draft_packet_v1(packet_input=packet_input)  # type: ignore[arg-type]

    assert result["status"] == "BLOCKED"
    assert "human_review_gate_not_passed" in result["blocked_reasons"]


def test_blocked_if_any_component_has_runtime_authorized_true() -> None:
    packet_input = _packet_input()
    packet_input["manifest_model_result"] = {
        **packet_input["manifest_model_result"],
        "runtime_authorized": True,
    }

    result = build_accounting_workpaper_draft_packet_v1(packet_input=packet_input)  # type: ignore[arg-type]

    assert result["status"] == "BLOCKED"
    assert "runtime_authorization_forbidden" in result["blocked_reasons"]


def test_blocked_if_any_component_has_production_allowed_true() -> None:
    packet_input = _packet_input()
    packet_input["workpaper_contract_result"] = {
        **packet_input["workpaper_contract_result"],
        "production_allowed": True,
    }

    result = build_accounting_workpaper_draft_packet_v1(packet_input=packet_input)  # type: ignore[arg-type]

    assert result["status"] == "BLOCKED"
    assert "production_use_forbidden" in result["blocked_reasons"]


def test_delivery_input_is_compatible_with_build_service_1_xlsx_delivery_v1(tmp_path: Path) -> None:
    result = build_accounting_workpaper_draft_packet_v1(packet_input=_packet_input())
    output_path = tmp_path / "accounting_workpaper_draft_packet.xlsx"

    delivery = build_service_1_xlsx_delivery_v1(
        delivery_input=result["delivery_input"],
        output_path=output_path,
    )
    workbook = load_workbook(output_path)

    assert output_path.exists()
    assert delivery["capability_ref"] == CAPABILITY_REF
    assert workbook["Resumen"]["B3"].value == CAPABILITY_REF
    assert ("next_allowed_action", "prepare_owner_operator_workpaper_draft_review") in list(
        workbook["Resultados"].iter_rows(values_only=True)
    )


def test_productive_module_does_not_import_forbidden_dependencies() -> None:
    import pymia.smartpyme.accounting_workpaper_draft_packet_v1 as module

    source = inspect.getsource(module)

    forbidden_tokens = [
        "openpyxl",
        "pandas",
        "Path(",
        "open(",
        "requests",
        "httpx",
        "subprocess",
        "vertical_slice",
    ]
    for token in forbidden_tokens:
        assert token not in source

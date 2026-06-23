from __future__ import annotations

import inspect
from pathlib import Path

import pytest
from openpyxl import load_workbook

from pymia.smartpyme.service_1_xlsx_delivery_v1 import (
    DELIVERY_SCHEMA_VERSION,
    SERVICE_NAME,
    SHEET_NAMES,
    Service1XlsxDeliveryInputV1,
    build_service_1_xlsx_delivery_v1,
)


def _delivery_input() -> Service1XlsxDeliveryInputV1:
    return {
        "service_name": "SERVICE_1",
        "capability_ref": "excel_lab_curated_table",
        "status": "OK",
        "owner_summary": "Archivo curado generado sobre datos declarados.",
        "inputs_used": {"source_file": "ventas.xlsx", "rows": 10},
        "computed_results": {"normalized_rows": 10, "warnings": ["header ambiguous"]},
        "missing_inputs": [],
        "limitations": ["No confirma verdad de negocio."],
        "forbidden_claims": ["No confirma archivo normalizado."],
        "technical_notes": ["Generic Service 1 delivery input."],
        "runtime_authorized": False,
    }


def test_generates_xlsx_from_generic_service_1_delivery_input(tmp_path: Path) -> None:
    output_path = tmp_path / "service_1_delivery.xlsx"

    delivery = build_service_1_xlsx_delivery_v1(
        delivery_input=_delivery_input(),
        output_path=output_path,
    )

    assert output_path.exists()
    assert delivery["schema_version"] == DELIVERY_SCHEMA_VERSION
    assert delivery["service_name"] == SERVICE_NAME
    assert delivery["capability_ref"] == "excel_lab_curated_table"
    assert delivery["runtime_authorized"] is False
    assert delivery["sheet_names"] == list(SHEET_NAMES)


def test_summary_sheet_uses_capability_ref_not_tool_ref(tmp_path: Path) -> None:
    output_path = tmp_path / "summary.xlsx"

    build_service_1_xlsx_delivery_v1(
        delivery_input=_delivery_input(),
        output_path=output_path,
    )

    workbook = load_workbook(output_path)
    summary_rows = list(workbook["Resumen"].iter_rows(values_only=True))

    assert summary_rows == [
        ("field", "value"),
        ("service_name", "SERVICE_1"),
        ("capability_ref", "excel_lab_curated_table"),
        ("status", "OK"),
        ("owner_summary", "Archivo curado generado sobre datos declarados."),
        ("runtime_authorized", "false"),
    ]


def test_generic_delivery_includes_all_expected_sheets(tmp_path: Path) -> None:
    output_path = tmp_path / "sheets.xlsx"

    build_service_1_xlsx_delivery_v1(
        delivery_input=_delivery_input(),
        output_path=output_path,
    )

    workbook = load_workbook(output_path)
    assert workbook.sheetnames == list(SHEET_NAMES)


def test_generic_delivery_serializes_nested_results_safely(tmp_path: Path) -> None:
    output_path = tmp_path / "nested.xlsx"

    build_service_1_xlsx_delivery_v1(
        delivery_input=_delivery_input(),
        output_path=output_path,
    )

    workbook = load_workbook(output_path)
    rows = list(workbook["Resultados"].iter_rows(values_only=True))

    assert ("normalized_rows", "10") in rows
    assert ("warnings", '["header ambiguous"]') in rows


def test_generic_delivery_escapes_formula_like_text(tmp_path: Path) -> None:
    delivery_input = _delivery_input()
    delivery_input["owner_summary"] = "=DANGEROUS"
    output_path = tmp_path / "safe.xlsx"

    build_service_1_xlsx_delivery_v1(
        delivery_input=delivery_input,
        output_path=output_path,
    )

    workbook = load_workbook(output_path)
    assert workbook.vba_archive is None
    assert workbook["Resumen"]["B5"].value == "'=DANGEROUS"
    for sheet_name in workbook.sheetnames:
        for row in workbook[sheet_name].iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                assert cell.data_type != "f"


def test_rejects_runtime_authorized_input(tmp_path: Path) -> None:
    delivery_input = _delivery_input()
    delivery_input["runtime_authorized"] = True

    with pytest.raises(ValueError, match="does not accept runtime_authorized=True"):
        build_service_1_xlsx_delivery_v1(
            delivery_input=delivery_input,
            output_path=tmp_path / "blocked.xlsx",
        )


def test_rejects_non_service_1_input(tmp_path: Path) -> None:
    delivery_input = _delivery_input()
    delivery_input["service_name"] = "SERVICE_2"

    with pytest.raises(ValueError, match="only accepts SERVICE_1 inputs"):
        build_service_1_xlsx_delivery_v1(
            delivery_input=delivery_input,
            output_path=tmp_path / "blocked.xlsx",
        )


def test_rejects_missing_output_dir() -> None:
    with pytest.raises(FileNotFoundError, match="Output directory does not exist"):
        build_service_1_xlsx_delivery_v1(
            delivery_input=_delivery_input(),
            output_path="/nonexistent/service1/xlsx/delivery.xlsx",
        )


def test_module_does_not_import_product_family_runtime() -> None:
    import pymia.smartpyme.service_1_xlsx_delivery_v1 as module

    source = inspect.getsource(module)

    assert "first_aid_" not in source
    assert "document_ingestion" not in source
    assert "exceland" not in source.lower()
    assert "vertical_pipeline" not in source
    assert "service_1_pipeline" not in source
    assert "service_1_fsm_decision_patch_v1" not in source
    assert "service_1_boundary_chain_v1" not in source
    assert "llm" not in source.lower()
    assert "chatbot" not in source.lower()

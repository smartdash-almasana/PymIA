from __future__ import annotations

import inspect
from pathlib import Path

import pytest
from openpyxl import load_workbook

from pymia.smartpyme.service_1_operator_harness_v1 import (
    SAMPLE_CASE_ID,
    SERVICE_NAME,
    build_service_1_operator_harness_sample_case_v1,
    run_service_1_operator_harness_v1,
)


def test_operator_harness_sample_case_creates_delivery_folder(tmp_path: Path) -> None:
    case = build_service_1_operator_harness_sample_case_v1()

    result = run_service_1_operator_harness_v1(case=case, output_root=tmp_path)

    delivery_dir = Path(result["delivery_dir"])
    assert result["schema_version"] == "1.0"
    assert result["service_name"] == SERVICE_NAME
    assert result["case_id"] == SAMPLE_CASE_ID
    assert result["runtime_authorized"] is False
    assert delivery_dir.exists()
    assert delivery_dir.is_dir()


def test_operator_harness_runs_pipeline_and_generates_five_xlsx_files(tmp_path: Path) -> None:
    result = run_service_1_operator_harness_v1(
        case=build_service_1_operator_harness_sample_case_v1(),
        output_root=tmp_path,
    )

    assert result["pipeline_result"]["requested_tool_count"] == 5
    assert result["pipeline_result"]["executed_tool_refs"] == [
        "precio_margen_basico",
        "caja_diaria_triage",
        "stock_alertas_basicas",
        "gastos_triage",
        "proveedores_precio_variacion_triage",
    ]
    assert len(result["generated_files"]) == 5
    for generated_file in result["generated_files"]:
        assert Path(generated_file).exists()
        assert Path(generated_file).suffix == ".xlsx"


def test_operator_harness_outputs_are_readable(tmp_path: Path) -> None:
    result = run_service_1_operator_harness_v1(
        case=build_service_1_operator_harness_sample_case_v1(),
        output_root=tmp_path,
    )

    for generated_file in result["generated_files"]:
        workbook = load_workbook(generated_file)
        assert workbook["Resumen"]["B2"].value == "SERVICE_1"
        assert "Resultados" in workbook.sheetnames
        assert "Claims prohibidos" in workbook.sheetnames


def test_operator_harness_writes_summary_and_operator_report(tmp_path: Path) -> None:
    result = run_service_1_operator_harness_v1(
        case=build_service_1_operator_harness_sample_case_v1(),
        output_root=tmp_path,
    )

    summary_path = Path(result["summary_path"])
    report_path = Path(result["operator_report_path"])

    assert summary_path.exists()
    assert report_path.exists()
    assert summary_path.read_text(encoding="utf-8") == result["summary_text"]
    assert "Entrega preliminar basada en datos declarados." in report_path.read_text(
        encoding="utf-8"
    )


def test_operator_harness_summary_contains_owner_relevant_limits(tmp_path: Path) -> None:
    result = run_service_1_operator_harness_v1(
        case=build_service_1_operator_harness_sample_case_v1(),
        output_root=tmp_path,
    )
    summary_text = result["summary_text"]

    assert "Resultados procesados: 5" in summary_text
    assert "precio_margen_basico: OK" in summary_text
    assert "caja_diaria_triage: OK" in summary_text
    assert "stock_alertas_basicas: OK" in summary_text
    assert "gastos_triage: OK" in summary_text
    assert "proveedores_precio_variacion_triage: OK" in summary_text
    assert "No es un diagnostico integral" in summary_text
    assert "No confirma saldo bancario real" in summary_text
    assert "No confirma stock fisico real" in summary_text


def test_operator_harness_preserves_custom_case_order(tmp_path: Path) -> None:
    case = {
        "case_id": "custom-case",
        "case_name": "Custom case",
        "tool_requests": [
            {
                "tool_ref": "stock_alertas_basicas",
                "inputs": {"producto": "SKU", "stock_actual": 2, "stock_minimo": 5},
            },
            {
                "tool_ref": "precio_margen_basico",
                "inputs": {"precio_venta": 100, "costo_unitario": 60},
            },
        ],
        "operator_notes": ["custom note"],
    }

    result = run_service_1_operator_harness_v1(case=case, output_root=tmp_path)  # type: ignore[arg-type]

    assert result["pipeline_result"]["executed_tool_refs"] == [
        "stock_alertas_basicas",
        "precio_margen_basico",
    ]
    assert result["pipeline_result"]["delivery_flow"]["tool_refs"] == [
        "stock_alertas_basicas",
        "precio_margen_basico",
    ]


def test_operator_harness_sanitizes_case_id_for_folder(tmp_path: Path) -> None:
    case = build_service_1_operator_harness_sample_case_v1()
    case["case_id"] = "case with/slash"

    result = run_service_1_operator_harness_v1(case=case, output_root=tmp_path)

    assert result["case_id"] == "case_with_slash"
    assert Path(result["delivery_dir"]).name == "case_with_slash"


def test_operator_harness_rejects_missing_output_root() -> None:
    with pytest.raises(FileNotFoundError, match="Output root does not exist"):
        run_service_1_operator_harness_v1(
            case=build_service_1_operator_harness_sample_case_v1(),
            output_root="/nonexistent/operator/harness/root",
        )


def test_operator_harness_rejects_empty_case_id(tmp_path: Path) -> None:
    case = build_service_1_operator_harness_sample_case_v1()
    case["case_id"] = "///"

    with pytest.raises(ValueError, match="requires a non-empty case_id"):
        run_service_1_operator_harness_v1(case=case, output_root=tmp_path)


def test_operator_harness_rejects_empty_tool_requests(tmp_path: Path) -> None:
    case = build_service_1_operator_harness_sample_case_v1()
    case["tool_requests"] = []

    with pytest.raises(ValueError, match="requires at least one tool request"):
        run_service_1_operator_harness_v1(case=case, output_root=tmp_path)


def test_operator_harness_does_not_import_first_aid_tools_directly() -> None:
    import pymia.smartpyme.service_1_operator_harness_v1 as module

    source = inspect.getsource(module)

    assert "first_aid_precio_margen_basico_v1" not in source
    assert "first_aid_caja_diaria_triage_v1" not in source
    assert "first_aid_stock_alertas_basicas_v1" not in source
    assert "first_aid_gastos_triage_v1" not in source
    assert "first_aid_proveedores_precio_variacion_triage_v1" not in source
    assert "run_precio_margen_basico_v1" not in source
    assert "run_caja_diaria_triage_v1" not in source
    assert "run_stock_alertas_basicas_v1" not in source
    assert "run_gastos_triage_v1" not in source
    assert "run_proveedores_precio_variacion_triage_v1" not in source


def test_operator_harness_does_not_depend_on_forbidden_product_layers() -> None:
    import pymia.smartpyme.service_1_operator_harness_v1 as module

    source = inspect.getsource(module)

    assert "vertical_pipeline" not in source
    assert "service_1_fsm_decision_patch_v1" not in source
    assert "service_1_boundary_chain_v1" not in source
    assert "document_ingestion" not in source
    assert "exceland" not in source.lower()
    assert "openai" not in source.lower()
    assert "chatbot" not in source.lower()

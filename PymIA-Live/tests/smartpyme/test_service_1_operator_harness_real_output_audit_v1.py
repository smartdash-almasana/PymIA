from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from pymia.smartpyme.service_1_operator_harness_v1 import (
    build_service_1_operator_harness_sample_case_v1,
    run_service_1_operator_harness_v1,
)


def test_operator_harness_real_output_audit_artifact_inventory(tmp_path: Path) -> None:
    run = run_service_1_operator_harness_v1(
        case=build_service_1_operator_harness_sample_case_v1(),
        output_root=tmp_path,
    )

    delivery_dir = Path(run["delivery_dir"])
    files = sorted(path.name for path in delivery_dir.iterdir() if path.is_file())

    assert files == [
        "first_aid_001_precio_margen_basico.xlsx",
        "first_aid_002_caja_diaria_triage.xlsx",
        "first_aid_003_stock_alertas_basicas.xlsx",
        "operator_report.txt",
        "summary.txt",
    ]


def test_operator_harness_real_output_audit_xlsx_contract(tmp_path: Path) -> None:
    run = run_service_1_operator_harness_v1(
        case=build_service_1_operator_harness_sample_case_v1(),
        output_root=tmp_path,
    )

    expected_sheets = [
        "Resumen",
        "Datos usados",
        "Resultados",
        "Faltantes",
        "Limitaciones",
        "Claims prohibidos",
        "Notas técnicas",
    ]

    for generated_file in run["generated_files"]:
        workbook = load_workbook(generated_file)
        assert workbook.sheetnames == expected_sheets
        assert workbook["Resumen"]["B2"].value == "SERVICE_1"
        assert workbook["Resumen"]["B4"].value == "OK"
        assert workbook["Claims prohibidos"].max_row >= 2
        assert workbook["Limitaciones"].max_row >= 2


def test_operator_harness_real_output_audit_summary_is_deliverable(tmp_path: Path) -> None:
    run = run_service_1_operator_harness_v1(
        case=build_service_1_operator_harness_sample_case_v1(),
        output_root=tmp_path,
    )

    summary_path = Path(run["summary_path"])
    summary = summary_path.read_text(encoding="utf-8")

    assert summary == run["summary_text"]
    assert "Resultados procesados: 3" in summary
    assert "precio_margen_basico: OK" in summary
    assert "caja_diaria_triage: OK" in summary
    assert "stock_alertas_basicas: OK" in summary
    assert "Limitaciones principales:" in summary
    assert "Aclaraciones conservadoras:" in summary
    assert "No es un diagnostico integral" in summary
    assert "No confirma saldo bancario real" in summary
    assert "No confirma stock fisico real" in summary
    assert "Entrega preliminar basada en datos declarados." in summary


def test_operator_harness_real_output_audit_operator_report_is_usable(tmp_path: Path) -> None:
    run = run_service_1_operator_harness_v1(
        case=build_service_1_operator_harness_sample_case_v1(),
        output_root=tmp_path,
    )

    report = Path(run["operator_report_path"]).read_text(encoding="utf-8")

    assert "Caso: Comercio minorista alimentos - First Aid demo" in report
    assert "Tools ejecutadas: 3" in report
    assert "precio_margen_basico: OK" in report
    assert "caja_diaria_triage: OK" in report
    assert "stock_alertas_basicas: OK" in report
    assert "Notas operador:" in report
    assert "Entrega preliminar basada en datos declarados." in report


def test_operator_harness_real_output_audit_metadata_matches_files(tmp_path: Path) -> None:
    run = run_service_1_operator_harness_v1(
        case=build_service_1_operator_harness_sample_case_v1(),
        output_root=tmp_path,
    )

    delivery_paths = [delivery["output_path"] for delivery in run["pipeline_result"]["delivery_flow"]["deliveries"]]

    assert run["generated_files"] == delivery_paths
    assert run["pipeline_result"]["delivery_flow"]["delivery_count"] == 3
    assert run["pipeline_result"]["delivery_flow"]["tool_refs"] == [
        "precio_margen_basico",
        "caja_diaria_triage",
        "stock_alertas_basicas",
    ]
    assert run["runtime_authorized"] is False
    assert run["pipeline_result"]["runtime_authorized"] is False
    assert run["pipeline_result"]["delivery_flow"]["runtime_authorized"] is False

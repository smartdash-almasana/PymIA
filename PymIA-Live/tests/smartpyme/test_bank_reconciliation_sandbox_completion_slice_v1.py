from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from pymia.smartpyme.bank_reconciliation_sandbox_completion_slice_v1 import (
    CAPABILITY_REF,
    SERVICE_NAME,
    SYNTHETIC_CASE_ID,
    run_bank_reconciliation_sandbox_completion_slice_v1,
)


def test_completion_slice_builds_ready_bank_sandbox_review_packet(tmp_path: Path) -> None:
    result = run_bank_reconciliation_sandbox_completion_slice_v1(tmp_path)

    assert result["schema_version"] == "1.0"
    assert result["service_name"] == SERVICE_NAME
    assert result["capability_ref"] == CAPABILITY_REF
    assert result["case_id"] == SYNTHETIC_CASE_ID
    assert result["synthetic_data"] is True
    assert result["real_client_data"] is False
    assert result["runtime_authorized"] is False
    assert result["production_allowed"] is False
    assert result["final_status"] == "READY"


def test_completion_slice_component_chain_is_ready_and_safe(tmp_path: Path) -> None:
    result = run_bank_reconciliation_sandbox_completion_slice_v1(tmp_path)

    assert result["base_contract"]["status"] == "READY_FOR_REVIEW"
    assert result["human_review_gate"]["status"] == "PASS"
    assert result["fixture_model"]["status"] == "VALID"
    assert result["fixture_handoff"]["status"] == "READY"
    assert result["sandbox_contract"]["status"] == "READY_FOR_SANDBOX_CONTRACT"
    assert result["review_packet"]["status"] == "READY"
    assert result["xlsx_delivery"]["runtime_authorized"] is False
    assert result["review_packet"]["production_allowed"] is False


def test_completion_slice_generates_three_reviewable_output_files(tmp_path: Path) -> None:
    result = run_bank_reconciliation_sandbox_completion_slice_v1(tmp_path)

    assert len(result["output_files"]) == 3
    for output_file in result["output_files"]:
        assert Path(output_file).exists()
        assert Path(output_file).stat().st_size > 0
        assert output_file in result["output_hashes"]
        assert len(result["output_hashes"][output_file]) == 64


def test_completion_slice_generates_readable_xlsx_delivery(tmp_path: Path) -> None:
    result = run_bank_reconciliation_sandbox_completion_slice_v1(tmp_path)
    workbook = load_workbook(result["xlsx_delivery"]["output_path"])

    assert workbook["Resumen"]["B2"].value == "SERVICE_1"
    assert workbook["Resumen"]["B3"].value == "service_1_bank_reconciliation_sandbox_review_packet_v1"
    assert workbook["Resumen"]["B4"].value == "READY"
    assert "Datos usados" in workbook.sheetnames
    assert "Resultados" in workbook.sheetnames
    assert "Limitaciones" in workbook.sheetnames
    assert "Claims prohibidos" in workbook.sheetnames


def test_completion_slice_owner_summary_is_conservative(tmp_path: Path) -> None:
    result = run_bank_reconciliation_sandbox_completion_slice_v1(tmp_path)
    owner_summary = Path(result["owner_summary_path"]).read_text(encoding="utf-8")

    assert "Paquete sandbox de revisión bancaria" in owner_summary
    assert "No confirma saldo conciliado." in owner_summary
    assert "No confirma diferencia final." in owner_summary
    assert "No genera asientos contables." in owner_summary
    assert "No certifica exactitud contable o fiscal." in owner_summary
    assert "No lee extractos bancarios reales." in owner_summary
    assert "No usa API bancaria." in owner_summary


def test_completion_slice_operator_notes_explain_limits(tmp_path: Path) -> None:
    result = run_bank_reconciliation_sandbox_completion_slice_v1(tmp_path)
    notes = Path(result["operator_notes_path"]).read_text(encoding="utf-8")

    assert "base_contract_status=READY_FOR_REVIEW" in notes
    assert "human_review_gate_status=PASS" in notes
    assert "fixture_model_status=VALID" in notes
    assert "fixture_handoff_status=READY" in notes
    assert "sandbox_contract_status=READY_FOR_SANDBOX_CONTRACT" in notes
    assert "review_packet_status=READY" in notes
    assert "No bank API was called." in notes
    assert "No source files were read or parsed" in notes


def test_completion_slice_rejects_missing_output_dir() -> None:
    missing_dir = Path("/nonexistent/service1/bank-reconciliation-sandbox-completion")

    try:
        run_bank_reconciliation_sandbox_completion_slice_v1(missing_dir)
    except FileNotFoundError as exc:
        assert "Output directory does not exist" in str(exc)
    else:
        raise AssertionError("Expected FileNotFoundError")


def test_completion_slice_source_does_not_open_forbidden_external_layers() -> None:
    module_path = Path(__file__).parents[2] / "pymia" / "smartpyme" / "bank_reconciliation_sandbox_completion_slice_v1.py"
    source = module_path.read_text(encoding="utf-8").lower()

    forbidden_fragments = (
        "openai",
        "langchain",
        "requests",
        "httpx",
        "vertical_pipeline",
        "servicio_2",
        "chatbot",
        "ocr",
        "parser_runtime",
        "api_call",
    )
    for fragment in forbidden_fragments:
        assert fragment not in source

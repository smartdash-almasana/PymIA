from __future__ import annotations

import inspect
from pathlib import Path

from openpyxl import load_workbook

from pymia.smartpyme.excel_treatment_lab_v1 import (
    CAPABILITY_REF,
    SCHEMA_VERSION,
    build_excel_treatment_lab_v1,
)
from pymia.smartpyme.service_1_xlsx_delivery_v1 import build_service_1_xlsx_delivery_v1


def _lab_input() -> dict[str, object]:
    return {
        "source_file": "ventas_mayo.xlsx",
        "detected_columns": [
            {
                "original_column_name": "Cantidad",
                "suggested_semantic_role": "cantidad",
                "confidence": "mapped",
            },
            {
                "original_column_name": "PrecioVenta",
                "suggested_semantic_role": "precio_venta",
                "confidence": "mapped",
            },
        ],
        "confirmed_columns": [
            {
                "original_column_name": "Cantidad",
                "confirmed_semantic_role": "cantidad",
            },
            {
                "original_column_name": "PrecioVenta",
                "confirmed_semantic_role": "precio_venta",
            },
        ],
        "rows_processed": 12,
        "warnings": ["Header ambiguo resuelto manualmente."],
        "limitations": ["No ejecuta normalización real todavía."],
        "technical_notes": ["Builds a logical contract only."],
    }


def test_returns_ok_when_all_detected_columns_are_confirmed() -> None:
    result = build_excel_treatment_lab_v1(lab_input=_lab_input())

    assert result["schema_version"] == SCHEMA_VERSION
    assert result["service_name"] == "SERVICE_1"
    assert result["capability_ref"] == CAPABILITY_REF
    assert result["status"] == "OK"
    assert result["runtime_authorized"] is False
    assert result["missing_inputs"] == []
    assert result["computed_results"]["rows_processed"] == 12
    assert result["computed_results"]["pending_confirmation_columns"] == []


def test_returns_missing_confirmation_when_detected_columns_are_not_fully_confirmed() -> None:
    lab_input = _lab_input()
    lab_input["confirmed_columns"] = [
        {
            "original_column_name": "Cantidad",
            "confirmed_semantic_role": "cantidad",
        }
    ]

    result = build_excel_treatment_lab_v1(lab_input=lab_input)

    assert result["status"] == "MISSING_CONFIRMATION"
    assert result["computed_results"]["pending_confirmation_columns"] == ["PrecioVenta"]
    assert "PrecioVenta" in result["owner_summary"]


def test_returns_missing_inputs_when_source_file_or_detected_columns_are_missing() -> None:
    result_without_source = build_excel_treatment_lab_v1(
        lab_input={
            "source_file": None,
            "detected_columns": _lab_input()["detected_columns"],
            "confirmed_columns": [],
            "rows_processed": 0,
            "warnings": [],
        }
    )
    result_without_columns = build_excel_treatment_lab_v1(
        lab_input={
            "source_file": "ventas_mayo.xlsx",
            "detected_columns": [],
            "confirmed_columns": [],
            "rows_processed": 0,
            "warnings": [],
        }
    )

    assert result_without_source["status"] == "MISSING_INPUTS"
    assert "source_file" in result_without_source["missing_inputs"]
    assert result_without_columns["status"] == "MISSING_INPUTS"
    assert "detected_columns" in result_without_columns["missing_inputs"]


def test_returns_invalid_input_when_rows_processed_is_negative() -> None:
    lab_input = _lab_input()
    lab_input["rows_processed"] = -1

    result = build_excel_treatment_lab_v1(lab_input=lab_input)

    assert result["status"] == "INVALID_INPUT"
    assert result["owner_summary"] == (
        "El input estructurado del Laboratorio Excel es inválido y debe corregirse antes de continuar."
    )


def test_result_is_compatible_with_service_1_xlsx_delivery_input_v1(tmp_path: Path) -> None:
    result = build_excel_treatment_lab_v1(lab_input=_lab_input())
    output_path = tmp_path / "excel_treatment_lab.xlsx"

    delivery = build_service_1_xlsx_delivery_v1(
        delivery_input=result,
        output_path=output_path,
    )

    workbook = load_workbook(output_path)

    assert output_path.exists()
    assert delivery["capability_ref"] == CAPABILITY_REF
    assert workbook["Resumen"]["B3"].value == CAPABILITY_REF
    assert ("rows_processed", "12") in list(workbook["Resultados"].iter_rows(values_only=True))


def test_module_does_not_depend_on_first_aid_or_runtime_or_do_io() -> None:
    import pymia.smartpyme.excel_treatment_lab_v1 as module

    source = inspect.getsource(module)

    assert "openpyxl" not in source
    assert "document_ingestion" not in source
    assert "first_aid" not in source
    assert "vertical_pipeline" not in source
    assert "service_1_pipeline" not in source
    assert "fsm" not in source.lower()
    assert "llm" not in source.lower()
    assert "chatbot" not in source.lower()
    assert "open(" not in source
    assert ".save(" not in source
    assert "read_excel" not in source

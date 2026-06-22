from __future__ import annotations

import hashlib
import json
from datetime import datetime
from pathlib import Path
from typing import Final, TypedDict

from openpyxl import Workbook

from pymia.smartpyme.first_aid_tool_result_v1 import FirstAidToolResultV1

DELIVERY_SCHEMA_VERSION: Final[str] = "1.0"
SERVICE_NAME: Final[str] = "SERVICE_1"
SHEET_NAMES: Final[tuple[str, ...]] = (
    "Resumen",
    "Datos usados",
    "Resultados",
    "Faltantes",
    "Limitaciones",
    "Claims prohibidos",
    "Notas técnicas",
)
_FIXED_WORKBOOK_TIMESTAMP: Final[datetime] = datetime(2000, 1, 1, 0, 0, 0)


class FirstAidXlsxDeliveryV1(TypedDict):
    delivery_id: str
    schema_version: str
    service_name: str
    tool_ref: str
    output_path: str
    sheet_names: list[str]
    runtime_authorized: bool
    notes: list[str]


def build_first_aid_xlsx_delivery_v1(
    *,
    tool_result: FirstAidToolResultV1,
    output_path: str | Path,
) -> FirstAidXlsxDeliveryV1:
    if tool_result["runtime_authorized"]:
        raise ValueError("FIRST_AID_XLSX_DELIVERY_V1 does not accept runtime_authorized=True.")

    output_file = Path(output_path)
    parent_dir = output_file.parent
    if not parent_dir.exists():
        raise FileNotFoundError(f"Output directory does not exist: {parent_dir}")

    workbook = Workbook()
    _configure_workbook(workbook)
    _populate_workbook(workbook, tool_result)
    workbook.save(output_file)

    return {
        "delivery_id": _build_delivery_id(tool_result),
        "schema_version": DELIVERY_SCHEMA_VERSION,
        "service_name": SERVICE_NAME,
        "tool_ref": str(tool_result["tool_ref"]),
        "output_path": str(output_file.resolve()),
        "sheet_names": list(SHEET_NAMES),
        "runtime_authorized": False,
        "notes": [
            "Deterministic XLSX delivery generated from FirstAidToolResultV1 only.",
            "No formulas, macros, or runtime execution were used.",
        ],
    }


def _configure_workbook(workbook: Workbook) -> None:
    workbook.properties.creator = SERVICE_NAME
    workbook.properties.lastModifiedBy = SERVICE_NAME
    workbook.properties.created = _FIXED_WORKBOOK_TIMESTAMP
    workbook.properties.modified = _FIXED_WORKBOOK_TIMESTAMP


def _populate_workbook(workbook: Workbook, tool_result: FirstAidToolResultV1) -> None:
    summary_sheet = workbook.active
    summary_sheet.title = SHEET_NAMES[0]
    _write_key_value_sheet(
        summary_sheet,
        header=("field", "value"),
        rows=[
            ("service_name", tool_result["service_name"]),
            ("tool_ref", tool_result["tool_ref"]),
            ("status", tool_result["status"]),
            ("owner_summary", tool_result["owner_summary"]),
            ("runtime_authorized", tool_result["runtime_authorized"]),
        ],
    )

    _write_key_value_sheet(
        workbook.create_sheet(SHEET_NAMES[1]),
        header=("key", "value"),
        rows=list(tool_result["inputs_used"].items()),
    )
    _write_key_value_sheet(
        workbook.create_sheet(SHEET_NAMES[2]),
        header=("key", "value"),
        rows=list(tool_result["computed_results"].items()),
    )
    _write_single_column_sheet(
        workbook.create_sheet(SHEET_NAMES[3]),
        header="missing_input",
        rows=tool_result["missing_inputs"],
    )
    _write_single_column_sheet(
        workbook.create_sheet(SHEET_NAMES[4]),
        header="limitation",
        rows=tool_result["limitations"],
    )
    _write_single_column_sheet(
        workbook.create_sheet(SHEET_NAMES[5]),
        header="forbidden_claim",
        rows=tool_result["forbidden_claims"],
    )
    _write_single_column_sheet(
        workbook.create_sheet(SHEET_NAMES[6]),
        header="technical_note",
        rows=tool_result["technical_notes"],
    )


def _write_key_value_sheet(
    worksheet,
    *,
    header: tuple[str, str],
    rows: list[tuple[object, object]],
) -> None:
    worksheet.append([header[0], header[1]])
    for key, value in rows:
        worksheet.append([_to_safe_text(key), _to_safe_text(value)])


def _write_single_column_sheet(
    worksheet,
    *,
    header: str,
    rows: list[object],
) -> None:
    worksheet.append([header])
    for value in rows:
        worksheet.append([_to_safe_text(value)])


def _to_safe_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, str):
        if value.startswith(("=", "+", "-", "@")):
            return f"'{value}"
        return value

    text = json.dumps(value, ensure_ascii=False, sort_keys=True)
    if text.startswith(("=", "+", "-", "@")):
        return f"'{text}"
    return text


def _build_delivery_id(tool_result: FirstAidToolResultV1) -> str:
    canonical_payload = json.dumps(
        {
            "tool_ref": tool_result["tool_ref"],
            "status": tool_result["status"],
            "inputs_used": tool_result["inputs_used"],
            "computed_results": tool_result["computed_results"],
            "missing_inputs": tool_result["missing_inputs"],
            "limitations": tool_result["limitations"],
            "forbidden_claims": tool_result["forbidden_claims"],
            "owner_summary": tool_result["owner_summary"],
            "technical_notes": tool_result["technical_notes"],
        },
        ensure_ascii=False,
        sort_keys=True,
    )
    payload_hash = hashlib.sha256(canonical_payload.encode("utf-8")).hexdigest()[:16]
    return f"first_aid_xlsx_delivery_v1:{payload_hash}"

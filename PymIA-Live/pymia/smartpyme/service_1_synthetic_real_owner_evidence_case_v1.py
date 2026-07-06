from __future__ import annotations

import json
import os
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from pymia.cli.service_1_operator import main as run_service_1_operator_cli
from pymia.smartpyme.service_1_question_bundle_v1 import build_service_1_question_bundle_v1

SCHEMA_VERSION = "SERVICE_1_SYNTHETIC_REAL_OWNER_EVIDENCE_CASE_V1"
SERVICE_NAME = "SERVICE_1"
CASE_ID = "synthetic_real_owner_evidence_case_v1"
TENANT_ID = "tenant_synthetic_textil_perales"


def run_service_1_synthetic_real_owner_evidence_case_v1(output_root: str | Path) -> dict[str, Any]:
    """Run a synthetic-real Servicio 1 case through the current CLI flow.

    This harness creates local synthetic evidence, asks one owner question,
    injects one owner answer, runs deterministic First Aid tools, and returns
    the generated case folder state. It does not use real client data.
    """
    root = Path(output_root).resolve()
    root.mkdir(parents=True, exist_ok=True)

    inputs_dir = root / "inputs"
    inputs_dir.mkdir(parents=True, exist_ok=True)

    xlsx_path = inputs_dir / "textil_perales_marzo_2026.xlsx"
    bundle_path = inputs_dir / "question_bundle.json"
    tools_path = inputs_dir / "tool_requests.json"
    reentry_store = root / "owner_reentry_store"

    _write_synthetic_xlsx(xlsx_path)
    question_ref = _write_question_bundle(bundle_path)
    _write_tool_requests(tools_path)

    cwd = Path.cwd()
    try:
        os.chdir(root)
        exit_code = run_service_1_operator_cli(
            [
                "--file",
                str(xlsx_path),
                "--question-bundle",
                str(bundle_path),
                "--question-ref",
                question_ref,
                "--owner-answer",
                "Marzo 2026. El archivo principal representa ventas minoristas del local y los totales son importes de venta final.",
                "--owner-reentry-storage-dir",
                str(reentry_store),
                "--run-tools",
                str(tools_path),
            ]
        )
    finally:
        os.chdir(cwd)

    case_root = root / ".tmp" / "service_1_cases"
    case_dirs = sorted([path for path in case_root.iterdir() if path.is_dir()]) if case_root.exists() else []
    case_dir = case_dirs[-1] if case_dirs else None

    manifest = _read_json(case_dir / "manifest.json") if case_dir else None
    product_gate = _read_json(case_dir / "product_gate.json") if case_dir else None
    operator_packet = _read_json(case_dir / "operator_packet.json") if case_dir else None

    return {
        "schema_version": SCHEMA_VERSION,
        "service_name": SERVICE_NAME,
        "case_id": CASE_ID,
        "synthetic_data": True,
        "real_client_data": False,
        "exit_code": exit_code,
        "output_root": str(root),
        "input_files": {
            "xlsx": str(xlsx_path),
            "question_bundle": str(bundle_path),
            "tool_requests": str(tools_path),
        },
        "question_ref_answered": question_ref,
        "case_dir": str(case_dir) if case_dir else None,
        "manifest_status": manifest.get("delivery_status") if isinstance(manifest, dict) else None,
        "product_gate_status": product_gate.get("status") if isinstance(product_gate, dict) else None,
        "runtime_authorized": False,
        "delivery_authorized": False,
        "artifacts_present": _artifact_presence(case_dir),
        "operator_packet_status": {
            "has_question_bundle": isinstance(operator_packet, dict) and isinstance(operator_packet.get("question_bundle"), dict),
            "has_owner_reentry_bridge": isinstance(operator_packet, dict) and isinstance(operator_packet.get("owner_reentry_bridge"), dict),
            "has_next_owner_question": isinstance(operator_packet, dict) and isinstance(operator_packet.get("next_owner_question"), dict),
            "has_pipeline_result": isinstance(operator_packet, dict) and isinstance(operator_packet.get("pipeline_result"), dict),
            "has_product_gate": isinstance(operator_packet, dict) and isinstance(operator_packet.get("product_gate"), dict),
        },
        "notes": [
            "Synthetic-real rehearsal only; no real client data is present.",
            "The owner answer is synthetic and used to exercise the evidence reentry path.",
            "The CLI remains the executable product path for this rehearsal.",
        ],
    }


def _write_synthetic_xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "ventas_marzo"
    ws.append(["Fecha", "Producto", "Cantidad", "Precio", "Costo", "Total"])
    ws.append(["2026-03-01", "Remera basica", 12, 8500, 5200, 102000])
    ws.append(["2026-03-02", "Jean clasico", 5, 26000, 18400, 130000])
    ws.append(["2026-03-03", "Campera liviana", 2, 58000, 49000, 116000])
    ws.append(["2026-03-04", "Remera basica", 9, 8500, 5200, 76500])

    stock = wb.create_sheet("stock")
    stock.append(["Producto", "Stock sistema", "Stock fisico", "Diferencia"])
    stock.append(["Remera basica", 48, 42, -6])
    stock.append(["Jean clasico", 17, 17, 0])
    stock.append(["Campera liviana", 8, 6, -2])

    caja = wb.create_sheet("caja")
    caja.append(["Concepto", "Importe", "Categoria"])
    caja.append(["saldo inicial", 180000, "saldo"])
    caja.append(["ingresos ventas", 424500, "ingreso"])
    caja.append(["egresos proveedores", 286750, "egreso"])
    wb.save(path)


def _write_question_bundle(path: Path) -> str:
    bundle = build_service_1_question_bundle_v1(
        case_id=CASE_ID,
        tenant_id=TENANT_ID,
        intake_id="intake_synthetic_textil_perales_marzo_2026",
        run_id="run_synthetic_textil_perales_marzo_2026",
        report={
            "next_questions": [
                {
                    "text": "Confirmá qué período querés revisar y qué representa el archivo principal.",
                    "target_ref": "period_and_primary_file_role",
                }
            ]
        },
        metadata={"case_type": "synthetic_real_owner_evidence_case"},
    )
    path.write_text(json.dumps(bundle.to_dict(), indent=2, ensure_ascii=False), encoding="utf-8")
    if not bundle.selected_next_question_ref:
        raise RuntimeError("Synthetic question bundle did not produce a selected question ref")
    return bundle.selected_next_question_ref


def _write_tool_requests(path: Path) -> None:
    requests = [
        {
            "tool_ref": "precio_margen_basico",
            "inputs": {"precio_venta": 8500, "costo_unitario": 5200},
        },
        {
            "tool_ref": "caja_diaria_triage",
            "inputs": {"saldo_inicial": 180000, "ingresos": 424500, "egresos": 286750},
        },
        {
            "tool_ref": "stock_alertas_basicas",
            "inputs": {
                "producto": "Remera basica",
                "stock_actual": 42,
                "stock_minimo": 50,
                "ventas_diarias_promedio": 7,
            },
        },
        {
            "tool_ref": "gastos_triage",
            "inputs": {
                "concepto": ["alquiler", "luz", "insumos"],
                "importe": [100000, 32000, 78000],
                "categoria": ["fijo", "fijo", "variable"],
            },
        },
        {
            "tool_ref": "proveedores_precio_variacion_triage",
            "inputs": {
                "proveedor": ["Proveedor Norte", "Proveedor Sur"],
                "producto_o_insumo": ["Algodon", "Algodon"],
                "precio_o_costo": [3900, 4550],
            },
        },
    ]
    path.write_text(json.dumps(requests, indent=2, ensure_ascii=False), encoding="utf-8")


def _read_json(path: Path | None) -> dict[str, Any] | None:
    if path is None or not path.exists():
        return None
    return json.loads(path.read_text(encoding="utf-8"))


def _artifact_presence(case_dir: Path | None) -> dict[str, bool]:
    expected = [
        "question_bundle.json",
        "owner_reentry_bridge.json",
        "pipeline_result.json",
        "next_owner_question.md",
        "evidence_loop_status.json",
        "case_record.json",
        "owner_delivery_packet.json",
        "product_gate.json",
        "manifest.json",
        "operator_packet.json",
    ]
    if case_dir is None:
        return {name: False for name in expected}
    return {name: (case_dir / name).exists() for name in expected}


__all__ = [
    "SCHEMA_VERSION",
    "SERVICE_NAME",
    "CASE_ID",
    "TENANT_ID",
    "run_service_1_synthetic_real_owner_evidence_case_v1",
]

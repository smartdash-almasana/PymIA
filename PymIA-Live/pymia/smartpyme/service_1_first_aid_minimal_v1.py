"""
Service 1 First Aid Minimal V1

Minimal First Aid execution for Servicio 1 local asistido.
Only descriptive weak findings on confirmed numeric columns.
No diagnosis, no accounting, no strong recommendations.
runtime_authorized is always False.
human_review_required is always True.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import openpyxl


SCHEMA_VERSION = "1.0"
SERVICE_NAME = "SERVICE_1"

_NUMERIC_HINTS = ("importe", "total", "cantidad", "precio", "monto", "valor", "saldo", "neto", "bruto")


# ---------------------------------------------------------------------------
# 1. load_confirmed_columns_v1
# ---------------------------------------------------------------------------

def load_confirmed_columns_v1(path: str | Path) -> dict[str, Any]:
    """Load and validate a confirmed_columns JSON file.

    Returns a dict with schema_version, service_name, confirmed_columns,
    runtime_authorized=False, and warnings.
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"confirmed_columns file not found: {p}")

    with open(p, encoding="utf-8") as f:
        data = json.load(f)

    if not isinstance(data, dict):
        raise ValueError("confirmed_columns JSON root must be an object")

    confirmed_columns = data.get("confirmed_columns", {})
    if not isinstance(confirmed_columns, dict):
        raise ValueError("confirmed_columns must be a dict")

    warnings: list[str] = []
    if not confirmed_columns:
        warnings.append("confirmed_columns is empty; First Aid will have limited scope.")

    return {
        "schema_version": SCHEMA_VERSION,
        "service_name": SERVICE_NAME,
        "confirmed_columns": confirmed_columns,
        "runtime_authorized": False,
        "warnings": warnings,
    }


# ---------------------------------------------------------------------------
# 2. evaluate_first_aid_minimal_eligibility_v1
# ---------------------------------------------------------------------------

def _scan_runtime_authorized_true(obj: Any) -> bool:
    """Recursively scan for any runtime_authorized=True in nested dicts/lists."""
    if isinstance(obj, dict):
        for key, value in obj.items():
            if key == "runtime_authorized" and value is True:
                return True
            if _scan_runtime_authorized_true(value):
                return True
    elif isinstance(obj, list):
        for item in obj:
            if _scan_runtime_authorized_true(item):
                return True
    return False


def evaluate_first_aid_minimal_eligibility_v1(packet: dict[str, Any]) -> dict[str, Any]:
    """Evaluate whether the packet is eligible for minimal First Aid.

    Returns a gate dict with status ELIGIBLE or BLOCKED.
    """
    blockers: list[str] = []
    warnings: list[str] = []

    # Check 1: detected_structure present
    detected_structure = packet.get("detected_structure")
    if detected_structure is None:
        blockers.append("missing detected_structure")

    # Check 2: column_confirmation_packet present
    column_confirmation = packet.get("column_confirmation_packet")
    if column_confirmation is None:
        blockers.append("missing column_confirmation_packet")

    # Check 3: confirmed_columns present
    confirmed_columns = packet.get("confirmed_columns")
    if confirmed_columns is None:
        blockers.append("missing confirmed_columns")
    elif isinstance(confirmed_columns, dict) and not confirmed_columns.get("confirmed_columns"):
        blockers.append("confirmed_columns.confirmed_columns is empty")

    # Check 4: qa_delivery_gate exists and status == PASS
    qa_gate = packet.get("qa_delivery_gate")
    if qa_gate is None:
        blockers.append("missing qa_delivery_gate")
    elif qa_gate.get("status") != "PASS":
        blockers.append(f"qa_delivery_gate status is {qa_gate.get('status')}, not PASS")

    # Check 5: no runtime_authorized=True anywhere
    if _scan_runtime_authorized_true(packet):
        blockers.append("runtime_authorized=True found in packet")

    # Check 6: headers present in detected_structure
    if detected_structure is not None and isinstance(detected_structure, dict):
        workbook = detected_structure.get("workbook", {})
        sheets = workbook.get("sheets", [])
        has_headers = False
        for sheet in sheets:
            if sheet.get("headers"):
                has_headers = True
                break
        if not has_headers:
            blockers.append("no headers found in detected_structure")

    status = "ELIGIBLE" if not blockers else "BLOCKED"

    return {
        "schema_version": SCHEMA_VERSION,
        "service_name": SERVICE_NAME,
        "gate_type": "FIRST_AID_MINIMAL_ELIGIBILITY",
        "status": status,
        "runtime_authorized": False,
        "human_review_required": True,
        "allowed_actions": ["table_profile", "missing_values", "simple_numeric_totals"],
        "warnings": warnings,
        "blockers": blockers,
    }


# ---------------------------------------------------------------------------
# 3. run_first_aid_minimal_v1
# ---------------------------------------------------------------------------

def _is_numeric_column(col_name: str, confirmed_columns: dict[str, Any]) -> bool:
    """Determine if a column is numeric based on confirmed_columns or name hints."""
    # Check confirmed_columns mapping
    col_lower = col_name.lower()
    for key, meta in confirmed_columns.items():
        if key.lower() == col_lower:
            if isinstance(meta, dict):
                role = (meta.get("role") or meta.get("semantic_role") or "").lower()
                if role in ("numeric", "amount", "quantity", "price", "importe", "total", "cantidad", "precio"):
                    return True
            elif isinstance(meta, str) and meta.lower() in ("numeric", "amount", "quantity", "price"):
                return True

    # Fall back to name hints
    for hint in _NUMERIC_HINTS:
        if hint in col_lower:
            return True

    return False


def run_first_aid_minimal_v1(
    packet: dict[str, Any],
    xlsx_path: str | Path,
) -> dict[str, Any]:
    """Execute minimal First Aid on the XLSX file.

    Only counts rows/columns, detects empty values, and computes simple
    totals on confirmed numeric columns. No diagnosis, no recommendations.

    Does NOT mutate the packet. Does NOT modify the XLSX.
    """
    path = Path(xlsx_path)
    warnings: list[str] = []

    confirmed_columns_block = packet.get("confirmed_columns", {})
    if isinstance(confirmed_columns_block, dict):
        confirmed_cols_map = confirmed_columns_block.get("confirmed_columns", {})
    else:
        confirmed_cols_map = {}

    # Load workbook
    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    except Exception as exc:
        warnings.append(f"Could not open XLSX for First Aid: {exc}")
        return {
            "schema_version": SCHEMA_VERSION,
            "service_name": SERVICE_NAME,
            "result_type": "FIRST_AID_MINIMAL",
            "status": "DRAFT_REVIEW_REQUIRED",
            "runtime_authorized": False,
            "human_review_required": True,
            "summary": {
                "sheet_count": 0,
                "sheets_profiled": 0,
                "total_findings": 0,
            },
            "findings": [],
            "warnings": warnings,
        }

    findings: list[dict[str, Any]] = []
    sheets_profiled = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.rows)

        if not rows:
            continue

        sheets_profiled += 1

        # Extract headers from first row
        headers: list[str] = []
        for cell in rows[0]:
            val = cell.value
            headers.append(str(val) if val is not None else "")

        data_rows = rows[1:]
        row_count = len(data_rows)
        col_count = len(headers)

        # Detect empty values per column
        empty_per_col: dict[str, int] = {}
        numeric_totals: dict[str, float] = {}
        numeric_counts: dict[str, int] = {}

        for col_idx, col_name in enumerate(headers):
            if not col_name:
                continue

            empty_count = 0
            is_numeric = _is_numeric_column(col_name, confirmed_cols_map)
            running_total = 0.0
            numeric_count = 0

            for row in data_rows:
                if col_idx < len(row):
                    cell_val = row[col_idx].value
                else:
                    cell_val = None

                if cell_val is None or (isinstance(cell_val, str) and not cell_val.strip()):
                    empty_count += 1
                elif is_numeric:
                    try:
                        num = float(cell_val)
                        running_total += num
                        numeric_count += 1
                    except (ValueError, TypeError):
                        pass

            empty_per_col[col_name] = empty_count

            if is_numeric and numeric_count > 0:
                numeric_totals[col_name] = running_total
                numeric_counts[col_name] = numeric_count

        # Generate findings for this sheet
        sheet_finding: dict[str, Any] = {
            "finding_type": "sheet_profile",
            "sheet_name": sheet_name,
            "row_count": row_count,
            "col_count": col_count,
            "headers": headers,
            "empty_values_per_column": empty_per_col,
        }

        if numeric_totals:
            simple_totals: dict[str, dict[str, Any]] = {}
            for col_name, total in numeric_totals.items():
                simple_totals[col_name] = {
                    "sum": round(total, 2),
                    "numeric_values_count": numeric_counts[col_name],
                }
            sheet_finding["simple_numeric_totals"] = simple_totals

        findings.append(sheet_finding)

    wb.close()

    return {
        "schema_version": SCHEMA_VERSION,
        "service_name": SERVICE_NAME,
        "result_type": "FIRST_AID_MINIMAL",
        "status": "DRAFT_REVIEW_REQUIRED",
        "runtime_authorized": False,
        "human_review_required": True,
        "summary": {
            "sheet_count": len(wb.sheetnames),
            "sheets_profiled": sheets_profiled,
            "total_findings": len(findings),
        },
        "findings": findings,
        "warnings": warnings,
    }


# ---------------------------------------------------------------------------
# Helper: render first_aid_owner_summary.md
# ---------------------------------------------------------------------------

def render_first_aid_owner_summary_v1(
    first_aid_result: dict[str, Any],
) -> str:
    """Render a human-readable First Aid summary in Markdown."""
    lines: list[str] = []
    lines.append("# First Aid mínimo — borrador para revisión")
    lines.append("")
    lines.append("**Estado:** DRAFT_REVIEW_REQUIRED")
    lines.append("**Revisión humana requerida:** Sí")
    lines.append("**Runtime autorizado:** No")
    lines.append("")

    summary = first_aid_result.get("summary", {})
    lines.append(f"- Hojas en el archivo: {summary.get('sheet_count', 0)}")
    lines.append(f"- Hojas perfiladas: {summary.get('sheets_profiled', 0)}")
    lines.append("")

    findings = first_aid_result.get("findings", [])
    for f in findings:
        if f.get("finding_type") == "sheet_profile":
            lines.append(f"## Hoja: {f.get('sheet_name', 'N/A')}")
            lines.append("")
            lines.append(f"- Filas de datos: {f.get('row_count', 0)}")
            lines.append(f"- Columnas: {f.get('col_count', 0)}")
            lines.append("")

            # Empty values
            empty_map = f.get("empty_values_per_column", {})
            cols_with_empties = {k: v for k, v in empty_map.items() if v > 0}
            if cols_with_empties:
                lines.append("### Columnas con valores vacíos")
                lines.append("")
                for col, count in cols_with_empties.items():
                    lines.append(f"- {col}: {count} vacíos")
                lines.append("")

            # Simple totals
            totals = f.get("simple_numeric_totals", {})
            if totals:
                lines.append("### Totales simples (solo columnas numéricas confirmadas)")
                lines.append("")
                for col, info in totals.items():
                    lines.append(
                        f"- {col}: suma = {info['sum']} "
                        f"({info['numeric_values_count']} valores numéricos)"
                    )
                lines.append("")

    warnings = first_aid_result.get("warnings", [])
    if warnings:
        lines.append("## Advertencias")
        lines.append("")
        for w in warnings:
            lines.append(f"- {w}")
        lines.append("")

    lines.append("---")
    lines.append(
        "*Este borrador es descriptivo y no constituye diagnóstico, "
        "cálculo contable ni recomendación de negocio.*"
    )
    lines.append("")

    return "\n".join(lines)

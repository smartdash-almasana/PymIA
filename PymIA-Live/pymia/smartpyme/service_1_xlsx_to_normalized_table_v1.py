from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from pymia.smartpyme.service_1_normalized_table_v1 import (
    NormalizedTableV1,
    build_normalized_table_v1,
)


def read_xlsx_to_normalized_table_v1(
    xlsx_path: str | Path,
    *,
    sheet_name: str | None = None,
) -> NormalizedTableV1:
    path = Path(xlsx_path)
    source_path = str(path)

    if not path.exists() or not path.is_file():
        return _blocked(source_path, sheet_name, "File not found.")
    if path.suffix.lower() != ".xlsx":
        return _blocked(source_path, sheet_name, "Only .xlsx files are accepted.")

    try:
        workbook = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        return _blocked(source_path, sheet_name, str(exc))

    if sheet_name is not None and sheet_name not in workbook.sheetnames:
        return _blocked(source_path, sheet_name, f"Sheet not found: {sheet_name}")

    worksheet = workbook[sheet_name] if sheet_name is not None else _select_default_sheet(workbook)
    selected_sheet_name = worksheet.title
    materialized_rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
    header_index = _first_non_empty_row_index(materialized_rows)

    if header_index is None:
        return _blocked(source_path, selected_sheet_name, "XLSX workbook has no non-empty rows.")

    raw_headers = _trim_trailing_empty(materialized_rows[header_index])
    headers = [_clean(value) for value in raw_headers]
    if not headers or any(not header for header in headers):
        return build_normalized_table_v1(
            source_kind="xlsx",
            source_path=source_path,
            sheet_name=selected_sheet_name,
            headers=headers,
            rows=[],
            blocking_errors=["XLSX headers are missing or incomplete."],
        )

    rows: list[dict[str, Any]] = []
    warnings: list[str] = []
    width = len(headers)

    for row_number, raw_row in enumerate(materialized_rows[header_index + 1 :], start=header_index + 2):
        if _is_empty_row(raw_row):
            continue
        if _last_non_empty_index(raw_row[:width]) < width - 1:
            warnings.append(f"Row {row_number} has fewer cells than headers.")
        if any(_clean(value) for value in raw_row[width:]):
            warnings.append(f"Row {row_number} has more cells than headers; extra cells ignored.")
        fitted = _fit_width(raw_row, width)
        rows.append({headers[index]: _clean(fitted[index]) for index in range(width)})

    return build_normalized_table_v1(
        source_kind="xlsx",
        source_path=source_path,
        sheet_name=selected_sheet_name,
        headers=headers,
        rows=rows,
        warnings=warnings,
    )


def _blocked(source_path: str, sheet_name: str | None, message: str) -> NormalizedTableV1:
    return build_normalized_table_v1(
        source_kind="xlsx",
        source_path=source_path,
        sheet_name=sheet_name,
        headers=[],
        rows=[],
        blocking_errors=[message],
    )


def _select_default_sheet(workbook: Any) -> Any:
    for worksheet in workbook.worksheets:
        rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
        if _first_non_empty_row_index(rows) is not None:
            return worksheet
    return workbook.worksheets[0]


def _first_non_empty_row_index(rows: list[list[Any]]) -> int | None:
    for index, row in enumerate(rows):
        if not _is_empty_row(row):
            return index
    return None


def _is_empty_row(row: list[Any]) -> bool:
    return all(not _clean(value) for value in row)


def _trim_trailing_empty(row: list[Any]) -> list[Any]:
    last_index = _last_non_empty_index(row)
    if last_index < 0:
        return []
    return row[: last_index + 1]


def _last_non_empty_index(row: list[Any]) -> int:
    for index in range(len(row) - 1, -1, -1):
        if _clean(row[index]):
            return index
    return -1


def _fit_width(row: list[Any], width: int) -> list[Any]:
    if len(row) < width:
        return row + [""] * (width - len(row))
    return row[:width]


def _clean(value: Any) -> str:
    return "" if value is None else str(value).strip()

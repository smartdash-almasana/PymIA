from __future__ import annotations

from pathlib import Path
from openpyxl import load_workbook

from pymia.contracts.presentation_labels_v1 import (
    load_operational_terms,
)
from pymia.contracts.vertical_slice_copy_v1 import vertical_slice_copy_for
from pymia.smartpyme.owner_facing_report import build_owner_facing_report
from pymia.smartpyme.owner_output import build_owner_simple_view
from pymia.smartpyme.pipeline_registration import (
    register_anamnesis_record,
    register_evidence_record,
    register_evidence_request_record,
    register_investigation_record,
    register_owner_answer_record,
    register_pipeline_run_record,
)
from pymia.smartpyme.question_alignment_gate import align_next_question
from pymia.smartpyme.question_resolution import (
    _requested_evidence_from_report,
    _resolve_owner_question_and_reference,
)
from pymia.smartpyme.diagnostic_operator_adapter import (
    _diagnostic_pipeline_result_for_report,
    _diagnostic_operator_summary_from_report,
)
from pymia.rendering.owner_markdown_renderer import render_markdown_from_report


def inspect_excel(path: Path) -> dict:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        headers = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        normalized_headers = [str(value).strip().lower() for value in headers if value]
        return {
            "sheet": worksheet.title,
            "rows": int(worksheet.max_row or 0),
            "columns": int(worksheet.max_column or 0),
            "headers": normalized_headers,
        }
    finally:
        workbook.close()


def has_operational_columns(headers: list[str]) -> bool:
    joined = " ".join(headers)
    return any(term in joined for term in load_operational_terms())


def build_structured_summary(
    path: Path,
    tenant_id: str,
    *,
    intake_id: str = "intake_cli_local",
    formula_ids: list[str] | None = None,
) -> dict:
    try:
        from pymia.smartpyme.structured_evidence_builder import build_structured_evidence_context
        from pymia.contracts.evidence_v1 import StructuredEvidence
        from pymia.audit_result.evidence_requirement_matcher import match_evidence_requirements

        original_formula_ids = formula_ids
        formula_ids = formula_ids or []
        payload = build_structured_evidence_context(
            excel_path=path,
            tenant_id=tenant_id,
            intake_record={"evidence_requests": [{"formula_ids": formula_ids}]},
        )
        evidence = payload["structured_evidence"]
        computed = evidence.get("computed_variables") or {}
        tables = evidence.get("tables") or []
        summary = {
            "status": "available",
            "computed_variables_count": len(computed),
            "computed_variable_names": sorted(computed.keys()),
            "tables_count": len(tables),
            "table_sheets": [
                {"name": t.get("sheet_name", "?"), "columns": len(t.get("columns", [])), "rows": len(t.get("rows", []))}
                for t in tables
            ],
            "case_id": intake_id,
            "sufficiency": [],
            "unsupported_formula_ids": [],
            "catalog_reconciliation": [],
        }

        structured_ev = StructuredEvidence.model_validate(evidence)
        matches = match_evidence_requirements(structured_ev)
        reconciliation_list = []
        for m in matches:
            reconciliation_list.append({
                "formula_id": m.formula_id,
                "pathology_code": m.pathology_code,
                "status": m.status,
                "available_evidence": m.available_evidence,
                "missing_evidence": m.missing_evidence,
                "matched_sources": m.matched_sources,
                "required_evidence": m.required_evidence,
                "required_variables": m.required_variables,
                "next_audit_questions": m.next_audit_questions,
            })

        if original_formula_ids:
            reconciliation_list = [
                d for d in reconciliation_list if d["formula_id"] in original_formula_ids
            ]
        summary["catalog_reconciliation"] = reconciliation_list

        if formula_ids:
            from pymia.contracts.formula_rules_v1 import load_formula_rules
            from pymia.diagnostic_core.evidence_sufficiency import build_evidence_sufficiency_report_from_structured_evidence

            rules = load_formula_rules()
            rules_by_formula = rules.get("rules_by_formula", {})
            supported_formula_ids = [formula_id for formula_id in formula_ids if formula_id in rules_by_formula]
            unsupported_formula_ids = [formula_id for formula_id in formula_ids if formula_id not in rules_by_formula]
            summary["unsupported_formula_ids"] = unsupported_formula_ids
            if supported_formula_ids:
                sufficiency = build_evidence_sufficiency_report_from_structured_evidence(
                    structured_ev,
                    case_id=intake_id,
                    tenant_id=tenant_id,
                    formula_ids=supported_formula_ids,
                )
                summary["sufficiency"] = [item.model_dump(mode="json") for item in sufficiency]
        return summary
    except Exception as exc:
        raise exc


def build_report(
    path: Path,
    message: str,
    profile: dict,
    *,
    tenant_id: str = "tenant_cli_local",
    intake_id: str = "intake_cli_local",
    formula_ids: list[str] | None = None,
    storage_dir: Path | None = None,
    business_taxonomy: dict | None = None,
    owner_answer: str | None = None,
    owner_answer_question_ref: str | None = None,
) -> dict:
    has_rows = profile["rows"] > 1 and profile["columns"] > 0
    has_columns = has_operational_columns(profile["headers"])
    actual_storage_dir = storage_dir or Path(".tmp/vertical_slice_storage")
    anamnesis_record = register_anamnesis_record(
        message,
        tenant_id,
        intake_id,
        actual_storage_dir,
        business_taxonomy=business_taxonomy,
    )
    investigation_record = register_investigation_record(
        message,
        tenant_id,
        intake_id,
        anamnesis_record["anamnesis_id"],
        actual_storage_dir,
    )
    owner_answer_record = None
    if owner_answer:
        owner_answer_record = register_owner_answer_record(
            tenant_id=tenant_id,
            intake_id=intake_id,
            anamnesis_id=anamnesis_record["anamnesis_id"],
            investigation_id=investigation_record["investigation_id"],
            question_ref=owner_answer_question_ref or "owner_answer:unspecified_question_ref",
            raw_owner_answer=owner_answer,
            storage_dir=actual_storage_dir,
        )
    structured_summary = build_structured_summary(
        path,
        tenant_id,
        intake_id=intake_id,
        formula_ids=formula_ids,
    )
    missing = []
    questions = []
    if not has_rows:
        missing.append("filas_de_datos")
        questions.append(vertical_slice_copy_for("missing_data_rows_question"))
    if not has_columns:
        missing.append("columnas_operativas")
        questions.append(vertical_slice_copy_for("missing_operational_columns_question"))
    blocked = bool(missing)
    summary = (
        vertical_slice_copy_for("blocked_summary")
        if blocked
        else vertical_slice_copy_for("candidate_summary")
    )
    report = build_owner_facing_report(
        operational_audit_result={"tenant_id": tenant_id, "status": "blocked" if blocked else "candidate", "evidence_used": ["excel_file_readable"], "missing_evidence": missing},
        render_contract={
            "tenant_id": tenant_id,
            "summary": summary,
            "blocked_message": summary if blocked else "",
            "next_questions": questions,
            "next_steps": [vertical_slice_copy_for("next_step_review_with_owner")],
            "references": [str(path)],
            "forbidden_inferences": [vertical_slice_copy_for("forbidden_inference_from_column_names")],
        },
        delivery_package={"tenant_id": tenant_id, "intake_id": intake_id, "status": "BLOCKED" if blocked else "DELIVERED", "summary": summary, "output_refs": ["stdout"], "warnings": ["Slice local; no es canal productivo."]},
    ).to_dict()
    report["structured_evidence_summary"] = structured_summary
    requested_evidence = _requested_evidence_from_report(report)
    evidence_request_record = None
    if requested_evidence:
        evidence_request_record = register_evidence_request_record(
            tenant_id=tenant_id,
            intake_id=intake_id,
            anamnesis_id=anamnesis_record["anamnesis_id"],
            investigation_id=investigation_record["investigation_id"],
            owner_answer_id=owner_answer_record["answer_id"] if owner_answer_record else None,
            requested_evidence=requested_evidence,
            request_reason=vertical_slice_copy_for("evidence_request_reason"),
            storage_dir=actual_storage_dir,
        )
    evidence_record = register_evidence_record(
        path,
        tenant_id,
        intake_id,
        actual_storage_dir,
        request_id=evidence_request_record["request_id"] if evidence_request_record else None,
    )
    pipeline_run_record = register_pipeline_run_record(
        tenant_id=tenant_id,
        intake_id=intake_id,
        message=message,
        anamnesis_record=anamnesis_record,
        investigation_record=investigation_record,
        owner_answer_record=owner_answer_record,
        evidence_request_record=evidence_request_record,
        evidence_record=evidence_record,
        structured_summary=structured_summary,
        blocked=blocked,
        storage_dir=actual_storage_dir,
    )
    report["anamnesis_record"] = anamnesis_record
    report["investigation_record"] = investigation_record
    report["owner_answer_record"] = owner_answer_record
    report["evidence_request_record"] = evidence_request_record
    report["evidence_record"] = evidence_record
    report["pipeline_run_record"] = pipeline_run_record
    report["diagnostic_pipeline_result"] = _diagnostic_pipeline_result_for_report(
        path=path,
        tenant_id=tenant_id,
        intake_id=intake_id,
        cliente_id=tenant_id,
        structured_summary=structured_summary,
    )
    owner_question, tech_reference = _resolve_owner_question_and_reference(report, message)
    report["owner_question"] = owner_question
    report["owner_question_technical_reference"] = tech_reference
    report["owner_simple"] = build_owner_simple_view(
        report=report,
        message=message,
        profile=profile,
        owner_question=owner_question,
    )
    request_reconciliation = report.get("structured_evidence_summary", {}).get("catalog_reconciliation") or []
    request_question_candidates = [e for e in request_reconciliation if e.get("next_audit_questions")]
    report["evidence_request_alignment"] = (
        align_next_question(message, request_question_candidates)
        if request_question_candidates
        else {"status": "ALIGNED"}
    )
    return report


def build_markdown(
    path: Path,
    message: str,
    profile: dict,
    *,
    tenant_id: str = "tenant_cli_local",
    intake_id: str = "intake_cli_local",
    formula_ids: list[str] | None = None,
    storage_dir: Path | None = None,
    business_taxonomy: dict | None = None,
    owner_answer: str | None = None,
    owner_answer_question_ref: str | None = None,
    audience: str = "owner",
) -> str:
    report = build_report(
        path,
        message,
        profile,
        tenant_id=tenant_id,
        intake_id=intake_id,
        formula_ids=formula_ids,
        storage_dir=storage_dir,
        business_taxonomy=business_taxonomy,
        owner_answer=owner_answer,
        owner_answer_question_ref=owner_answer_question_ref,
    )
    diagnostic_operator_summary = _diagnostic_operator_summary_from_report(report)
    return render_markdown_from_report(
        path,
        message,
        profile,
        report,
        audience=audience,
        diagnostic_operator_summary=diagnostic_operator_summary,
    )


def build_pipeline(
    path: Path,
    message: str,
    *,
    tenant_id: str = "tenant_cli_local",
    intake_id: str = "intake_cli_local",
    formula_ids: list[str] | None = None,
    storage_dir: Path | None = None,
    business_taxonomy: dict | None = None,
    owner_answer: str | None = None,
    owner_answer_question_ref: str | None = None,
    audience: str = "owner",
) -> dict:
    profile = inspect_excel(path)
    report = build_report(
        path,
        message,
        profile,
        tenant_id=tenant_id,
        intake_id=intake_id,
        formula_ids=formula_ids,
        storage_dir=storage_dir,
        business_taxonomy=business_taxonomy,
        owner_answer=owner_answer,
        owner_answer_question_ref=owner_answer_question_ref,
    )
    diagnostic_operator_summary = _diagnostic_operator_summary_from_report(report)
    markdown = render_markdown_from_report(
        path,
        message,
        profile,
        report,
        audience=audience,
        diagnostic_operator_summary=diagnostic_operator_summary,
    )
    evidence_record = report["evidence_record"]
    pipeline_run_record = report["pipeline_run_record"]
    structured_summary = report.get("structured_evidence_summary") or {}
    catalog_reconciliation = structured_summary.get("catalog_reconciliation") or []
    return {
        "status": report["status"],
        "profile": profile,
        "report": report,
        "markdown": markdown,
        "diagnostic_operator_summary": diagnostic_operator_summary,
        "evidence_id": evidence_record["evidence_id"],
        "evidence_hash": evidence_record["content_hash"],
        "run_id": pipeline_run_record["run_id"],
        "output_hash": pipeline_run_record["output_hash"],
        "missing_evidence": report["missing_evidence"],
        "next_questions": report["next_questions"],
        "structured_summary": report["structured_evidence_summary"],
        "catalog_reconciliation": catalog_reconciliation,
    }

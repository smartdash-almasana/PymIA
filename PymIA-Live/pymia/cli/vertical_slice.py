from __future__ import annotations

import argparse
import hashlib
from pathlib import Path

from openpyxl import load_workbook

from pymia.contracts.language_corpus_v1 import load_language_corpus_seed, owner_label_for_variable_id
from pymia.contracts.presentation_labels_v1 import (
    label_for_field,
    label_for_pathology,
    load_operational_terms,
)
from pymia.contracts.vertical_slice_copy_v1 import vertical_slice_copy_for
from pymia.contracts.vertical_slice_copy_v1 import (
    owner_simple_readable_areas,
    owner_simple_understanding_by_axis,
)
from pymia.smartpyme.owner_facing_report import build_owner_facing_report
from pymia.smartpyme.question_alignment_gate import (
    AXIS_AUTOMATIZACION_MANUAL,
    AXIS_CAJA_LIQUIDEZ,
    AXIS_COSTOS_PROVEEDORES,
    AXIS_DESCONOCIDO,
    AXIS_PRODUCCION,
    AXIS_RRHH,
    AXIS_STOCK_REPOSICION,
    AXIS_VENTAS_MARGEN,
    align_next_question,
    detect_owner_axis,
)


def _humanize_field(name: str) -> str:
    return label_for_field(name)


def _owner_label_for_variable(name: str) -> str:
    corpus = load_language_corpus_seed()
    label = owner_label_for_variable_id(name, corpus)
    if label == name:
        return name
    return f"{label} ({name})"


def _build_owner_question(entry: dict) -> tuple[str | None, str | None]:
    qs = entry.get("next_audit_questions", [])
    if not qs:
        return None, None
    pathology_code = entry.get("pathology_code", "")
    formula_id = entry.get("formula_id", "")
    missing = entry.get("missing_evidence", [])
    pathology_label = label_for_pathology(pathology_code)
    humanized_fields = [_humanize_field(f) for f in missing]
    if humanized_fields:
        if len(humanized_fields) == 1:
            field_text = humanized_fields[0]
        elif len(humanized_fields) == 2:
            field_text = f"{humanized_fields[0]} y {humanized_fields[1]}"
        else:
            field_text = ", ".join(humanized_fields[:-1]) + f" y {humanized_fields[-1]}"
        owner_q = vertical_slice_copy_for("owner_question_missing_field").format(
            pathology_label=pathology_label,
            field_text=field_text,
        )
    else:
        owner_q = vertical_slice_copy_for("owner_question_missing_generic").format(
            pathology_label=pathology_label,
        )
    tech_parts = [f"Referencia técnica: {formula_id}"]
    if missing:
        tech_parts.append(f"inputs faltantes: {', '.join(missing)}")
    tech_ref = "; ".join(tech_parts)
    return owner_q, tech_ref


def _requested_evidence_from_report(report: dict) -> list[str]:
    requested: list[str] = []
    structured_summary = report.get("structured_evidence_summary") or {}
    if structured_summary.get("status") == "available":
        for entry in structured_summary.get("catalog_reconciliation") or []:
            for item in entry.get("missing_evidence") or []:
                if item not in requested:
                    requested.append(item)
    for item in report.get("missing_evidence") or []:
        if item not in requested:
            requested.append(item)
    return requested


def _resolve_owner_question_and_reference(report: dict, message: str) -> tuple[str | None, str | None]:
    structured_summary = report.get("structured_evidence_summary") or {}
    reconciliation = structured_summary.get("catalog_reconciliation") if structured_summary.get("status") == "available" else None
    owner_question = None
    tech_reference = None
    if reconciliation:
        question_candidates = [e for e in reconciliation if e.get("next_audit_questions")]
        alignment = align_next_question(message, question_candidates)
        if alignment["status"] == "MISALIGNED":
            owner_question = alignment["final_question_text"]
            tech_reference = alignment["technical_reference"]
        else:
            for entry in question_candidates:
                owner_q, tech_ref = _build_owner_question(entry)
                if owner_q:
                    owner_question = owner_q
                    tech_reference = tech_ref
                    break
    if not owner_question and report.get("next_questions"):
        owner_question = str(report["next_questions"][0])
    return owner_question, tech_reference


def _owner_understanding_text(message: str) -> str:
    owner_axis = detect_owner_axis(message)
    axis_messages = owner_simple_understanding_by_axis()
    return axis_messages[owner_axis]


def _owner_readable_summary(profile: dict) -> str:
    headers = [str(item).lower() for item in profile.get("headers", [])]
    joined = " ".join(headers)
    readable_areas: list[str] = []

    for label, payload in owner_simple_readable_areas().items():
        keywords = payload.get("keywords") or []
        if any(keyword in joined for keyword in keywords):
            readable_areas.append(label)

    if readable_areas:
        if len(readable_areas) == 1:
            area_text = readable_areas[0]
        elif len(readable_areas) == 2:
            area_text = f"{readable_areas[0]} y {readable_areas[1]}"
        else:
            area_text = ", ".join(readable_areas[:-1]) + f" y {readable_areas[-1]}"
        return vertical_slice_copy_for("owner_simple_readable_summary_template").format(area_text=area_text)

    if profile.get("rows", 0) > 1 and profile.get("columns", 0) > 0:
        return vertical_slice_copy_for("owner_simple_minimal_signals")

    return vertical_slice_copy_for("owner_simple_unreadable")


def _build_owner_simple_view(report: dict, message: str, profile: dict) -> dict[str, Any]:
    owner_question, _ = _resolve_owner_question_and_reference(report, message)
    return {
        "que_entendimos": _owner_understanding_text(message),
        "que_pudimos_leer": _owner_readable_summary(profile),
        "que_todavia_no_podemos_afirmar": vertical_slice_copy_for("owner_simple_unknown_assertion"),
        "proxima_pregunta": owner_question or vertical_slice_copy_for("next_question_fallback"),
        "limites": list(report.get("limit_warnings") or []) + [vertical_slice_copy_for("final_limit_warning")],
    }


def _serializable_diagnostic_pipeline_result(result) -> dict:
    return {
        "core_input": result.core_input.model_dump(mode="json"),
        "gate_decisions": [item.model_dump(mode="json") for item in result.gate_decisions],
        "formula_results": [item.model_dump(mode="json") for item in result.formula_results],
        "pathology_findings": [item.model_dump(mode="json") for item in result.pathology_findings],
        "finding_records": [item.model_dump(mode="json") for item in result.finding_records],
        "report": result.report.model_dump(mode="json") if result.report else None,
    }


def _diagnostic_pipeline_result_for_report(
    *,
    path: Path,
    tenant_id: str,
    intake_id: str,
    cliente_id: str,
    structured_summary: dict,
) -> dict | None:
    reconciliation = structured_summary.get("catalog_reconciliation") or []
    relevant_reconciliation = [
        entry for entry in reconciliation
        if isinstance(entry, dict) and str(entry.get("status") or "").lower() in ("calculable", "pending_data", "blocked")
    ]
    if not relevant_reconciliation:
        return None

    from pymia.contracts.evidence_v1 import StructuredEvidence
    from pymia.contracts.formula_rules_v1 import load_formula_rules
    from pymia.services.diagnostic_pipeline import (
        formula_pathology_map_from_catalog_reconciliation,
        run_diagnostic_pipeline_from_structured_evidence,
    )
    from pymia.smartpyme.structured_evidence_builder import build_structured_evidence_context

    rules = load_formula_rules()
    rules_by_formula = rules.get("rules_by_formula", {})
    formula_to_pathology = {
        formula_id: pathology_code
        for formula_id, pathology_code in formula_pathology_map_from_catalog_reconciliation(relevant_reconciliation).items()
        if formula_id in rules_by_formula
    }
    if not formula_to_pathology:
        return None

    payload = build_structured_evidence_context(
        excel_path=path,
        tenant_id=tenant_id,
        intake_record={"evidence_requests": [{"formula_ids": list(formula_to_pathology.keys())}]},
    )
    evidence = StructuredEvidence.model_validate(payload["structured_evidence"])
    result = run_diagnostic_pipeline_from_structured_evidence(
        evidence,
        case_id=intake_id,
        cliente_id=cliente_id,
        formula_to_pathology=formula_to_pathology,
    )
    return _serializable_diagnostic_pipeline_result(result)


def _diagnostic_operator_summary_from_report(report: dict) -> dict | None:
    diagnostic = report.get("diagnostic_pipeline_result")
    if not isinstance(diagnostic, dict):
        return None

    diagnostic_report = diagnostic.get("report")
    if not isinstance(diagnostic_report, dict):
        return None

    gate_decisions = diagnostic.get("gate_decisions") or []
    pathology_findings = diagnostic.get("pathology_findings") or []
    structured_summary = report.get("structured_evidence_summary") or {}
    reconciliation = structured_summary.get("catalog_reconciliation") or []

    has_gate_block = any(
        isinstance(d, dict) and d.get("decision") == "BLOCK_MISSING_INPUTS"
        for d in gate_decisions
    )
    gate_status = "blocked" if has_gate_block else "ready"

    blocked_formulas = [
        str(d.get("formula_id"))
        for d in gate_decisions
        if isinstance(d, dict) and d.get("decision") == "BLOCK_MISSING_INPUTS"
    ]

    missing_variables = {}
    for d in gate_decisions:
        if isinstance(d, dict) and d.get("decision") == "BLOCK_MISSING_INPUTS":
            f_id = str(d.get("formula_id"))
            missing_vars = d.get("missing_variables") or []
            missing_variables[f_id] = list(missing_vars)

    pending_pathologies = []
    for f in pathology_findings:
        if isinstance(f, dict) and f.get("status") == "PENDING_DATA":
            p_id = f.get("pathology_id")
            if p_id and p_id not in pending_pathologies:
                pending_pathologies.append(p_id)

    unsupported_pathologies = []
    for f in pathology_findings:
        if isinstance(f, dict) and f.get("status") == "PENDING_DATA":
            meta = f.get("metadata") or {}
            if meta.get("blocking_reason") == "PATHOLOGY_NOT_SUPPORTED":
                p_id = f.get("pathology_id")
                if p_id and p_id not in unsupported_pathologies:
                    unsupported_pathologies.append(p_id)

    owner_safe_question_candidates = []
    for entry in reconciliation:
        if isinstance(entry, dict):
            owner_q, _ = _build_owner_question(entry)
            if owner_q and owner_q not in owner_safe_question_candidates:
                owner_safe_question_candidates.append(owner_q)

    suggested_operator_next_step = "Solicitar evidencia faltante antes de reintentar diagnóstico."

    return {
        "status": "available",
        "diagnosis_status": diagnostic_report.get("diagnosis_status"),
        "kernel_state": diagnostic_report.get("kernel_state"),
        "blocking_reason": diagnostic_report.get("blocking_reason"),
        "finding_types": [
            finding.get("finding_type")
            for finding in diagnostic_report.get("findings", [])
            if isinstance(finding, dict) and finding.get("finding_type")
        ],
        "formulas_used": list(diagnostic_report.get("formulas_used") or []),
        "evidence_used": list(diagnostic_report.get("evidence_used") or []),
        "gate_status": gate_status,
        "blocked_formulas": blocked_formulas,
        "missing_variables": missing_variables,
        "pending_pathologies": pending_pathologies,
        "unsupported_pathologies": unsupported_pathologies,
        "owner_safe_question_candidates": owner_safe_question_candidates,
        "suggested_operator_next_step": suggested_operator_next_step,
    }


def inspect_excel(path: Path) -> dict:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        headers = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        normalized_headers = [str(value).strip().lower() for value in headers if value]
        return {
            "sheet": worksheet.title,
            "rows": int(worksheet.max_row or 0),
            "columns": int(worksheet.max_column or 0),
            "headers": normalized_headers,
        }
    finally:
        workbook.close()


def has_operational_columns(headers: list[str]) -> bool:
    joined = " ".join(headers)
    return any(term in joined for term in load_operational_terms())


def calculate_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while chunk := handle.read(1024 * 1024):
            digest.update(chunk)
    return digest.hexdigest()


def _write_jsonl_line(target: Path, payload: dict) -> Path:
    target.parent.mkdir(parents=True, exist_ok=True)
    import json

    with target.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(payload, sort_keys=True, ensure_ascii=False) + "\n")
    return target


def register_anamnesis_record(
    message: str,
    tenant_id: str,
    intake_id: str,
    storage_dir: Path,
    *,
    business_taxonomy: dict | None = None,
) -> dict:
    from pymia.smartpyme.anamnesis import create_anamnesis_record
    from pymia.smartpyme.storage import save_anamnesis_record

    record = create_anamnesis_record(
        tenant_id=tenant_id,
        intake_id=intake_id,
        raw_owner_message=message,
        business_taxonomy=business_taxonomy,
        metadata={"registered_by": "vertical_slice_cli"},
    )
    save_anamnesis_record(tenant_id, record, base_dir=storage_dir)
    return record.to_dict()


def register_investigation_record(message: str, tenant_id: str, intake_id: str, anamnesis_id: str, storage_dir: Path) -> dict:
    from pymia.smartpyme.investigation import create_investigation_record
    from pymia.smartpyme.storage import save_investigation_record

    record = create_investigation_record(
        tenant_id=tenant_id,
        intake_id=intake_id,
        anamnesis_id=anamnesis_id,
        owner_prompt=message,
        metadata={"registered_by": "vertical_slice_cli"},
    )
    save_investigation_record(tenant_id, record, base_dir=storage_dir)
    return record.to_dict()


def register_owner_answer_record(
    *,
    tenant_id: str,
    intake_id: str,
    anamnesis_id: str,
    investigation_id: str,
    question_ref: str,
    raw_owner_answer: str,
    storage_dir: Path,
) -> dict:
    from pymia.smartpyme.owner_answer import create_owner_answer_record
    from pymia.smartpyme.storage import save_owner_answer_record

    record = create_owner_answer_record(
        tenant_id=tenant_id,
        intake_id=intake_id,
        anamnesis_id=anamnesis_id,
        investigation_id=investigation_id,
        question_ref=question_ref,
        raw_owner_answer=raw_owner_answer,
        metadata={"registered_by": "vertical_slice_cli"},
    )
    save_owner_answer_record(tenant_id, record, base_dir=storage_dir)
    return record.to_dict()


def register_evidence_request_record(
    *,
    tenant_id: str,
    intake_id: str,
    anamnesis_id: str,
    investigation_id: str,
    owner_answer_id: str | None,
    requested_evidence: list[str],
    request_reason: str,
    storage_dir: Path,
) -> dict:
    from pymia.smartpyme.evidence_request import (
        EVIDENCE_REQUEST_STATUS_WAITING_UPLOAD,
        create_evidence_request_record,
    )
    from pymia.smartpyme.storage import save_evidence_request_record

    record = create_evidence_request_record(
        tenant_id=tenant_id,
        intake_id=intake_id,
        anamnesis_id=anamnesis_id,
        investigation_id=investigation_id,
        owner_answer_id=owner_answer_id,
        requested_evidence=requested_evidence,
        request_reason=request_reason,
        status=EVIDENCE_REQUEST_STATUS_WAITING_UPLOAD,
        metadata={"registered_by": "vertical_slice_cli"},
    )
    save_evidence_request_record(tenant_id, record, base_dir=storage_dir)
    return record.to_dict()


def register_evidence_record(path: Path, tenant_id: str, intake_id: str, storage_dir: Path, request_id: str | None = None) -> dict:
    from pymia.smartpyme.evidence import (
        EVIDENCE_STATUS_REGISTERED,
        SOURCE_KIND_UPLOADED_FILE,
        create_evidence_record,
    )
    from pymia.smartpyme.storage import save_evidence_record

    record = create_evidence_record(
        tenant_id=tenant_id,
        intake_id=intake_id,
        request_id=request_id,
        evidence_type="xlsx_upload",
        source_kind=SOURCE_KIND_UPLOADED_FILE,
        source_ref=str(path),
        original_filename=path.name,
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        size_bytes=path.stat().st_size,
        content_hash=calculate_sha256(path),
        status=EVIDENCE_STATUS_REGISTERED,
        metadata={"registered_by": "vertical_slice_cli"},
    )
    save_evidence_record(tenant_id, record, base_dir=storage_dir)
    return record.to_dict()


def register_pipeline_run_record(
    *,
    tenant_id: str,
    intake_id: str,
    message: str,
    anamnesis_record: dict,
    investigation_record: dict,
    owner_answer_record: dict | None,
    evidence_request_record: dict | None,
    evidence_record: dict,
    structured_summary: dict,
    blocked: bool,
    storage_dir: Path,
) -> dict:
    from pymia.contracts.pipeline_run_v1 import build_pipeline_run_record

    output_payload = {
        "tenant_id": tenant_id,
        "intake_id": intake_id,
        "anamnesis_id": anamnesis_record["anamnesis_id"],
        "investigation_id": investigation_record["investigation_id"],
        "evidence_id": evidence_record["evidence_id"],
        "structured_evidence_status": structured_summary["status"],
        "blocked": blocked,
    }
    if owner_answer_record:
        output_payload["owner_answer_id"] = owner_answer_record["answer_id"]
    if evidence_request_record:
        output_payload["evidence_request_id"] = evidence_request_record["request_id"]
    record = build_pipeline_run_record(
        tenant_id=tenant_id,
        intake_id=intake_id,
        message=message,
        evidence_ids=[evidence_record["evidence_id"]],
        status="BLOCKED" if blocked else "COMPLETED",
        output_payload=output_payload,
        steps_executed=[
            "evidence_record_registered",
            "structured_evidence_built",
            "evidence_sufficiency_checked",
            "owner_facing_report_built",
        ],
    )
    payload = record.model_dump(mode="json")
    payload["metadata"]["anamnesis_id"] = anamnesis_record["anamnesis_id"]
    payload["metadata"]["investigation_id"] = investigation_record["investigation_id"]
    if owner_answer_record:
        payload["metadata"]["owner_answer_id"] = owner_answer_record["answer_id"]
    if evidence_request_record:
        payload["metadata"]["evidence_request_id"] = evidence_request_record["request_id"]
    _write_jsonl_line(storage_dir / tenant_id / "pipeline_runs.jsonl", payload)
    return payload


def build_structured_summary(
    path: Path,
    tenant_id: str,
    *,
    intake_id: str = "intake_cli_local",
    formula_ids: list[str] | None = None,
) -> dict:
    try:
        from pymia.smartpyme.structured_evidence_builder import build_structured_evidence_context
        from pymia.contracts.evidence_v1 import StructuredEvidence
        from pymia.audit_result.evidence_requirement_matcher import match_evidence_requirements

        original_formula_ids = formula_ids
        formula_ids = formula_ids or []
        payload = build_structured_evidence_context(
            excel_path=path,
            tenant_id=tenant_id,
            intake_record={"evidence_requests": [{"formula_ids": formula_ids}]},
        )
        evidence = payload["structured_evidence"]
        computed = evidence.get("computed_variables") or {}
        tables = evidence.get("tables") or []
        summary = {
            "status": "available",
            "computed_variables_count": len(computed),
            "computed_variable_names": sorted(computed.keys()),
            "tables_count": len(tables),
            "table_sheets": [
                {"name": t.get("sheet_name", "?"), "columns": len(t.get("columns", [])), "rows": len(t.get("rows", []))}
                for t in tables
            ],
            "case_id": intake_id,
            "sufficiency": [],
            "unsupported_formula_ids": [],
            "catalog_reconciliation": [],
        }

        structured_ev = StructuredEvidence.model_validate(evidence)
        matches = match_evidence_requirements(structured_ev)
        reconciliation_list = []
        for m in matches:
            reconciliation_list.append({
                "formula_id": m.formula_id,
                "pathology_code": m.pathology_code,
                "status": m.status,
                "available_evidence": m.available_evidence,
                "missing_evidence": m.missing_evidence,
                "matched_sources": m.matched_sources,
                "required_evidence": m.required_evidence,
                "required_variables": m.required_variables,
                "next_audit_questions": m.next_audit_questions,
            })

        if original_formula_ids:
            reconciliation_list = [
                d for d in reconciliation_list if d["formula_id"] in original_formula_ids
            ]
        summary["catalog_reconciliation"] = reconciliation_list

        if formula_ids:
            from pymia.contracts.formula_rules_v1 import load_formula_rules
            from pymia.diagnostic_core.evidence_sufficiency import build_evidence_sufficiency_report_from_structured_evidence

            rules = load_formula_rules()
            rules_by_formula = rules.get("rules_by_formula", {})
            supported_formula_ids = [formula_id for formula_id in formula_ids if formula_id in rules_by_formula]
            unsupported_formula_ids = [formula_id for formula_id in formula_ids if formula_id not in rules_by_formula]
            summary["unsupported_formula_ids"] = unsupported_formula_ids
            if supported_formula_ids:
                sufficiency = build_evidence_sufficiency_report_from_structured_evidence(
                    structured_ev,
                    case_id=intake_id,
                    tenant_id=tenant_id,
                    formula_ids=supported_formula_ids,
                )
                summary["sufficiency"] = [item.model_dump(mode="json") for item in sufficiency]
        return summary
    except Exception as exc:
        raise exc


def build_report(
    path: Path,
    message: str,
    profile: dict,
    *,
    tenant_id: str = "tenant_cli_local",
    intake_id: str = "intake_cli_local",
    formula_ids: list[str] | None = None,
    storage_dir: Path | None = None,
    business_taxonomy: dict | None = None,
    owner_answer: str | None = None,
    owner_answer_question_ref: str | None = None,
) -> dict:
    has_rows = profile["rows"] > 1 and profile["columns"] > 0
    has_columns = has_operational_columns(profile["headers"])
    actual_storage_dir = storage_dir or Path(".tmp/vertical_slice_storage")
    anamnesis_record = register_anamnesis_record(
        message,
        tenant_id,
        intake_id,
        actual_storage_dir,
        business_taxonomy=business_taxonomy,
    )
    investigation_record = register_investigation_record(
        message,
        tenant_id,
        intake_id,
        anamnesis_record["anamnesis_id"],
        actual_storage_dir,
    )
    owner_answer_record = None
    if owner_answer:
        owner_answer_record = register_owner_answer_record(
            tenant_id=tenant_id,
            intake_id=intake_id,
            anamnesis_id=anamnesis_record["anamnesis_id"],
            investigation_id=investigation_record["investigation_id"],
            question_ref=owner_answer_question_ref or "owner_answer:unspecified_question_ref",
            raw_owner_answer=owner_answer,
            storage_dir=actual_storage_dir,
        )
    structured_summary = build_structured_summary(
        path,
        tenant_id,
        intake_id=intake_id,
        formula_ids=formula_ids,
    )
    missing = []
    questions = []
    if not has_rows:
        missing.append("filas_de_datos")
        questions.append(vertical_slice_copy_for("missing_data_rows_question"))
    if not has_columns:
        missing.append("columnas_operativas")
        questions.append(vertical_slice_copy_for("missing_operational_columns_question"))
    blocked = bool(missing)
    summary = (
        vertical_slice_copy_for("blocked_summary")
        if blocked
        else vertical_slice_copy_for("candidate_summary")
    )
    report = build_owner_facing_report(
        operational_audit_result={"tenant_id": tenant_id, "status": "blocked" if blocked else "candidate", "evidence_used": ["excel_file_readable"], "missing_evidence": missing},
        render_contract={
            "tenant_id": tenant_id,
            "summary": summary,
            "blocked_message": summary if blocked else "",
            "next_questions": questions,
            "next_steps": [vertical_slice_copy_for("next_step_review_with_owner")],
            "references": [str(path)],
            "forbidden_inferences": [vertical_slice_copy_for("forbidden_inference_from_column_names")],
        },
        delivery_package={"tenant_id": tenant_id, "intake_id": intake_id, "status": "BLOCKED" if blocked else "DELIVERED", "summary": summary, "output_refs": ["stdout"], "warnings": ["Slice local; no es canal productivo."]},
    ).to_dict()
    report["structured_evidence_summary"] = structured_summary
    requested_evidence = _requested_evidence_from_report(report)
    evidence_request_record = None
    if requested_evidence:
        evidence_request_record = register_evidence_request_record(
            tenant_id=tenant_id,
            intake_id=intake_id,
            anamnesis_id=anamnesis_record["anamnesis_id"],
            investigation_id=investigation_record["investigation_id"],
            owner_answer_id=owner_answer_record["answer_id"] if owner_answer_record else None,
            requested_evidence=requested_evidence,
            request_reason=vertical_slice_copy_for("evidence_request_reason"),
            storage_dir=actual_storage_dir,
        )
    evidence_record = register_evidence_record(
        path,
        tenant_id,
        intake_id,
        actual_storage_dir,
        request_id=evidence_request_record["request_id"] if evidence_request_record else None,
    )
    pipeline_run_record = register_pipeline_run_record(
        tenant_id=tenant_id,
        intake_id=intake_id,
        message=message,
        anamnesis_record=anamnesis_record,
        investigation_record=investigation_record,
        owner_answer_record=owner_answer_record,
        evidence_request_record=evidence_request_record,
        evidence_record=evidence_record,
        structured_summary=structured_summary,
        blocked=blocked,
        storage_dir=actual_storage_dir,
    )
    report["anamnesis_record"] = anamnesis_record
    report["investigation_record"] = investigation_record
    report["owner_answer_record"] = owner_answer_record
    report["evidence_request_record"] = evidence_request_record
    report["evidence_record"] = evidence_record
    report["pipeline_run_record"] = pipeline_run_record
    report["diagnostic_pipeline_result"] = _diagnostic_pipeline_result_for_report(
        path=path,
        tenant_id=tenant_id,
        intake_id=intake_id,
        cliente_id=tenant_id,
        structured_summary=structured_summary,
    )
    report["owner_simple"] = _build_owner_simple_view(report, message, profile)
    return report


def build_markdown(
    path: Path,
    message: str,
    profile: dict,
    *,
    tenant_id: str = "tenant_cli_local",
    intake_id: str = "intake_cli_local",
    formula_ids: list[str] | None = None,
    storage_dir: Path | None = None,
    business_taxonomy: dict | None = None,
    owner_answer: str | None = None,
    owner_answer_question_ref: str | None = None,
) -> str:
    report = build_report(
        path,
        message,
        profile,
        tenant_id=tenant_id,
        intake_id=intake_id,
        formula_ids=formula_ids,
        storage_dir=storage_dir,
        business_taxonomy=business_taxonomy,
        owner_answer=owner_answer,
        owner_answer_question_ref=owner_answer_question_ref,
    )
    return render_markdown_from_report(path, message, profile, report)


def render_markdown_from_report(path: Path, message: str, profile: dict, report: dict) -> str:
    owner_simple = report.get("owner_simple") or _build_owner_simple_view(report, message, profile)
    lines = [
        "# Reporte owner-facing local",
        f"Estado: {report['status']}",
        f"Archivo: {path.name}",
        f"Mensaje: {message}",
        f"Tenant: {report['tenant_id']}",
        f"Intake: {report['intake_id']}",
        f"Anamnesis ID: {report['anamnesis_record']['anamnesis_id']}",
        f"Investigation ID: {report['investigation_record']['investigation_id']}",
        f"Evidence ID: {report['evidence_record']['evidence_id']}",
        f"Evidence SHA-256: {report['evidence_record']['content_hash']}",
        f"Run ID: {report['pipeline_run_record']['run_id']}",
        f"Hoja: {profile['sheet']}",
        f"Filas: {profile['rows']}",
        f"Columnas: {profile['columns']}",
        f"Resumen: {report['summary']}",
        "",
        "Qué entendimos:",
        owner_simple["que_entendimos"],
        "",
        "Qué pudimos leer:",
        owner_simple["que_pudimos_leer"],
        "",
        "Qué todavía no podemos afirmar:",
        owner_simple["que_todavia_no_podemos_afirmar"],
        "",
        "Próxima pregunta:",
        owner_simple["proxima_pregunta"],
        "",
        "Límites:",
        "",
    ]
    for item in owner_simple["limites"]:
        lines.append(f"- {item}")
    lines.extend(["", "## Anamnesis"])
    owner_answer_record = report.get("owner_answer_record")
    evidence_request_record = report.get("evidence_request_record")
    if owner_answer_record:
        lines.insert(10, f"Owner Answer ID: {owner_answer_record['answer_id']}")
    if evidence_request_record:
        lines.insert(11, f"Evidence Request ID: {evidence_request_record['request_id']}")
    taxonomy = report["anamnesis_record"].get("business_taxonomy", {})
    lines.append(f"- Empresa tipo: {taxonomy.get('empresa_tipo', 'desconocido')}")
    lines.append(f"- Industria: {taxonomy.get('industria', 'desconocido')}")
    lines.append(f"- Modelo comercial: {taxonomy.get('modelo_comercial', 'desconocido')}")
    if taxonomy.get("canales_venta"):
        lines.append(f"- Canales de venta: {', '.join(taxonomy['canales_venta'])}")
    if taxonomy.get("areas_criticas"):
        lines.append(f"- Áreas críticas: {', '.join(taxonomy['areas_criticas'])}")
    if owner_answer_record:
        lines.extend([
            "",
            "## Respuesta del dueño",
            f"- Pregunta referida: {owner_answer_record['question_ref']}",
            f"- Tipo: {owner_answer_record['answer_kind']}",
        ])
    if evidence_request_record:
        lines.extend([
            "",
            "## Solicitud de evidencia",
            f"- Estado: {evidence_request_record['status']}",
            f"- Motivo: {evidence_request_record['request_reason']}",
        ])
        request_reconciliation = report.get("structured_evidence_summary", {}).get("catalog_reconciliation") or []
        request_question_candidates = [e for e in request_reconciliation if e.get("next_audit_questions")]
        request_alignment = (
            align_next_question(message, request_question_candidates)
            if request_question_candidates
            else {"status": "ALIGNED"}
        )
        if request_alignment["status"] == "MISALIGNED":
            lines.append("- Pendiente: reconducir con el dueño antes de solicitar evidencia.")
        else:
            for item in evidence_request_record["requested_evidence"]:
                lines.append(f"- {_humanize_field(item)}")
    lines.extend([
        "",
        "## Evidencia usada",
    ])
    for item in report["evidence_used"]:
        lines.append(f"- {item}")
    lines.append("")
    lines.append("## Evidencia faltante")
    if report["missing_evidence"]:
        for item in report["missing_evidence"]:
            lines.append(f"- {item}")
    else:
        lines.append("- Sin faltantes mínimos detectados en este slice.")
    lines.append("")
    lines.append("## Evidencia estructurada")
    structured_summary = report["structured_evidence_summary"]
    lines.append(f"- Estado: {structured_summary['status']}")
    if structured_summary["status"] == "available":
        var_count = structured_summary["computed_variables_count"]
        var_names = structured_summary.get("computed_variable_names", [])
        lines.append(f"- Variables computables: {var_count}")
        for name in var_names:
            lines.append(f"  - {_owner_label_for_variable(name)}")
        table_sheets = structured_summary.get("table_sheets", [])
        lines.append(f"- Tablas estructuradas: {structured_summary['tables_count']}")
        for t in table_sheets:
            lines.append(f"  - {t['name']} ({t['rows']} filas, {t['columns']} columnas)")
        lines.append(f"- Case ID: {structured_summary['case_id']}")
        if structured_summary["sufficiency"] or structured_summary["unsupported_formula_ids"]:
            lines.append("")
            lines.append("## Suficiencia de evidencia")
            for item in structured_summary["sufficiency"]:
                lines.append(f"- {item['formula_id']}: {item['status']}")
                if item["missing_variables"]:
                    lines.append(f"  - Faltan: {', '.join(item['missing_variables'])}")
            for formula_id in structured_summary["unsupported_formula_ids"]:
                lines.append(f"- {formula_id}: UNSUPPORTED_FORMULA")
    else:
        lines.append(f"- Motivo: {structured_summary['reason']}")
    lines.append("")
    lines.append("## Próxima pregunta")
    owner_question, tech_reference = _resolve_owner_question_and_reference(report, message)
    if owner_question:
        lines.append(f"- {owner_question}")
        if tech_reference:
            lines.append(f"  - {tech_reference}")
    else:
        lines.append(f"- {vertical_slice_copy_for('next_question_fallback')}")
    lines.append("")
    lines.append("## Límites")
    for warning in report["limit_warnings"]:
        lines.append(f"- {warning}")
    lines.append(f"- {vertical_slice_copy_for('final_limit_warning')}")
    return "\n".join(lines) + "\n"


def build_pipeline(
    path: Path,
    message: str,
    *,
    tenant_id: str = "tenant_cli_local",
    intake_id: str = "intake_cli_local",
    formula_ids: list[str] | None = None,
    storage_dir: Path | None = None,
    business_taxonomy: dict | None = None,
    owner_answer: str | None = None,
    owner_answer_question_ref: str | None = None,
) -> dict:
    profile = inspect_excel(path)
    report = build_report(
        path,
        message,
        profile,
        tenant_id=tenant_id,
        intake_id=intake_id,
        formula_ids=formula_ids,
        storage_dir=storage_dir,
        business_taxonomy=business_taxonomy,
        owner_answer=owner_answer,
        owner_answer_question_ref=owner_answer_question_ref,
    )
    markdown = render_markdown_from_report(path, message, profile, report)
    evidence_record = report["evidence_record"]
    pipeline_run_record = report["pipeline_run_record"]
    structured_summary = report.get("structured_evidence_summary") or {}
    catalog_reconciliation = structured_summary.get("catalog_reconciliation") or []
    diagnostic_operator_summary = _diagnostic_operator_summary_from_report(report)
    return {
        "status": report["status"],
        "profile": profile,
        "report": report,
        "markdown": markdown,
        "diagnostic_operator_summary": diagnostic_operator_summary,
        "evidence_id": evidence_record["evidence_id"],
        "evidence_hash": evidence_record["content_hash"],
        "run_id": pipeline_run_record["run_id"],
        "output_hash": pipeline_run_record["output_hash"],
        "missing_evidence": report["missing_evidence"],
        "next_questions": report["next_questions"],
        "structured_summary": report["structured_evidence_summary"],
        "catalog_reconciliation": catalog_reconciliation,
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", required=True)
    parser.add_argument("--message", required=True)
    parser.add_argument("--output")
    parser.add_argument("--tenant-id", default="tenant_cli_local")
    parser.add_argument("--intake-id", default="intake_cli_local")
    parser.add_argument("--formula-id", action="append", default=[])
    parser.add_argument("--storage-dir", default=".tmp/vertical_slice_storage")
    parser.add_argument("--empresa-tipo", default=None)
    parser.add_argument("--industria", default=None)
    parser.add_argument("--modelo-comercial", default=None)
    parser.add_argument("--canal-venta", action="append", default=[])
    parser.add_argument("--area-critica", action="append", default=[])
    parser.add_argument("--owner-answer", default=None)
    parser.add_argument("--owner-answer-question-ref", default=None)
    args = parser.parse_args(argv)
    path = Path(args.excel)
    if not path.exists():
        raise FileNotFoundError(path)
    business_taxonomy = {
        key: value
        for key, value in {
            "empresa_tipo": args.empresa_tipo,
            "industria": args.industria,
            "modelo_comercial": args.modelo_comercial,
            "canales_venta": args.canal_venta,
            "areas_criticas": args.area_critica,
        }.items()
        if value not in (None, [])
    }
    pipeline = build_pipeline(
        path,
        args.message,
        tenant_id=args.tenant_id,
        intake_id=args.intake_id,
        formula_ids=args.formula_id,
        storage_dir=Path(args.storage_dir),
        business_taxonomy=business_taxonomy or None,
        owner_answer=args.owner_answer,
        owner_answer_question_ref=args.owner_answer_question_ref,
    )
    markdown = pipeline["markdown"]
    if args.output:
        output_path = Path(args.output)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(markdown, encoding="utf-8")
    else:
        print(markdown)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

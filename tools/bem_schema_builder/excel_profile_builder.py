from __future__ import annotations

import math
import re
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from .models import ColumnProfile, ExcelProfile, SheetProfile


class ColumnSemanticClassifier:
    _LABEL_KEYWORDS: dict[str, tuple[str, ...]] = {
        "producto": ("producto", "articulo", "item", "descripcion"),
        "sku": ("sku", "codigo", "cod", "ean"),
        "cantidad": ("cantidad", "cant", "unidades", "qty"),
        "precio_venta": ("precio venta", "p venta", "pv", "precio_vta", "precio final"),
        "costo_unitario": ("costo unit", "costo", "coste unit", "c unit", "precio compra"),
        "venta_total": ("venta total", "ventas", "facturacion", "ingreso"),
        "costo_total": ("costo total", "costos", "egreso costo"),
        "fecha": ("fecha", "day", "date", "periodo"),
        "cliente": ("cliente", "razon social", "customer"),
        "proveedor": ("proveedor", "supplier"),
        "stock": ("stock", "inventario", "existencia"),
        "pago": ("pago", "cobro", "abono", "cuota"),
        "gasto": ("gasto", "egreso", "expense"),
        "saldo": ("saldo", "balance"),
        "impuesto": ("iva", "impuesto", "retencion", "percepcion"),
        "descuento": ("descuento", "bonificacion"),
        "moneda": ("moneda", "currency", "divisa", "usd", "ars"),
    }

    _AMBIGUOUS = {
        "importe",
        "monto",
        "total",
        "precio",
        "valor",
        "estado",
        "cantidad",
        "saldo",
        "diferencia",
        "cuenta",
        "concepto",
    }

    def classify(self, column_name: str) -> tuple[str, bool, str | None]:
        normalized = self._normalize(column_name)
        compact = normalized.replace("_", " ")
        for label, keywords in self._LABEL_KEYWORDS.items():
            if any(kw in compact for kw in keywords):
                is_ambiguous = any(token in compact for token in self._AMBIGUOUS)
                reason = "keyword_is_ambiguous" if is_ambiguous else None
                return label, is_ambiguous, reason

        is_ambiguous = any(token in compact for token in self._AMBIGUOUS)
        reason = "column_name_is_ambiguous" if is_ambiguous else None
        return "unknown", is_ambiguous, reason

    @staticmethod
    def _normalize(text: str) -> str:
        t = str(text or "").strip().lower()
        t = re.sub(r"\s+", " ", t)
        return t


class ExcelProfileBuilder:
    def __init__(self) -> None:
        self.classifier = ColumnSemanticClassifier()

    def build_profile(self, excel_path: str | Path) -> ExcelProfile:
        path = Path(excel_path)
        wb = load_workbook(path, data_only=False, read_only=False)
        sheets: list[SheetProfile] = []

        for ws in wb.worksheets:
            merged_ranges = [str(rng) for rng in ws.merged_cells.ranges]
            formula_cells_count = 0
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_cells_count += 1

            sheet_profile = SheetProfile(
                sheet_name=ws.title,
                max_row=ws.max_row or 0,
                max_column=ws.max_column or 0,
                merged_ranges=merged_ranges,
                formula_cells_count=formula_cells_count,
            )

            if ws.max_row and ws.max_column:
                df = pd.read_excel(path, sheet_name=ws.title, header=None, dtype=object)
                header_idx = self._detect_header_row(df)
                sheet_profile.probable_header_row = None if header_idx is None else header_idx + 1
                if header_idx is not None:
                    table_df = self._build_table(df, header_idx)
                    sheet_profile.columns, sheet_profile.empty_columns = self._profile_columns(table_df)
                    sheet_profile.first_useful_rows = self._sample_rows(table_df, limit=5)
                    sheet_profile.tabular_likelihood = self._tabular_likelihood(table_df)

            sheet_profile.sheet_kind = self._classify_sheet_kind(sheet_profile)
            sheets.append(sheet_profile)

        likely_tabular = [s.sheet_name for s in sheets if s.sheet_kind == "tabular"]
        likely_summary = [s.sheet_name for s in sheets if s.sheet_kind == "summary"]
        likely_auxiliary = [s.sheet_name for s in sheets if s.sheet_kind == "auxiliary"]

        return ExcelProfile(
            file_name=path.name,
            file_path=str(path.resolve()),
            sheets=sheets,
            likely_tabular_sheets=likely_tabular,
            likely_summary_sheets=likely_summary,
            likely_auxiliary_sheets=likely_auxiliary,
        )

    def _detect_header_row(self, df: pd.DataFrame) -> int | None:
        max_scan = min(len(df), 20)
        best_idx: int | None = None
        best_score = -math.inf
        for idx in range(max_scan):
            row = df.iloc[idx]
            non_null = row.notna().sum()
            if non_null == 0:
                continue
            text_like = sum(1 for v in row.tolist() if isinstance(v, str) and v.strip())
            unique_ratio = row.nunique(dropna=True) / max(non_null, 1)
            score = (text_like * 2.0) + unique_ratio - (idx * 0.1)
            if score > best_score and non_null >= 2:
                best_score = score
                best_idx = idx
        return best_idx

    def _build_table(self, df: pd.DataFrame, header_idx: int) -> pd.DataFrame:
        header_values = [self._safe_column_name(v, i + 1) for i, v in enumerate(df.iloc[header_idx].tolist())]
        body = df.iloc[header_idx + 1 :].copy()
        body.columns = header_values
        body = body.reset_index(drop=True)
        return body

    def _profile_columns(self, table_df: pd.DataFrame) -> tuple[list[ColumnProfile], list[str]]:
        columns: list[ColumnProfile] = []
        empty_columns: list[str] = []
        for idx, col_name in enumerate(table_df.columns):
            series = table_df[col_name]
            cleaned = series.dropna()
            inferred_type = self._infer_series_type(cleaned)
            null_pct = round(float(series.isna().mean() * 100), 2) if len(series) else 100.0
            unique_count = int(cleaned.astype(str).nunique()) if len(cleaned) else 0
            sample_values = cleaned.head(5).tolist()

            semantic_label, is_ambiguous, ambiguity_reason = self.classifier.classify(str(col_name))
            col_profile = ColumnProfile(
                name=str(col_name),
                index=idx,
                inferred_type=inferred_type,
                null_pct=null_pct,
                unique_count_sample=unique_count,
                sample_values=sample_values,
                semantic_label=semantic_label,
                is_ambiguous=is_ambiguous,
                ambiguity_reason=ambiguity_reason,
            )
            if inferred_type == "empty":
                empty_columns.append(str(col_name))
            columns.append(col_profile)

        return columns, empty_columns

    def _sample_rows(self, df: pd.DataFrame, limit: int = 5) -> list[dict[str, Any]]:
        useful = df.dropna(how="all").head(limit)
        records = useful.to_dict(orient="records")
        return records

    def _tabular_likelihood(self, df: pd.DataFrame) -> float:
        if df.empty:
            return 0.0
        density = 1.0 - float(df.isna().mean().mean())
        enough_columns = min(len(df.columns) / 8.0, 1.0)
        enough_rows = min(len(df) / 30.0, 1.0)
        return round(max(0.0, min((density * 0.5) + (enough_columns * 0.3) + (enough_rows * 0.2), 1.0)), 3)

    def _classify_sheet_kind(self, sheet: SheetProfile) -> str:
        name = sheet.sheet_name.lower()
        if any(token in name for token in ("resumen", "summary", "dashboard", "kpi")):
            return "summary"
        if sheet.tabular_likelihood >= 0.45 and sheet.columns:
            return "tabular"
        return "auxiliary"

    def _infer_series_type(self, cleaned: pd.Series) -> str:
        if cleaned.empty:
            return "empty"

        bool_like = cleaned.astype(str).str.lower().isin(["true", "false", "si", "no", "0", "1"]).mean()
        if bool_like >= 0.9:
            return "boolean"

        dt = pd.to_datetime(cleaned, errors="coerce", dayfirst=True)
        date_ratio = dt.notna().mean()
        if date_ratio >= 0.8:
            return "date"

        numeric = pd.to_numeric(cleaned, errors="coerce")
        num_ratio = numeric.notna().mean()
        if num_ratio >= 0.8:
            return "number"

        text_ratio = cleaned.map(lambda x: isinstance(x, str)).mean()
        if text_ratio >= 0.8:
            return "text"

        return "mixed"

    def _safe_column_name(self, value: Any, index: int) -> str:
        if value is None:
            return f"unnamed_{index}"
        text = str(value).strip()
        if not text:
            return f"unnamed_{index}"
        return text

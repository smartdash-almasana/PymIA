from __future__ import annotations

import argparse
import ast
import inspect
import json
import tempfile
from pathlib import Path
from typing import Any, Final

from openpyxl import Workbook

from pymia.smartpyme import service_1_product_pipeline_v1 as product_root
from pymia.smartpyme.service_1_capability_registry_v1 import (
    get_capability_definition_v1,
    list_capability_refs_v1,
)
from pymia.smartpyme.service_1_computability_v1 import (
    STATUS_COMPUTABLE,
    STATUS_NEEDS_EVIDENCE,
    build_service_1_composite_governed_computation_input_v1,
    build_service_1_computability_decision_v1,
)
from pymia.smartpyme.service_1_canonical_ingestion_output_to_semantic_bridge_v1 import (
    build_service_1_semantic_bridge_from_canonical_ingestion_output_v1,
)
from pymia.smartpyme.service_1_p6_approval_decision_v1 import (
    build_service_1_p6_approval_decisions_v1,
)
from pymia.smartpyme.service_1_variable_family_bindings_v1 import (
    build_service_1_requirement_matches_v1,
)
from pymia.smartpyme.service_1_deterministic_semantic_pipeline_v1 import (
    build_computability_decision_from_confirmed_bindings_v1,
    run_initial_pass,
    run_owner_reentry,
)
from pymia.smartpyme.service_1_generic_capability_engine_v1 import (
    STATUS_EVALUATED,
    execute_generic_capability_v1,
)
from pymia.smartpyme.service_1_owner_confirmation_to_canonical_ingestion_output_v1 import (
    build_service_1_canonical_ingestion_output_from_owner_confirmation_v1,
)
from pymia.smartpyme.service_1_web_column_confirmation_intake_boundary_v1 import (
    build_service_1_web_column_confirmation_intake_boundary_v1,
)
from tools.service_1_bounded_six_physical_computable_controls_v1 import (
    NEGATIVES as BOUNDED_NEGATIVES,
    POSITIVES as BOUNDED_POSITIVES,
    evaluate_service_1_bounded_six_physical_computable_controls_v1,
)
from tools.service_1_physical_computable_positive_controls_v1 import (
    CONTROLS as INITIAL_CONTROLS,
    evaluate_physical_computable_positive_controls_v1,
)

SCHEMA_VERSION: Final[str] = "CAPABILITY_PHYSICAL_COVERAGE_GATE_V1"
VERDICT_PASS: Final[str] = "PASS_CAPABILITY_PHYSICAL_COVERAGE_GATE_V1"
VERDICT_FAIL: Final[str] = "FAIL_CAPABILITY_PHYSICAL_COVERAGE_GATE_V1"

PHYSICAL_E2E_PASS: Final[str] = "PHYSICAL_E2E_PASS"
PHYSICAL_PARTIAL: Final[str] = "PHYSICAL_PARTIAL"
EXPECTED_NEEDS_EVIDENCE: Final[str] = "EXPECTED_NEEDS_EVIDENCE"
DEFERRED_BY_CONTRACT: Final[str] = "DEFERRED_BY_CONTRACT"
FAILED_UNSAFE: Final[str] = "FAILED_UNSAFE"
PASS: Final[str] = "PASS"
FAIL: Final[str] = "FAIL"

ALLOWED_COVERAGE_STATUSES: Final[frozenset[str]] = frozenset(
    {PHYSICAL_E2E_PASS, PHYSICAL_PARTIAL, EXPECTED_NEEDS_EVIDENCE, DEFERRED_BY_CONTRACT, FAILED_UNSAFE}
)
SPECIALIZED_REQUIRED_VARIABLES: Final[dict[str, tuple[str, ...]]] = {
    "sold_vs_collected_gap": ("sold_amount", "collected_amount"),
    "net_margin_real": ("sale_price", "costs", "taxes"),
}
FIXTURE_PYME_026_POSITIVE: Final[str] = "SERVICE_1_PHYSICAL_POSITIVE_PYME_026_ADJUSTED_OPERATING_CASH_FLOW.xlsx"
FIXTURE_PYME_026_NEGATIVE: Final[str] = "SERVICE_1_PHYSICAL_NEGATIVE_PYME_026_ADJUSTED_OPERATING_CASH_FLOW.xlsx"
FIXTURE_DPO_POSITIVE: Final[str] = "SERVICE_1_PHYSICAL_POSITIVE_DPO.xlsx"
FIXTURE_DPO_NEGATIVE: Final[str] = "SERVICE_1_PHYSICAL_NEGATIVE_DPO.xlsx"
SAFETY_FLAGS: Final[tuple[str, ...]] = (
    "runtime_authorized",
    "tool_execution_authorized",
    "delivery_authorized",
    "diagnosis_generated",
)


def write_required_fixtures_v1(root: Path | None = None) -> tuple[str, ...]:
    repo = root or Path(__file__).resolve().parents[1]
    target = repo / "prueba_excels"
    specs = (
        (FIXTURE_PYME_026_POSITIVE, "CashFlow", ("net_income", "depreciation", "amortization", "working_capital_change"), (100, 20, 10, 30)),
        (FIXTURE_PYME_026_NEGATIVE, "CashFlow", ("net_income", "depreciation", "amortization"), (100, 20, 10)),
        (FIXTURE_DPO_POSITIVE, "DPO", ("accounts_payable", "purchases", "days"), (3000, 9000, 30)),
        (FIXTURE_DPO_NEGATIVE, "DPO", ("accounts_payable", "days"), (3000, 30)),
    )
    written: list[str] = []
    for filename, sheet_name, headers, values in specs:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        sheet.append(list(headers))
        sheet.append(list(values))
        workbook.save(target / filename)
        written.append(filename)
    return tuple(written)


def _literal_assignment(path: Path, name: str) -> object:
    tree = ast.parse(path.read_text(encoding="utf-8"), filename=str(path))
    assignments: dict[str, ast.AST] = {}
    for node in tree.body:
        if not isinstance(node, ast.Assign):
            continue
        for target in node.targets:
            if isinstance(target, ast.Name):
                assignments[target.id] = node.value

    def evaluate(node: ast.AST, resolving: frozenset[str] = frozenset()) -> object:
        if isinstance(node, ast.Name):
            if node.id in resolving or node.id not in assignments:
                raise ValueError(f"unresolvable external closure authority: {node.id}")
            return evaluate(assignments[node.id], resolving | {node.id})
        if isinstance(node, ast.BinOp) and isinstance(node.op, ast.BitOr):
            left = evaluate(node.left, resolving)
            right = evaluate(node.right, resolving)
            if not isinstance(left, set) or not isinstance(right, set):
                raise ValueError("external closure union must combine sets")
            return left | right
        return ast.literal_eval(node)

    if name not in assignments:
        raise ValueError(f"missing external closure authority: {name}")
    return evaluate(assignments[name], frozenset({name}))


def _derive_inventory_authority_v1(repo: Path) -> dict[str, Any]:
    closure_path = repo / "tests" / "smartpyme" / "test_service_1_cycle_053_global_12_pathology_closure_v1.py"
    closure_pathologies = set(_literal_assignment(closure_path, "EXPECTED_PRODUCTIVE_PATHOLOGIES"))
    closure_generic_refs = set(_literal_assignment(closure_path, "EXPECTED_GENERIC_PRODUCTIVE_REFS"))
    closure_root_refs = set(_literal_assignment(closure_path, "EXPECTED_ROOT_REFS"))
    registry_refs = set(list_capability_refs_v1())
    dpo_definition = get_capability_definition_v1("dpo")
    generic_refs = {
        ref
        for ref in registry_refs
        if ref != "dpo"
        and (definition := get_capability_definition_v1(ref)) is not None
        and definition.pathology_code in closure_pathologies
    }
    specialized_refs = {
        product_root.LIQ_001_CAPABILITY_REF,
        product_root.REN_001_CAPABILITY_REF,
    }
    derived_root_refs = generic_refs | specialized_refs
    rows: list[tuple[str, str]] = []
    for ref in sorted(derived_root_refs):
        if ref == product_root.LIQ_001_CAPABILITY_REF:
            pathology = "LIQ_001"
        elif ref == product_root.REN_001_CAPABILITY_REF:
            pathology = "REN_001"
        else:
            definition = get_capability_definition_v1(ref)
            pathology = "" if definition is None else definition.pathology_code
        rows.append((ref, pathology))
    authority_match = (
        len(rows) == 12
        and len({ref for ref, _ in rows}) == 12
        and {ref for ref, _ in rows} == closure_root_refs
        and {pathology for _, pathology in rows} == closure_pathologies
        and dpo_definition is not None
        and dpo_definition.pathology_code == "PYME_013_PREREQUISITE_DPO"
        and "dpo" not in closure_root_refs
    )
    return {
        "rows": tuple(rows),
        "registry_refs": tuple(sorted(registry_refs)),
        "specialized_refs": tuple(sorted(specialized_refs)),
        "closure_root_refs": tuple(sorted(closure_root_refs)),
        "closure_pathologies": tuple(sorted(closure_pathologies)),
        "authority_match": authority_match,
    }


def _structural_guards_v1() -> dict[str, str]:
    module_source = Path(__file__).read_text(encoding="utf-8")
    tree = ast.parse(module_source)
    imported_modules = {
        alias.name
        for node in ast.walk(tree)
        if isinstance(node, (ast.Import, ast.ImportFrom))
        for alias in node.names
    }
    imported_from = {
        node.module or ""
        for node in ast.walk(tree)
        if isinstance(node, ast.ImportFrom)
    }
    chain_source = inspect.getsource(_physical_chain)
    ren_source = (
        inspect.getsource(_ren_001_control)
        + inspect.getsource(_ren_001_semantic_case)
        + inspect.getsource(_run_product_counted_v1)
    )
    product_source = inspect.getsource(product_root.run_service_1_product_pipeline_v1)
    ingestion_checks = {
        "boundary_module": build_service_1_web_column_confirmation_intake_boundary_v1.__module__
        == "pymia.smartpyme.service_1_web_column_confirmation_intake_boundary_v1",
        "connector_module": build_service_1_canonical_ingestion_output_from_owner_confirmation_v1.__module__
        == "pymia.smartpyme.service_1_owner_confirmation_to_canonical_ingestion_output_v1",
        "chain_calls_boundary": "build_service_1_web_column_confirmation_intake_boundary_v1" in chain_source,
        "chain_calls_connector": "build_service_1_canonical_ingestion_output_from_owner_confirmation_v1" in chain_source,
        "ren_calls_boundary": "build_service_1_web_column_confirmation_intake_boundary_v1" in ren_source,
        "no_second_xlsx_adapter_import": "pymia.smartpyme.service_1_xlsx_to_normalized_table_v1" not in imported_from,
        "no_load_workbook_import": "load_workbook" not in imported_modules,
        "no_openpyxl_module_import": "openpyxl" not in imported_modules,
    }
    product_checks = {
        "root_module": product_root.run_service_1_product_pipeline_v1.__module__
        == "pymia.smartpyme.service_1_product_pipeline_v1",
        "chain_calls_root": "run_service_1_product_pipeline_v1" in chain_source,
        "ren_calls_root": "run_service_1_product_pipeline_v1" in ren_source,
        "root_uses_generic_engine": "execute_generic_capability_v1" in product_source,
    }
    ingestion_ok = all(ingestion_checks.values())
    product_ok = all(product_checks.values())
    return {
        "canonical_ingestion_guard": PASS if ingestion_ok else FAIL,
        "canonical_product_root_guard": PASS if product_ok else FAIL,
        "structural_guard_evidence": {
            "ingestion": ingestion_checks,
            "product_root": product_checks,
        },
    }


def _owner_answers(boundary: dict[str, Any]) -> dict[str, str]:
    return {
        str(question["field_id"]): f"La columna {question['column_name']} representa {question['column_name']}"
        for question in boundary.get("owner_questions") or []
    }


def _semantic_owner_answers(semantic_run: dict[str, Any]) -> dict[str, str]:
    return {
        str(question["column_name"]): next(
            str(option["option_id"])
            for option in question.get("options") or []
            if option.get("option_id") not in {"OTHER", "IGNORE"}
        )
        for question in semantic_run.get("owner_questions") or []
    }


def _safety_snapshot(payload: dict[str, Any], *, governed_input_present: bool, executed: bool) -> dict[str, Any]:
    flags = {flag: payload.get(flag) for flag in SAFETY_FLAGS}
    missing_flags = [flag for flag in SAFETY_FLAGS if flag not in payload]
    executed_without_governed_input = executed and not governed_input_present
    unsafe = any(flag in payload and payload[flag] is not False for flag in SAFETY_FLAGS) or executed_without_governed_input
    complete = not missing_flags
    return {
        "flags": flags,
        "missing_flags": missing_flags,
        "complete": complete,
        "unsafe": unsafe,
        "executed_without_governed_input": executed_without_governed_input,
    }


def _physical_chain(repo: Path, filename: str, sheet_name: str, capability: str) -> dict[str, Any]:
    boundary = build_service_1_web_column_confirmation_intake_boundary_v1(
        local_xlsx_path=repo / "prueba_excels" / filename,
        sheet_name=sheet_name,
    )
    if boundary.get("status") != "NEEDS_OWNER_CONFIRMATION":
        return {"status": PHYSICAL_PARTIAL, "blocker": f"INTAKE:{boundary.get('blocked_reason')}"}
    connector = build_service_1_canonical_ingestion_output_from_owner_confirmation_v1(
        owner_question_packet=boundary,
        owner_answers=_owner_answers(boundary),
    )
    if connector.get("status") != "INGESTION_OUTPUT_READY":
        return {"status": PHYSICAL_PARTIAL, "blocker": f"CONNECTOR:{connector.get('blocked_reason')}"}
    ingestion = dict(connector["ingestion_output"])
    ingestion["normalized_tables"] = boundary.get("normalized_tables")
    semantic = run_initial_pass(ingestion_output=ingestion, sheet_name=sheet_name)
    owner_answers = _semantic_owner_answers(semantic)
    if semantic.get("status") == "OWNER_QUESTIONS":
        semantic = run_owner_reentry(
            previous_run=semantic,
            owner_answers=owner_answers,
        )
    if semantic.get("status") != "CONFIRMED_BINDINGS":
        return {
            "status": PHYSICAL_PARTIAL,
            "blocker": f"P6:{semantic.get('status')}",
            "p6": semantic.get("status"), "p7": None, "p8": None,
            "governed_input": False, "p9": None,
            "safety": {"unsafe": False, "executed_without_governed_input": False},
        }
    decision = build_computability_decision_from_confirmed_bindings_v1(
        confirmed_bindings=semantic,
        requested_capability=capability,
    )
    if decision.status != STATUS_COMPUTABLE or decision.governed_computation_input is None:
        return {
            "status": EXPECTED_NEEDS_EVIDENCE if decision.status == STATUS_NEEDS_EVIDENCE else PHYSICAL_PARTIAL,
            "blocker": decision.reason or decision.status,
            "p6": "APPROVED", "p7": "OBSERVED", "p8": decision.status,
            "governed_input": False, "p9": None,
            "safety": {"unsafe": False, "executed_without_governed_input": False},
        }
    product = product_root.run_service_1_product_pipeline_v1(
        ingestion_output=ingestion,
        tool_requests=(),
        output_dir=repo / ".tmp" / "capability_physical_coverage_gate" / capability,
        sheet_name=sheet_name,
        requested_capability=capability,
        deliver_result=False,
        owner_answers=owner_answers,
    )
    computation = product.get("computation_result") or {}
    executed = product.get("status") == product_root.STATUS_COMPUTATION_PLAN_READY and computation.get("status") == STATUS_EVALUATED
    safety = _safety_snapshot(product, governed_input_present=True, executed=executed)
    status = (
        FAILED_UNSAFE
        if safety["unsafe"]
        else PHYSICAL_E2E_PASS
        if executed and safety["complete"]
        else PHYSICAL_PARTIAL
    )
    return {
        "status": status,
        "blocker": (
            "UNSAFE_EXECUTION"
            if safety["unsafe"]
            else "MISSING_EXPLICIT_SAFETY_FLAGS"
            if executed and not safety["complete"]
            else None
            if executed
            else product.get("blocked_reason") or "P9_NOT_EVALUATED"
        ),
        "p6": "APPROVED", "p7": "REQUIREMENT_MATCHED", "p8": decision.status,
        "governed_input": True, "p9": computation.get("status"),
        "result": computation.get("computed"), "classification": computation.get("classification"),
        "safety": safety,
    }


def _ren_001_semantic_case(path: Path, *, include_taxes: bool) -> dict[str, Any]:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Resumen"
    headers = ["ventas_periodo", "cmv_total"]
    values = [100000, 60000]
    expected_roles = {
        "ventas_periodo": "period_sales_total" if include_taxes else "sales_amount",
        "cmv_total": "period_costs_total",
    }
    if include_taxes:
        headers.append("impuestos_periodo")
        values.append(10000)
        expected_roles["impuestos_periodo"] = "period_taxes_total"
    sheet.append(headers)
    sheet.append(values)
    workbook.save(path)
    boundary = build_service_1_web_column_confirmation_intake_boundary_v1(local_xlsx_path=path, sheet_name="Resumen")
    connector = build_service_1_canonical_ingestion_output_from_owner_confirmation_v1(
        owner_question_packet=boundary,
        owner_answers=_owner_answers(boundary),
    )
    ingestion = dict(connector["ingestion_output"])
    ingestion["normalized_tables"] = boundary.get("normalized_tables")
    semantic_first = run_initial_pass(ingestion_output=ingestion, sheet_name="Resumen")
    owner_answers = _semantic_owner_answers(semantic_first)
    bridge = build_service_1_semantic_bridge_from_canonical_ingestion_output_v1(
        ingestion_output=ingestion,
        sheet_name="Resumen",
    )
    candidates = tuple(bridge["column_candidates"])
    owner_events = tuple(
        {
            "question_ref": candidate.metadata["column_ref_id"],
            "sheet_ref": candidate.sheet_name,
            "column_ref": candidate.source_column_name,
            "confirmed_by_owner": True,
            "confirmation_scope": "SEMANTIC_ROLE",
            "confirmed_role": expected_roles[candidate.source_column_name],
        }
        for candidate in candidates
        if candidate.source_column_name in expected_roles
    )
    p6 = build_service_1_p6_approval_decisions_v1(
        case_id=bridge["case_id"],
        candidates=candidates,
        owner_confirmation_events=owner_events,
    )
    p7 = build_service_1_requirement_matches_v1(p6)
    decision = build_service_1_computability_decision_v1(
        case_id=str(bridge["case_id"]),
        requested_capability="net_margin_real",
        p6_decisions=[item.to_dict() for item in p6],
        requirement_matches=[item.to_dict() for item in p7],
    )
    semantic = {
        "status": "CONFIRMED_BINDINGS" if all(item.status == "APPROVED" for item in p6) else "OWNER_QUESTIONS",
        "p6": p6,
        "p7": p7,
    }
    return {
        "ingestion": ingestion,
        "semantic": semantic,
        "decision": decision,
        "owner_answers": owner_answers,
    }


def _run_product_counted_v1(counters: dict[str, int], **kwargs: Any) -> dict[str, Any]:
    counters["product_root_calls"] += 1
    product = product_root.run_service_1_product_pipeline_v1(**kwargs)
    computation = product.get("computation_result") or {}
    if computation.get("status") == STATUS_EVALUATED:
        counters["p9_calls"] += 1
    return product


def _ren_001_control(repo: Path) -> dict[str, Any]:
    positive_counters = {"product_root_calls": 0, "p9_calls": 0}
    negative_counters = {"product_root_calls": 0, "p9_calls": 0}
    with tempfile.TemporaryDirectory() as tmp:
        positive_case = _ren_001_semantic_case(Path(tmp) / "ren_001_positive.xlsx", include_taxes=True)
        negative_case = _ren_001_semantic_case(Path(tmp) / "ren_001_missing_taxes.xlsx", include_taxes=False)
        decision = positive_case["decision"]
        product = _run_product_counted_v1(
            positive_counters,
            ingestion_output=positive_case["ingestion"],
            tool_requests=(),
            output_dir=repo / ".tmp" / "capability_physical_coverage_gate" / "net_margin_real",
            sheet_name="Resumen",
            requested_capability="net_margin_real",
            deliver_result=False,
            owner_answers=positive_case["owner_answers"],
        )
    computation = product.get("computation_result") or {}
    executed = product.get("status") == product_root.STATUS_COMPUTATION_PLAN_READY and computation.get("status") == STATUS_EVALUATED
    governed = decision.status == STATUS_COMPUTABLE and decision.governed_computation_input is not None
    safety = _safety_snapshot(product, governed_input_present=governed, executed=executed)
    negative_decision = negative_case["decision"]
    negative = {
        "p8": negative_decision.status,
        "governed_input": negative_decision.governed_computation_input is not None,
        "p9": None,
        "execution_attempted": negative_counters["product_root_calls"] > 0,
        "product_root_calls": negative_counters["product_root_calls"],
        "p9_calls": negative_counters["p9_calls"],
        "unsafe": False,
    }
    negative_ok = (
        negative["p8"] == STATUS_NEEDS_EVIDENCE
        and negative["governed_input"] is False
        and negative["p9"] is None
        and negative["product_root_calls"] == 0
        and negative["p9_calls"] == 0
    )
    status = (
        FAILED_UNSAFE
        if safety["unsafe"]
        else PHYSICAL_E2E_PASS
        if executed and governed and negative_ok and safety["complete"]
        else PHYSICAL_PARTIAL
    )
    return {
        "status": status,
        "blocker": (
            "UNSAFE_EXECUTION"
            if safety["unsafe"]
            else "MISSING_EXPLICIT_SAFETY_FLAGS"
            if executed and not safety["complete"]
            else None
            if status == PHYSICAL_E2E_PASS
            else "REN_001_CONTROL_INCOMPLETE"
        ),
        "p6": "APPROVED" if positive_case["semantic"].get("status") == "CONFIRMED_BINDINGS" else "NOT_APPROVED",
        "p7": "REQUIREMENT_MATCHED" if governed else "NOT_MATCHED",
        "p8": decision.status,
        "governed_input": governed,
        "p9": computation.get("status"),
        "product_root_calls": positive_counters["product_root_calls"],
        "p9_calls": positive_counters["p9_calls"],
        "result": computation.get("computed"),
        "classification": computation.get("classification"),
        "negative": negative,
        "safety": safety,
    }


def _composite_pyme_013(dso_result: dict[str, Any], dpo_result: dict[str, Any]) -> dict[str, Any]:
    if dso_result.get("coverage_status") != PHYSICAL_E2E_PASS or dpo_result.get("coverage_status") != PHYSICAL_E2E_PASS:
        return {
            "status": DEFERRED_BY_CONTRACT, "blocker": "PHYSICAL_DSO_AND_DPO_RESULTS_REQUIRED",
            "p6": "NOT_APPLICABLE_COMPOSITE", "p7": "PREREQUISITES_NOT_READY", "p8": "DEFERRED",
            "governed_input": False, "p9": None,
            "safety": {"unsafe": False, "executed_without_governed_input": False},
        }
    governed = build_service_1_composite_governed_computation_input_v1(
        case_id="case_capability_physical_coverage_gate", capability_ref="payment_collection_gap"
    )
    sources = []
    for capability_ref, result_key, row in (("dso", "dso_days", dso_result), ("dpo", "dpo_days", dpo_result)):
        value = float((row.get("result") or {})[result_key])
        sources.append({
            "status": "EVALUATED", "capability_ref": capability_ref,
            "computed": {result_key: value, "typed_result": {"value": value, "unit": "days", "provenance": "governed_physical_gate"}},
            "outcome": {"causal_diagnosis_generated": False},
            "runtime_authorized": False, "tool_execution_authorized": False,
            "product_ready": False, "delivery_authorized": False, "diagnosis_generated": False,
        })
    computation = execute_generic_capability_v1(
        capability_ref="payment_collection_gap", computation_plan=None,
        governed_computation_input=governed, normalized_tables=None, column_refs=None, governed_results=sources,
    )
    executed = computation.get("status") == STATUS_EVALUATED
    safety = _safety_snapshot(computation, governed_input_present=True, executed=executed)
    status = (
        FAILED_UNSAFE
        if safety["unsafe"]
        else PHYSICAL_E2E_PASS
        if executed and safety["complete"]
        else DEFERRED_BY_CONTRACT
    )
    return {
        "status": status, "blocker": "UNSAFE_EXECUTION" if safety["unsafe"] else None if executed else computation.get("errors"),
        "p6": "NOT_APPLICABLE_COMPOSITE", "p7": "GOVERNED_PREREQUISITES_READY", "p8": "COMPUTABLE",
        "governed_input": True, "p9": computation.get("status"), "result": computation.get("computed"),
        "classification": computation.get("classification"), "safety": safety,
    }


def evaluate_service_1_capability_physical_coverage_gate_v1(root: Path | None = None) -> dict[str, Any]:
    repo = root or Path(__file__).resolve().parents[1]
    authority = _derive_inventory_authority_v1(repo)
    guards = _structural_guards_v1()
    authority_rows = list(authority["rows"])
    rows: dict[str, dict[str, Any]] = {}
    for capability, pathology in authority_rows:
        definition = get_capability_definition_v1(capability)
        variables = SPECIALIZED_REQUIRED_VARIABLES.get(capability) or tuple(variable.name for variable in definition.variables)  # type: ignore[union-attr]
        rows[capability] = {"capability": capability, "pathology": pathology, "required_variables": list(variables)}

    initial = evaluate_physical_computable_positive_controls_v1(root=repo)
    initial_rows = {row["capability"]: row for row in initial.get("rows") or []}
    for control in INITIAL_CONTROLS:
        capability = control.capability
        if capability not in rows:
            continue
        source = initial_rows.get(capability) or {}
        individual = _physical_chain(repo, control.filename, control.sheet_name, capability)
        baseline_ok = all(
            (
                source.get("p6_ok") is True,
                source.get("p7_ok") is True,
                source.get("p8_ok") is True,
                source.get("execution_ok") is True,
            )
        )
        if individual.get("status") == PHYSICAL_E2E_PASS and not baseline_ok:
            individual["status"] = PHYSICAL_PARTIAL
            individual["blocker"] = "INITIAL_PHYSICAL_CONTROL_FAILED"
        rows[capability].update({
            "positive_fixture": control.filename,
            "negative_fixture": None,
            **{("coverage_status" if key == "status" else key): value for key, value in individual.items()},
        })

    bounded = evaluate_service_1_bounded_six_physical_computable_controls_v1(root=repo)
    bounded_rows = {row["capability"]: row for row in bounded.get("positive_rows") or []}
    bounded_negatives = {row["capability"]: row for row in bounded.get("negative_rows") or []}
    negative_sheets = {capability: sheet for sheet, capability in BOUNDED_NEGATIVES}
    for spec in BOUNDED_POSITIVES:
        capability = spec.capability
        if capability not in rows:
            continue
        source = bounded_rows.get(capability) or {}
        negative_source = bounded_negatives.get(capability) or {}
        individual = _physical_chain(
            repo,
            "SERVICE_1_BOUNDED_SIX_PHYSICAL_CONTROLS.xlsx",
            spec.sheet,
            capability,
        )
        negative = _physical_chain(
            repo,
            "SERVICE_1_BOUNDED_SIX_PHYSICAL_CONTROLS.xlsx",
            negative_sheets[capability],
            capability,
        )
        baseline_ok = bool(source.get("ok")) and bool(negative_source.get("ok"))
        negative_ok = (
            negative.get("status") == EXPECTED_NEEDS_EVIDENCE
            and negative.get("governed_input") is False
            and negative.get("p9") is None
            and negative.get("safety", {}).get("unsafe") is False
        )
        if individual.get("status") == PHYSICAL_E2E_PASS and not (baseline_ok and negative_ok):
            individual["status"] = PHYSICAL_PARTIAL
            individual["blocker"] = "BOUNDED_PHYSICAL_CONTROL_FAILED"
        rows[capability].update({
            "positive_fixture": "SERVICE_1_BOUNDED_SIX_PHYSICAL_CONTROLS.xlsx",
            "negative_fixture": "SERVICE_1_BOUNDED_SIX_PHYSICAL_CONTROLS.xlsx",
            **{("coverage_status" if key == "status" else key): value for key, value in individual.items()},
            "negative_p8": negative.get("p8"),
            "negative_status": negative.get("status"),
        })

    ren = _ren_001_control(repo)
    rows["net_margin_real"].update({
        "positive_fixture": "generated physical XLSX", "negative_fixture": "generated missing-taxes XLSX",
        **{("coverage_status" if key == "status" else key): value for key, value in ren.items()},
    })

    pyme026_positive = _physical_chain(repo, FIXTURE_PYME_026_POSITIVE, "CashFlow", "adjusted_operating_cash_flow")
    pyme026_negative = _physical_chain(repo, FIXTURE_PYME_026_NEGATIVE, "CashFlow", "adjusted_operating_cash_flow")
    rows["adjusted_operating_cash_flow"].update({
        "positive_fixture": FIXTURE_PYME_026_POSITIVE, "negative_fixture": FIXTURE_PYME_026_NEGATIVE,
        **{("coverage_status" if key == "status" else key): value for key, value in pyme026_positive.items()},
        "negative_p8": pyme026_negative.get("p8"), "negative_status": pyme026_negative.get("status"),
    })

    dpo_positive = _physical_chain(repo, FIXTURE_DPO_POSITIVE, "DPO", "dpo")
    dpo_negative = _physical_chain(repo, FIXTURE_DPO_NEGATIVE, "DPO", "dpo")
    dpo = {
        "capability": "dpo", "pathology": "PYME_013_PREREQUISITE_DPO", "prerequisite_only": True,
        "positive_fixture": FIXTURE_DPO_POSITIVE, "negative_fixture": FIXTURE_DPO_NEGATIVE,
        **{("coverage_status" if key == "status" else key): value for key, value in dpo_positive.items()},
        "negative_p8": dpo_negative.get("p8"), "negative_status": dpo_negative.get("status"),
    }
    composite = _composite_pyme_013(rows["dso"], dpo)
    rows["payment_collection_gap"].update({
        "positive_fixture": f"{rows['dso'].get('positive_fixture')} + {FIXTURE_DPO_POSITIVE}",
        "negative_fixture": FIXTURE_DPO_NEGATIVE,
        **{("coverage_status" if key == "status" else key): value for key, value in composite.items()},
    })

    ordered_rows = [rows[capability] for capability, _ in authority_rows if capability in rows]
    failures: list[str] = []
    if not authority["authority_match"]:
        failures.append("INVENTORY_AUTHORITY_DIVERGENCE")
    if len(ordered_rows) != 12 or len({row["capability"] for row in ordered_rows}) != 12:
        failures.append("PRODUCTIVE_CAPABILITY_COUNT_MUST_BE_EXACTLY_12")
    if any(row["capability"] == "dpo" for row in ordered_rows):
        failures.append("DPO_MUST_NOT_BE_A_THIRTEENTH_CAPABILITY")
    if any(row.get("coverage_status") not in ALLOWED_COVERAGE_STATUSES for row in ordered_rows):
        failures.append("INVALID_OR_MISSING_COVERAGE_STATUS")
    unsafe_executions = sum(row.get("coverage_status") == FAILED_UNSAFE for row in ordered_rows)
    if unsafe_executions:
        failures.append("UNSAFE_EXECUTIONS_MUST_BE_ZERO")
    if guards["canonical_ingestion_guard"] != PASS:
        failures.append("CANONICAL_INGESTION_GUARD_FAILED")
    if guards["canonical_product_root_guard"] != PASS:
        failures.append("CANONICAL_PRODUCT_ROOT_GUARD_FAILED")

    pass_conditions = (
        len(ordered_rows) == 12
        and unsafe_executions == 0
        and not any(row.get("coverage_status") == FAILED_UNSAFE for row in ordered_rows)
        and authority["authority_match"] is True
        and guards["canonical_ingestion_guard"] == PASS
        and guards["canonical_product_root_guard"] == PASS
        and not failures
    )
    return {
        "schema_version": SCHEMA_VERSION,
        "verdict": VERDICT_PASS if pass_conditions else VERDICT_FAIL,
        "productive_capability_count": len(ordered_rows),
        "productive_capabilities": ordered_rows,
        "dpo_prerequisite": dpo,
        "dpo_counted_as_productive_capability": any(row["capability"] == "dpo" for row in ordered_rows),
        "inventory_authority_match": authority["authority_match"],
        "inventory_authority": authority,
        **guards,
        "physical_e2e_pass_count": sum(row.get("coverage_status") == PHYSICAL_E2E_PASS for row in ordered_rows),
        "physical_partial_count": sum(row.get("coverage_status") == PHYSICAL_PARTIAL for row in ordered_rows),
        "expected_needs_evidence_count": sum(row.get("coverage_status") == EXPECTED_NEEDS_EVIDENCE for row in ordered_rows),
        "deferred_by_contract_count": sum(row.get("coverage_status") == DEFERRED_BY_CONTRACT for row in ordered_rows),
        "failed_unsafe_count": sum(row.get("coverage_status") == FAILED_UNSAFE for row in ordered_rows),
        "unsafe_executions": unsafe_executions,
        "failures": failures,
        "runtime_authorized": False, "delivery_authorized": False, "product_ready": False,
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--write-fixtures", action="store_true")
    args = parser.parse_args()
    if args.write_fixtures:
        print(json.dumps({"written": write_required_fixtures_v1()}, ensure_ascii=False, indent=2))
        return 0
    result = evaluate_service_1_capability_physical_coverage_gate_v1()
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0 if result["verdict"] == VERDICT_PASS else 2


if __name__ == "__main__":
    raise SystemExit(main())

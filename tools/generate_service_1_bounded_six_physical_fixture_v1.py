"""Deterministic generator for the bounded-six physical controls fixture.

Reproduces prueba_excels/SERVICE_1_BOUNDED_SIX_PHYSICAL_CONTROLS.xlsx
with exactly the twelve sheets consumed by
tools/service_1_bounded_six_physical_computable_controls_v1.py.

No product logic is imported or executed; column names ARE the semantic
variable names, matching the owner-answer convention used by the tool
("La columna X representa X").
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

FIXTURE_RELATIVE = Path("prueba_excels") / "SERVICE_1_BOUNDED_SIX_PHYSICAL_CONTROLS.xlsx"

POSITIVE_SHEETS: tuple[dict, ...] = (
    {
        "sheet": "POS_REORDER",
        "columns": ("average_sales", "lead_time", "safety_stock"),
        "row": (10, 5, 20),
    },
    {
        "sheet": "POS_INV_TURN",
        "columns": ("cost_of_goods_sold", "average_stock"),
        "row": (12000, 3000),
    },
    {
        "sheet": "POS_CURRENT_RATIO",
        "columns": ("current_assets", "current_liabilities"),
        "row": (15000, 10000),
    },
    {
        "sheet": "POS_CONCENTRATION",
        "columns": ("main_sku_sales", "total_sales"),
        "row": (4000, 10000),
    },
    {
        "sheet": "POS_INTEREST",
        "columns": ("interest_expense", "ebitda"),
        "row": (1000, 5000),
    },
    {
        "sheet": "POS_INDEX",
        "columns": ("closing_index", "origin_index"),
        "row": (150, 100),
    },
)

# Each negative omits exactly one required variable of its positive pair.
NEGATIVE_SHEETS: tuple[dict, ...] = (
    {"sheet": "NEG_REORDER", "columns": ("average_sales", "lead_time"), "row": (10, 5)},
    {"sheet": "NEG_INV_TURN", "columns": ("cost_of_goods_sold",), "row": (12000,)},
    {"sheet": "NEG_CURRENT_RATIO", "columns": ("current_assets",), "row": (15000,)},
    {"sheet": "NEG_CONCENTRATION", "columns": ("main_sku_sales",), "row": (4000,)},
    {"sheet": "NEG_INTEREST", "columns": ("interest_expense",), "row": (1000,)},
    {"sheet": "NEG_INDEX", "columns": ("closing_index",), "row": (150,)},
)

EXPECTED_SHEETS: tuple[str, ...] = tuple(
    spec["sheet"] for spec in POSITIVE_SHEETS
) + tuple(spec["sheet"] for spec in NEGATIVE_SHEETS)


def generate_fixture(output_path: Path) -> Path:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for spec in POSITIVE_SHEETS + NEGATIVE_SHEETS:
        worksheet = workbook.create_sheet(title=spec["sheet"])
        worksheet.append(list(spec["columns"]))
        worksheet.append(list(spec["row"]))
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def _verify(path: Path) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        actual = tuple(workbook.sheetnames)
        if actual != EXPECTED_SHEETS:
            raise SystemExit(f"SHEET MISMATCH: {actual}")
    finally:
        workbook.close()
    print(f"FIXTURE_OK {path} sheets={len(actual)}")


def main() -> int:
    repo_root = Path(__file__).resolve().parents[1]
    target = repo_root / FIXTURE_RELATIVE
    generate_fixture(target)
    _verify(target)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
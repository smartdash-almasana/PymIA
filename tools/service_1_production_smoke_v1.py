from __future__ import annotations

import json
import os
import re
import sys
from io import BytesIO
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, build_opener

from openpyxl import Workbook


BASE_URL_ENV = "PYMIA_PRODUCTION_BASE_URL"
SUPABASE_URL_ENV = "PYMIA_SUPABASE_URL"
SUPABASE_PUBLISHABLE_KEY_ENV = "PYMIA_SUPABASE_PUBLISHABLE_KEY"
SMOKE_EMAIL_ENV = "PYMIA_SMOKE_EMAIL"
SMOKE_PASSWORD_ENV = "PYMIA_SMOKE_PASSWORD"


class SmokeFailure(RuntimeError):
    pass


def _required_env(name: str) -> str:
    value = str(os.environ.get(name) or "").strip()
    if not value:
        raise SmokeFailure(f"missing required environment variable: {name}")
    return value


def _xlsx_bytes() -> bytes:
    stream = BytesIO()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Ventas"
    sheet.append(["fecha", "venta_total", "cobrado"])
    sheet.append(["2026-08-01", 1000, 800])
    sheet.append(["2026-08-02", 2000, 1500])
    workbook.save(stream)
    return stream.getvalue()


def _ren_001_xlsx_bytes(*, include_taxes: bool) -> bytes:
    stream = BytesIO()
    workbook = Workbook()
    ventas = workbook.active
    ventas.title = "Ventas"
    ventas.append([
        "VentaID", "Fecha", "ProductoID", "Cantidad", "PrecioUnitario", "Descuento"
    ])
    ventas.append(["V0001", "2026-01-01", "P008", 1, 60, 0])
    ventas.append(["V0002", "2026-01-01", "P008", 2, 60, 0.10])

    productos = workbook.create_sheet("Productos")
    productos.append(["ProductoID", "Producto", "Costo"])
    productos.append(["P008", "Brownie", 28])

    if include_taxes:
        resumen = workbook.create_sheet("Resumen")
        resumen.append(["impuestos_periodo"])
        resumen.append([20])

    workbook.save(stream)
    return stream.getvalue()


def _request(opener, method: str, url: str, *, body: bytes = b"", headers: dict[str, str] | None = None):
    request = Request(url, data=body if method != "GET" else None, method=method, headers=headers or {})
    try:
        response = opener.open(request, timeout=30)
        return response.status, response.headers, response.read()
    except HTTPError as exc:
        return exc.code, exc.headers, exc.read()
    except URLError as exc:
        raise SmokeFailure(f"network failure for {url}: {exc.reason}") from None


def _supabase_access_token(opener, *, url: str, publishable_key: str, email: str, password: str) -> str:
    endpoint = url.rstrip("/") + "/auth/v1/token?grant_type=password"
    payload = json.dumps({"email": email, "password": password}).encode("utf-8")
    status, _, content = _request(
        opener,
        "POST",
        endpoint,
        body=payload,
        headers={
            "apikey": publishable_key,
            "Authorization": f"Bearer {publishable_key}",
            "Content-Type": "application/json",
        },
    )
    if status != 200:
        raise SmokeFailure(f"Supabase login failed with HTTP {status}")
    try:
        token = str(json.loads(content.decode("utf-8"))["access_token"]).strip()
    except Exception:
        raise SmokeFailure("Supabase login response did not contain access_token") from None
    if not token:
        raise SmokeFailure("Supabase returned an empty access_token")
    return token


def _multipart(filename: str, content: bytes, *, launch_review: str) -> tuple[bytes, str]:
    boundary = "PymIAProductionSmokeBoundary"
    body = (
        f"--{boundary}\r\n"
        'Content-Disposition: form-data; name="launch_review"\r\n\r\n'
        f"{launch_review}\r\n"
        f"--{boundary}\r\n"
        f'Content-Disposition: form-data; name="file"; filename="{filename}"\r\n'
        "Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n\r\n"
    ).encode("utf-8") + content + f"\r\n--{boundary}--\r\n".encode("utf-8")
    return body, f"multipart/form-data; boundary={boundary}"


def _cookie(headers) -> str:
    raw = headers.get("Set-Cookie")
    if not raw:
        raise SmokeFailure("application did not issue service session cookie")
    return raw.split(";", 1)[0]


def _answers(page: str) -> dict[str, str]:
    answers: dict[str, str] = {}
    for question_id, option_id in re.findall(r'name="answer_([^"]+)" value="([^"]+)"', page):
        if option_id not in {"OTHER", "IGNORE", "not_sure"}:
            answers.setdefault(f"answer_{question_id}", option_id)
    if answers:
        return answers
    for decision_id in re.findall(r'name="action_([^"]+)" value="ACCEPT"', page):
        answers.setdefault(f"action_{decision_id}", "ACCEPT")
    if not answers:
        raise SmokeFailure("owner confirmation page exposed no acceptable semantic answers")
    return answers


def _unit_answers(page: str) -> dict[str, str]:
    question_ids = re.findall(
        r'name="unit_([^"]+)" value="DISCOUNT_FRACTION_0_1"',
        page,
    )
    if not question_ids:
        raise SmokeFailure("REN_001 unit confirmation page exposed no discount fraction option")
    return {f"unit_{question_id}": "DISCOUNT_FRACTION_0_1" for question_id in question_ids}


def _assert(condition: bool, message: str) -> None:
    if not condition:
        raise SmokeFailure(message)


def _durable_case_link(cases_page: str) -> str:
    case_links = re.findall(r'href="(/case\?case_ref=case_[^"]+)"', cases_page)
    durable_links = [link for link in case_links if "::" not in link]
    if not durable_links:
        raise SmokeFailure("persisted cases exposed no durable case_* reentry link")
    return durable_links[0]


def _post_form(opener, *, url: str, fields: dict[str, str], cookie: str, auth: dict[str, str]):
    body = urlencode(fields).encode("utf-8")
    return _request(
        opener,
        "POST",
        url,
        body=body,
        headers={
            "Content-Type": "application/x-www-form-urlencoded",
            "Cookie": cookie,
            **auth,
        },
    )


def _run_ren_001_journey(
    opener,
    *,
    base_url: str,
    auth: dict[str, str],
    include_taxes: bool,
) -> tuple[str, str]:
    xlsx = _ren_001_xlsx_bytes(include_taxes=include_taxes)
    upload_body, content_type = _multipart(
        "production_smoke_ren_001.xlsx",
        xlsx,
        launch_review="net_margin_real",
    )
    status, response_headers, content = _request(
        opener,
        "POST",
        base_url + "/upload",
        body=upload_body,
        headers={"Content-Type": content_type, **auth},
    )
    page = content.decode("utf-8", errors="replace")
    _assert(status == 200, f"REN_001 authenticated upload failed: HTTP {status}")
    _assert(
        "Confirmá la interpretación material" in page and "SEM-8 · Confirmación empresarial" in page,
        "REN_001 upload did not reach SEM-8 owner confirmation",
    )
    cookie = _cookie(response_headers)

    status, _, content = _post_form(
        opener,
        url=base_url + "/confirm-meanings",
        fields=_answers(page),
        cookie=cookie,
        auth=auth,
    )
    page = content.decode("utf-8", errors="replace")
    _assert(status == 200, f"REN_001 semantic confirmation failed: HTTP {status}")
    _assert(
        "Confirmá cómo está expresado el descuento" in page,
        "REN_001 did not request governed discount-unit evidence",
    )

    status, _, content = _post_form(
        opener,
        url=base_url + "/confirm-meanings",
        fields=_unit_answers(page),
        cookie=cookie,
        auth=auth,
    )
    return cookie, content.decode("utf-8", errors="replace")


def run() -> dict[str, object]:
    base_url = _required_env(BASE_URL_ENV).rstrip("/")
    supabase_url = _required_env(SUPABASE_URL_ENV)
    publishable_key = _required_env(SUPABASE_PUBLISHABLE_KEY_ENV)
    email = _required_env(SMOKE_EMAIL_ENV)
    password = _required_env(SMOKE_PASSWORD_ENV)
    _assert(base_url.startswith("https://"), "production base URL must use HTTPS")

    opener = build_opener()
    checks: dict[str, str] = {}

    status, _, body = _request(opener, "GET", base_url + "/")
    _assert(status == 200, f"root health failed: HTTP {status}")
    _assert(b"Revisar informaci" in body or b"html" in body.lower(), "root health did not return the application page")
    checks["health_root"] = "PASS"

    xlsx = _xlsx_bytes()
    upload_body, content_type = _multipart("production_smoke_ventas.xlsx", xlsx, launch_review="sold_vs_collected_gap")
    status, _, _ = _request(
        opener,
        "POST",
        base_url + "/upload",
        body=upload_body,
        headers={"Content-Type": content_type},
    )
    _assert(status == 400, f"unauthenticated upload must fail closed, got HTTP {status}")
    checks["unauthenticated_fail_closed"] = "PASS"

    token = _supabase_access_token(
        opener,
        url=supabase_url,
        publishable_key=publishable_key,
        email=email,
        password=password,
    )
    auth = {"Authorization": f"Bearer {token}"}
    checks["supabase_login"] = "PASS"

    status, response_headers, content = _request(
        opener,
        "POST",
        base_url + "/upload",
        body=upload_body,
        headers={"Content-Type": content_type, **auth},
    )
    page = content.decode("utf-8", errors="replace")
    _assert(status == 200, f"authenticated upload failed: HTTP {status}")
    _assert(
        "Confirmá la interpretación material" in page and "SEM-8 · Confirmación empresarial" in page,
        "upload did not reach the SEM-8 owner confirmation flow",
    )
    cookie = _cookie(response_headers)
    checks["authenticated_upload"] = "PASS"

    confirm_body = urlencode(_answers(page)).encode("utf-8")
    status, _, content = _request(
        opener,
        "POST",
        base_url + "/confirm-meanings",
        body=confirm_body,
        headers={
            "Content-Type": "application/x-www-form-urlencoded",
            "Cookie": cookie,
            **auth,
        },
    )
    page = content.decode("utf-8", errors="replace")
    _assert(status == 200, f"owner confirmation failed: HTTP {status}")
    checks["owner_confirmation"] = "PASS"

    if "Ventas y cobranzas" not in page:
        review_body = urlencode({"review": "sold_vs_collected_gap"}).encode("utf-8")
        status, _, content = _request(
            opener,
            "POST",
            base_url + "/run-review",
            body=review_body,
            headers={
                "Content-Type": "application/x-www-form-urlencoded",
                "Cookie": cookie,
                **auth,
            },
        )
        page = content.decode("utf-8", errors="replace")
        _assert(status == 200, f"deterministic execution failed: HTTP {status}")

    _assert("Ventas y cobranzas" in page, "sellable result was not rendered")
    _assert("Total vendido" in page and "Diferencia" in page, "sellable result is incomplete")
    checks["deterministic_execution_and_persistence"] = "PASS"

    status, headers, content = _request(
        opener,
        "GET",
        base_url + "/download-sales-collections",
        headers={"Cookie": cookie},
    )
    _assert(status == 200 and content.startswith(b"PK"), f"XLSX download failed: HTTP {status}")
    disposition = str(headers.get("Content-Disposition") or "")
    _assert("service_1_liq_001_result.xlsx" in disposition, "unexpected delivery filename")
    checks["delivery_download"] = "PASS"

    status, _, content = _request(
        opener,
        "GET",
        base_url + "/cases",
        headers=auth,
    )
    cases = content.decode("utf-8", errors="replace")
    _assert(status == 200 and "EVIDENCIA PERSISTIDA" in cases, "durable case reentry listing failed")
    durable_link = _durable_case_link(cases)
    status, _, content = _request(
        opener,
        "GET",
        base_url + durable_link,
        headers=auth,
    )
    reopened = content.decode("utf-8", errors="replace")
    _assert(
        status == 200
        and "Reingreso durable del caso" in reopened
        and "Evidencia confirmada por el dueño" in reopened,
        "persisted case reentry failed",
    )
    checks["reentry_persisted_case"] = "PASS"

    _, incomplete_margin_page = _run_ren_001_journey(
        opener,
        base_url=base_url,
        auth=auth,
        include_taxes=False,
    )
    _assert(
        "FALTA INFORMACIÓN" in incomplete_margin_page
        and "/download-net-margin" not in incomplete_margin_page,
        "REN_001 must fail closed when taxes are missing",
    )
    checks["ren_001_missing_taxes_fail_closed"] = "PASS"

    margin_cookie, margin_page = _run_ren_001_journey(
        opener,
        base_url=base_url,
        auth=auth,
        include_taxes=True,
    )
    _assert(
        "Margen neto real" in margin_page
        and "/download-net-margin" in margin_page
        and "FALTA INFORMACIÓN" not in margin_page,
        "REN_001 complete derived-evidence journey did not produce a sellable result",
    )
    checks["ren_001_sem8_owner_flow"] = "PASS"
    checks["ren_001_discount_unit_confirmation"] = "PASS"
    checks["ren_001_derived_evidence_execution"] = "PASS"

    status, headers, content = _request(
        opener,
        "GET",
        base_url + "/download-net-margin",
        headers={"Cookie": margin_cookie},
    )
    _assert(status == 200 and content.startswith(b"PK"), f"REN_001 XLSX download failed: HTTP {status}")
    disposition = str(headers.get("Content-Disposition") or "")
    _assert("service_1_ren_001_result.xlsx" in disposition, "unexpected REN_001 delivery filename")
    checks["ren_001_delivery_download"] = "PASS"

    return {
        "verdict": "PASS",
        "base_url": base_url,
        "checks": checks,
        "non_claims": [
            "Durable reentry proves persisted owner evidence, not restoration of ephemeral XLSX result artifacts.",
            "Does not expose or print access tokens or passwords.",
        ],
    }


def main() -> int:
    try:
        result = run()
    except SmokeFailure as exc:
        print(json.dumps({"verdict": "BLOCKED", "reason": str(exc)}, ensure_ascii=False, indent=2))
        return 1
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
